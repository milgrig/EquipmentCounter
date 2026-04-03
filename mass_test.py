#!/usr/bin/env python3
"""mass_test.py - VOR generation mass-testing harness.

Runs VOR generation on multiple test cases, compares with reference VOR files,
and produces an Excel report with run history for tracking improvements.

Usage:
    python mass_test.py                          # run all cases, default report
    python mass_test.py --convert                # convert DWG->DXF first
    python mass_test.py --report results.xlsx    # custom report file
    python mass_test.py --only gpk_z3            # run single case
    python mass_test.py --only gpk_z3,test_3_12  # run multiple cases
"""

from __future__ import annotations

import argparse
import difflib
import math
import os
import re
import sys
import time
import traceback
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

# ---------------------------------------------------------------------------
# Resolve project root
# ---------------------------------------------------------------------------
_SCRIPT_DIR = Path(__file__).resolve().parent
_DATA = _SCRIPT_DIR / "Data"
_DBT = _DATA / "ДБТ разделы для ИИ"
_VORY = _DBT / "Ст.П" / "ВОРЫ"

# ---------------------------------------------------------------------------
# TEST_CASES - absolute Cyrillic paths
# ---------------------------------------------------------------------------
TEST_CASES: list[dict] = [
    {
        "name": "test_3_12",
        "dxf_dir": str(_DATA / "test" / "_converted_dxf"),
        "dwg_dir": None,  # already converted
        "ref_vor": str(_DATA / "test" / "ВОР ЭОМ_3.12.docx"),
    },
    {
        "name": "gpk_z3",
        "dwg_dir": str(_DBT / "03_ГПК_" / "3-я захватка" / "01_DWG"),
        "dxf_dir": str(_DBT / "03_ГПК_" / "3-я захватка" / "_converted_dxf" / "01_DWG"),
        "ref_vor": str(_DBT / "03_ГПК_" / "3-я захватка" / "ВОР ЭО, Захватка 3_ГПК.docx"),
    },
    {
        "name": "gpk_z4",
        "dwg_dir": str(_DBT / "03_ГПК_" / "4-я захватка"),
        "dxf_dir": str(_DBT / "03_ГПК_" / "4-я захватка" / "_converted_dxf"),
        "ref_vor": str(_DBT / "03_ГПК_" / "4-я захватка" / "ВОР ЭО, Захватка 4_ГПК.docx"),
    },
    {
        "name": "abk_eo",
        "dwg_dir": str(_DBT / "01_АБК" / "Обновленные файлы" / "DWG" / "ЭО"),
        "dxf_dir": str(_DBT / "01_АБК" / "Обновленные файлы" / "DWG" / "ЭО" / "_converted_dxf"),
        "ref_vor": str(_VORY / "1. Административно – бытовой корпус с КПП №1"
                       / "1Д-24-ИОС1.5.1 поз.1 ВОР.docx"),
    },
    {
        "name": "abk_em",
        "dwg_dir": str(_DBT / "01_АБК" / "Обновленные файлы" / "DWG" / "ЭМ"),
        "dxf_dir": str(_DBT / "01_АБК" / "Обновленные файлы" / "DWG" / "ЭМ" / "_converted_dxf"),
        "ref_vor": str(_VORY / "1. Административно – бытовой корпус с КПП №1"
                       / "1Д-24-ИОС1.5.1 поз.1 ВОР.docx"),
    },
    {
        "name": "abk_eg",
        "dwg_dir": str(_DBT / "01_АБК" / "Обновленные файлы" / "DWG" / "ЭГ"),
        "dxf_dir": str(_DBT / "01_АБК" / "Обновленные файлы" / "DWG" / "ЭГ" / "_converted_dxf"),
        "ref_vor": str(_VORY / "1. Административно – бытовой корпус с КПП №1"
                       / "1Д-24-ИОС1.5.1 поз.1 ВОР.docx"),
    },
    {
        "name": "zhd_16_2",
        "dwg_dir": str(_DBT / "16.2_ЖД КПП" / "Обновленные файлы 2" / "01_DWG"),
        "dxf_dir": str(_DBT / "16.2_ЖД КПП" / "Обновленные файлы 2" / "01_DWG" / "_converted_dxf"),
        "ref_vor": str(_DBT / "16.2_ЖД КПП" / "Обновленные файлы 2"
                       / "05-02-03 ЭОМ поз.16.2 ВОР.docx"),
    },
    {
        "name": "sklad_27",
        "dwg_dir": str(_DBT / "27_Склад вспомогательных материалов "
                       "с участком погрузки крытых вагонов"
                       / "Обновленные файлы" / "02_DWG"),
        "dxf_dir": str(_DBT / "27_Склад вспомогательных материалов "
                       "с участком погрузки крытых вагонов"
                       / "Обновленные файлы" / "02_DWG" / "_converted_dxf"),
        "ref_vor": str(_VORY / "27. Склад вспомогательных материалов "
                       "с участком погрузки крытых вагонов"
                       / "1Д-24-ИОС1.5.1 поз.27 ВОР.docx"),
    },
    {
        "name": "sulfat_8_2_em",
        "dwg_dir": str(_DBT / "8.2 Участок хранения сульфата аммония"
                       / "ЭМ" / "Обновленные файлы" / "01_DWG"),
        "dxf_dir": str(_DBT / "8.2 Участок хранения сульфата аммония"
                       / "ЭМ" / "Обновленные файлы" / "01_DWG" / "_converted_dxf"),
        "ref_vor": str(_DBT / "8.2 Участок хранения сульфата аммония"
                       / "ЭМ" / "Обновленные файлы"
                       / "ВОР ЭМ_8.2 рев.1.docx"),
    },
    {
        "name": "sulfat_8_2_eo",
        "dwg_dir": str(_DBT / "8.2 Участок хранения сульфата аммония"
                       / "ЭО" / "01_DWG"),
        "dxf_dir": str(_DBT / "8.2 Участок хранения сульфата аммония"
                       / "ЭО" / "01_DWG" / "_converted_dxf"),
        "ref_vor": str(_VORY / "8,8.1, 8.2 Участок хранения и  разгрузки сульфата аммония"
                       / "1Д-24-ИОС1.5.1 поз.8.2 ВОР.docx"),
    },
    {
        "name": "autovesy_28",
        "dwg_dir": str(_DBT / "28. Автовесы" / "01_DWG"),
        "dxf_dir": str(_DBT / "28. Автовесы" / "01_DWG" / "_converted_dxf"),
        "ref_vor": str(_VORY / "28. Автовесы"
                       / "1Д-24-ИОС1.5.1 поз.28 ВОР.docx"),
    },
    {
        "name": "kpp_30",
        "dwg_dir": str(_DBT / "30. КПП" / "02_DWG"),
        "dxf_dir": str(_DBT / "30. КПП" / "02_DWG" / "_converted_dxf"),
        "ref_vor": str(_VORY / "30. КПП"
                       / "1Д-24-ИОС1.5.1 поз.30 ВОР.docx"),
    },
    {
        "name": "nasosnaya_12_3",
        "dwg_dir": str(_DBT / "12.3 Насосная станция поверхностных стоков"
                       / "01_DWG"),
        "dxf_dir": str(_DBT / "12.3 Насосная станция поверхностных стоков"
                       / "01_DWG" / "_converted_dxf"),
        "ref_vor": str(_VORY / "12.3 Насосная станция поверхностных стоков"
                       / "1Д-24-ИОС1.5.1 поз.12.3 ВОР.docx"),
    },
]


# ---------------------------------------------------------------------------
# VOR .docx row extraction
# ---------------------------------------------------------------------------
_NUM_RE = re.compile(r"[\d]+(?:[.,]\d+)?")


@dataclass
class VorRow:
    """Single row from VOR table (work item or material)."""
    item_num: str  # row number or empty for material sub-rows
    description: str
    unit: str
    quantity: str
    section: str = ""  # section header this row belongs to

    @property
    def qty_float(self) -> float | None:
        """Parse quantity as float."""
        s = self.quantity.strip().replace(",", ".").replace("\xa0", "")
        if not s:
            return None
        m = _NUM_RE.search(s)
        return float(m.group()) if m else None

    @property
    def key(self) -> str:
        """Normalized description for fuzzy matching."""
        return re.sub(r"\s+", " ", self.description).strip().lower()


def extract_vor_rows(docx_path: str) -> list[VorRow]:
    """Extract rows with quantities from a VOR .docx file."""
    from docx import Document
    doc = Document(docx_path)
    rows: list[VorRow] = []
    current_section = ""
    for table in doc.tables:
        for row in table.rows:
            cells = [c.text.strip() for c in row.cells]
            if len(cells) < 4:
                continue
            # Detect section headers (merged cells, no quantity)
            text = cells[1] if len(cells) > 1 else cells[0]
            if len(cells) >= 3 and not cells[0] and not cells[2] and not cells[3]:
                if text and len(text) > 5:
                    current_section = text
                continue
            # Skip header row
            if cells[0] in ("", "№\nп/п", "№ п/п") and "Наименование" in text:
                continue
            # Extract data rows
            item_num = cells[0].strip()
            desc = cells[1].strip() if len(cells) > 1 else ""
            unit = cells[2].strip() if len(cells) > 2 else ""
            qty = cells[3].strip() if len(cells) > 3 else ""
            if not desc:
                continue
            # Only include rows that have some quantity or are numbered
            if qty or item_num:
                rows.append(VorRow(
                    item_num=item_num,
                    description=desc,
                    unit=unit,
                    quantity=qty,
                    section=current_section,
                ))
    return rows


# ---------------------------------------------------------------------------
# Fuzzy comparison
# ---------------------------------------------------------------------------
@dataclass
class MatchResult:
    """Result of matching a generated row to a reference row."""
    gen_desc: str
    gen_qty: str
    ref_desc: str
    ref_qty: str
    match_ratio: float
    qty_match: bool
    delta_pct: float | None  # (gen - ref) / ref * 100
    section: str = ""


@dataclass
class CaseResult:
    """Full comparison result for one test case."""
    name: str
    status: str  # "ok", "error", "skip"
    error_msg: str = ""
    dxf_count: int = 0
    gen_row_count: int = 0
    ref_row_count: int = 0
    matched: list[MatchResult] = field(default_factory=list)
    unmatched_gen: list[VorRow] = field(default_factory=list)
    unmatched_ref: list[VorRow] = field(default_factory=list)
    elapsed_sec: float = 0.0

    @property
    def match_pct(self) -> float:
        """Overall match percentage: matched rows with qty_match / total ref rows."""
        if self.ref_row_count == 0:
            return 0.0
        qty_ok = sum(1 for m in self.matched if m.qty_match)
        return qty_ok / self.ref_row_count * 100

    @property
    def fuzzy_match_pct(self) -> float:
        """Rows that fuzzy-matched regardless of qty."""
        if self.ref_row_count == 0:
            return 0.0
        return len(self.matched) / self.ref_row_count * 100

    def category_matches(self) -> dict[str, tuple[int, int]]:
        """Per-section: (qty_matched, total_ref)."""
        from collections import defaultdict
        by_sec: dict[str, list[bool]] = defaultdict(list)
        for m in self.matched:
            by_sec[m.section or "Без раздела"].append(m.qty_match)
        for r in self.unmatched_ref:
            by_sec[r.section or "Без раздела"].append(False)
        result = {}
        for sec, bools in by_sec.items():
            result[sec] = (sum(bools), len(bools))
        return result


def compare_rows(
    gen_rows: list[VorRow],
    ref_rows: list[VorRow],
    threshold: float = 0.6,
) -> tuple[list[MatchResult], list[VorRow], list[VorRow]]:
    """Fuzzy-match generated rows against reference rows."""
    matched: list[MatchResult] = []
    used_gen: set[int] = set()
    unmatched_ref: list[VorRow] = []

    for ref in ref_rows:
        best_ratio = 0.0
        best_idx = -1
        for i, gen in enumerate(gen_rows):
            if i in used_gen:
                continue
            ratio = difflib.SequenceMatcher(None, gen.key, ref.key).ratio()
            if ratio > best_ratio:
                best_ratio = ratio
                best_idx = i
        if best_ratio >= threshold and best_idx >= 0:
            gen = gen_rows[best_idx]
            used_gen.add(best_idx)
            gq = gen.qty_float
            rq = ref.qty_float
            qty_match = False
            delta_pct = None
            if gq is not None and rq is not None and rq != 0:
                delta_pct = (gq - rq) / rq * 100
                qty_match = abs(delta_pct) < 1.0  # exact or < 1% diff
            elif gq is None and rq is None:
                qty_match = True
            matched.append(MatchResult(
                gen_desc=gen.description,
                gen_qty=gen.quantity,
                ref_desc=ref.description,
                ref_qty=ref.quantity,
                match_ratio=best_ratio,
                qty_match=qty_match,
                delta_pct=delta_pct,
                section=ref.section,
            ))
        else:
            unmatched_ref.append(ref)

    unmatched_gen = [g for i, g in enumerate(gen_rows) if i not in used_gen]
    return matched, unmatched_gen, unmatched_ref


# ---------------------------------------------------------------------------
# DWG -> DXF conversion
# ---------------------------------------------------------------------------
def ensure_dxf(case: dict, oda_exe: str | None = None) -> Path | None:
    """Ensure _converted_dxf exists. Returns DXF folder path."""
    dxf_dir = case.get("dxf_dir")
    if dxf_dir and Path(dxf_dir).exists():
        return Path(dxf_dir)

    dwg_dir = case.get("dwg_dir")
    if not dwg_dir or not Path(dwg_dir).exists():
        return None

    if oda_exe is None:
        from batch_equipment import find_oda
        oda_exe = find_oda()

    if not oda_exe:
        print(f"  [!] ODA converter not found, cannot convert DWG->DXF")
        return None

    from batch_equipment import scan_files, convert_dwg_files
    dwg_path = Path(dwg_dir)
    dwg_files, _, _ = scan_files(dwg_path)
    if not dwg_files:
        return None

    out_dir = dwg_path / "_converted_dxf"
    print(f"  [convert] {len(dwg_files)} DWG -> {out_dir}")
    convert_dwg_files(oda_exe, dwg_files, out_dir)
    return out_dir if out_dir.exists() else None


# ---------------------------------------------------------------------------
# Run single test case
# ---------------------------------------------------------------------------
def run_case(case: dict, do_convert: bool = False) -> CaseResult:
    """Run VOR generation for a test case and compare with reference."""
    name = case["name"]
    ref_path = case.get("ref_vor", "")

    # Check reference exists
    if not ref_path or not Path(ref_path).exists():
        return CaseResult(
            name=name, status="skip",
            error_msg=f"Reference VOR not found: {ref_path}",
        )

    # Only .docx references supported
    if not ref_path.lower().endswith(".docx"):
        return CaseResult(
            name=name, status="skip",
            error_msg=f"Reference is not .docx (xlsx not supported): {ref_path}",
        )

    t0 = time.time()

    # Resolve DXF folder
    dxf_dir = case.get("dxf_dir")
    if dxf_dir and Path(dxf_dir).exists():
        dxf_path = Path(dxf_dir)
    elif do_convert:
        dxf_path = ensure_dxf(case)
        if dxf_path is None:
            return CaseResult(
                name=name, status="error",
                error_msg="Failed to convert DWG->DXF or no DWG files found",
                elapsed_sec=time.time() - t0,
            )
    else:
        dwg_dir = case.get("dwg_dir")
        if dwg_dir and Path(dwg_dir).exists():
            # Try _converted_dxf subfolder
            auto_dxf = Path(dwg_dir) / "_converted_dxf"
            if auto_dxf.exists():
                dxf_path = auto_dxf
            else:
                return CaseResult(
                    name=name, status="skip",
                    error_msg=f"No _converted_dxf in {dwg_dir} (run with --convert)",
                    elapsed_sec=time.time() - t0,
                )
        else:
            return CaseResult(
                name=name, status="skip",
                error_msg=f"DXF/DWG dir not found",
                elapsed_sec=time.time() - t0,
            )

    # Count DXF files
    dxf_files = list(dxf_path.rglob("*.dxf"))
    dxf_count = len(dxf_files)

    if dxf_count == 0:
        return CaseResult(
            name=name, status="error",
            error_msg=f"No .dxf files found in {dxf_path}",
            dxf_count=0,
            elapsed_sec=time.time() - t0,
        )

    # Run VOR pipeline
    try:
        from vor_generator import generate_vor
        output_path = str(_SCRIPT_DIR / f"_mass_test_output_{name}.docx")
        # Suppress noisy output
        logs: list[str] = []
        generate_vor(
            folder=dxf_path,
            output_path=output_path,
            project_name=name,
            log=lambda msg: logs.append(str(msg)),
        )
    except Exception as e:
        return CaseResult(
            name=name, status="error",
            error_msg=f"Pipeline error: {e}\n{traceback.format_exc()[-500:]}",
            dxf_count=dxf_count,
            elapsed_sec=time.time() - t0,
        )

    # Parse generated and reference
    try:
        gen_rows = extract_vor_rows(output_path)
        ref_rows = extract_vor_rows(ref_path)
    except Exception as e:
        return CaseResult(
            name=name, status="error",
            error_msg=f"DOCX parse error: {e}",
            dxf_count=dxf_count,
            elapsed_sec=time.time() - t0,
        )

    # Compare
    matched, unmatched_gen, unmatched_ref = compare_rows(gen_rows, ref_rows)

    # Cleanup temp file
    try:
        os.remove(output_path)
    except OSError:
        pass

    return CaseResult(
        name=name,
        status="ok",
        dxf_count=dxf_count,
        gen_row_count=len(gen_rows),
        ref_row_count=len(ref_rows),
        matched=matched,
        unmatched_gen=unmatched_gen,
        unmatched_ref=unmatched_ref,
        elapsed_sec=time.time() - t0,
    )


# ---------------------------------------------------------------------------
# Excel report generation with run history
# ---------------------------------------------------------------------------
def _make_detail_items(result: CaseResult) -> list[dict]:
    """Flatten a CaseResult into detail row dicts."""
    items = []
    for m in result.matched:
        items.append({
            "case": result.name,
            "section": m.section,
            "ref_desc": m.ref_desc[:80],
            "ref_qty": m.ref_qty,
            "gen_desc": m.gen_desc[:80],
            "gen_qty": m.gen_qty,
            "match_ratio": f"{m.match_ratio:.0%}",
            "qty_match": "YES" if m.qty_match else "NO",
            "delta_pct": f"{m.delta_pct:+.1f}%" if m.delta_pct is not None else "",
        })
    for r in result.unmatched_ref:
        items.append({
            "case": result.name,
            "section": r.section,
            "ref_desc": r.description[:80],
            "ref_qty": r.quantity,
            "gen_desc": "(not found)",
            "gen_qty": "",
            "match_ratio": "0%",
            "qty_match": "MISS",
            "delta_pct": "",
        })
    for g in result.unmatched_gen:
        items.append({
            "case": result.name,
            "section": g.section,
            "ref_desc": "(no match)",
            "ref_qty": "",
            "gen_desc": g.description[:80],
            "gen_qty": g.quantity,
            "match_ratio": "0%",
            "qty_match": "EXTRA",
            "delta_pct": "",
        })
    return items


def _history_row_key(case_name: str, section: str, ref_desc: str) -> str:
    """Create a stable row key for history tracking."""
    desc_short = re.sub(r"\s+", " ", ref_desc).strip()[:60]
    return f"{case_name} / {section[:30]} / {desc_short}"


def write_excel_report(
    results: list[CaseResult],
    report_path: str,
) -> str:
    """Write/append Excel report with History, Latest, Overview sheets."""
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    run_id = datetime.now().strftime("%Y-%m-%d_%H:%M")

    # ── Load or create workbook ──
    rpath = Path(report_path)
    if rpath.exists():
        try:
            wb = load_workbook(str(rpath))
        except Exception:
            wb = Workbook()
    else:
        wb = Workbook()

    bold = Font(bold=True)
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")
    wrap = Alignment(wrap_text=True, vertical="top")

    # ── Sheet: Overview ──
    if "Overview" in wb.sheetnames:
        del wb["Overview"]
    ws_ov = wb.create_sheet("Overview", 0)
    ov_headers = [
        "Case", "Status", "DXF files", "Gen rows", "Ref rows",
        "Fuzzy match %", "Exact qty %", "Key categories", "Time (s)", "Trend",
    ]
    for ci, h in enumerate(ov_headers, 1):
        c = ws_ov.cell(row=1, column=ci, value=h)
        c.font = bold

    for ri, res in enumerate(results, 2):
        ws_ov.cell(row=ri, column=1, value=res.name)
        ws_ov.cell(row=ri, column=2, value=res.status)
        status_cell = ws_ov.cell(row=ri, column=2)
        if res.status == "ok":
            status_cell.fill = green_fill
        elif res.status == "error":
            status_cell.fill = red_fill
        else:
            status_cell.fill = yellow_fill
        ws_ov.cell(row=ri, column=3, value=res.dxf_count)
        ws_ov.cell(row=ri, column=4, value=res.gen_row_count)
        ws_ov.cell(row=ri, column=5, value=res.ref_row_count)
        ws_ov.cell(row=ri, column=6, value=f"{res.fuzzy_match_pct:.1f}%")
        ws_ov.cell(row=ri, column=7, value=f"{res.match_pct:.1f}%")

        # Key categories
        cats = res.category_matches()
        cat_strs = []
        for sec, (ok, total) in sorted(cats.items()):
            cat_strs.append(f"{sec[:25]}: {ok}/{total}")
        ws_ov.cell(row=ri, column=8, value="\n".join(cat_strs[:8]))
        ws_ov.cell(row=ri, column=8).alignment = wrap

        ws_ov.cell(row=ri, column=9, value=f"{res.elapsed_sec:.1f}")

    # Set column widths
    col_widths_ov = [18, 8, 8, 8, 8, 12, 12, 45, 8, 8]
    for ci, w in enumerate(col_widths_ov, 1):
        ws_ov.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet: Latest (detailed per-row comparison) ──
    if "Latest" in wb.sheetnames:
        del wb["Latest"]
    ws_lat = wb.create_sheet("Latest", 1)
    lat_headers = [
        "Case", "Section", "Ref description", "Ref qty",
        "Gen description", "Gen qty", "Match", "Qty OK", "Delta",
    ]
    for ci, h in enumerate(lat_headers, 1):
        ws_lat.cell(row=1, column=ci, value=h).font = bold

    row_idx = 2
    for res in results:
        if res.status != "ok":
            continue
        items = _make_detail_items(res)
        for item in items:
            ws_lat.cell(row=row_idx, column=1, value=item["case"])
            ws_lat.cell(row=row_idx, column=2, value=item["section"])
            ws_lat.cell(row=row_idx, column=3, value=item["ref_desc"])
            ws_lat.cell(row=row_idx, column=4, value=item["ref_qty"])
            ws_lat.cell(row=row_idx, column=5, value=item["gen_desc"])
            ws_lat.cell(row=row_idx, column=6, value=item["gen_qty"])
            ws_lat.cell(row=row_idx, column=7, value=item["match_ratio"])
            qm = item["qty_match"]
            c = ws_lat.cell(row=row_idx, column=8, value=qm)
            if qm == "YES":
                c.fill = green_fill
            elif qm == "MISS":
                c.fill = red_fill
            elif qm == "NO":
                c.fill = yellow_fill
            ws_lat.cell(row=row_idx, column=9, value=item["delta_pct"])
            row_idx += 1

    col_widths_lat = [14, 28, 45, 10, 45, 10, 8, 8, 10]
    for ci, w in enumerate(col_widths_lat, 1):
        ws_lat.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet: History (append run column) ──
    if "History" not in wb.sheetnames:
        ws_hist = wb.create_sheet("History")
        ws_hist.cell(row=1, column=1, value="Item").font = bold
    else:
        ws_hist = wb["History"]

    # Find next free column for this run
    hist_col = ws_hist.max_column + 1
    if ws_hist.cell(row=1, column=1).value is None:
        hist_col = 2
        ws_hist.cell(row=1, column=1, value="Item").font = bold

    ws_hist.cell(row=1, column=hist_col, value=run_id).font = bold
    ws_hist.column_dimensions[get_column_letter(hist_col)].width = 22

    # Build existing row-key -> row-number map
    key_to_row: dict[str, int] = {}
    for ri in range(2, ws_hist.max_row + 1):
        key_val = ws_hist.cell(row=ri, column=1).value
        if key_val:
            key_to_row[str(key_val)] = ri

    next_row = ws_hist.max_row + 1

    for res in results:
        if res.status != "ok":
            # Write case-level status
            case_key = f"{res.name} / STATUS"
            if case_key not in key_to_row:
                key_to_row[case_key] = next_row
                ws_hist.cell(row=next_row, column=1, value=case_key)
                next_row += 1
            ws_hist.cell(row=key_to_row[case_key], column=hist_col,
                         value=res.status)
            continue

        # Case-level overview
        case_key = f"{res.name} / MATCH%"
        if case_key not in key_to_row:
            key_to_row[case_key] = next_row
            ws_hist.cell(row=next_row, column=1, value=case_key)
            next_row += 1
        ws_hist.cell(row=key_to_row[case_key], column=hist_col,
                     value=f"{res.match_pct:.1f}%")

        # Per-matched-item history
        for m in res.matched:
            key = _history_row_key(res.name, m.section, m.ref_desc)
            if key not in key_to_row:
                key_to_row[key] = next_row
                ws_hist.cell(row=next_row, column=1, value=key)
                next_row += 1
            gq = m.gen_qty or "-"
            rq = m.ref_qty or "-"
            delta = f"{m.delta_pct:+.0f}%" if m.delta_pct is not None else "="
            cell_val = f"{gq} / {rq} / {delta}"
            c = ws_hist.cell(row=key_to_row[key], column=hist_col,
                             value=cell_val)
            if m.qty_match:
                c.fill = green_fill
            elif m.delta_pct is not None:
                c.fill = yellow_fill

        for r in res.unmatched_ref:
            key = _history_row_key(res.name, r.section, r.description)
            if key not in key_to_row:
                key_to_row[key] = next_row
                ws_hist.cell(row=next_row, column=1, value=key)
                next_row += 1
            ws_hist.cell(row=key_to_row[key], column=hist_col,
                         value=f"MISS / {r.quantity}")

    # ── Trend column in Overview (compare with previous run) ──
    if hist_col >= 3:
        prev_col = hist_col - 1
        for ri, res in enumerate(results, 2):
            case_key = f"{res.name} / MATCH%"
            hist_row = key_to_row.get(case_key)
            if hist_row is None:
                continue
            curr_val = ws_hist.cell(row=hist_row, column=hist_col).value
            prev_val = ws_hist.cell(row=hist_row, column=prev_col).value
            if curr_val and prev_val:
                try:
                    curr_f = float(str(curr_val).replace("%", ""))
                    prev_f = float(str(prev_val).replace("%", ""))
                    if curr_f > prev_f + 0.5:
                        ws_ov.cell(row=ri, column=10, value="UP")
                        ws_ov.cell(row=ri, column=10).fill = green_fill
                    elif curr_f < prev_f - 0.5:
                        ws_ov.cell(row=ri, column=10, value="DOWN")
                        ws_ov.cell(row=ri, column=10).fill = red_fill
                    else:
                        ws_ov.cell(row=ri, column=10, value="=")
                except (ValueError, TypeError):
                    pass

    # ── Sheet: DataReadiness ──
    if "DataReadiness" in wb.sheetnames:
        del wb["DataReadiness"]
    ws_dr = wb.create_sheet("DataReadiness")
    dr_headers = ["Case", "DWG dir exists", "DXF dir exists",
                   "DXF count", "Ref VOR exists", "Ref format"]
    for ci, h in enumerate(dr_headers, 1):
        ws_dr.cell(row=1, column=ci, value=h).font = bold
    for ri, case in enumerate(TEST_CASES, 2):
        ws_dr.cell(row=ri, column=1, value=case["name"])
        dwg_ok = bool(case.get("dwg_dir") and Path(case["dwg_dir"]).exists())
        dxf_ok = bool(case.get("dxf_dir") and Path(case["dxf_dir"]).exists())
        ws_dr.cell(row=ri, column=2, value="YES" if dwg_ok else "NO")
        ws_dr.cell(row=ri, column=2).fill = green_fill if dwg_ok else red_fill
        ws_dr.cell(row=ri, column=3, value="YES" if dxf_ok else "NO")
        ws_dr.cell(row=ri, column=3).fill = green_fill if dxf_ok else yellow_fill
        if dxf_ok:
            dxf_n = len(list(Path(case["dxf_dir"]).rglob("*.dxf")))
        else:
            dxf_n = 0
        ws_dr.cell(row=ri, column=4, value=dxf_n)
        ref_ok = bool(case.get("ref_vor") and Path(case["ref_vor"]).exists())
        ws_dr.cell(row=ri, column=5, value="YES" if ref_ok else "NO")
        ws_dr.cell(row=ri, column=5).fill = green_fill if ref_ok else red_fill
        ref = case.get("ref_vor", "")
        ext = Path(ref).suffix.lower() if ref else ""
        ws_dr.cell(row=ri, column=6, value=ext or "?")

    col_widths_dr = [18, 14, 14, 10, 14, 10]
    for ci, w in enumerate(col_widths_dr, 1):
        ws_dr.column_dimensions[get_column_letter(ci)].width = w

    # ── Remove default empty sheet ──
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    # ── Save ──
    wb.save(str(rpath))
    return str(rpath)


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description="VOR mass-testing harness",
    )
    parser.add_argument(
        "--convert", action="store_true",
        help="Convert DWG->DXF if _converted_dxf doesn't exist",
    )
    parser.add_argument(
        "--report", default="mass_test_report.xlsx",
        help="Excel report file path (default: mass_test_report.xlsx)",
    )
    parser.add_argument(
        "--only", default="",
        help="Comma-separated list of case names to run",
    )
    args = parser.parse_args()

    # Filter cases
    cases = TEST_CASES
    if args.only:
        names = {n.strip() for n in args.only.split(",")}
        cases = [c for c in cases if c["name"] in names]
        if not cases:
            print(f"No matching cases for: {args.only}")
            print(f"Available: {', '.join(c['name'] for c in TEST_CASES)}")
            sys.exit(1)

    print(f"=" * 60)
    print(f"  VOR Mass Test - {len(cases)} cases")
    print(f"  Report: {args.report}")
    print(f"=" * 60)

    results: list[CaseResult] = []
    for i, case in enumerate(cases, 1):
        name = case["name"]
        print(f"\n[{i}/{len(cases)}] {name}")

        res = run_case(case, do_convert=args.convert)
        results.append(res)

        if res.status == "ok":
            print(f"  DXF: {res.dxf_count}, "
                  f"Gen: {res.gen_row_count}, Ref: {res.ref_row_count}")
            print(f"  Fuzzy match: {res.fuzzy_match_pct:.1f}%, "
                  f"Exact qty: {res.match_pct:.1f}%")
            cats = res.category_matches()
            for sec, (ok, total) in sorted(cats.items()):
                pct = ok / total * 100 if total else 0
                marker = "OK" if pct >= 80 else "WARN" if pct >= 50 else "LOW"
                print(f"    [{marker:4s}] {sec[:40]:40s} {ok:3d}/{total:3d} "
                      f"({pct:.0f}%)")
        elif res.status == "skip":
            print(f"  SKIPPED: {res.error_msg[:80]}")
        else:
            print(f"  ERROR: {res.error_msg[:120]}")

        print(f"  Time: {res.elapsed_sec:.1f}s")

    # Write report
    print(f"\n{'=' * 60}")
    report = write_excel_report(results, args.report)
    print(f"  Report saved: {report}")

    # Summary
    ok_count = sum(1 for r in results if r.status == "ok")
    skip_count = sum(1 for r in results if r.status == "skip")
    err_count = sum(1 for r in results if r.status == "error")
    print(f"  OK: {ok_count}, Skipped: {skip_count}, Errors: {err_count}")

    if ok_count > 0:
        avg_match = sum(r.match_pct for r in results if r.status == "ok") / ok_count
        avg_fuzzy = sum(r.fuzzy_match_pct for r in results if r.status == "ok") / ok_count
        print(f"  Avg fuzzy match: {avg_fuzzy:.1f}%")
        print(f"  Avg exact qty match: {avg_match:.1f}%")

    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
