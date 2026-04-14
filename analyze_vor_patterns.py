#!/usr/bin/env python3
"""Analyze reference VOR Excel files to understand naming patterns.

Extracts all unique work item names and categorizes them by type.
"""

import re
import sys
from pathlib import Path
from collections import defaultdict

sys.stdout.reconfigure(encoding="utf-8")

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required: pip install openpyxl")

PROJECT_ROOT = Path(__file__).resolve().parent
DATA_DIR = PROJECT_ROOT / "Data"
DBT_DIR = DATA_DIR / "ДБТ разделы для ИИ"

# All VOR files
VOR_FILES = [
    DBT_DIR / "01_АБК" / "Обновленные файлы" / "ВОР_" / "ВОР АБК ЭО.xlsx",
    DBT_DIR / "01_АБК" / "Обновленные файлы" / "ВОР_" / "ВОР АБК ЭМ.xlsx",
    DBT_DIR / "01_АБК" / "Обновленные файлы" / "ВОР_" / "ВОР АБК ЭГ.xlsx",
    DBT_DIR / "30. КПП" / "ВОР поз.30.xlsx",
    DBT_DIR / "28. Автовесы" / "ВОР поз.28.xlsx",
    DBT_DIR / "27_Склад вспомогательных материалов с участком погрузки крытых вагонов"
    / "Обновленные файлы" / "ВОР поз.27.xlsx",
    DBT_DIR / "16.2_ЖД КПП" / "Обновленные файлы 1" / "ВОР поз. 16.2.xlsx",
    DBT_DIR / "12.3 Насосная станция поверхностных стоков" / "ВОР поз. 12.3.xlsx",
    DBT_DIR / "8.2 Участок хранения сульфата аммония" / "ЭО" / "ВОР поз. 8.2.xlsx",
    DBT_DIR / "8.2 Участок хранения сульфата аммония" / "ЭМ" / "ВОР ЭМ поз. 8.2.xlsx",
]


def read_vor_items(xlsx_path):
    """Read all items from VOR excel."""
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)

    ws = None
    for sname in wb.sheetnames:
        if "заказчик" in sname.lower():
            ws = wb[sname]
            break
    if ws is None:
        for sname in wb.sheetnames:
            candidate = wb[sname]
            if candidate.max_row and candidate.max_row > 10:
                ws = candidate
                break
    if ws is None:
        ws = wb.active

    items = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if not row or len(row) < 4:
            continue
        col_a = row[0]
        col_b = row[1]
        col_c = row[2]
        col_d = row[3]

        name = str(col_b or "").strip()
        if not name:
            continue

        try:
            qty = float(col_d) if col_d is not None else 0
        except (ValueError, TypeError):
            qty = 0

        unit = str(col_c or "").strip()
        row_num = str(col_a or "").strip()

        items.append({
            "row_num": row_num,
            "name": name,
            "unit": unit,
            "qty": qty,
            "is_section": (col_d is None and col_a is None and name),
        })

    wb.close()
    return items


def classify_item(name):
    """Classify a work item by type."""
    lower = name.lower()

    if re.search(r'светильник', lower):
        return "luminaire"
    if re.search(r'розетк', lower):
        return "socket"
    if re.search(r'выключатель', lower):
        return "switch"
    if re.search(r'(щит|щр|що|вру|щэ)', lower):
        return "panel"
    if re.search(r'(кабел|ввг|nym|прокладк.*кабел)', lower):
        return "cable"
    if re.search(r'провод', lower):
        return "wire"
    if re.search(r'(трубы?|лот[оа]к|короб)', lower):
        return "conduit"
    if re.search(r'(коробк|распред)', lower):
        return "junction_box"
    if re.search(r'(указатель|exit|выход)', lower):
        return "exit_sign"
    if re.search(r'(датчик|извещатель)', lower):
        return "sensor"
    if re.search(r'(автомат|предохранитель|дифф)', lower):
        return "circuit_breaker"
    if re.search(r'заземл', lower):
        return "grounding"
    if re.search(r'монтаж', lower):
        return "work_item"
    return "other"


def main():
    categories = defaultdict(list)
    all_items = []
    work_items = []
    material_items = []

    for vor_path in VOR_FILES:
        if not vor_path.exists():
            print(f"SKIP: {vor_path.name} not found")
            continue

        print(f"\n{'='*80}")
        print(f"FILE: {vor_path.name}")
        print(f"{'='*80}")

        items = read_vor_items(vor_path)

        for item in items:
            if item["is_section"]:
                continue
            if item["qty"] <= 0:
                continue

            name = item["name"]
            cat = classify_item(name)
            categories[cat].append({
                "name": name,
                "unit": item["unit"],
                "qty": item["qty"],
                "file": vor_path.name,
            })
            all_items.append(item)

            # Distinguish work items (with "монтаж", "прокладка" etc) vs materials
            lower = name.lower()
            if re.match(r'(монтаж|установка|прокладка|забивка|крепление|окраска|подключение|пробивка|заделка|герметизация|испытание)', lower):
                work_items.append(item)
            else:
                material_items.append(item)

    print(f"\n\n{'='*80}")
    print("CATEGORY SUMMARY")
    print(f"{'='*80}")

    for cat in sorted(categories.keys()):
        items_in_cat = categories[cat]
        print(f"\n--- {cat.upper()} ({len(items_in_cat)} items) ---")
        # Show unique patterns
        seen = set()
        for it in items_in_cat:
            # Show first 10 unique
            if it["name"] not in seen and len(seen) < 15:
                seen.add(it["name"])
                print(f"  [{it['unit']:>5s}] {it['name'][:100]}")

    print(f"\n\n{'='*80}")
    print("WORK ITEMS (with action verbs)")
    print(f"{'='*80}")
    seen = set()
    for it in work_items:
        name = it["name"]
        if name not in seen:
            seen.add(name)
            print(f"  [{it['unit']:>5s}] {name[:120]}")

    print(f"\n\n{'='*80}")
    print("MATERIAL ITEMS (equipment/material names)")
    print(f"{'='*80}")
    seen = set()
    for it in material_items:
        name = it["name"]
        if name not in seen and len(seen) < 60:
            seen.add(name)
            print(f"  [{it['unit']:>5s}] {name[:120]}")

    print(f"\n\nTotals: {len(all_items)} items total, "
          f"{len(work_items)} work items, {len(material_items)} material items")


if __name__ == "__main__":
    main()
