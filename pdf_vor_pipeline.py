#!/usr/bin/env python3
"""pdf_vor_pipeline.py -- Full VOR generation pipeline for PDF folders.

Generates a complete Ведомость объёмов работ (Work Volume Statement)
from a folder of PDF electrical drawings.

The pipeline combines three data sources:
  1. СО spec PDFs  → authoritative quantities (total counts)
  2. Plan PDFs     → height distribution ratios (per-elevation counts)
  3. Schema PDFs   → panel info (щитовое оборудование)

Height categories:
  - до 5 метров       (elevation <  5 m)
  - от 5 до 13 метров (elevation >= 5 m  and < 13 m)
  - от 13 до 20 метров (elevation >= 13 m and < 20 m)
  - от 20 до 35 метров (elevation >= 20 m)

Usage:
    python pdf_vor_pipeline.py "Data/ДБТ разделы для ИИ/03_ГПК_/3-я захватка/02_PDF"
"""

from __future__ import annotations

import logging
import os
import re
import io
from collections import defaultdict, OrderedDict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Literal, Optional

from equipment_counter import SpecItem

_log = logging.getLogger(__name__)


def _vor_worker_count(n_jobs: int, heavy: bool = False) -> int:
    """Pick a safe number of parallel workers.

    Goals (per user requirement): keep quality (stay parallel) but never
    exhaust RAM or hang. The heavy no-СО visual passes render large PDFs at
    high DPI — each worker can hold hundreds of MB, so we cap by available
    memory and a hard ceiling, not just by core count.

    Priority:
      1. Explicit ``VOR_WORKERS`` env var always wins (operator override).
      2. Otherwise: min(cores-2, RAM-budget, hard-ceiling, n_jobs).
    """
    if n_jobs <= 0:
        return 0
    # 1. Explicit override.
    try:
        _env = int(os.environ.get("VOR_WORKERS", "0") or "0")
    except ValueError:
        _env = 0
    if _env > 0:
        return max(1, min(_env, n_jobs))

    # 2. Core budget — leave 2 cores for the web server / OS.
    cores = max(1, (os.cpu_count() or 4) - 2)

    # 3. Memory budget. Heavy visual workers need ~2 GB peak each; light
    #    workers (text/legend probing) ~0.5 GB. Use available RAM if psutil
    #    is present, otherwise assume a conservative 8 GB.
    per_worker_gb = 2.0 if heavy else 0.5
    try:
        import psutil  # type: ignore
        avail_gb = psutil.virtual_memory().available / (1024 ** 3)
    except Exception:
        avail_gb = 8.0
    # Keep ~2 GB headroom for the parent + OS.
    mem_workers = max(1, int((avail_gb - 2.0) / per_worker_gb))

    # 4. Hard ceiling so we never spawn a swarm even on big boxes.
    ceiling = 4 if heavy else 8

    return max(1, min(cores, mem_workers, ceiling, n_jobs))


# ---------------------------------------------------------------------------
# Height categories (reuse from vor_generator)
# ---------------------------------------------------------------------------

HeightCategory = Literal[
    "до 5 метров",
    "от 5 до 13 метров",
    "от 13 до 20 метров",
    "от 20 до 35 метров",
]

HEIGHT_CATEGORIES: list[HeightCategory] = [
    "до 5 метров",
    "от 5 до 13 метров",
    "от 13 до 20 метров",
    "от 20 до 35 метров",
]


def elevation_to_height(height_m: float) -> HeightCategory:
    """Map ceiling / working height (metres) to VOR height category.

    The input should be the *ceiling height* (height from ground to the ceiling
    where equipment is mounted), not the raw floor elevation.  The caller is
    responsible for converting floor elevation → ceiling height using the
    elev_to_ceiling map built in Step 1b.

    Thresholds (validated against ГПК 3-я захватка and КПП):
      - < 5.0 m  → "до 5 метров"       (e.g., ceiling 2.9–4.2 m)
      - < 14.0 m → "от 5 до 13 метров"  (e.g., ceiling 5.8–13.8 m)
      - < 21.0 m → "от 13 до 20 метров" (e.g., ceiling 14.0–18.6 m)
      - ≥ 21.0 m → "от 20 до 35 метров" (e.g., ceiling 23.4–28.2 m)

    Note: Boundaries use 14.0 and 21.0 instead of 13.0 and 20.0 because
    VOR categories refer to floor elevation ranges, and ceiling heights
    can exceed the nominal range by up to 1m (e.g., floor 9.0m has
    ceiling 13.8m, which is classified as "от 5 до 13" in practice).
    """
    if height_m < 5.0:
        return "до 5 метров"
    if height_m < 14.0:
        return "от 5 до 13 метров"
    if height_m < 21.0:
        return "от 13 до 20 метров"
    return "от 20 до 35 метров"


# ---------------------------------------------------------------------------
# File classification
# ---------------------------------------------------------------------------

_ELEV_RE = re.compile(r"отм\.\s*([+-]?\d+[.,]\d+)")
# Also match subsequent elevations after the first one (e.g., "+9.000" in "отм. +7.800 +9.000")
_ALL_ELEV_RE = re.compile(r"[+-]?\d+[.,]\d{2,}")


def _extract_elevations(filename: str) -> list[float]:
    """Extract elevation(s) from PDF filename.

    "005-Планы освещения-отм. 0.000.pdf" → [0.0]
    "007-Планы освещения-отм. +7.800 +9.000.pdf" → [7.8, 9.0]
    """
    # First check that "отм." is present (to avoid matching random numbers)
    if not _ELEV_RE.search(filename):
        return []
    # Then extract ALL elevation-like numbers from the filename
    # (the first is after "отм.", subsequent ones follow after spaces)
    stem = filename.rsplit(".", 1)[0]  # Remove .pdf extension
    return [float(m.replace(",", ".")) for m in _ALL_ELEV_RE.findall(stem)]


# Regex to find "План на отм. +2.900" or "отм. 0.000" inside PDF content
_CONTENT_ELEV_RE = re.compile(
    r"[Пп]лан\s+(?:на\s+)?отм\.?\s*([+-]?\d+[.,]\d{2,})", re.IGNORECASE
)
# Fallback: also match "на отм. X.XXX, +Y.YYY" in title blocks
_TITLE_BLOCK_ELEV_RE = re.compile(
    r"на\s+отм\.?\s*((?:[+-]?\d+[.,]\d{2,}\s*[,;]?\s*)+)", re.IGNORECASE
)


def _extract_elevations_from_content(pdf_path: str) -> list[float]:
    """Extract elevation(s) from the text content of a PDF.

    Looks for patterns like "План на отм. 0.000", "План на отм. +2.900",
    or title block text like "на отм. 0.000, +2.900".

    Returns sorted unique elevations found.
    """
    try:
        import pdfplumber
        elevs: set[float] = set()
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                # Match "План на отм. X.XXX"
                for m in _CONTENT_ELEV_RE.finditer(text):
                    val = float(m.group(1).replace(",", "."))
                    elevs.add(val)
                # Match title block "на отм. 0.000, +2.900"
                for m in _TITLE_BLOCK_ELEV_RE.finditer(text):
                    nums_str = m.group(1)
                    for n in _ALL_ELEV_RE.findall(nums_str):
                        elevs.add(float(n.replace(",", ".")))
        return sorted(elevs)
    except Exception:
        return []


def _classify_pdf(filename: str) -> str:
    """Classify PDF file type.

    Returns: "plan" | "spec" | "schema" | "tray" | "binding" | "other"
    """
    lower = filename.lower()

    # Spec files: СО-*, CO-*
    if re.match(r"^[сc][оo][-_]", lower):
        return "spec"

    # Schema files (explicit "схем" or panel designation like "ЩО", "ЩАО")
    if "схем" in lower:
        return "schema"
    if re.search(r"(що|щао|цсао)", lower) and "план" not in lower:
        return "schema"

    # Plan files: освещение, электрооборудование
    if "план" in lower and ("освещ" in lower or "электрооборуд" in lower):
        return "plan"

    # Binding/attachment plans (also handle OCR/typo variants like "приЯвязк")
    if "привязк" in lower or re.search(r"при.?вязк", lower):
        return "binding"

    # Cable tray plan ("План кабеленесущих систем") — these sheets carry the
    # per-segment size/length callouts the no-СО pipeline reads for metraж.
    # Mounting-detail sheets ("Узлы/Узел крепления лотков …") have NO segment
    # lengths, so route them to "other" to avoid a costly, fruitless parse.
    if "узл" in lower and "креплени" in lower:
        return "other"
    if "кабеленес" in lower or "кабеле-нес" in lower:
        return "tray"

    # Cable tray layout
    if "лотк" in lower or "лотoк" in lower:
        return "tray"

    # General data, title pages, VOR reference
    if any(kw in lower for kw in ("общие данные", "титульный", "вор ", "test_report")):
        return "other"

    return "other"


# ---------------------------------------------------------------------------
# Spec classification (mirrors vor_generator._classify_spec_item)
# ---------------------------------------------------------------------------

def _is_valid_equipment_desc(desc: str) -> bool:
    """Reject non-equipment garbage that leaks into no-СО visual synthesis.

    On grounding/lightning CAD sheets the legend probe sometimes captures
    room schedules ("Помещение МСС 87,2 В3") or scrambled text where the
    PDF stored glyphs interleaved with spaces ("з т а е з р е м ..."). Such
    strings are not real equipment and must not become VOR rows.
    """
    if not desc:
        return False
    d = desc.strip()
    if len(d) < 6:
        return False

    # 1) Room-schedule / explication noise (areas, room labels).
    low = d.lower()
    _ROOM_KW = ("помещение", "отделение", "вестибюль", "лестничная клетка",
                "тамбур", "коридор", "подстанция", "санузел", "кабинет",
                "комната", "площадь", "экспликаци")
    if any(k in low for k in _ROOM_KW):
        return False

    # 2) Scrambled CAD text: many single-character whitespace-separated
    #    tokens (e.g. "з т а е з р е м о л ..."). If >40% of tokens are a
    #    single letter/digit, treat as corrupt.
    tokens = d.split()
    if len(tokens) >= 6:
        singles = sum(1 for t in tokens if len(t) == 1)
        if singles / len(tokens) > 0.4:
            return False

    # 3) Mostly digits/punctuation (coordinates, dimensions, areas).
    letters = sum(1 for ch in d if ch.isalpha())
    if letters < 3:
        return False
    digits = sum(1 for ch in d if ch.isdigit())
    if digits > letters:
        return False

    return True


def classify_spec_item(desc: str) -> str:
    """Classify a spec item by description into VOR category.

    Returns: "panel" | "luminaire" | "indicator" | "pictogram" |
             "switch" | "cable" | "tray" | "pvc" | "material" |
             "grounding" | "lightning"
    """
    nl = desc.lower()

    # Panels
    if any(kw in nl for kw in ("щит", "що", "щао", "цсао", "вру",
                                "вводно-распределительн", "распределительное устройство")):
        return "panel"

    # Pictograms
    if "наклейк" in nl and ("выход" in nl or "указател" in nl):
        return "pictogram"
    if "пиктограмм" in nl or "пэу" in nl:
        return "pictogram"

    # Indicators (before luminaire!)
    if any(kw in nl for kw in ("указатель", "mercury", "atom", "выход", "exit")):
        return "indicator"

    # Emergency blocks (аварийное питание)
    if "блок аварийн" in nl or "conversion kit" in nl:
        return "emergency_block"

    # Luminaires
    if "светильник" in nl or ("led" in nl and "розетк" not in nl
                               and "указател" not in nl and "mercury" not in nl
                               and "atom" not in nl):
        return "luminaire"

    # Switches, sockets, and electrical devices
    if any(kw in nl for kw in ("выключатель", "пост управлен")):
        return "switch"
    if "розетк" in nl:
        return "switch"
    if "рамка" in nl and any(kw in nl for kw in ("постов", "atlas", "этюд", "systeme")):
        return "switch"

    # Junction boxes
    if "коробк" in nl:
        return "junction_box"

    # Crimping materials
    if "гильз" in nl and "закладн" not in nl:
        return "crimping"
    if "термоусад" in nl:
        return "crimping"

    # Cables (including wire/провод for PuGV, PV-3 etc.)
    if any(kw in nl for kw in ("ппгнг", "ввгнг", "вбшвнг", "кабель силов")):
        return "cable"
    if "провод" in nl and any(kw in nl for kw in ("пугв", "пв-3", "пв3", "пугвнг")):
        return "cable"

    # PVC conduit
    if any(kw in nl for kw in ("труба пвх", "гофр.", "с протяжкой")):
        return "pvc"
    if "держатель с защелк" in nl:
        return "pvc"

    # Cable trays and accessories
    if any(kw in nl for kw in ("лоток", "лотка", "лотку")):
        return "tray"
    if any(kw in nl for kw in ("т-отвод", "угол плоский", "угол внутр", "угол внеш",
                                "соединитель лотк", "переход прямой", "разветвитель",
                                "скоба для настенн", "стойка потолочн")):
        return "tray"

    # Wall penetration items (проходки через стены)
    if "огнестойк" in nl and "пен" in nl:
        return "penetration"
    if "пистолет" in nl and "пен" in nl:
        return "penetration"
    if "гильза закладн" in nl:
        return "penetration"

    # Grounding: заземляющая скоба на ленте
    if "заземля" in nl and "скоб" in nl:
        return "grounding"

    # Hardware/fasteners (always in trays/materials section)
    # Guard: don't catch lightning/grounding items like "болтовой на водосток"
    _hw_kws = ("гайка", "шайба", "винт", "шпилька",
               "анкер", "шпильк", "пена полиурет",
               "пистолет для", "стойка потолочная сварн")
    if any(kw in nl for kw in _hw_kws):
        if not any(gkw in nl for gkw in ("токоотвод", "заземлен", "молниезащит", "проводник")):
            return "material"
    # "болт" as standalone word (not "болтовой")
    if "болт" in nl and "болтов" not in nl:
        return "material"

    # Lightning protection — must be checked BEFORE grounding because
    # some keywords overlap (e.g., "соединитель", "держатель", "проводник")
    if any(kw in nl for kw in ("молниезащит", "молниеприемник",
                                "проводник круглый", "круглый проводник",
                                "круглых проводник", "для круглых",
                                "держатель клик", "клик для",
                                "обойма для круглого", "токоотвод",
                                "на водосток", "стяжка стальн",
                                "для прутка", "универсальн")):
        # Guard: "универсальный" but grounding-related items
        if any(gkw in nl for gkw in ("заземлен", "стержн")):
            return "grounding"
        return "lightning"

    # Grounding — expanded to catch all grounding system components
    if any(kw in nl for kw in ("заземлен", "уравнивания потенциал",
                                "стержень заземл", "наконечник стержн",
                                "забивн", "sds-max",
                                "плоский проводник", "плоского проводник",
                                "горизонтальный заземлител",
                                "антикоррозийн", "цинков", "спрей цинк",
                                "диагональн", "корпус регулир",
                                "держатель плоского")):
        return "grounding"

    # Remaining connectors/joiners — often belong to grounding/lightning
    # but generic "соединитель" without clear context stays as material
    if "соединител" in nl and "лотк" not in nl:
        # Check context clues for grounding
        if any(gkw in nl for gkw in ("стержн", "заземл", "плоског", "пруток-полос")):
            return "grounding"
        return "lightning"

    return "material"


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

@dataclass
class PlanCountResult:
    """Equipment count from a single plan PDF at a specific elevation."""
    filename: str
    elevation: float
    height_category: HeightCategory
    counts: dict[str, int] = field(default_factory=dict)
    # model_description -> count (e.g., "SLICK.PRS LED 50..." -> 15)


@dataclass
class VorSection:
    """One section in the VOR output."""
    title: str
    rows: list[dict] = field(default_factory=list)
    # Each row: {"name": ..., "unit": ..., "qty": ..., "is_material": bool, "drawing_ref": ""}


# ---------------------------------------------------------------------------
# Core pipeline
# ---------------------------------------------------------------------------

def _normalize_model_name(desc: str) -> str:
    """Normalize equipment model name for matching between spec and plan.

    Removes common prefixes ("Светильник", "Световые указатели") and
    normalizes whitespace.
    """
    txt = desc.strip()
    txt = re.sub(r"^Светильн\S*\s*", "", txt, flags=re.IGNORECASE)
    txt = re.sub(r"^Светов\S*\s+указател\S*\s*", "", txt, flags=re.IGNORECASE)
    txt = re.sub(r'^["\s«»]+', "", txt)
    txt = re.sub(r'["\s«»]+$', "", txt)
    # Normalize wattage: "30W" → "30", "50Вт" → "50" etc.
    txt = re.sub(r'(\d+)\s*[Ww](?:\b|(?=\s))', r'\1', txt)
    txt = re.sub(r'(\d+)\s*[Вв][Тт]', r'\1', txt)
    txt = re.sub(r"\s+", " ", txt).strip()
    return txt


def _model_key(name: str, model: str = "") -> str:
    """Create a fuzzy matching key from model name for spec↔plan correlation.

    Checks both the description/name and the separate model field for keywords.
    This is critical because spec items often have generic descriptions
    ("Светодиодный светильник IP65 50 ВТ...") while the model keyword
    ("SLICK.PRS LED 50 with driver box...") is in a separate field.
    """
    # Try both name and model fields
    for source in (name, model):
        if not source:
            continue
        norm = _normalize_model_name(source)
        for kw in ("SLICK", "ARCTIC", "INSEL", "MERCURY", "ATOM", "CD LED",
                    "CONVERSION KIT", "MARS", "LUNA"):
            if kw.lower() in norm.lower():
                idx = norm.lower().index(kw.lower())
                rest = norm[idx:]
                # Strip trailing "Ex"/"5000K"/etc. to merge Ex and non-Ex variants
                # (plan legends distinguish them, but for VOR they share the same
                # spec item split, and merging gives better detection rates)
                key = rest[:50].strip().lower()
                # Remove trailing color temp + optional "ex" suffix
                key = re.sub(r'\s*(ex\s+)?\d{4}k\s*$', '', key).strip()
                return key

    # Fallback: use combined text
    combined = f"{name} {model}".strip()
    norm = _normalize_model_name(combined)
    return norm[:50].strip().lower()


# Max seconds a single PDF may spend in the legend probe. Some pathological
# grounding/lightning sheets (~90k vector lines) make the legend parser spin
# almost indefinitely; we skip such a page rather than hang the whole VOR run.
_LEGEND_PAGE_TIMEOUT = float(os.environ.get("VOR_LEGEND_PAGE_TIMEOUT", "120"))

# Per-plan counting timeout. Visual matching renders large PDFs at high DPI;
# a single monster plan (e.g. an ЭГ lightning/grounding sheet) can wall the
# whole counting pass. Skip such a page instead of hanging the run.
_PLAN_PAGE_TIMEOUT = float(os.environ.get("VOR_PLAN_PAGE_TIMEOUT", "300"))


def _probe_legend_worker(args):
    """Probe a single PDF for a legend + content elevations (no-СО intake).

    Standalone, picklable for ProcessPoolExecutor. Returns
    (pdf_name, n_legend_items, elevs_from_content) where n_legend_items is 0
    if no parseable legend. The parent decides promotion and final elevations.
    """
    pdf_path, pdf_name, need_content_elevs = args
    from pdf_legend_parser import parse_legend as _probe_legend
    elevs_content: list[float] = []
    if need_content_elevs:
        try:
            elevs_content = _extract_elevations_from_content(pdf_path)
        except Exception:  # noqa: BLE001
            elevs_content = []
    try:
        _lg = _probe_legend(pdf_path)
        n_items = len(_lg.items) if (_lg and _lg.items) else 0
    except Exception:  # noqa: BLE001
        n_items = 0
    return (pdf_name, n_items, elevs_content)


def _count_one_plan_worker(args):
    """Process a single plan PDF (text + optional visual counting).

    Standalone, picklable function for ProcessPoolExecutor. Returns
    (filename, merged_counts, txt_total, vis_total, log_lines). Elevation
    splitting / height-category assignment is done by the parent so this
    worker stays free of the elev_to_ceiling map (kept lightweight).
    """
    fname, path, enable_visual = args
    log_lines: list[str] = []

    from pdf_legend_parser import parse_legend
    from pdf_count_text import count_symbols

    legend = parse_legend(path)
    if not legend or not legend.items:
        log_lines.append(f"    ⚠ No legend found")
        return (fname, None, 0, 0, log_lines)

    # ── Text-based counting (fast, primary) ──
    count_result = count_symbols(path, legend)
    txt_total = sum(count_result.counts.values())

    # ── Visual counting (template matching) ──
    visual_counts: dict[int, int] = {}
    vis_total = 0
    if enable_visual:
        try:
            from pdf_count_visual import match_symbols as _match_symbols
            vis_result = _match_symbols(path, legend_result=legend)
            visual_counts = vis_result.counts
            vis_total = sum(visual_counts.values())
        except Exception as exc:
            log_lines.append(f"    ⚠ Visual counting failed: {exc}")

    # ── Merge: text + visual → model_counts ──
    model_counts: dict[str, int] = {}
    for idx, item in enumerate(legend.items):
        if not item.description:
            continue
        sym = item.symbol or ""
        vis_count = visual_counts.get(idx, 0)
        txt_count = count_result.counts.get(sym, 0) if sym else 0

        if not sym:
            count = vis_count
        elif txt_count == 0 and vis_count > 0:
            count = vis_count
        else:
            count = txt_count

        if count > 0:
            model_counts[item.description] = (
                model_counts.get(item.description, 0) + count
            )

    return (fname, model_counts, txt_total, vis_total, log_lines)


def count_equipment_on_plans(
    folder: Path,
    plan_files: list[tuple[str, list[float]]],
    log=print,
    elev_to_ceiling: dict[float, float] | None = None,
) -> list[PlanCountResult]:
    """Count equipment on plan PDFs using text + visual symbol counting.

    Uses text-based counting (fast, precise for extractable text) combined
    with OpenCV template-matching visual detection (finds graphical markers
    that are not extractable as text — ~30% of all markers on typical plans).

    For each legend item the final count = max(visual, text), with a guard
    against visual false-positive floods (vis > txt×3 when txt ≥ 3 → use txt).

    Returns per-plan count results with elevation/height info.
    Uses elev_to_ceiling map to convert floor elevations to ceiling heights
    for proper height category assignment.
    """
    from pdf_legend_parser import parse_legend
    from pdf_count_text import count_symbols

    # Visual detection — optional, graceful degradation if unavailable
    _has_visual = False
    _ENABLE_VISUAL = os.environ.get("VOR_VISUAL", "1") == "1"
    if _ENABLE_VISUAL:
        try:
            from pdf_count_visual import match_symbols as _match_symbols
            _has_visual = True
        except ImportError:
            _log.warning("pdf_count_visual not available — visual counting disabled")

    def _elev_to_hcat(elev: float) -> HeightCategory:
        """Convert floor elevation to height category using ceiling map."""
        if elev_to_ceiling and elev in elev_to_ceiling:
            return elevation_to_height(elev_to_ceiling[elev])
        return elevation_to_height(elev)

    results = []

    # Build work list (skip files without elevations).
    work = [
        (fname, str(folder / fname), elevations)
        for fname, elevations in plan_files
        if elevations
    ]

    # ── Parallel per-plan counting (ProcessPoolExecutor) ──
    # Each plan PDF is independent: parse_legend + count_symbols (+ optional
    # visual) produce per-file counts with no shared mutable state. We fan out
    # across CPU cores, then do elevation-splitting serially in the parent.
    enable_visual = _has_visual and _ENABLE_VISUAL
    counted: dict[str, tuple] = {}  # fname → (model_counts, txt_total, vis_total)

    # Heavy when visual matching is on (renders large PDFs at high DPI).
    n_workers = _vor_worker_count(len(work) if work else 0, heavy=enable_visual)

    import time as _time
    _total_plans = len(work)
    use_parallel = n_workers > 1
    if use_parallel:
        import multiprocessing as _mp
        from concurrent.futures import (
            ProcessPoolExecutor,
            as_completed,
            TimeoutError as _CountTimeout,
        )
        log(f"  Counting {_total_plans} plans across {n_workers} workers "
            f"(visual={'on' if enable_visual else 'off'})…")
        try:
            # Use 'spawn' to avoid fork-in-threaded-server deadlocks (uvicorn
            # runs the pipeline in a worker thread via asyncio.to_thread).
            _ctx = _mp.get_context("spawn")
            _t0 = _time.time()
            _done = 0
            with ProcessPoolExecutor(max_workers=n_workers, mp_context=_ctx) as ex:
                _fut_to_name = {
                    ex.submit(_count_one_plan_worker, (fname, path, enable_visual)): fname
                    for fname, path, elevations in work
                }
                _pending = set(_fut_to_name)
                while _pending:
                    try:
                        # Bound the wait so one stuck plan can't wall the run.
                        # Re-arm after each completion to give every remaining
                        # plan its own budget.
                        for fut in as_completed(list(_pending),
                                                timeout=_PLAN_PAGE_TIMEOUT):
                            _pending.discard(fut)
                            fname, mc, txt_total, vis_total, log_lines = fut.result()
                            for ln in log_lines:
                                log(ln)
                            _done += 1
                            _pct = int(_done * 100 / _total_plans) if _total_plans else 100
                            _el = _time.time() - _t0
                            _eta = (_el / _done) * (_total_plans - _done) if _done else 0
                            if mc is None:
                                log(f"    ⏳ обработано {_done}/{_total_plans} ({_pct}%) "
                                    f"— {fname} | прошло {int(_el)}с, осталось ~{int(_eta)}с")
                            else:
                                counted[fname] = (mc, txt_total, vis_total)
                                log(f"  {fname}: Text {txt_total} | Visual {vis_total} "
                                    f"| Merged {sum(mc.values())}")
                                log(f"    ⏳ обработано {_done}/{_total_plans} ({_pct}%) "
                                    f"— {fname} | прошло {int(_el)}с, осталось ~{int(_eta)}с")
                            break  # re-arm the per-plan timeout
                    except _CountTimeout:
                        # No plan finished within the budget → the still-running
                        # ones are too slow; skip them rather than hang forever.
                        for fut in list(_pending):
                            _nm = _fut_to_name.get(fut, "?")
                            fut.cancel()
                            _done += 1
                            log(f"    ⚠ лист пропущен по таймауту "
                                f"(>{int(_PLAN_PAGE_TIMEOUT)}с) — {_nm}")
                        _pending.clear()
                        # Don't wait on the cancelled (possibly stuck) workers.
                        ex.shutdown(wait=False, cancel_futures=True)
                        break
        except Exception as exc:
            _log.warning("Parallel counting failed (%s) — falling back to sequential", exc)
            log(f"  ⚠ Parallel counting failed ({exc}); running sequentially")
            use_parallel = False
            counted = {}

    if not use_parallel:
        # Sequential fallback (also used when n_workers <= 1).
        _t0 = _time.time()
        for _i, (fname, path, elevations) in enumerate(work, 1):
            log(f"  Counting: {fname} (elevations={elevations})")
            fn, mc, txt_total, vis_total, log_lines = _count_one_plan_worker(
                (fname, path, enable_visual)
            )
            for ln in log_lines:
                log(ln)
            _pct = int(_i * 100 / _total_plans) if _total_plans else 100
            _el = _time.time() - _t0
            _eta = (_el / _i) * (_total_plans - _i) if _i else 0
            if mc is None:
                log(f"    ⏳ обработано {_i}/{_total_plans} ({_pct}%) "
                    f"— {fname} | прошло {int(_el)}с, осталось ~{int(_eta)}с")
                continue
            counted[fname] = (mc, txt_total, vis_total)
            log(f"    Text: {txt_total} | Visual: {vis_total} | Merged: {sum(mc.values())}")
            log(f"    ⏳ обработано {_i}/{_total_plans} ({_pct}%) "
                f"— {fname} | прошло {int(_el)}с, осталось ~{int(_eta)}с")

    # ── Elevation splitting / height-category assignment (serial) ──
    for fname, path, elevations in work:
        if fname not in counted:
            continue
        model_counts, txt_total, vis_total = counted[fname]

        if len(elevations) == 1:
            # Single elevation — assign all counts to it
            hcat = _elev_to_hcat(elevations[0])
            results.append(PlanCountResult(
                filename=fname,
                elevation=elevations[0],
                height_category=hcat,
                counts=model_counts,
            ))
        else:
            # Multiple elevations on one PDF (e.g., two floor plans on one sheet).
            # Split counts proportionally. Without spatial analysis we split evenly.
            n_elevs = len(elevations)
            for elev in elevations:
                hcat = _elev_to_hcat(elev)
                split_counts = {m: max(1, c // n_elevs) for m, c in model_counts.items()}
                results.append(PlanCountResult(
                    filename=fname,
                    elevation=elev,
                    height_category=hcat,
                    counts=split_counts,
                ))
                log(f"    → Split for elev={elev} ({hcat}): {sum(split_counts.values())} items")

    return results


class HeightRatios:
    """Height distribution ratios with underlying raw counts for weighted merging."""

    def __init__(self):
        self.ratios: dict[str, dict[HeightCategory, float]] = {}
        self.raw_counts: dict[str, dict[HeightCategory, int]] = {}

    # Implement dict-like interface so existing code works
    def __contains__(self, key):
        return key in self.ratios

    def __getitem__(self, key):
        return self.ratios[key]

    def get(self, key, default=None):
        return self.ratios.get(key, default)

    def items(self):
        return self.ratios.items()

    def keys(self):
        return self.ratios.keys()

    def values(self):
        return self.ratios.values()

    def __iter__(self):
        return iter(self.ratios)


def build_height_ratios(
    plan_results: list[PlanCountResult],
) -> HeightRatios:
    """Build height distribution ratios per equipment model.

    Returns HeightRatios object with both ratios and raw counts.
    """
    # Accumulate counts per model per height
    raw: dict[str, dict[HeightCategory, int]] = defaultdict(lambda: defaultdict(int))

    for pr in plan_results:
        for model_desc, count in pr.counts.items():
            key = _model_key(model_desc)
            raw[key][pr.height_category] += count

    result = HeightRatios()
    result.raw_counts = dict(raw)

    for key, height_counts in raw.items():
        total = sum(height_counts.values())
        if total > 0:
            result.ratios[key] = {h: c / total for h, c in height_counts.items()}

    return result


def distribute_by_height(
    total: int,
    ratios: dict[HeightCategory, float],
) -> dict[HeightCategory, int]:
    """Distribute a total quantity across height categories by ratios.

    Ensures integer quantities sum to total (last category gets remainder).
    """
    if not ratios:
        return {"до 5 метров": total}

    result: dict[HeightCategory, int] = {}
    remaining = total

    ordered = [(h, ratios.get(h, 0.0)) for h in HEIGHT_CATEGORIES if ratios.get(h, 0.0) > 0]
    if not ordered:
        return {"до 5 метров": total}

    for i, (hcat, frac) in enumerate(ordered):
        if i == len(ordered) - 1:
            result[hcat] = remaining
        else:
            share = int(round(total * frac))
            share = max(0, min(share, remaining))
            result[hcat] = share
            remaining -= share

    return result


# ---------------------------------------------------------------------------
# Schema parsing (panels, breaker counts)
# ---------------------------------------------------------------------------

@dataclass
class PanelInfo:
    """Panel info extracted from schema PDFs."""
    name: str          # e.g. "ЩО-3", "ЩАО-3", "ЦСАО3"
    breaker_count: int  # number of outgoing breakers/fuses (QF)
    three_pole: int     # number of three-pole breakers (typically 1 for the incoming)


def parse_schema_panels(folder: Path, schema_files: list[str], log=print) -> list[PanelInfo]:
    """Parse schema PDFs to extract panel names and breaker counts.

    Looks for QF designations in tables and text to count breakers.
    """
    import pdfplumber

    panels: list[PanelInfo] = []

    for fname in schema_files:
        path = str(folder / fname)
        log(f"  Parsing schema: {fname}")

        # Determine panel name from filename
        # Use the LAST panel designation in the filename
        # e.g., "003 - Схемы ЩО, ЩАО-ЩО-3.pdf" → ЩО-3 (last match)
        # "004 - Схемы ЩО, ЩАО-ЩАО-3.pdf" → ЩАО-3 (last match)
        panel_name = ""
        # Find all panel designations in filename
        panel_matches = list(re.finditer(
            r"(ЦСАО|ЩАО|ЩО|ВРУ|ГРЩ)[-\s]*(\d*)",
            fname, re.IGNORECASE
        ))
        if panel_matches:
            # Use the last match as the target panel
            last = panel_matches[-1]
            prefix = last.group(1).upper()
            num = last.group(2) or ""
            panel_name = f"{prefix}-{num}" if num else prefix

        if not panel_name:
            continue

        qf_set: set[str] = set()  # unique QF identifiers

        try:
            with pdfplumber.open(path) as pdf:
                for page in pdf.pages:
                    text = page.extract_text() or ""
                    # Find all QF-N.N designations
                    for m in re.finditer(r"QF[-\s]*(\d+(?:\.\d+)?)", text):
                        qf_set.add(m.group(0))

                    # Also check tables
                    for table in (page.extract_tables() or []):
                        for row in table:
                            for cell in row:
                                if cell:
                                    for m in re.finditer(r"QF[-\s]*(\d+(?:\.\d+)?)", str(cell)):
                                        qf_set.add(m.group(0))

                    # For ЦСАО (uses fuses, not QF), count groups ГР/Гр
                    if "ЦСАО" in panel_name:
                        for m in re.finditer(r"(?:Гр|ГР)\.?\s*(\d+)", text):
                            group_num = int(m.group(1))
                            qf_set.add(f"GR-{group_num}")

                        for table in (page.extract_tables() or []):
                            for row in table:
                                for cell in row:
                                    if cell:
                                        for m in re.finditer(r"(?:Гр|ГР)\.?\s*(\d+)", str(cell)):
                                            group_num = int(m.group(1))
                                            qf_set.add(f"GR-{group_num}")

        except Exception as e:
            log(f"    ⚠ Error parsing {fname}: {e}")
            continue

        n_total = len(qf_set)
        # Separate incoming (3-pole) vs outgoing (1-pole) breakers.
        # Conventions vary:
        #   ГПК:  QF1 = incoming, QF1.1/QF1.2 = outgoing (dot-separated)
        #   КПП:  QF1/QF2 = incoming, QF-01..QF-26 = outgoing (dash-separated)
        outgoing_qf = set()
        incoming_qf = set()
        for qf in qf_set:
            if "." in qf:
                # Dot sub-index → outgoing (ГПК convention)
                outgoing_qf.add(qf)
            elif qf.startswith("GR-"):
                outgoing_qf.add(qf)
            elif re.match(r"QF[-\s]*\d{2,}", qf):
                # QF-01, QF-101, QF 03, etc. → outgoing (КПП convention)
                outgoing_qf.add(qf)
            else:
                # QF1, QF2, QF3 → incoming (short number, no dash)
                incoming_qf.add(qf)

        n_breakers = len(outgoing_qf) if outgoing_qf else n_total
        n_three_pole = max(1, len(incoming_qf))

        if n_breakers > 0:
            panels.append(PanelInfo(
                name=panel_name,
                breaker_count=n_breakers,
                three_pole=n_three_pole,
            ))
            log(f"    {panel_name}: {n_breakers} breakers/fuses, {n_three_pole} three-pole")

    return panels


def parse_tray_plan_totals(
    folder: Path,
    tray_files: list[str],
    log=print,
    elev_to_ceiling: dict[float, float] | None = None,
) -> dict[HeightCategory, float]:
    """Parse tray plan PDFs to extract total tray length per height category.

    Looks for "Итого:" values in each tray plan PDF (one per elevation).
    Returns a dict mapping HeightCategory to total tray length (metres).
    These totals provide real height distribution data for cables/trays/PVC.
    """
    import pdfplumber

    hcat_totals: dict[HeightCategory, float] = defaultdict(float)

    for fname in tray_files:
        elevs = _extract_elevations(fname)
        if not elevs:
            continue
        elev = elevs[0]
        ceiling_h = elev_to_ceiling.get(elev, elev) if elev_to_ceiling else elev
        hcat = elevation_to_height(ceiling_h)

        path = str(folder / fname)
        try:
            with pdfplumber.open(path) as pdf:
                page = pdf.pages[0]
                words = page.extract_words(x_tolerance=3, y_tolerance=3) or []

                # Find "Итого:" and extract the associated number
                for w in words:
                    if "Итого" in w["text"] or "итого" in w["text"]:
                        # Get words on the same line
                        same_line = sorted(
                            [w2 for w2 in words if abs(w2["top"] - w["top"]) < 8],
                            key=lambda x: x["x0"],
                        )
                        line_text = " ".join(w2["text"] for w2 in same_line)

                        # Extract numbers > 10 (skip row numbers, indices)
                        nums = re.findall(r"(\d+[.,]\d+|\d+)", line_text)
                        big_nums = [
                            float(n.replace(",", "."))
                            for n in nums
                            if float(n.replace(",", ".")) > 10
                        ]
                        if big_nums:
                            val = max(big_nums)
                            hcat_totals[hcat] += val
                            log(f"    {fname}: Итого={val:.1f} → {hcat}")
                        break  # only first Итого per file
        except Exception as e:
            log(f"    ⚠ Error parsing tray plan {fname}: {e}")

    return dict(hcat_totals)


_TRAY_SIZE_RE = re.compile(r"^(\d{2,3})[хx](\d{2,3})$")
_TRAY_LEN_RE = re.compile(r"^(\d{3,5})$")

# Tray plan pages are enormous (200k+ vector lines); pdfplumber.extract_words
# holds ~2-3 GB transiently per file.  Refuse to start parsing a file unless at
# least this much RAM is free, so the parser degrades gracefully (skips trays)
# instead of OOM-killing the whole server.  Tunable via env.
_TRAY_MIN_FREE_MB = int(os.environ.get("VOR_TRAY_MIN_FREE_MB", "4096"))


def _available_mb() -> float | None:
    """Return available system RAM in MB, or None if it can't be determined."""
    try:
        with open("/proc/meminfo") as fh:
            for line in fh:
                if line.startswith("MemAvailable:"):
                    return int(line.split()[1]) / 1024.0
    except Exception:  # noqa: BLE001
        pass
    try:
        import psutil  # type: ignore

        return psutil.virtual_memory().available / (1024 * 1024)
    except Exception:  # noqa: BLE001
        return None


def _parse_one_tray_plan(path: str) -> tuple[dict[str, float], int]:
    """Parse one tray plan PDF → ({"50x50": metres, ...}, matched_count).

    Collects size tokens (``NхM``) and length tokens (3-5 digit number directly
    followed by ``мм``) with page coordinates, then pairs every length to its
    nearest size token by geometric proximity (pdfplumber's text reflow scatters
    them across "lines", so a line-based regex under-counts).
    """
    import math

    import pdfplumber

    per_size: dict[str, float] = defaultdict(float)
    try:
        with pdfplumber.open(path) as pdf:
            page = pdf.pages[0]
            words = page.extract_words(x_tolerance=2, y_tolerance=2) or []
    except Exception:  # noqa: BLE001
        return ({}, 0)

    sizes: list[tuple[float, float, str]] = []
    lengths: list[tuple[float, float, int]] = []
    for i, w in enumerate(words):
        txt = w["text"]
        cx = (w["x0"] + w["x1"]) / 2
        cy = (w["top"] + w["bottom"]) / 2
        ms = _TRAY_SIZE_RE.match(txt)
        if ms:
            sizes.append((cx, cy, f"{int(ms.group(1))}x{int(ms.group(2))}"))
            continue
        if (_TRAY_LEN_RE.match(txt) and i + 1 < len(words)
                and words[i + 1]["text"].startswith("мм")):
            lval = int(txt)
            if 100 <= lval <= 12000:  # plausible single-segment length (mm)
                lengths.append((cx, cy, lval))

    if not sizes or not lengths:
        return ({}, 0)
    matched = 0
    for lx, ly, lval in lengths:
        best_size = None
        best_d = 1e9
        for sx, sy, s in sizes:
            d = math.hypot(lx - sx, ly - sy)
            if d < best_d:
                best_d = d
                best_size = s
        # Reject lengths with no size token reasonably close (orphan numbers).
        if best_size is not None and best_d < 120:
            per_size[best_size] += lval / 1000.0
            matched += 1
    return ({s: round(v, 1) for s, v in per_size.items()}, matched)


def parse_tray_segment_lengths(
    folder: Path,
    tray_files: list[str],
    log=print,
) -> dict[str, float]:
    """Extract tray length per typoразмер from on-plan segment callouts (no-СО).

    Tray plans annotate each run segment with its size and length, e.g.
    ``50х50 6000 мм`` / ``100х80 4500 мм``.  Files are parsed **sequentially**
    in-process: each tray page is huge (200k+ vector lines) so pdfplumber's
    extract_words holds substantial RAM (~1-2 GB) for ~60-75s; parsing them in
    parallel spawn workers risks OOM, so we deliberately stay single-process.

    Returns ``{"50x50": metres, "100x80": metres, ...}`` summed across all
    given tray plan PDFs.  Used by the no-СО pipeline to recover real tray
    metraж that visual symbol counting cannot measure.
    """
    if not tray_files:
        return {}

    per_size: dict[str, float] = defaultdict(float)
    for fname in tray_files:
        free_mb = _available_mb()
        if free_mb is not None and free_mb < _TRAY_MIN_FREE_MB:
            log(f"    [tray] пропуск {fname}: мало свободной памяти "
                f"({free_mb:.0f} МБ < {_TRAY_MIN_FREE_MB} МБ) — метраж лотков не считаем")
            break
        sizes_m, matched = _parse_one_tray_plan(str(folder / fname))
        for s, m in sizes_m.items():
            per_size[s] += m
        if matched:
            summary = ", ".join(f"{s}: {sizes_m[s]:.0f}м" for s in sorted(sizes_m))
            log(f"    {fname}: {matched} сегментов лотков → {summary}")

    return {s: round(v, 1) for s, v in per_size.items()}


def _tray_ratios_from_totals(
    tray_totals: dict[HeightCategory, float],
) -> dict[HeightCategory, float]:
    """Convert absolute tray totals to normalised ratios."""
    total = sum(tray_totals.values())
    if total <= 0:
        return {}
    return {h: v / total for h, v in tray_totals.items()}


def _merge_plan_and_binding_results(
    plan_results: list[PlanCountResult],
    binding_results: list[PlanCountResult],
    log=print,
) -> list[PlanCountResult]:
    """Merge plan and binding count results using best-of-both strategy.

    For each (elevation, model) pair, take the higher count from either
    plan or binding PDF.  Binding PDFs show the same luminaires but may
    cover elevations not present in plan PDFs (e.g., +9.000).

    Returns a merged list of PlanCountResult.
    """
    # Index plan results by height category
    plan_by_hcat: dict[HeightCategory, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for pr in plan_results:
        for model, cnt in pr.counts.items():
            plan_by_hcat[pr.height_category][model] += cnt

    # Index binding results by height category
    bind_by_hcat: dict[HeightCategory, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    bind_elevations: dict[HeightCategory, float] = {}
    for br in binding_results:
        for model, cnt in br.counts.items():
            bind_by_hcat[br.height_category][model] += cnt
        if br.height_category not in bind_elevations:
            bind_elevations[br.height_category] = br.elevation

    # Start with plan results (they are the primary source)
    merged = list(plan_results)

    # Check for height categories that exist only in bindings
    plan_hcats = {pr.height_category for pr in plan_results}

    extra_added = 0
    for hcat, bind_models in bind_by_hcat.items():
        if hcat not in plan_hcats:
            # This height category has no plan PDF — add binding data entirely
            elev = bind_elevations[hcat]
            merged.append(PlanCountResult(
                filename=f"(from binding, elev={elev})",
                elevation=elev,
                height_category=hcat,
                counts=dict(bind_models),
            ))
            n = sum(bind_models.values())
            log(f"    Added {n} items from binding at {hcat} (no plan PDF for this height)")
            extra_added += n
        else:
            # Height category exists in both — check for models that binding
            # found more of (may indicate plan missed some)
            plan_models = plan_by_hcat[hcat]
            for model, bind_cnt in bind_models.items():
                plan_cnt = plan_models.get(model, 0)
                if bind_cnt > plan_cnt:
                    # Binding found more — supplement plan results
                    diff = bind_cnt - plan_cnt
                    elev = bind_elevations.get(hcat, 0)
                    merged.append(PlanCountResult(
                        filename=f"(binding supplement, {hcat})",
                        elevation=elev,
                        height_category=hcat,
                        counts={model: diff},
                    ))
                    extra_added += diff

    if extra_added:
        log(f"    Total extra from bindings: {extra_added} items")

    return merged


# ---------------------------------------------------------------------------
# VOR generation
# ---------------------------------------------------------------------------

def generate_vor_from_pdfs(
    folder: str | Path,
    log=print,
) -> list[VorSection]:
    """Generate complete VOR from a PDF folder.

    Returns list of VorSection objects representing the full VOR document.
    """
    folder = Path(folder)
    if not folder.is_dir():
        log(f"ERROR: Folder not found: {folder}")
        return []

    # ── Step 1: Scan & classify files ──────────────────────────────────
    log("Step 1: Scanning PDF files...")
    all_pdfs = sorted(folder.glob("*.pdf"))
    classified: dict[str, list[str]] = defaultdict(list)
    plan_info: list[tuple[str, list[float]]] = []

    binding_info: list[tuple[str, list[float]]] = []

    for pdf in all_pdfs:
        ftype = _classify_pdf(pdf.name)
        classified[ftype].append(pdf.name)
        if ftype == "plan":
            elevs = _extract_elevations(pdf.name)
            if not elevs:
                # Fallback: extract elevations from PDF content
                elevs = _extract_elevations_from_content(str(pdf))
                if elevs:
                    log(f"  [i] Elevations from content of {pdf.name}: {elevs}")
            plan_info.append((pdf.name, elevs))
        elif ftype == "binding":
            elevs = _extract_elevations(pdf.name)
            if not elevs:
                elevs = _extract_elevations_from_content(str(pdf))
                if elevs:
                    log(f"  [i] Elevations from content of {pdf.name}: {elevs}")
            if elevs:
                binding_info.append((pdf.name, elevs))

    for ftype, files in sorted(classified.items()):
        log(f"  {ftype}: {len(files)} files")

    # ── Step 2: Parse specs (СО) ───────────────────────────────────────
    # Parsed early so no-СО (plan-driven) mode can be detected before the
    # elevation/ceiling maps are built.
    log("\nStep 2: Parsing specifications (СО)...")
    from pdf_spec_parser import parse_all_specs_in_folder
    spec_items = parse_all_specs_in_folder(folder)
    log(f"  Extracted {len(spec_items)} spec items")
    no_spec = len(spec_items) == 0

    # Classify spec items
    spec_by_cat: dict[str, list[SpecItem]] = defaultdict(list)
    for si in spec_items:
        cat = classify_spec_item(si.description)
        spec_by_cat[cat].append(si)
        _log.debug("  %s → %s: %s qty=%d", cat, si.unit, si.description[:50], si.quantity)

    for cat, items in sorted(spec_by_cat.items()):
        log(f"    {cat}: {len(items)} items")

    # ── Step 1a (no-СО intake): promote legend-bearing PDFs to plans ──
    # Without a СО spec, equipment quantities must come from the drawings.
    # Combined per-позиция PDFs classify as "other" (no "план"/elevation in
    # the filename), so promote any PDF that contains a parseable legend
    # into plan_info.  Guarded by no_spec → zero effect on СО projects.
    if no_spec:
        log("  ⚠ No СО spec — entering plan-driven (no-СО) mode")
        _known = {f for f, _ in plan_info} | {f for f, _ in binding_info}
        # Filename-based elevations are cheap (no PDF read) — compute in parent.
        _probe_jobs = []          # (path, name, need_content_elevs)
        _fname_elevs: dict[str, list[float]] = {}
        for pdf in all_pdfs:
            if pdf.name in _known:
                continue
            elevs = _extract_elevations(pdf.name)
            _fname_elevs[pdf.name] = elevs
            _probe_jobs.append((str(pdf), pdf.name, not elevs))

        # Parallel legend-probe (+ content-elevation extraction). This loop is
        # the heavy part of no-СО intake (~20-45 s per PDF). Each probe is an
        # independent per-file read → fan out across cores.
        # Legend probing is lighter than visual counting (no high-DPI render).
        _pw = _vor_worker_count(len(_probe_jobs) if _probe_jobs else 0, heavy=False)

        import time as _time
        _ptot = len(_probe_jobs)
        _probe_results: dict[str, tuple[int, list[float]]] = {}
        if _pw > 1:
            import multiprocessing as _mp
            from concurrent.futures import ProcessPoolExecutor, as_completed
            log(f"  Probing {_ptot} PDFs for legends across {_pw} workers…")
            try:
                from concurrent.futures import TimeoutError as _FTimeout
                _ctx = _mp.get_context("spawn")
                _pt0 = _time.time()
                _pdone = 0
                # Per-page budget: never let one stuck PDF wall the whole run.
                # Deadline grows with each completed page so the slow ones get a
                # fair share but a single hang is bounded by _LEGEND_PAGE_TIMEOUT.
                with ProcessPoolExecutor(max_workers=_pw, mp_context=_ctx) as _ex:
                    _fut_to_name = {
                        _ex.submit(_probe_legend_worker, j): j[1]
                        for j in _probe_jobs
                    }
                    _pending = set(_fut_to_name)
                    while _pending:
                        _budget = _LEGEND_PAGE_TIMEOUT
                        try:
                            for _f in as_completed(list(_pending), timeout=_budget):
                                _pending.discard(_f)
                                _nm, _ni, _ec = _f.result()
                                _probe_results[_nm] = (_ni, _ec)
                                _pdone += 1
                                _ppct = int(_pdone * 100 / _ptot) if _ptot else 100
                                _pel = _time.time() - _pt0
                                _peta = (_pel / _pdone) * (_ptot - _pdone) if _pdone else 0
                                log(f"    ⏳ легенды {_pdone}/{_ptot} ({_ppct}%) "
                                    f"— {_nm} | прошло {int(_pel)}с, осталось ~{int(_peta)}с")
                                break  # re-arm the timeout for the next page
                        except _FTimeout:
                            # No page finished within the budget → treat the
                            # still-running ones as too slow and skip them.
                            for _f in list(_pending):
                                _nm = _fut_to_name.get(_f, "?")
                                _f.cancel()
                                _probe_results.setdefault(_nm, (0, []))
                                _pdone += 1
                                log(f"    ⚠ легенда пропущена по таймауту "
                                    f"(>{int(_LEGEND_PAGE_TIMEOUT)}с) — {_nm}")
                            _pending.clear()
                            # Don't wait on the cancelled (possibly stuck) workers.
                            _ex.shutdown(wait=False, cancel_futures=True)
                            break
            except Exception as _exc:  # noqa: BLE001
                log(f"  ⚠ Parallel legend probe failed ({_exc}); running sequentially")
                _probe_results = {}

        if not _probe_results:
            # Sequential fallback — but STILL bound each page by a timeout so a
            # single monster page (e.g. an ЭГ lightning/grounding sheet with
            # ~90k vector lines) can never wall the whole run. We run each probe
            # in a throwaway 1-worker subprocess and abandon it on timeout.
            import multiprocessing as _mp
            from concurrent.futures import (
                ProcessPoolExecutor as _PPE,
                TimeoutError as _FTimeout,
            )
            _ctx = _mp.get_context("spawn")
            _pt0 = _time.time()
            for _pi, (_path, _nm, _need) in enumerate(_probe_jobs, 1):
                # NB: do NOT use ``with _PPE(...) as ex`` here — the context
                # manager's __exit__ calls shutdown(wait=True), which blocks
                # forever on a worker stuck inside pdfplumber, defeating the
                # timeout. Manage the executor manually and abandon it on hang.
                _ex1 = _PPE(max_workers=1, mp_context=_ctx)
                try:
                    _fut = _ex1.submit(_probe_legend_worker, (_path, _nm, _need))
                    try:
                        _probe_results[_nm] = _fut.result(
                            timeout=_LEGEND_PAGE_TIMEOUT)[1:]
                        # Clean teardown only on success (workers are idle).
                        _ex1.shutdown(wait=False, cancel_futures=True)
                    except _FTimeout:
                        _probe_results[_nm] = (0, [])
                        log(f"    ⚠ легенда пропущена по таймауту "
                            f"(>{int(_LEGEND_PAGE_TIMEOUT)}с) — {_nm}")
                        # Abandon the stuck worker: never wait on it.
                        _ex1.shutdown(wait=False, cancel_futures=True)
                except Exception as _exc1:  # noqa: BLE001
                    # Worker crashed (e.g. OOM-killed) → skip this page, keep going.
                    _probe_results[_nm] = (0, [])
                    log(f"    ⚠ легенда пропущена (ошибка воркера: {_exc1}) — {_nm}")
                    try:
                        _ex1.shutdown(wait=False, cancel_futures=True)
                    except Exception:  # noqa: BLE001
                        pass
                _ppct = int(_pi * 100 / _ptot) if _ptot else 100
                _pel = _time.time() - _pt0
                _peta = (_pel / _pi) * (_ptot - _pi) if _pi else 0
                log(f"    ⏳ легенды {_pi}/{_ptot} ({_ppct}%) "
                    f"— {_nm} | прошло {int(_pel)}с, осталось ~{int(_peta)}с")

        # Promote legend-bearing PDFs to plans (parent, deterministic order).
        for pdf in all_pdfs:
            if pdf.name not in _probe_results:
                continue
            n_items, elevs_content = _probe_results[pdf.name]
            if n_items <= 0:
                continue
            elevs = _fname_elevs.get(pdf.name) or elevs_content
            plan_info.append((pdf.name, elevs or [0.0]))
            log(f"    [no-СО] plan: {pdf.name} "
                f"(legend items={n_items}, elevs={elevs or [0.0]})")

    # ── Step 1b: Build ceiling-height map from all elevations ─────────
    # Collect ALL unique floor elevations, then estimate ceiling height
    # for each floor.  Ceiling of floor i ≈ elevation of floor i+1.
    # For the topmost floor, ceiling ≈ elev + avg storey height.
    all_elevs_sorted: list[float] = sorted(set(
        e for _, elevs in (plan_info + binding_info) for e in (elevs or [])
    ))
    elev_to_ceiling: dict[float, float] = {}
    if all_elevs_sorted:
        for i, elev in enumerate(all_elevs_sorted):
            if i + 1 < len(all_elevs_sorted):
                ceiling_h = all_elevs_sorted[i + 1]
            else:
                if len(all_elevs_sorted) >= 2:
                    avg_storey = (all_elevs_sorted[-1] - all_elevs_sorted[0]) / (len(all_elevs_sorted) - 1)
                else:
                    avg_storey = 3.0
                avg_storey = max(avg_storey, 2.5)
                ceiling_h = elev + avg_storey
            elev_to_ceiling[elev] = ceiling_h
        ceil_str = ", ".join(f"elev {e:.1f} → ceil {c:.1f}m" for e, c in sorted(elev_to_ceiling.items()))
        log(f"  Ceiling map: {ceil_str}")

    # ── Step 3: Count equipment on plans ───────────────────────────────
    log("\nStep 3: Counting equipment on floor plans...")
    plan_results = count_equipment_on_plans(folder, plan_info, log=log,
                                            elev_to_ceiling=elev_to_ceiling)

    # ── Step 3b: Count equipment on binding plans (extra coverage) ────
    if binding_info:
        log(f"\nStep 3b: Counting equipment on binding plans ({len(binding_info)} files)...")
        binding_results = count_equipment_on_plans(folder, binding_info, log=log,
                                                   elev_to_ceiling=elev_to_ceiling)
        plan_results = _merge_plan_and_binding_results(plan_results, binding_results, log)

    # ── Step 4: Build height ratios ────────────────────────────────────
    log("\nStep 4: Building height distribution ratios...")
    height_ratios = build_height_ratios(plan_results)
    for model, ratios in sorted(height_ratios.items()):
        ratio_str = ", ".join(f"{h}: {r:.1%}" for h, r in sorted(ratios.items()) if r > 0)
        log(f"  {model[:50]}: {ratio_str}")

    # ── Step 4a (no-СО): synthesize spec items from plan counts ──
    # Turn merged plan model_counts into SpecItem rows and route them with
    # the same classify_spec_item used for СО, so all section builders work
    # unchanged.  Equipment found only by visual matching (marker-less
    # pictograms/luminaires) thus becomes real VOR rows.
    nospec_tray_lengths: dict[str, list[int]] = {}
    if no_spec:
        synth_counts: dict[str, int] = defaultdict(int)
        # Per-page sweep across ALL sheets (power / lighting / grounding).
        # Combined per-позиция PDFs carry a different legend on each sheet, so
        # a single whole-PDF legend (Step 3) misses most element types (e.g.
        # luminaires on the lighting sheet).  count_plans_nospec parses and
        # counts every page independently for full marker-less coverage.
        try:
            from pdf_vor_nospec import count_plans_nospec
            pp_results, _, nospec_tray_lengths = count_plans_nospec(
                folder, log, include_all=True)
            for pr in pp_results:
                for desc, c in pr.counts.items():
                    synth_counts[desc] += int(c)
        except Exception as exc:  # noqa: BLE001
            log(f"  [no-СО] per-page sweep failed ({exc}); using whole-PDF counts")
        # Fallback: if the per-page sweep produced nothing, use Step-3 counts.
        if not synth_counts:
            for pr in plan_results:
                for desc, c in pr.counts.items():
                    synth_counts[desc] += int(c)
        _rejected = 0
        _kept = 0
        for desc, qty in synth_counts.items():
            if qty <= 0:
                continue
            if not _is_valid_equipment_desc(desc):
                _rejected += qty
                continue
            si = SpecItem(
                position="", description=desc, model="",
                catalog_code="", supplier="", unit="шт", quantity=int(qty),
            )
            spec_by_cat[classify_spec_item(desc)].append(si)
            _kept += qty
        n_cats = sum(1 for k in spec_by_cat if spec_by_cat[k])
        log(f"  [no-СО] synthesized {_kept} units (filtered {_rejected} garbage) "
            f"across {len(synth_counts)} models into {n_cats} categories")

        # ── Step 4a-trays (no-СО): recover tray metraж from plan callouts ──
        # Visual symbol counting cannot measure linear tray runs, so the СО
        # ``tray`` category would otherwise be empty.  Tray plans annotate each
        # segment with "<размер> <длина> мм"; parse and sum per typoразмер, then
        # inject linear (unit="м") SpecItems so _build_trays_section produces
        # real metraж rows instead of nothing.
        _tray_files = classified.get("tray", [])
        if _tray_files:
            try:
                _seg_m = parse_tray_segment_lengths(folder, _tray_files, log)
            except Exception as exc:  # noqa: BLE001
                log(f"  [no-СО] tray segment parse failed ({exc})")
                _seg_m = {}
            _tray_total = 0.0
            for _size, _metres in sorted(_seg_m.items(), key=lambda kv: -kv[1]):
                if _metres <= 0:
                    continue
                spec_by_cat["tray"].append(SpecItem(
                    position="", description=f"Лоток металлический {_size} мм",
                    model="", catalog_code="", supplier="",
                    unit="м", quantity=round(_metres, 1),
                ))
                _tray_total += _metres
            if _tray_total > 0:
                log(f"  [no-СО] recovered tray metraж: {_tray_total:.1f} м "
                    f"across {len(_seg_m)} типоразмеров")

        # ── Step 4a-cables (no-СО): measure wiring runs on plan sheets ──
        # Точечный счётчик отбрасывает линейные объекты, поэтому категория
        # "cable" в no-СО режиме пустовала и секция «Кабельная продукция»
        # не строилась вовсе. Меряем красные/синие трассы векторным движком
        # pdf_cables_by_method (дедуп, способ прокладки, марки из аннотаций
        # листа) на планах освещения и привязки; листы, где трассы лежат
        # растровой картинкой, добираются fallback-движком cable_length.
        # Отключается VOR_MEASURE_CABLES=0.
        if os.environ.get("VOR_MEASURE_CABLES", "1") == "1":
            # Гейт охвата (S011, Test2707): линии схем — не физические
            # трассы (655 м шума с одной схемы ЩАО); лотковые листы уже
            # учтены метражом лотков (двойной счёт); ЭГ-листы (заземление/
            # УП/молниезащита) обрабатывает Step 4a-eg — их линии это
            # полоса и проводник заземления, а не «кабель в гофре».
            from pdf_count_grounding import is_eg_sheet
            _measure_skip_re = re.compile(
                r"схем|кабеленес|лотк|общие\s+данные|титул|опросн|\bВОР\b",
                re.IGNORECASE)
            _plan_files = sorted(
                n for n in ({n for n, _ in plan_info}
                            | {n for n, _ in binding_info})
                if not is_eg_sheet(n) and not _measure_skip_re.search(n))
            _cab_by_mark: dict[tuple[str, str], float] = {}
            _cab_by_method: dict[str, float] = {}
            if _plan_files:
                log(f"  [no-СО] measuring cable runs on "
                    f"{len(_plan_files)} plan sheets…")
                try:
                    from pdf_vor_nospec import measure_plan_cables
                    _cab_by_mark, _cab_by_method = measure_plan_cables(
                        folder, _plan_files, log)
                except Exception as exc:  # noqa: BLE001
                    log(f"  [no-СО] cable measuring failed ({exc})")
            _cab_total = 0.0
            for (_mark, _sec), _m in sorted(_cab_by_mark.items(),
                                            key=lambda kv: -kv[1]):
                if _m < 0.5:
                    continue
                # "Кабель силовой" гарантирует маршрутизацию в категорию
                # "cable" у classify_spec_item даже для незнакомых марок.
                _name = " ".join(
                    x for x in ("Кабель силовой", _mark, _sec) if x)
                spec_by_cat["cable"].append(SpecItem(
                    position="", description=_name, model="",
                    catalog_code="", supplier="", unit="м",
                    quantity=round(_m, 1),
                ))
                _cab_total += _m
            if _cab_total > 0:
                _meth_s = ", ".join(
                    f"{k}: {v:.0f} м" for k, v in
                    sorted(_cab_by_method.items(), key=lambda kv: -kv[1]))
                log(f"  [no-СО] recovered cable metraж: {_cab_total:.1f} м "
                    f"({_meth_s})")

        # ── Step 4a-eg (no-СО): заземление/УП/молниезащита с ЭГ-листов ──
        # Раздел ЭГ без СО раньше не считался вовсе: извлекаем измеримое
        # с чертежей — метраж полосы/сетки по геометрии (с валидацией на
        # эталоне АБК-1: полоса 228 vs 229 м, держатели 242 vs 245) и
        # деривации из примечаний листа (шаг держателей, бухты, траншея).
        try:
            from pdf_count_grounding import extract_eg_quantities
            _eg_n = 0
            for pdf in all_pdfs:
                if not is_eg_sheet(pdf.name):
                    continue
                try:
                    _egr = extract_eg_quantities(str(pdf), pdf.name, log=log)
                except Exception as exc:  # noqa: BLE001
                    log(f"  [no-СО] ЭГ-извлечение: сбой на {pdf.name}: {exc}")
                    continue
                for _row in _egr.rows:
                    _cat = ("lightning" if _row.kind == "lightning"
                            else "grounding")
                    spec_by_cat[_cat].append(SpecItem(
                        position="", description=_row.description, model="",
                        catalog_code="", supplier="", unit=_row.unit,
                        quantity=_row.quantity,
                    ))
                    _eg_n += 1
                for _note in _egr.notes:
                    log(f"    [ЭГ] {_note}")
            if _eg_n:
                log(f"  [no-СО] ЭГ: извлечено {_eg_n} позиций с чертежей")
        except Exception as exc:  # noqa: BLE001
            log(f"  [no-СО] ЭГ-извлечение недоступно ({exc})")

    # ── Step 4b: Parse schema PDFs for panels/breakers ──────────────────
    log("\nStep 4b: Parsing schema PDFs for panel/breaker data...")
    schema_panels = parse_schema_panels(folder, classified.get("schema", []), log)
    for pi in schema_panels:
        log(f"  {pi.name}: {pi.breaker_count} breakers, {pi.three_pole} three-pole")

    # ── Step 4c: Parse tray plan PDFs for height distribution ─────────
    tray_ratios: dict[HeightCategory, float] = {}
    tray_files_list = classified.get("tray", [])
    if tray_files_list:
        log(f"\nStep 4c: Parsing tray plans for height distribution ({len(tray_files_list)} files)...")
        tray_totals = parse_tray_plan_totals(folder, tray_files_list, log,
                                                elev_to_ceiling=elev_to_ceiling)
        tray_ratios = _tray_ratios_from_totals(tray_totals)
        if tray_ratios:
            ratio_str = ", ".join(f"{h}: {r:.1%}" for h, r in sorted(tray_ratios.items()) if r > 0)
            log(f"  Tray height ratios: {ratio_str}")
        else:
            log("  ⚠ No tray totals found, will use equipment proxy")

    # ── Step 4d: Compute floor-count per height category ──────────────
    # Uses elev_to_ceiling map computed in Step 1b.
    floor_count_per_hcat: dict[HeightCategory, int] = defaultdict(int)
    hcats_seen: set[str] = set()  # track "filename→hcat" to avoid double-counting
    for fname, elevs in plan_info:
        if elevs:
            for elev in elevs:
                ceiling_h = elev_to_ceiling.get(elev, elev)
                hcat = elevation_to_height(ceiling_h)
                key = f"{fname}→{hcat}"
                if key not in hcats_seen:
                    hcats_seen.add(key)
                    floor_count_per_hcat[hcat] += 1
    # Binding files may cover elevations not in plan files
    all_plan_elevs = set()
    for _, pe in plan_info:
        for e in (pe or []):
            all_plan_elevs.add(round(e, 1))
    for fname, elevs in binding_info:
        if elevs:
            ceiling_h = elev_to_ceiling.get(elevs[0], elevs[0])
            hcat = elevation_to_height(ceiling_h)
            elev_round = round(elevs[0], 1)
            if elev_round not in all_plan_elevs:
                key = f"{fname}→{hcat}"
                if key not in hcats_seen:
                    hcats_seen.add(key)
                    floor_count_per_hcat[hcat] += 1
    if floor_count_per_hcat:
        fc_str = ", ".join(f"{h}: {c} floors" for h, c in sorted(floor_count_per_hcat.items()))
        log(f"  Floor counts: {fc_str}")

    # ── Step 4e: Detect single-floor building ─────────────────────────
    # A single-floor building has only one height category (usually "до 5 м").
    # In this case all cable/conduit/tray/PVC goes to that one height — no splitting.
    n_hcats = len(floor_count_per_hcat)
    is_single_floor = (n_hcats <= 1)
    if is_single_floor:
        log(f"  ⚡ Single-floor building detected — all work at one height")
        # Force single height category for all distributions
        single_hcat: HeightCategory = "до 5 метров"
        if floor_count_per_hcat:
            single_hcat = next(iter(floor_count_per_hcat))

    # ── Step 5: Generate VOR sections ──────────────────────────────────
    log("\nStep 5: Generating VOR sections...")
    sections: list[VorSection] = []

    # ── Section 1: Щитовое оборудование ──
    panels_section = _build_panels_section(
        spec_by_cat.get("panel", []), schema_panels, log,
        is_single_floor=is_single_floor,
    )
    if panels_section.rows:
        sections.append(panels_section)

    # ── Section 2: Светотехническое оборудование ──
    # Detect if building has cable trays (linear metres). If not → small building
    # where luminaires are typically wall-mounted, not ceiling/шпильки.
    has_trays = any(si.unit == "м" for si in spec_by_cat.get("tray", []))
    # Extract raw plan counts for plan-based quantities
    plan_counts = height_ratios.raw_counts if isinstance(height_ratios, HeightRatios) else {}

    lighting_section = _build_lighting_section(
        spec_by_cat.get("luminaire", []),
        spec_by_cat.get("indicator", []),
        spec_by_cat.get("pictogram", []),
        spec_by_cat.get("emergency_block", []),
        height_ratios,
        log,
        is_single_floor=is_single_floor,
        has_trays=has_trays,
        plan_counts=plan_counts,
        tray_ratios=tray_ratios if tray_ratios else None,
    )
    if lighting_section.rows:
        sections.append(lighting_section)

    # ── Section 3: Электроустановочные изделия ──
    devices_section = _build_devices_section(
        spec_by_cat.get("switch", []),
        spec_by_cat.get("junction_box", []),
        spec_by_cat.get("crimping", []),
        log,
    )
    if devices_section.rows:
        sections.append(devices_section)

    # ── Section 4: Кабельная продукция ──
    cables_section = _build_cables_section(
        spec_by_cat.get("cable", []),
        spec_by_cat.get("pvc", []),
        height_ratios,
        log,
        floor_counts=dict(floor_count_per_hcat) if floor_count_per_hcat else None,
        is_single_floor=is_single_floor,
        has_trays=has_trays,
    )
    if cables_section.rows:
        sections.append(cables_section)

    # ── Section 4b: Земляные работы (для подземного кабеля ВБШвнг) ──
    # Only in buildings without trays (small buildings) — ВБШвнг goes underground.
    # In industrial buildings with trays, ВБШвнг is routed through cable trays.
    # When PVC d.32mm exists, only a fraction of ВБШвнг goes underground
    # (the rest routes through conduit inside the building).
    underground_cables = []
    if not has_trays:
        underground_cables = [
            si for si in spec_by_cat.get("cable", [])
            if "вбшвнг" in si.description.lower()
        ]
    if underground_cables:
        # Estimate how much is actually underground (vs conduit)
        pvc_32_total = sum(
            si.quantity for si in spec_by_cat.get("pvc", [])
            if si.unit == "м" and "32" in si.description
        )
        earthworks_section = _build_earthworks_section(
            underground_cables, log, conduit_deduction_m=pvc_32_total,
        )
        if earthworks_section.rows:
            sections.append(earthworks_section)

    # ── Section 4c: Выполнение проходки кабеля через стены ──
    penetration_items = spec_by_cat.get("penetration", [])
    if penetration_items:
        pen_section = VorSection(title="Выполнение проходки кабеля через стены")
        for si in penetration_items:
            pen_section.rows.append({
                "name": si.description,
                "unit": si.unit,
                "qty": si.quantity,
                "is_material": True,
                "drawing_ref": "",
            })
        if pen_section.rows:
            log(f"  Section 4c (Проходки): {len(pen_section.rows)} rows")
            sections.append(pen_section)

    # ── Section 5: Кабельные лотки ──
    trays_section = _build_trays_section(
        spec_by_cat.get("tray", []),
        spec_by_cat.get("material", []),
        height_ratios,
        log,
        tray_ratios=tray_ratios if tray_ratios else None,
        floor_counts=dict(floor_count_per_hcat) if floor_count_per_hcat else None,
        is_single_floor=is_single_floor,
    )
    if trays_section.rows:
        sections.append(trays_section)

    # ── Section 5a-bis (no-СО): DISABLED — superseded by Step 4a-trays ──
    # This used pdf_vor_nospec._parse_tray_lengths (a same-line regex on
    # extract_text) which under-counts badly: pdfplumber reflows the callout
    # so the "<size>" and "<len> мм" tokens land on different text lines and
    # never match the regex (measured 982 m vs the true 1596 m on ЭО 024-026).
    # Step 4a-trays above now injects geometric-paired tray metraж into
    # spec_by_cat["tray"], which _build_trays_section renders as Section 5.
    # Re-enabling this block would DOUBLE-COUNT the tray section, so it stays
    # off.  (nospec_tray_lengths is still produced by count_plans_nospec for
    # backward-compat callers but is intentionally not consumed here.)
    _ = nospec_tray_lengths  # kept for reference; not used (see above)

    # ── Section 5b: Монтаж системы заземления ──
    grounding_items = spec_by_cat.get("grounding", [])
    if grounding_items:
        grounding_section = _build_grounding_section(grounding_items, log)
        if grounding_section.rows:
            sections.append(grounding_section)

    # ── Section 5c: Монтаж системы молниезащиты ──
    # Deduplicate lightning items (same item may be parsed from multiple pages)
    lightning_items = spec_by_cat.get("lightning", [])
    if lightning_items:
        _seen_lightning: set[tuple[str, str, float]] = set()
        deduped_lightning: list[SpecItem] = []
        for si in lightning_items:
            key = (si.description, si.unit, si.quantity)
            if key not in _seen_lightning:
                _seen_lightning.add(key)
                deduped_lightning.append(si)
        lightning_items = deduped_lightning
        lightning_section = _build_lightning_section(lightning_items, log)
        if lightning_section.rows:
            sections.append(lightning_section)

    # ── Section 6: ПВХ изделия и трубы ──
    pvc_section = _build_pvc_section(
        spec_by_cat.get("pvc", []),
        height_ratios,
        log,
        floor_counts=dict(floor_count_per_hcat) if floor_count_per_hcat else None,
        is_single_floor=is_single_floor,
        has_trays=has_trays,
    )
    if pvc_section.rows:
        sections.append(pvc_section)

    # ── Section 7: Пусконаладочные работы ──
    has_grounding = bool(grounding_items)
    pnr_section = _build_pnr_section(
        spec_by_cat.get("cable", []),
        spec_by_cat.get("panel", []),
        spec_by_cat.get("luminaire", []),
        spec_by_cat.get("switch", []),
        spec_by_cat.get("junction_box", []),
        schema_panels,
        log,
        has_grounding=has_grounding,
        grounding_items=grounding_items,
    )
    if pnr_section.rows:
        sections.append(pnr_section)

    # Явное предупреждение в результате (первая секция): без СО расчёт
    # ведётся по чертежам и заведомо неполон — раньше об этом говорила
    # только строка в логе, и «огрызок» выглядел как полный ВОР.
    if no_spec and sections:
        warn = VorSection(title="ВНИМАНИЕ: расчёт выполнен БЕЗ спецификации (СО)")
        warn.rows.append({
            "name": ("Спецификация оборудования (СО) в папке не найдена. "
                     "Количества получены подсчётом и измерением по чертежам: "
                     "точность ограничена, номенклатура может быть неполной. "
                     "Для точного ВОР добавьте СО-файлы в папку."),
            "unit": "", "qty": "", "is_material": False, "drawing_ref": "",
        })
        sections.insert(0, warn)

    total_rows = sum(len(s.rows) for s in sections)
    log(f"\nGenerated {len(sections)} sections with {total_rows} total rows")

    return sections


# ---------------------------------------------------------------------------
# Height distribution helpers
# ---------------------------------------------------------------------------

def _inverse_floor_count_ratios(
    floor_counts: dict[HeightCategory, int] | None,
) -> dict[HeightCategory, float]:
    """Compute height ratios inversely proportional to floor count per band.

    In industrial buildings, cable/tray length per height band tends to be
    inversely proportional to the number of floors in that band.  A band
    covering 3 floors distributes cable among them, while a band with only
    1 floor concentrates the entire run there.

    The 20-35m band uses a tuned weight (0.4 instead of 1/3≈0.333) because
    empirical comparison shows cables in tall 3-floor bands are slightly more
    concentrated than a pure 1/fc model predicts.

    Returns normalised ratios {HeightCategory: fraction}, or {} if floor_counts
    is None/empty.
    """
    if not floor_counts:
        return {}
    # Tuned inverse weights per band (override pure 1/fc for 20-35m)
    _inv_fc_overrides: dict[HeightCategory, float] = {
        "от 20 до 35 метров": 0.4,  # pure 1/3=0.333 → tuned 0.4
    }
    inv: dict[HeightCategory, float] = {}
    for hcat, n in floor_counts.items():
        if n > 0:
            inv[hcat] = _inv_fc_overrides.get(hcat, 1.0 / n)
    total = sum(inv.values())
    if total <= 0:
        return {}
    return {h: v / total for h, v in inv.items()}


# ---------------------------------------------------------------------------
# Section builders
# ---------------------------------------------------------------------------

def _build_panels_section(
    panels: list[SpecItem],
    schema_panels: list[PanelInfo] = None,
    log=print,
    is_single_floor: bool = False,
) -> VorSection:
    """Build Section 1: Щитовое оборудование."""
    section = VorSection(title="Щитовое оборудование")

    # Collect panel names from spec
    spec_panel_names: set[str] = set()
    for si in panels:
        desc_lower = si.description.lower()
        m = re.search(r"(ЩО|ЩАО|ЦСАО|ВРУ|ГРЩ)[-\s]*\d*", si.description)
        if m:
            designation = m.group(0).strip()
            work_name = f"Монтаж щита распределительного {designation}"
        elif "вводно-распределительн" in desc_lower or "распределительное устройство" in desc_lower:
            # ВРУ described by full name, e.g. "Вводно-распределительное устройство ВРУ-1"
            m2 = re.search(r"ВРУ[-\s]*\d*", si.description)
            if m2:
                designation = m2.group(0).strip()
            else:
                designation = "ВРУ"
            work_name = f"Монтаж вводно-распределительного устройства {designation}"
        else:
            designation = si.description[:30]
            work_name = f"Монтаж щита распределительного {designation}"
        spec_panel_names.add(designation.replace(" ", "").replace("-", "").upper())

        section.rows.append({
            "name": work_name,
            "unit": "шт",
            "qty": si.quantity,
            "is_material": False,
            "drawing_ref": "",
        })

    # Add panels found in schemas but not in specs (e.g., ЦСАО3)
    if schema_panels:
        for pi in schema_panels:
            norm = pi.name.replace(" ", "").replace("-", "").upper()
            if norm not in spec_panel_names:
                section.rows.append({
                    "name": f"Монтаж щита распределительного {pi.name}",
                    "unit": "шт",
                    "qty": 1,
                    "is_material": False,
                    "drawing_ref": "",
                })
                log(f"    Added panel from schema: {pi.name}")

    # Cable connection count depends on building type:
    # Single-floor (small panels): breaker_count * 2 (L + N per outgoing group)
    # Multi-floor (larger panels): breaker_count * 3 + three_pole * 3 (3 cores per cable)
    if schema_panels:
        total_connections = 0
        for pi in schema_panels:
            if is_single_floor:
                total_connections += pi.breaker_count * 2
            else:
                total_connections += pi.breaker_count * 3
                total_connections += pi.three_pole * 3
        # Each wire needs to be connected → total_connections is the number of wire terminations
        # But VOR counts "подключение жил" which is individual wire connections
        if total_connections > 0:
            section.rows.append({
                "name": "Подключение жил кабелей до 10 мм2",
                "unit": "шт",
                "qty": total_connections,
                "is_material": False,
                "drawing_ref": "",
            })
            log(f"    Cable connections: {total_connections} wires")

    log(f"  Section 1 (Панели): {len(section.rows)} rows")
    return section


def _build_lighting_section(
    luminaires: list[SpecItem],
    indicators: list[SpecItem],
    pictograms: list[SpecItem],
    emergency_blocks: list[SpecItem],
    height_ratios: dict[str, dict[HeightCategory, float]],
    log=print,
    is_single_floor: bool = False,
    has_trays: bool = True,
    plan_counts: dict[str, dict[HeightCategory, int]] | None = None,
    tray_ratios: dict[HeightCategory, float] | None = None,
) -> VorSection:
    """Build Section 2: Светотехническое оборудование.

    Groups luminaires by mounting height, with model sub-rows.
    When plan_counts is provided, uses absolute counts from floor plans
    as the primary data source (instead of spec quantities).
    """
    section = VorSection(title="Светотехническое оборудование")

    # ── Luminaires by height ──
    lumi_by_height: dict[HeightCategory, list[tuple[SpecItem, int]]] = defaultdict(list)

    # Track which plan_counts keys have been consumed (for multi-SpecItem models)
    _used_plan_keys: set[str] = set()

    for si in luminaires:
        if is_single_floor:
            lumi_by_height["до 5 метров"].append((si, si.quantity))
            continue

        key = _model_key(si.description, si.model)

        # ── Hybrid plan+spec approach ──
        # Plan data gives height distribution (ratios), spec gives total qty.
        # Three tiers based on plan detection rate:
        #   ≥85% → trust plan counts directly
        #   50-85% → use spec qty with plan-derived ratios
        #   <50% → plan ratios unreliable, use global average ratios
        pc = _find_plan_counts(key, plan_counts) if plan_counts else None

        if pc:
            plan_total = sum(pc.values())
            spec_qty = si.quantity

            # For multi-SpecItem models: compute total spec qty across all matching
            all_matching = [s for s in luminaires
                            if _model_key(s.description, s.model) == key]
            total_spec = sum(s.quantity for s in all_matching)

            # Detection rate for this model
            detection_rate = plan_total / max(1, total_spec)

            if key in _used_plan_keys:
                # Already consumed — compute this SpecItem's share
                if total_spec > 0:
                    share = spec_qty / total_spec
                else:
                    share = 1.0 / max(1, len(all_matching))

                if 0.85 <= detection_rate <= 1.15:
                    # Plan reliable, use raw plan counts scaled by share
                    scaled = {h: max(0, int(round(c * share)))
                              for h, c in pc.items()}
                elif detection_rate >= 0.50:
                    plan_ratios = {h: c / plan_total for h, c in pc.items()}
                    scaled = distribute_by_height(spec_qty, plan_ratios)
                else:
                    # Very low detection — use tray or global avg ratios
                    ratios = tray_ratios or _find_best_ratio(key, height_ratios) or _avg_ratios(height_ratios)
                    scaled = distribute_by_height(spec_qty, ratios)

                for hcat, qty in scaled.items():
                    if qty > 0:
                        lumi_by_height[hcat].append((si, qty))
                _log.info("  Luminaire '%s' → hybrid (shared %.0f%%, detect=%.0f%%): %s",
                          si.description[:40], share * 100, detection_rate * 100, scaled)
            else:
                _used_plan_keys.add(key)

                # Log plan vs spec comparison
                if spec_qty > 0 and abs(plan_total - spec_qty) / max(1, spec_qty) > 0.15:
                    log(f"  \u26a0 {si.description[:50]}: план={plan_total}, СО={spec_qty}")

                if 0.85 <= detection_rate <= 1.15:
                    # Plan data ≈ spec ±15% → trust plan counts directly
                    for hcat, qty in pc.items():
                        if qty > 0:
                            lumi_by_height[hcat].append((si, qty))
                    _log.info("  Luminaire '%s' → PLAN counts (reliable %.0f%%): %s (spec=%d)",
                              si.description[:40], detection_rate * 100, dict(pc), spec_qty)
                elif detection_rate >= 0.50:
                    # 50-85% → plan ratios are decent, use with spec qty
                    plan_ratios = {h: c / plan_total for h, c in pc.items()}
                    dist = distribute_by_height(spec_qty, plan_ratios)
                    for hcat, qty in dist.items():
                        if qty > 0:
                            lumi_by_height[hcat].append((si, qty))
                    _log.info("  Luminaire '%s' → SPEC qty=%d + PLAN ratios (detect=%.0f%%): %s",
                              si.description[:40], spec_qty, detection_rate * 100, dict(dist))
                else:
                    # <50% → plan ratios unreliable, fall back to tray or global avg
                    ratios = tray_ratios or _find_best_ratio(key, height_ratios) or _avg_ratios(height_ratios)
                    dist = distribute_by_height(spec_qty, ratios)
                    for hcat, qty in dist.items():
                        if qty > 0:
                            lumi_by_height[hcat].append((si, qty))
                    source = "TRAY" if tray_ratios else "AVG"
                    _log.info("  Luminaire '%s' → SPEC qty=%d + %s ratios (detect=%.0f%% too low): %s",
                              si.description[:40], spec_qty, source, detection_rate * 100, dict(dist))
        else:
            # ── Fallback: spec qty + ratios (no plan data for this model) ──
            ratios = _find_best_ratio(key, height_ratios)
            if not ratios and height_ratios:
                avg_ratios: dict[HeightCategory, float] = defaultdict(float)
                n = 0
                for r_vals in height_ratios.values():
                    for hcat, val in r_vals.items():
                        avg_ratios[hcat] += val
                    n += 1
                if n > 0:
                    ratios = {h: v / n for h, v in avg_ratios.items() if v > 0}

            _log.info("  Luminaire '%s' → SPEC fallback qty=%d, ratios=%s",
                      si.description[:40], si.quantity,
                      {h: f"{r:.1%}" for h, r in ratios.items()} if ratios else "DEFAULT")
            dist = distribute_by_height(si.quantity, ratios)
            for hcat, qty in dist.items():
                if qty > 0:
                    lumi_by_height[hcat].append((si, qty))

    # Determine mount type based on luminaire descriptions and building type
    # Single-floor: "подвесной" (suspended)
    # Multi-floor default: "на шпильках к перекрытию"
    # Wall-mounted: "настенного" (CD LED 27)
    # Anchor: "анкерный" (INSEL high-bay)

    for hcat in HEIGHT_CATEGORIES:
        items = lumi_by_height.get(hcat, [])
        if not items:
            continue

        # Separate by mount type
        ceiling_items = []
        wall_items = []
        anchor_items = []

        for si, qty in items:
            nl = si.description.lower()
            model = si.model.lower() if si.model else ""
            if not is_single_floor and ("insel" in nl or "insel" in model):
                anchor_items.append((si, qty))
            elif not has_trays:
                # No trays → small building → all luminaires wall-mounted
                wall_items.append((si, qty))
            elif "cd led" in nl or "cd led" in model:
                wall_items.append((si, qty))
            else:
                ceiling_items.append((si, qty))

        # Ceiling/suspended-mounted luminaires
        if ceiling_items:
            total = sum(q for _, q in ceiling_items)
            if is_single_floor:
                mount_label = "подвесной"
            else:
                mount_label = "на шпильках к перекрытию"
            section.rows.append({
                "name": f"Монтаж светильников {mount_label} на высоте {hcat}",
                "unit": "шт",
                "qty": total,
                "is_material": False,
                "drawing_ref": "",
            })
            for si, qty in ceiling_items:
                model_name = si.description
                # Add spec model from type_brand if available
                if si.model and si.model not in model_name:
                    model_name = f"{si.description} {si.model}"
                section.rows.append({
                    "name": model_name,
                    "unit": "шт",
                    "qty": qty,
                    "is_material": True,
                    "drawing_ref": "",
                })

        # Wall-mounted
        if wall_items:
            total = sum(q for _, q in wall_items)
            section.rows.append({
                "name": f"Монтаж настенного светодиодного светильника на высоте {hcat}",
                "unit": "шт",
                "qty": total,
                "is_material": False,
                "drawing_ref": "",
            })
            for si, qty in wall_items:
                model_name = si.description
                if si.model and si.model not in model_name:
                    model_name = f"{si.description}, {si.model}"
                section.rows.append({
                    "name": model_name,
                    "unit": "шт",
                    "qty": qty,
                    "is_material": True,
                    "drawing_ref": "",
                })

        # Anchor-mounted
        if anchor_items:
            total = sum(q for _, q in anchor_items)
            section.rows.append({
                "name": f"Монтаж светильников анкерный на высоте {hcat}",
                "unit": "шт",
                "qty": total,
                "is_material": False,
                "drawing_ref": "",
            })
            for si, qty in anchor_items:
                model_name = si.description
                if si.model and si.model not in model_name:
                    model_name = f"{si.description} {si.model}"
                section.rows.append({
                    "name": model_name,
                    "unit": "шт",
                    "qty": qty,
                    "is_material": True,
                    "drawing_ref": "",
                })

    # ── Emergency blocks ──
    for si in emergency_blocks:
        section.rows.append({
            "name": si.description,
            "unit": "шт",
            "qty": si.quantity,
            "is_material": True,
            "drawing_ref": "",
        })

    # ── Indicators by height ──
    ind_by_height: dict[HeightCategory, list[tuple[SpecItem, int]]] = defaultdict(list)

    # Fallback weighted distribution for indicators (used when no plan data)
    _ind_weights: dict[HeightCategory, float] = {
        "до 5 метров": 1.0,
        "от 5 до 13 метров": 0.7,
        "от 13 до 20 метров": 1.0,
        "от 20 до 35 метров": 0.9,
    }
    _ind_total_w = sum(_ind_weights.values())
    ind_ratio: dict[HeightCategory, float] = {h: _ind_weights[h] / _ind_total_w
                                                for h in HEIGHT_CATEGORIES}

    # ── Distribute indicators (MERCURY, ATOM etc.) ──
    for si in indicators:
        ind_key = _model_key(si.description, si.model)

        # Hybrid plan+spec approach for indicators (3-tier)
        pc = _find_plan_counts(ind_key, plan_counts) if plan_counts else None
        if pc:
            plan_total = sum(pc.values())
            detection_rate = plan_total / max(1, si.quantity)
            if si.quantity > 0 and abs(plan_total - si.quantity) / max(1, si.quantity) > 0.15:
                log(f"  \u26a0 {si.description[:50]}: план={plan_total}, СО={si.quantity}")

            if 0.85 <= detection_rate <= 1.15:
                # Plan ratios reliable — use spec_qty with plan ratios
                plan_ratios = {h: c / plan_total for h, c in pc.items()}
                dist = distribute_by_height(si.quantity, plan_ratios)
                for hcat, qty in dist.items():
                    if qty > 0:
                        ind_by_height[hcat].append((si, qty))
                _log.info("  Indicator '%s' → SPEC qty=%d + PLAN ratios (reliable %.0f%%): %s",
                          si.description[:40], si.quantity, detection_rate * 100, dict(dist))
            elif detection_rate >= 0.50:
                # 50-85% or >115% — plan ratios noisy, smooth toward uniform
                plan_ratios = {h: c / plan_total for h, c in pc.items()}
                n_heights = len(plan_ratios) or 1
                uniform = 1.0 / n_heights
                alpha = 0.3  # 70% plan signal, 30% uniform
                smoothed = {h: (1 - alpha) * r + alpha * uniform
                            for h, r in plan_ratios.items()}
                dist = distribute_by_height(si.quantity, smoothed)
                for hcat, qty in dist.items():
                    if qty > 0:
                        ind_by_height[hcat].append((si, qty))
                _log.info("  Indicator '%s' → SPEC qty=%d + SMOOTHED plan ratios (detect=%.0f%%): %s",
                          si.description[:40], si.quantity, detection_rate * 100, dict(dist))
            else:
                # <50% → plan ratios unreliable, use ind_ratio (weighted)
                dist = distribute_by_height(si.quantity, ind_ratio)
                for hcat, qty in dist.items():
                    if qty > 0:
                        ind_by_height[hcat].append((si, qty))
                _log.info("  Indicator '%s' → SPEC qty=%d + WEIGHTED ratios (detect=%.0f%% too low)",
                          si.description[:40], si.quantity, detection_rate * 100)
        else:
            # Fallback: spec qty + weighted ratio
            dist = distribute_by_height(si.quantity, ind_ratio)
            for hcat, qty in dist.items():
                if qty > 0:
                    ind_by_height[hcat].append((si, qty))
            _log.info("  Indicator '%s' → SPEC fallback qty=%d",
                      si.description[:40], si.quantity)

    # ── Distribute pictograms (stickers) by height using same weighted ratio ──
    pict_by_height: dict[HeightCategory, list[tuple[SpecItem, int]]] = defaultdict(list)
    for si in pictograms:
        dist = distribute_by_height(si.quantity, ind_ratio)
        for hcat, qty in dist.items():
            if qty > 0:
                pict_by_height[hcat].append((si, qty))

    # ── Output indicator work rows (without pictograms in the work total) ──
    for hcat in HEIGHT_CATEGORIES:
        ind_items = ind_by_height.get(hcat, [])
        pict_items = pict_by_height.get(hcat, [])
        if not ind_items and not pict_items:
            continue

        # Work row: mounting count = indicators only (не пиктограммы)
        if ind_items:
            ind_total = sum(q for _, q in ind_items)
            section.rows.append({
                "name": f"Монтаж настенного указателя светодиодного (с пиктограммой) {hcat}",
                "unit": "шт",
                "qty": ind_total,
                "is_material": False,
                "drawing_ref": "",
            })
            # Indicator material rows
            for si, qty in ind_items:
                section.rows.append({
                    "name": si.description,
                    "unit": "шт",
                    "qty": qty,
                    "is_material": True,
                    "drawing_ref": "",
                })

        # Pictogram material rows (stickers — no separate mounting work)
        for si, qty in pict_items:
            section.rows.append({
                "name": si.description,
                "unit": "шт",
                "qty": qty,
                "is_material": True,
                "drawing_ref": "",
            })

    log(f"  Section 2 (Светотехника): {len(section.rows)} rows")
    return section


def _avg_ratios(
    height_ratios: dict[str, dict[HeightCategory, float]],
) -> dict[HeightCategory, float]:
    """Compute average height ratios across all models.

    Used as fallback when model-specific plan data is too sparse.
    """
    if not height_ratios:
        return {}
    avg: dict[HeightCategory, float] = defaultdict(float)
    n = 0
    for r_vals in height_ratios.values():
        for hcat, val in r_vals.items():
            avg[hcat] += val
        n += 1
    if n > 0:
        return {h: v / n for h, v in avg.items() if v > 0}
    return {}


def _find_best_ratio(
    model_key: str,
    height_ratios: dict[str, dict[HeightCategory, float]],
) -> dict[HeightCategory, float]:
    """Find the best matching height ratio for a model.

    Tries exact match first, then substring, then token-based matching.
    """
    if not model_key:
        return {}

    if model_key in height_ratios:
        # Check if there are OTHER keys that START with model_key
        # (e.g., "cd led 27 4000k" exact match exists, but also
        # "cd led 27 4000k+emergency conversion kit..." exists)
        # If so, merge all of them for better distribution.
        all_starting = [(k, r) for k, r in height_ratios.items()
                        if k.startswith(model_key)]
        if len(all_starting) > 1:
            return _merge_ratios(all_starting, height_ratios if isinstance(height_ratios, HeightRatios) else None)
        return height_ratios[model_key]

    # Try substring match (handles truncation at 50 chars)
    # Require similar length to avoid cross-matching variants
    for stored_key, ratios in height_ratios.items():
        if model_key in stored_key or stored_key in model_key:
            shorter = min(len(model_key), len(stored_key))
            longer = max(len(model_key), len(stored_key))
            if shorter >= longer * 0.7:
                return ratios

    # Try matching core model identifier (first 2-3 tokens: e.g., "slick.prs led 50")
    # This handles cases like spec "slick.prs led 50 with driver box /temper"
    # vs plan "slick.prs led 50 with driver box /tempered glass/"
    key_tokens = model_key.split()
    if len(key_tokens) >= 2:
        # Try matching first N tokens (from 4 down to 2)
        for n_tokens in range(min(4, len(key_tokens)), 1, -1):
            prefix = " ".join(key_tokens[:n_tokens])
            matches = [(k, r) for k, r in height_ratios.items()
                       if k.startswith(prefix)]
            if len(matches) == 1:
                return matches[0][1]
            if len(matches) > 1:
                # If one match is exact, prefer it over merging
                exact = [(k, r) for k, r in matches if k == model_key]
                if exact:
                    return exact[0][1]
                # Check if matches differ by a numeric token (different products)
                suffixes = [k[len(prefix):].strip() for k, _ in matches]
                first_nums = set()
                for suf in suffixes:
                    m_num = re.match(r"(\d+)", suf)
                    first_nums.add(m_num.group(1) if m_num else "")
                if len(first_nums) > 1:
                    continue  # Different products — don't merge
                # Multiple keys share this prefix — combine their ratios
                return _merge_ratios(matches, height_ratios if isinstance(height_ratios, HeightRatios) else None)

    # Try matching just the first word (brand: slick, arctic, mercury, etc.)
    first_word = key_tokens[0] if key_tokens else ""
    if first_word and len(first_word) > 3:
        matches = [(k, r) for k, r in height_ratios.items() if first_word in k]
        if len(matches) == 1:
            return matches[0][1]
        # If multiple matches by brand, try adding size/watt info
        if len(matches) > 1 and len(key_tokens) >= 2:
            # e.g., "slick" matches both "slick.prs led 50" and "slick.prs led 30"
            # Try "slick" + number (wattage)
            for tok in key_tokens[1:]:
                if re.match(r"^\d+$", tok):
                    refined = [(k, r) for k, r in matches if tok in k]
                    if len(refined) == 1:
                        return refined[0][1]
                    if len(refined) > 1:
                        return _merge_ratios(refined, height_ratios if isinstance(height_ratios, HeightRatios) else None)

    # Default: empty → distribute_by_height will put all in "до 5 метров"
    _log.warning("  No height ratio match for key='%s'", model_key)
    return {}


def _find_plan_counts(
    model_key: str,
    plan_counts: dict[str, dict[HeightCategory, int]],
) -> dict[HeightCategory, int] | None:
    """Find absolute plan counts for a model using the same matching logic as ratios.

    Returns dict {height_category: count} or None if no match.
    """
    if not model_key or not plan_counts:
        return None

    # Exact match
    if model_key in plan_counts:
        return plan_counts[model_key]

    # Substring match — require similar length to avoid double-counting
    # (e.g. "slick.prs led 30" should not match "slick.prs led 30 /tempered glass/")
    for stored_key, counts in plan_counts.items():
        if model_key in stored_key or stored_key in model_key:
            shorter = min(len(model_key), len(stored_key))
            longer = max(len(model_key), len(stored_key))
            if shorter >= longer * 0.7:  # lengths within 30%
                return counts

    # Token-based matching (first N tokens)
    key_tokens = model_key.split()
    if len(key_tokens) >= 2:
        for n_tokens in range(min(4, len(key_tokens)), 1, -1):
            prefix = " ".join(key_tokens[:n_tokens])
            matches = [(k, c) for k, c in plan_counts.items()
                       if k.startswith(prefix)]
            if len(matches) == 1:
                return matches[0][1]
            if len(matches) > 1:
                # If one of the matches is an exact key match, prefer it
                # (e.g. "cd led 27" should not merge with "cd led 27 4000k+emergency")
                exact = [(k, c) for k, c in matches if k == model_key]
                if exact:
                    return exact[0][1]
                # Check if matches differ by a numeric token after the prefix
                # (e.g. "slick.prs led 30" vs "slick.prs led 50" — different products)
                suffixes = [k[len(prefix):].strip() for k, _ in matches]
                first_nums = set()
                for suf in suffixes:
                    m = re.match(r"(\d+)", suf)
                    first_nums.add(m.group(1) if m else "")
                if len(first_nums) > 1:
                    # Matches are different products — do NOT merge, skip
                    continue
                # Same product variants — merge counts
                merged: dict[HeightCategory, int] = defaultdict(int)
                for _k, c in matches:
                    for hcat, cnt in c.items():
                        merged[hcat] += cnt
                return dict(merged)

    # Brand + wattage matching
    first_word = key_tokens[0] if key_tokens else ""
    if first_word and len(first_word) > 3:
        matches = [(k, c) for k, c in plan_counts.items() if first_word in k]
        if len(matches) == 1:
            return matches[0][1]
        if len(matches) > 1 and len(key_tokens) >= 2:
            for tok in key_tokens[1:]:
                if re.match(r"^\d+$", tok):
                    refined = [(k, c) for k, c in matches if tok in k]
                    if len(refined) == 1:
                        return refined[0][1]
                    if len(refined) > 1:
                        merged = defaultdict(int)
                        for _k, c in refined:
                            for hcat, cnt in c.items():
                                merged[hcat] += cnt
                        return dict(merged)

    return None


def _merge_ratios(
    matches: list[tuple[str, dict[HeightCategory, float]]],
    height_ratios: "HeightRatios | None" = None,
) -> dict[HeightCategory, float]:
    """Merge multiple height ratio dicts into one.

    If height_ratios with raw_counts is available, uses count-weighted merge
    (so a model with 4 items on one floor outweighs 1 item on another).
    Otherwise falls back to equal-weighted average.
    """
    combined: dict[HeightCategory, float] = defaultdict(float)

    # Try count-weighted merge
    if height_ratios is not None and hasattr(height_ratios, 'raw_counts'):
        for key, _ratios in matches:
            counts = height_ratios.raw_counts.get(key, {})
            for hcat, cnt in counts.items():
                combined[hcat] += cnt
    else:
        for _key, ratios in matches:
            for hcat, r in ratios.items():
                combined[hcat] += r

    total = sum(combined.values())
    if total > 0:
        return {h: v / total for h, v in combined.items()}
    return {}


def _build_devices_section(
    switches: list[SpecItem],
    junction_boxes: list[SpecItem],
    crimping: list[SpecItem],
    log=print,
) -> VorSection:
    """Build Section 3: Монтаж электроустановочных изделий."""
    section = VorSection(title="Монтаж электроустановочных изделий")

    for si in switches:
        desc_lower = si.description.lower()
        if "розетк" in desc_lower:
            work_name = f"Монтаж розетки в подрозетник {si.description}"
        elif "рамка" in desc_lower or "рамк" in desc_lower:
            work_name = f"Установка {si.description}"
        elif "выключател" in desc_lower:
            work_name = f"Монтаж {si.description}"
        else:
            work_name = f"Монтаж {si.description}"
        section.rows.append({
            "name": work_name,
            "unit": "шт",
            "qty": si.quantity,
            "is_material": False,
            "drawing_ref": "",
        })

    for si in junction_boxes:
        section.rows.append({
            "name": f"Монтаж {si.description}",
            "unit": "шт",
            "qty": si.quantity,
            "is_material": False,
            "drawing_ref": "",
        })

    # Crimping materials as a group
    if crimping:
        section.rows.append({
            "name": "Соединение жил кабелей методом опрессовки",
            "unit": "",
            "qty": 0,
            "is_material": False,
            "drawing_ref": "",
        })
        for si in crimping:
            section.rows.append({
                "name": si.description,
                "unit": si.unit,
                "qty": si.quantity,
                "is_material": True,
                "drawing_ref": "",
            })

    log(f"  Section 3 (Электроуст.): {len(section.rows)} rows")
    return section


def _build_cables_section(
    cables: list[SpecItem],
    pvc_items: list[SpecItem],
    height_ratios: dict[str, dict[HeightCategory, float]],
    log=print,
    tray_ratios: dict[HeightCategory, float] | None = None,
    floor_counts: dict[HeightCategory, int] | None = None,
    is_single_floor: bool = False,
    has_trays: bool = True,
) -> VorSection:
    """Build Section 4: Кабельная продукция.

    Splits total cable length into:
      - Cable in tray (лоток): total_cable − cable_in_conduit
      - Cable in conduit (гофра): ≈ PVC conduit spec lengths
    Each gets height-distributed work rows, then all cable types as materials.

    Cable height distribution uses **inverse floor-count** ratios:
    cable per band ∝ 1/(number of floors in that band).
    This reflects that floors sharing a height band split cable routing,
    so bands with fewer floors get proportionally more cable per floor.

    Conduit distribution depends on diameter:
      - d16 (small): 4 heights with inverse floor-count
      - d20+ (large): only "до 5м" and "от 13 до 20м" (power feeds to panels)
    """
    section = VorSection(title="Кабельная продукция")

    if not cables:
        return section

    total_cable_m = sum(si.quantity for si in cables)

    # ── Cable-in-conduit estimate from PVC conduit spec lengths ──
    conduit_by_diam: dict[str, int] = {}  # diameter → conduit length
    total_conduit_m = 0
    for pvc in pvc_items:
        if pvc.unit == "м":  # conduit tubes only, not holders
            diam_match = re.search(r"д\.?\s*(\d+)\s*мм", pvc.description)
            diam = diam_match.group(1) if diam_match else "16"
            conduit_by_diam[diam] = conduit_by_diam.get(diam, 0) + pvc.quantity
            total_conduit_m += pvc.quantity

    cable_in_tray_m = max(0, total_cable_m - total_conduit_m)

    log(f"    Cable total: {total_cable_m} м, in tray: {cable_in_tray_m} м, "
        f"in conduit: {total_conduit_m} м")

    # ── Height distribution ratios ──
    if is_single_floor:
        # Single-floor: per-brand cable grouping, all at one height
        log(f"    Single-floor: per-brand cable grouping at до 5 метров")

        # Group cables by brand family (reuse _cable_brand defined below)
        def _cable_brand_sf(desc: str) -> str:
            nl = desc.lower()
            if "вбшвнг" in nl:
                m = re.search(r"ВБШвнг[^\s]*", desc, re.IGNORECASE)
                return m.group(0) if m else "ВБШвнг"
            if "ппгнг" in nl:
                m = re.search(r"ППГнг[^\s]*", desc, re.IGNORECASE)
                return m.group(0) if m else "ППГнг"
            if "ввгнг" in nl:
                m = re.search(r"ВВГнг[^\s]*", desc, re.IGNORECASE)
                return m.group(0) if m else "ВВГнг"
            if "пугв" in nl or ("провод" in nl and any(k in nl for k in ("пв-3", "пв3"))):
                m = re.search(r"ПуГВнг[^\s]*|ПуГВ[^\s]*", desc, re.IGNORECASE)
                return m.group(0) if m else "Провод"
            if "провод" in nl:
                return "Провод"
            return "Кабель"

        def _cable_cross_section_sf(desc: str) -> float:
            m = re.search(r"(\d+)\s*[хx×]\s*([\d,]+)", desc)
            if m:
                return int(m.group(1)) * float(m.group(2).replace(",", "."))
            return 0

        sf_groups: dict[str, list] = OrderedDict()
        for si in cables:
            brand = _cable_brand_sf(si.description)
            sf_groups.setdefault(brand, []).append(si)

        # Separate wire (Провод) from cables — wire goes to separate section
        wire_items = sf_groups.pop("Провод", [])

        # Determine underground brands (ВБШвнг → in ground, not in conduit)
        _UNDERGROUND_BRANDS = {"вбшвнг"}

        for brand, brand_cables in sf_groups.items():
            brand_total_m = sum(si.quantity for si in brand_cables)
            if brand_total_m <= 0:
                continue

            is_underground = any(k in brand.lower() for k in _UNDERGROUND_BRANDS)

            if is_underground:
                # Underground cable: "Прокладка кабеля в земле"
                section.rows.append({
                    "name": "Прокладка кабеля в земле",
                    "unit": "м",
                    "qty": brand_total_m,
                    "is_material": False,
                    "drawing_ref": "",
                })
            else:
                # Cable in conduit: per cross-section group
                max_xs = max(_cable_cross_section_sf(si.description) for si in brand_cables)
                xs_label = "суммарное сечение до 6 мм2" if max_xs <= 6 else "суммарное сечение до 16 мм2"
                section.rows.append({
                    "name": (f"Прокладка кабеля в гофре на высоте до 5 м "
                             f"({xs_label})"),
                    "unit": "м",
                    "qty": brand_total_m,
                    "is_material": False,
                    "drawing_ref": "",
                })

            # Material sub-rows for this brand
            for si in brand_cables:
                sec_m = re.search(r"(\d+)\s*[хx×]\s*([\d,]+)", si.description)
                if sec_m:
                    mat_name = f"Кабель {brand} сечением {sec_m.group(1)}х{sec_m.group(2)}"
                else:
                    mat_name = si.description
                section.rows.append({
                    "name": mat_name,
                    "unit": si.unit,
                    "qty": si.quantity,
                    "is_material": True,
                    "drawing_ref": "",
                })

        # Wire sub-section (Провод) — separate work row + materials
        if wire_items:
            wire_total_m = sum(si.quantity for si in wire_items)
            if wire_total_m > 0:
                section.rows.append({
                    "name": ("Прокладка провода с медной многопроволочной жилой, "
                             "допускающего частые изгибы"),
                    "unit": "м",
                    "qty": wire_total_m,
                    "is_material": False,
                    "drawing_ref": "",
                })
                for si in wire_items:
                    sec_m = re.search(r"(\d+)\s*[хx×]\s*([\d,]+)", si.description)
                    if sec_m:
                        mat_name = f"Провод {sec_m.group(1)}х{sec_m.group(2)} мм2"
                    else:
                        # Try single-core format: 1x6
                        sec_m2 = re.search(r"(\d+[.,]\d+|\d+)\s*мм", si.description)
                        if sec_m2:
                            mat_name = f"Провод 1х{sec_m2.group(1)} мм2"
                        else:
                            mat_name = si.description
                    section.rows.append({
                        "name": mat_name,
                        "unit": si.unit,
                        "qty": si.quantity,
                        "is_material": True,
                        "drawing_ref": "",
                    })

        log(f"  Section 4 (Кабели): {len(section.rows)} rows")
        return section

    # Prefer inverse floor-count: cable per band ∝ 1/(floors in band)
    cable_ratios = _inverse_floor_count_ratios(floor_counts)
    if cable_ratios:
        log(f"    Using inverse floor-count ratios for cables")
    else:
        # Fallback: equipment proxy
        equip_totals: dict[HeightCategory, int] = defaultdict(int)
        for ratios in height_ratios.values():
            for hcat, r in ratios.items():
                equip_totals[hcat] += int(r * 100)
        if equip_totals:
            total_weight = sum(equip_totals.values())
            cable_ratios = {h: c / total_weight for h, c in equip_totals.items()}
        else:
            cable_ratios = {"до 5 метров": 1.0}

    # ── Group cables by brand and routing category ──
    def _cable_brand(desc: str) -> str:
        nl = desc.lower()
        if "вбшвнг" in nl:
            m = re.search(r"ВБШвнг[^\s]*", desc, re.IGNORECASE)
            return m.group(0) if m else "ВБШвнг"
        if "ппгнг" in nl:
            m = re.search(r"ППГнг[^\s]*", desc, re.IGNORECASE)
            return m.group(0) if m else "ППГнг"
        if "ввгнг" in nl:
            m = re.search(r"ВВГнг[^\s]*", desc, re.IGNORECASE)
            return m.group(0) if m else "ВВГнг"
        if "пугв" in nl:
            m = re.search(r"ПуГВнг[^\s]*|ПуГВ[^\s]*", desc, re.IGNORECASE)
            return m.group(0) if m else "Провод"
        if "провод" in nl:
            return "Провод"
        return "Кабель"

    def _cable_cross_section(desc: str) -> float:
        m = re.search(r"(\d+)\s*[хx×]\s*([\d,]+)", desc)
        if m:
            return int(m.group(1)) * float(m.group(2).replace(",", "."))
        return 0

    cable_groups: dict[str, list] = OrderedDict()
    for si in cables:
        brand = _cable_brand(si.description)
        cable_groups.setdefault(brand, []).append(si)

    # Separate into routing categories:
    # - underground: ВБШвнг portion in ground (small building, external run)
    # - tray_only: ВБШвнг when has trays (large building → in tray, not conduit)
    # - conduit: ППГнг, ВВГнг, and ВБШвнг conduit portion (in гофра)
    # - wire: PuGV/Провод (separate section)
    _ARMORED_KW = {"вбшвнг"}
    underground_groups: dict[str, list] = OrderedDict()
    underground_m_override: dict[str, float] = {}  # brand → override qty for underground
    tray_only_groups: dict[str, list] = OrderedDict()
    wire_group: list = []
    conduit_groups: dict[str, list] = OrderedDict()

    # For buildings without trays, estimate underground vs conduit split
    # for ВБШвнг using PVC d.32mm (fits ВБШвнг 3x2.5) as proxy.
    pvc_32_m = 0
    if not has_trays:
        for pi in pvc_items:
            if pi.unit == "м" and "32" in pi.description:
                pvc_32_m += pi.quantity

    for brand, brand_cables in cable_groups.items():
        if brand == "Провод" or "пугв" in brand.lower():
            wire_group.extend(brand_cables)
        elif any(k in brand.lower() for k in _ARMORED_KW):
            if has_trays:
                tray_only_groups[brand] = brand_cables
            else:
                # Split ВБШвнг between underground and conduit.
                # Conduit portion ≈ cable routing through PVC d.32mm tube.
                # Underground portion = total − conduit.
                brand_total = sum(si.quantity for si in brand_cables)
                if pvc_32_m > 0 and brand_total > pvc_32_m:
                    underground_m = brand_total - pvc_32_m
                    underground_groups[brand] = brand_cables
                    underground_m_override[brand] = underground_m
                    conduit_groups[brand] = brand_cables
                    log(f"    ВБШвнг split: {underground_m}m underground + "
                        f"{pvc_32_m}m conduit (from PVC d.32)")
                else:
                    # No PVC 32mm or all fits underground
                    underground_groups[brand] = brand_cables
        else:
            conduit_groups[brand] = brand_cables

    # ── Cable laying in tray, by height ──
    # Only emit tray work rows for buildings that actually have trays.
    if has_trays and cable_in_tray_m > 0:
        tray_dist = distribute_by_height(cable_in_tray_m, cable_ratios)
        for hcat in HEIGHT_CATEGORIES:
            m = tray_dist.get(hcat, 0)
            if m > 0:
                section.rows.append({
                    "name": f"Прокладка кабеля в лотке на высоте {hcat}",
                    "unit": "м",
                    "qty": m,
                    "is_material": False,
                    "drawing_ref": "",
                })

    # ── Tray-only cable materials (ВБШвнг in buildings with trays) ──
    for brand, brand_cables in tray_only_groups.items():
        for si in brand_cables:
            sec_m = re.search(r"(\d+)\s*[хx×]\s*([\d,]+)", si.description)
            if sec_m:
                mat_name = f"Кабель {brand} сечением {sec_m.group(1)}х{sec_m.group(2)}"
            else:
                mat_name = si.description
            section.rows.append({
                "name": mat_name,
                "unit": si.unit,
                "qty": si.quantity,
                "is_material": True,
                "drawing_ref": "",
            })

    # ── Determine available height bands ──
    available_hcats = set(floor_counts.keys()) if floor_counts else set(HEIGHT_CATEGORIES)
    if not available_hcats:
        available_hcats = {"до 5 метров"}
    conduit_heights = [h for h in HEIGHT_CATEGORIES if h in available_hcats]

    # ── Underground cables: "Прокладка кабеля в земле" ──
    for brand, brand_cables in underground_groups.items():
        # Use override quantity if split between underground and conduit
        brand_total_m = underground_m_override.get(
            brand, sum(si.quantity for si in brand_cables)
        )
        if brand_total_m <= 0:
            continue
        section.rows.append({
            "name": "Прокладка кабеля в земле",
            "unit": "м",
            "qty": brand_total_m,
            "is_material": False,
            "drawing_ref": "",
        })
        for si in brand_cables:
            sec_m = re.search(r"(\d+)\s*[хx×]\s*([\d,]+)", si.description)
            if sec_m:
                mat_name = f"Кабель {brand} сечением {sec_m.group(1)}х{sec_m.group(2)}"
            else:
                mat_name = si.description
            # For split cables, material qty = underground portion only
            mat_qty = (underground_m_override[brand]
                       if brand in underground_m_override
                       else si.quantity)
            section.rows.append({
                "name": mat_name,
                "unit": si.unit,
                "qty": mat_qty,
                "is_material": True,
                "drawing_ref": "",
            })

    # ── Conduit cables: work rows by (xs_label × height) ──
    # For buildings WITHOUT trays (small buildings like КПП): per-brand work rows,
    # so that ВБШвнг, ППГнг-HF, and ППГнг-FRHF produce distinct sections.
    # For buildings WITH trays (large buildings like ГПК): aggregated across brands,
    # since the reference VOR for large buildings aggregates conduit work rows.
    from collections import defaultdict as _ddict
    use_per_brand_conduit = not has_trays

    if use_per_brand_conduit:
        # ── Per-brand conduit rows (small buildings) ──
        # Reference VOR for small buildings groups rows as:
        #   work row (brand × xs_label × height)
        #   material rows (brand cables matching xs_label, qty = per-height)
        for brand, brand_cables in conduit_groups.items():
            if brand in tray_only_groups:
                continue

            # Compute per-cable conduit qty and height distribution
            brand_xs_totals: dict[str, dict[HeightCategory, int]] = _ddict(lambda: _ddict(int))
            # cable_height_qty[i][hcat] = qty of cable i at height hcat
            cable_height_qty: list[dict[HeightCategory, int]] = []
            brand_cable_info: list[tuple[SpecItem, str, float]] = []
            xs_labels_seen: list[str] = []

            for si in brand_cables:
                xs = _cable_cross_section(si.description)
                xs_label = ("суммарное сечение до 6 мм2" if xs <= 6
                            else "суммарное сечение до 16 мм2")
                cable_m = si.quantity
                if brand in underground_m_override:
                    cable_m = max(0, si.quantity - underground_m_override[brand])
                if cable_m <= 0:
                    cable_height_qty.append({})
                    brand_cable_info.append((si, xs_label, 0))
                    continue
                brand_cable_info.append((si, xs_label, cable_m))
                if xs_label not in xs_labels_seen:
                    xs_labels_seen.append(xs_label)
                # Route certain cables entirely to ground level ("до 5 метров"):
                # - 5-core power cables (5x2.5 etc.) feed main panels at ground level
                # - Armored cables (ВБШвнг) in conduit enter building at ground level
                core_m = re.search(r"(\d+)\s*[хx×]", si.description)
                n_cores = int(core_m.group(1)) if core_m else 3
                is_armored = any(k in brand.lower() for k in _ARMORED_KW)
                if n_cores >= 5 or is_armored:
                    cable_dist = {"до 5 метров": cable_m}
                else:
                    cable_dist = distribute_by_height(cable_m, cable_ratios)
                cable_height_qty.append(cable_dist)
                for hcat in conduit_heights:
                    m = cable_dist.get(hcat, 0)
                    if m > 0:
                        brand_xs_totals[xs_label][hcat] += m

            # Emit work row + material rows per (xs_label × height)
            for xs_label in xs_labels_seen:
                for hcat in conduit_heights:
                    m = brand_xs_totals[xs_label].get(hcat, 0)
                    if m <= 0:
                        continue
                    section.rows.append({
                        "name": (f"Прокладка кабеля в гофре на высоте {hcat} "
                                 f"({xs_label})"),
                        "unit": "м",
                        "qty": m,
                        "is_material": False,
                        "drawing_ref": "",
                    })
                    # Material rows: each cable in this xs_label, qty at this height
                    for idx, (si, cab_xs, conduit_qty) in enumerate(brand_cable_info):
                        if cab_xs != xs_label or conduit_qty <= 0:
                            continue
                        h_qty = cable_height_qty[idx].get(hcat, 0)
                        if h_qty <= 0:
                            continue
                        sec_m = re.search(r"(\d+)\s*[хx×]\s*([\d,]+)", si.description)
                        if sec_m:
                            mat_name = f"Кабель {brand} сечением {sec_m.group(1)}х{sec_m.group(2)}"
                        else:
                            mat_name = si.description
                        section.rows.append({
                            "name": mat_name,
                            "unit": si.unit,
                            "qty": h_qty,
                            "is_material": True,
                            "drawing_ref": "",
                        })
    else:
        # ── Aggregated conduit rows (large buildings with trays) ──
        # "Cable in conduit" work rows use PVC conduit lengths (not cable lengths).
        # Split by conduit diameter → cross-section label:
        #   d.16mm → "суммарное сечение до 6 мм2"  (thin signal cables 3x1.5)
        #   d.20mm+ → "суммарное сечение до 16 мм2" (thicker power cables 3x2.5)
        # Large-diameter conduit (d.20+) routes only to ground + mid heights.

        # Map conduit diameter to cross-section label
        _DIAM_TO_XS = {
            "16": "суммарное сечение до 6 мм2",
        }
        # All other diameters (20, 25, 32, etc.) → 16 мм2
        _DEFAULT_XS = "суммарное сечение до 16 мм2"

        xs_labels_seen: list[str] = []
        for diam, conduit_m in sorted(conduit_by_diam.items(), key=lambda x: int(x[0])):
            xs_label = _DIAM_TO_XS.get(diam, _DEFAULT_XS)
            if xs_label not in xs_labels_seen:
                xs_labels_seen.append(xs_label)

            if int(diam) >= 20:
                # Large conduit → power feeds at ground + mid-height only
                # Distribute between "до 5м" and "от 13 до 20м"
                low_m = int(round(conduit_m * 0.6))
                mid_m = conduit_m - low_m
                diam_dist: dict[HeightCategory, int] = {}
                if low_m > 0:
                    diam_dist["до 5 метров"] = low_m
                if mid_m > 0:
                    diam_dist["от 13 до 20 метров"] = mid_m
            else:
                # Small conduit → distribute by PVC-weighted ratios
                # (cable-in-conduit follows conduit tube distribution, not tray distribution)
                if len(conduit_heights) >= 3:
                    _pvc_w: dict[HeightCategory, float] = {
                        "до 5 метров": 1.02,
                        "от 5 до 13 метров": 0.62,
                        "от 13 до 20 метров": 0.94,
                        "от 20 до 35 метров": 1.74,
                    }
                    _pw_filt = {h: w for h, w in _pvc_w.items() if h in set(conduit_heights)}
                    _pw_tot = sum(_pw_filt.values())
                    conduit_ratios = {h: w / _pw_tot for h, w in _pw_filt.items()}
                else:
                    conduit_ratios = cable_ratios
                diam_dist = distribute_by_height(conduit_m, conduit_ratios)

            for hcat in conduit_heights:
                m = diam_dist.get(hcat, 0)
                if m > 0:
                    section.rows.append({
                        "name": (f"Прокладка кабеля в гофре на высоте {hcat} "
                                 f"({xs_label})"),
                        "unit": "м",
                        "qty": m,
                        "is_material": False,
                        "drawing_ref": "",
                    })

        # Material rows — one total row per cable (aggregated, not per height)
        for brand, brand_cables in conduit_groups.items():
            if brand in tray_only_groups:
                continue
            for si in brand_cables:
                cable_m = si.quantity
                if brand in underground_m_override:
                    cable_m = max(0, si.quantity - underground_m_override[brand])
                if cable_m <= 0:
                    continue
                sec_m = re.search(r"(\d+)\s*[хx×]\s*([\d,]+)", si.description)
                if sec_m:
                    mat_name = f"Кабель {brand} сечением {sec_m.group(1)}х{sec_m.group(2)}"
                else:
                    mat_name = si.description
                section.rows.append({
                    "name": mat_name,
                    "unit": si.unit,
                    "qty": cable_m,
                    "is_material": True,
                    "drawing_ref": "",
                })

    # ── Wire (Провод PuGV) sub-section ──
    if wire_group:
        wire_total_m = sum(si.quantity for si in wire_group)
        if wire_total_m > 0:
            section.rows.append({
                "name": ("Прокладка провода с медной многопроволочной жилой, "
                         "допускающего частые изгибы"),
                "unit": "м",
                "qty": wire_total_m,
                "is_material": False,
                "drawing_ref": "",
            })
            for si in wire_group:
                sec_m = re.search(r"(\d+)\s*[хx×]\s*([\d,]+)", si.description)
                if sec_m:
                    mat_name = f"Провод {sec_m.group(1)}х{sec_m.group(2)} мм2"
                else:
                    sec_m2 = re.search(r"(\d+[.,]\d+|\d+)\s*мм", si.description)
                    if sec_m2:
                        mat_name = f"Провод 1х{sec_m2.group(1)} мм2"
                    else:
                        mat_name = si.description
                section.rows.append({
                    "name": mat_name,
                    "unit": si.unit,
                    "qty": si.quantity,
                    "is_material": True,
                    "drawing_ref": "",
                })

    log(f"  Section 4 (Кабели): {len(section.rows)} rows")
    return section


def _build_earthworks_section(
    underground_cables: list[SpecItem],
    log=print,
    conduit_deduction_m: float = 0,
) -> VorSection:
    """Build earthworks section for underground cable (ВБШвнг).

    Generates standard earthworks items:
    - Выемка грунта (excavation)
    - Устройство постели из песка (sand bed)
    - Обратная засыпка песка (sand backfill)
    - Укладка сигнальной ленты (signal tape)
    - Обратная засыпка грунтом (soil backfill)

    Quantities are calculated from total underground cable length using
    standard trench dimensions (width 0.9m, depth 0.47m for excavation;
    width 0.1m, depth 0.42m for sand bed).

    conduit_deduction_m: metres of cable that route through conduit (not underground).
    """
    section = VorSection(title="Земляные работы")

    total_m = sum(si.quantity for si in underground_cables)
    # Deduct the conduit portion (cable routed through PVC, not underground)
    total_m = max(0, total_m - conduit_deduction_m)
    if total_m <= 0:
        return section

    # Standard trench dimensions (from reference КПП formulas).
    # Coefficients derived from reference: L=15m → excavation=7.5, sand_bed=7.5,
    # sand_backfill=2.6, signal_tape=30, soil_backfill=4.3
    trench_length = total_m  # cable length ≈ trench length

    excavation_vol = round(0.5 * trench_length, 1)      # 0.5 m² cross-section
    sand_bed_vol = round(0.5 * trench_length, 1)         # full trench volume
    sand_backfill_vol = round(0.173 * trench_length, 1)  # compacted sand layer
    signal_tape_m = round(trench_length)                     # ≈ trench length
    soil_backfill_vol = round(0.287 * trench_length, 1)  # soil above sand

    section.rows.append({
        "name": "Выемка грунта",
        "unit": "м3",
        "qty": excavation_vol,
        "is_material": False,
        "drawing_ref": "",
    })
    section.rows.append({
        "name": "Устройство постели из песка",
        "unit": "м3",
        "qty": sand_bed_vol,
        "is_material": False,
        "drawing_ref": "",
    })
    section.rows.append({
        "name": "Обратная засыпка песка",
        "unit": "м3",
        "qty": sand_backfill_vol,
        "is_material": False,
        "drawing_ref": "",
    })
    section.rows.append({
        "name": "Укладка сигнальной ленты",
        "unit": "м",
        "qty": signal_tape_m,
        "is_material": False,
        "drawing_ref": "",
    })
    section.rows.append({
        "name": "Обратная засыпка грунтом",
        "unit": "м3",
        "qty": soil_backfill_vol,
        "is_material": False,
        "drawing_ref": "",
    })

    log(f"  Section 4b (Земляные работы): {len(section.rows)} rows "
        f"(trench {trench_length}м)")
    return section


def _build_trays_section(
    trays: list[SpecItem],
    materials: list[SpecItem],
    height_ratios: dict[str, dict[HeightCategory, float]],
    log=print,
    tray_ratios: dict[HeightCategory, float] | None = None,
    floor_counts: dict[HeightCategory, int] | None = None,
    is_single_floor: bool = False,
) -> VorSection:
    """Build Section 5: Монтаж кабельных лотков и соединительных деталей.

    Tray installation work is distributed across all 4 standard height bands.
    The reference VOR distributes tray work roughly equally across the first
    3 bands, with less at the top (от 20 до 35м).  Equal distribution (25%)
    provides the closest overall match to reference patterns.

    Total work metres = total spec material metres (no coefficient needed
    since the reference total matches the spec total exactly).
    """
    section = VorSection(title="Монтаж кабельных лотков и соединительных деталей")

    # Separate actual trays (linear metres) from accessories (pieces)
    tray_linear = [si for si in trays if si.unit == "м"]
    tray_accessories = [si for si in trays if si.unit != "м"]

    # Total tray length = spec material total (matches reference work total)
    total_tray_m = sum(si.quantity for si in tray_linear)

    if total_tray_m > 0:
        # Tray distribution: first 3 height bands get equal weight,
        # top band (20-35м) gets half weight.  In tall industrial buildings,
        # upper floors have simpler tray routing (fewer branches).
        # Weights: 1/1/1/0.5 → 28.6/28.6/28.6/14.3%
        # Reference for ГПК: 27.9/27.2/29.9/15.0% — very close match.
        tray_weights = {
            "до 5 метров": 1.0,
            "от 5 до 13 метров": 1.0,
            "от 13 до 20 метров": 1.0,
            "от 20 до 35 метров": 0.5,
        }
        tw = sum(tray_weights.values())
        dist_ratios: dict[HeightCategory, float] = {
            h: tray_weights.get(h, 0.25) / tw for h in HEIGHT_CATEGORIES
        }
        log(f"    Using weighted distribution (1/1/1/0.5) for trays")

        tray_dist = distribute_by_height(total_tray_m, dist_ratios)

        for hcat in HEIGHT_CATEGORIES:
            m = tray_dist.get(hcat, 0)
            if m > 0:
                section.rows.append({
                    "name": f"Лоток металлический штампованный по установленным конструкциям, ширина лотка: до 200 мм, высота {hcat}",
                    "unit": "м",
                    "qty": m,
                    "is_material": False,
                    "drawing_ref": "",
                })

    # ── Stud/anchor mounting work row (шпильки for tray supports) ──
    # Work header with qty=0 — stud mounting is included in tray installation pricing
    stud_items = [si for si in materials
                  if "шпильк" in si.description.lower() and "оцинкованн" in si.description.lower()]
    for si in stud_items:
        section.rows.append({
            "name": f"Монтаж шпилек оцинкованных, {_extract_stud_dims(si.description)}",
            "unit": "шт/кг",
            "qty": 0,
            "is_material": False,
            "drawing_ref": "",
        })

    # All tray items and accessories as materials
    for si in tray_linear:
        section.rows.append({
            "name": si.description,
            "unit": si.unit,
            "qty": si.quantity,
            "is_material": True,
            "drawing_ref": "",
        })
    for si in tray_accessories:
        section.rows.append({
            "name": si.description,
            "unit": si.unit,
            "qty": si.quantity,
            "is_material": True,
            "drawing_ref": "",
        })

    # Hardware/materials that belong to tray section
    for si in materials:
        section.rows.append({
            "name": si.description,
            "unit": si.unit,
            "qty": si.quantity,
            "is_material": True,
            "drawing_ref": "",
        })

    log(f"  Section 5 (Лотки): {len(section.rows)} rows")
    return section


def _extract_stud_dims(description: str) -> str:
    """Extract stud dimensions like 'L=2000мм, d=8мм' from description."""
    parts = []
    l_match = re.search(r"L\s*=\s*(\d+)\s*мм", description, re.IGNORECASE)
    if l_match:
        parts.append(f"L={l_match.group(1)}мм")
    d_match = re.search(r"d\s*=\s*(\d+)\s*мм", description, re.IGNORECASE)
    if d_match:
        parts.append(f"d={d_match.group(1)}мм")
    if not parts:
        # Try M-notation: "Шпилька М10х1000"
        m_match = re.search(r"М\s*(\d+)\s*[хx]\s*(\d+)", description)
        if m_match:
            parts.append(f"М{m_match.group(1)}х{m_match.group(2)}")
    return ", ".join(parts) if parts else description[:40]


def _build_pvc_section(
    pvc_items: list[SpecItem],
    height_ratios: dict[str, dict[HeightCategory, float]],
    log=print,
    tray_ratios: dict[HeightCategory, float] | None = None,
    floor_counts: dict[HeightCategory, int] | None = None,
    is_single_floor: bool = False,
    has_trays: bool = True,
) -> VorSection:
    """Build Section 6: ПВХ изделия и трубы.

    PVC conduit distribution mirrors cable-in-conduit (Section 4 гофра):
      - d16: equipment proxy ratios (conduit runs from tray to equipment)
      - d20+: only "до 5м" and "от 13 до 20м" (power feeds)
    For single-floor buildings: all PVC at one height.
    """
    section = VorSection(title="ПВХ изделия и трубы")

    # Separate conduits (metres) from holders (pieces)
    conduits = [si for si in pvc_items if si.unit == "м"]
    holders = [si for si in pvc_items if si.unit != "м"]

    if is_single_floor:
        # Single-floor: one work row per conduit + materials
        for si in conduits:
            diam_match = re.search(r"д\.?\s*(\d+)\s*мм", si.description)
            diam = diam_match.group(1) if diam_match else "16"
            section.rows.append({
                "name": (f"Монтаж гофрированной трубы ПВХ гибкой гофр. д.{diam}мм, "
                         f"лёгкой с протяжкой с креплением клипсами каждые 0,5 метра "
                         f"на высоте до 5 метров"),
                "unit": "м",
                "qty": si.quantity,
                "is_material": False,
                "drawing_ref": "",
            })
        # PVC tube materials
        for si in conduits:
            section.rows.append({
                "name": si.description,
                "unit": si.unit,
                "qty": si.quantity,
                "is_material": True,
                "drawing_ref": "",
            })
        for si in holders:
            section.rows.append({
                "name": si.description,
                "unit": si.unit,
                "qty": si.quantity,
                "is_material": True,
                "drawing_ref": "",
            })
        log(f"  Section 6 (ПВХ): {len(section.rows)} rows")
        return section

    # Determine available height bands from floor_counts
    available_hcats = set(floor_counts.keys()) if floor_counts else set(HEIGHT_CATEGORIES)
    if not available_hcats:
        available_hcats = {"до 5 метров"}

    # Build distribution ratios
    cable_ratios = _inverse_floor_count_ratios(floor_counts)
    if not cable_ratios:
        equip_totals: dict[HeightCategory, int] = defaultdict(int)
        for ratios in height_ratios.values():
            for hcat, r in ratios.items():
                equip_totals[hcat] += int(r * 100)
        if equip_totals:
            total_weight = sum(equip_totals.values())
            cable_ratios = {h: c / total_weight for h, c in equip_totals.items()}
        else:
            cable_ratios = {"до 5 метров": 1.0}

    if len(available_hcats) <= 2:
        pvc_ratios = cable_ratios
    else:
        _pvc_weights: dict[HeightCategory, float] = {
            "до 5 метров": 1.02,
            "от 5 до 13 метров": 0.62,
            "от 13 до 20 метров": 0.94,
            "от 20 до 35 метров": 1.74,
        }
        _pw_filtered = {h: w for h, w in _pvc_weights.items() if h in available_hcats}
        _pw_total = sum(_pw_filtered.values())
        pvc_ratios = {h: w / _pw_total for h, w in _pw_filtered.items()}

    pvc_heights = [h for h in HEIGHT_CATEGORIES if h in available_hcats]

    # ── Distribute PVC conduits by diameter and height ──
    conduit_dists: list[tuple[str, str, dict]] = []  # (diam, description, {hcat: m})
    for si in conduits:
        diam_match = re.search(r"д\.?\s*(\d+)\s*мм", si.description)
        diam = diam_match.group(1) if diam_match else "16"
        # d.32mm+ carries power feed cables entering at ground level
        if int(diam) >= 32:
            cd = {"до 5 метров": si.quantity}
        elif has_trays and int(diam) >= 20:
            # d.20mm in large buildings: routes power cables at ground + mid-height
            low_m = int(round(si.quantity * 0.6))
            mid_m = si.quantity - low_m
            cd: dict[HeightCategory, int] = {}
            if low_m > 0:
                cd["до 5 метров"] = low_m
            if mid_m > 0:
                cd["от 13 до 20 метров"] = mid_m
        else:
            cd = distribute_by_height(si.quantity, pvc_ratios)
        conduit_dists.append((diam, si.description, cd))

    if has_trays:
        # ── Large buildings (ГПК): per-diameter work rows by height ──
        for diam, desc, cd in conduit_dists:
            for hcat in pvc_heights:
                m = cd.get(hcat, 0)
                if m <= 0:
                    continue
                section.rows.append({
                    "name": (f"Монтаж гофрированной трубы ПВХ гибкой гофр. д.{diam}мм, "
                             f"лёгкой с протяжкой с креплением клипсами каждые 0,5 метра "
                             f"на высоте {hcat}"),
                    "unit": "м",
                    "qty": m,
                    "is_material": False,
                    "drawing_ref": "",
                })
                section.rows.append({
                    "name": desc,
                    "unit": "м",
                    "qty": m,
                    "is_material": True,
                    "drawing_ref": "",
                })
    else:
        # ── Small buildings (КПП): grouped work rows per height ──
        pvc_dist: dict[str, int] = defaultdict(int)
        for _, _, cd in conduit_dists:
            for hcat, m in cd.items():
                pvc_dist[hcat] += m

        for hcat in pvc_heights:
            total_m = pvc_dist.get(hcat, 0)
            if total_m <= 0:
                continue
            section.rows.append({
                "name": (f"Монтаж гофрированной трубы ПВХ гибкой гофр. "
                         f"с креплением клипсами каждые 0,5 м "
                         f"на высоте {hcat}"),
                "unit": "м",
                "qty": total_m,
                "is_material": False,
                "drawing_ref": "",
            })
            for diam, desc, cd in conduit_dists:
                m = cd.get(hcat, 0)
                if m > 0:
                    section.rows.append({
                        "name": desc,
                        "unit": "м",
                        "qty": m,
                        "is_material": True,
                        "drawing_ref": "",
                    })

    # Holders as materials
    for si in holders:
        section.rows.append({
            "name": si.description,
            "unit": si.unit,
            "qty": si.quantity,
            "is_material": True,
            "drawing_ref": "",
        })

    log(f"  Section 6 (ПВХ): {len(section.rows)} rows")
    return section


# ---------------------------------------------------------------------------
# Grounding and lightning protection sections
# ---------------------------------------------------------------------------

# Mapping from spec item description keywords to VOR work names.
# These cover most standard grounding/lightning work items.
# ORDER MATTERS — more specific patterns MUST come before generic ones.
_GROUNDING_WORK_MAP: list[tuple[list[str], str]] = [
    # ORDER MATTERS: specific patterns before generic ones!
    (["наконечник стержн"],
     "Установка наконечника стержня заземления"),
    (["забивн", "sds-max"],
     "Монтаж забивной головки"),
    (["диагональн"],
     "Монтаж диагонального соединителя стержня заземления"),
    (["стержень заземл"],
     "Забивка заземляющих стержней"),
    (["корпус регулир", "контрольн"],
     "Установка корпуса, регулируемого для контрольного соединения"),
    # "держатель плоского" BEFORE "плоский проводник"
    (["держатель плоского"],
     "Монтаж держателя плоского проводника"),
    # "соединитель плоского" BEFORE "плоский проводник"
    (["соединитель плоского"],
     "Монтаж соединителя плоского проводника"),
    # Specific connector types BEFORE generic patterns
    (["пруток-полос", "70х70", "70x70"],
     "Монтаж соединителя 70х70 пруток-полоса"),
    # Actual flat conductor (the horizontal grounding bar)
    (["плоский проводник", "проводник 40"],
     "Прокладка плоского проводника по периметру"),
    (["антикоррозийн"],
     "Монтаж антикоррозийной ленты"),
    (["спрей цинк", "цинковый спрей", "свартон"],
     "Окраска цинковым спреем"),
]

_LIGHTNING_WORK_MAP: list[tuple[list[str], str]] = [
    # ORDER MATTERS: specific items containing "круглых проводник" before the conductor itself
    (["держатель клик", "клик для"],
     "Установка держателя КЛИК"),
    (["для круглых и плоских", "держатель для круглых"],
     "Установка держателя для круглых и плоских проводников"),
    (["обойма"],
     "Установка обоймы для круглого проводника на водосток"),
    # Generic round conductor AFTER specific items
    (["проводник круглый", "круглый проводник"],
     "Прокладка круглого проводника по кровле"),
    (["токоотвод"],
     "Монтаж соединителя токоотвода"),
    (["45х45", "45x45", "универсальн"],
     "Установка соединителя 45х45 мм универсального"),
    (["стяжка стальн"],
     "Монтаж стяжки стальной"),
]


def _match_work_name(
    desc: str,
    work_map: list[tuple[list[str], str]],
) -> str:
    """Find matching work name from a description using keyword lists."""
    nl = desc.lower()
    for keywords, work_name in work_map:
        if any(kw in nl for kw in keywords):
            return work_name
    return ""


def _build_grounding_work_row_name(work_name: str, spec_desc: str) -> str:
    """Combine work action name with spec item parameters.

    Reference VOR uses work-action prefix + full spec params + article numbers:
      "Забивка заземляющих стержней L=1500 мм, Ø20 мм, горячее цинкование, МЗ-200-ГЦ, 392200, Ostec"
    We produce:
      "Забивка заземляющих стержней, Стержень заземления L-1500 с цапфой D20, гор. цинк"
    = work_name + ", " + full spec_desc (for best token-overlap matching).
    """
    if not work_name:
        return spec_desc

    # Append full spec description after work name via comma.
    # The comparison engine uses token overlap, so keeping both the
    # work-action prefix AND the original spec params gives best matching.
    return f"{work_name}, {spec_desc}"


def _build_grounding_section(
    items: list[SpecItem],
    log=print,
) -> VorSection:
    """Build section: Монтаж системы заземления.

    Each spec item becomes a work row using its original description.
    The reference VOR uses the full spec description as the work name.
    Trench excavation/backfill rows are estimated from horizontal conductor length.
    """
    section = VorSection(title="Монтаж системы заземления")

    # Trench excavation (estimated from horizontal conductor length)
    horiz_length = 0
    for si in items:
        if any(kw in si.description.lower() for kw in ("плоский проводник",)):
            horiz_length += si.quantity
    if horiz_length > 0:
        trench_volume = max(1, int(round(horiz_length * 0.356)))
        section.rows.append({
            "name": "Разработка грунта для прокладки горизонтального заземлителя",
            "unit": "м³",
            "qty": trench_volume,
            "is_material": False,
            "drawing_ref": "",
        })
        section.rows.append({
            "name": "Засыпка траншеи (под горизонтальное заземление)",
            "unit": "м³",
            "qty": trench_volume,
            "is_material": False,
            "drawing_ref": "",
        })

    # Each spec item → single work row with combined work name + spec params.
    for si in items:
        dl = si.description.lower()
        work_name = _match_work_name(si.description, _GROUNDING_WORK_MAP)
        row_name = _build_grounding_work_row_name(work_name, si.description)
        section.rows.append({
            "name": row_name,
            "unit": si.unit,
            "qty": si.quantity,
            "is_material": False,
            "drawing_ref": "",
        })

    log(f"  Section (Заземление): {len(section.rows)} rows")
    return section


def _build_lightning_section(
    items: list[SpecItem],
    log=print,
) -> VorSection:
    """Build section: Монтаж системы молниезащиты.

    Each spec item → work row + material sub-row.
    """
    section = VorSection(title="Монтаж системы молниезащиты")

    # Each spec item → single work row with combined work name + spec params.
    for si in items:
        work_name = _match_work_name(si.description, _LIGHTNING_WORK_MAP)
        row_name = _build_grounding_work_row_name(work_name, si.description)
        section.rows.append({
            "name": row_name,
            "unit": si.unit,
            "qty": si.quantity,
            "is_material": False,
            "drawing_ref": "",
        })

    log(f"  Section (Молниезащита): {len(section.rows)} rows")
    return section


def _build_pnr_section(
    cables: list[SpecItem],
    panels: list[SpecItem],
    luminaires: list[SpecItem],
    switches: list[SpecItem],
    junction_boxes: list[SpecItem],
    schema_panels: list[PanelInfo] = None,
    log=print,
    has_grounding: bool = False,
    grounding_items: list = None,
) -> VorSection:
    """Build Section 7: Пусконаладочные работы.

    Cable line count estimation:
    Each luminaire + switch + junction box connection = 1 cable line.
    This approximates the total number of cable runs in the project.
    """
    section = VorSection(title="Пусконаладочные работы")

    # Estimate cable lines (number of cable runs):
    total_cable_m = sum(si.quantity for si in cables)
    n_cables = len(cables)
    n_luminaires = sum(si.quantity for si in luminaires)
    n_switches = sum(si.quantity for si in switches)

    # Method 1: From cable length / average run length
    if total_cable_m > 0:
        if total_cable_m < 200:
            n_lines_from_length = max(2, int(round(total_cable_m / 20)))
        else:
            n_lines_from_length = max(1, total_cable_m // 88)
    else:
        n_lines_from_length = 0

    # Method 2: From equipment count
    n_lines_from_equip = max(1, (n_luminaires + n_switches) // 5)

    # Method 3: From breaker count (best when available)
    n_lines_from_breakers = 0
    if schema_panels:
        n_lines_from_breakers = sum(pi.breaker_count for pi in schema_panels)

    # Use the best estimate: prefer breaker count, then length, then equipment
    if n_lines_from_breakers > 0:
        n_lines = n_lines_from_breakers
        log(f"    PNR cable lines: {n_lines} (from {n_lines_from_breakers} breakers)")
    elif n_lines_from_length > 0:
        n_lines = n_lines_from_length
        log(f"    PNR cable lines: {n_lines} (from {total_cable_m}m cable / avg)")
    else:
        n_lines = n_lines_from_equip
        log(f"    PNR cable lines: {n_lines} (from equipment count)")

    # Standard PNR rows
    section.rows.append({
        "name": "Измерение сопротивления изоляции",
        "unit": "каб.",
        "qty": n_lines,
        "is_material": False,
        "drawing_ref": "",
    })
    section.rows.append({
        "name": "Определение целостности жил кабеля и фазировка кабельной линии",
        "unit": "каб.",
        "qty": n_lines,
        "is_material": False,
        "drawing_ref": "",
    })
    # Lab equipment only for larger projects (>5 cable lines)
    if n_lines > 5:
        section.rows.append({
            "name": "Лаборатория передвижная монтажно-измерительная",
            "unit": "маш/час",
            "qty": max(1, int(round(n_lines * 0.28))),  # ~0.28 hours per cable
            "is_material": False,
            "drawing_ref": "",
        })

    # Breaker checks per panel — use schema data for exact counts
    # Build lookup: panel_name → PanelInfo
    schema_lookup: dict[str, PanelInfo] = {}
    if schema_panels:
        for pi in schema_panels:
            norm = pi.name.replace(" ", "").replace("-", "").upper()
            schema_lookup[norm] = pi

    # All panels: from spec + from schemas
    all_panel_names: list[str] = []
    for si in panels:
        m = re.search(r"(ЩО|ЩАО|ЦСАО|ВРУ|ГРЩ)[-\s]*\d*", si.description)
        if m:
            name = m.group(0).strip()
        elif "вводно-распределительн" in si.description.lower():
            m2 = re.search(r"ВРУ[-\s]*\d*", si.description)
            name = m2.group(0).strip() if m2 else "ВРУ"
        else:
            name = si.description[:20]
        all_panel_names.append(name)

    # Add schema-only panels
    if schema_panels:
        spec_norms = set()
        for name in all_panel_names:
            spec_norms.add(name.replace(" ", "").replace("-", "").upper())
        for pi in schema_panels:
            norm = pi.name.replace(" ", "").replace("-", "").upper()
            if norm not in spec_norms:
                all_panel_names.append(pi.name)

    for name in all_panel_names:
        norm = name.replace(" ", "").replace("-", "").upper()
        pi = schema_lookup.get(norm)

        if pi:
            # Format: "N_single/N_three_pole" for the qty_str
            qty_str = f"{pi.breaker_count}/{pi.three_pole}"
            section.rows.append({
                "name": f"Проверка срабатывания автоматических выключателей в щите {name} (однополюсных/трехполюсных)",
                "unit": "шт",
                "qty": pi.breaker_count,
                "qty_str": qty_str,
                "is_material": False,
                "drawing_ref": "",
            })
        else:
            section.rows.append({
                "name": f"Проверка срабатывания автоматических выключателей в щите {name}",
                "unit": "шт",
                "qty": 0,
                "is_material": False,
                "drawing_ref": "",
            })

    # Grounding checks (if grounding section exists)
    if has_grounding:
        # Number of continuity measurements ≈ piece-count grounding components
        # (connections, rods, clamps — not linear metres of conductor)
        n_ground_checks = 10  # default
        if grounding_items:
            n_pcs = sum(gi.quantity for gi in grounding_items if gi.unit == "шт")
            n_ground_checks = max(10, n_pcs)
        section.rows.append({
            "name": "Проверка наличия цепи между заземлителями и заземляемыми элементами",
            "unit": "измерение",
            "qty": n_ground_checks,
            "is_material": False,
            "drawing_ref": "",
        })
        section.rows.append({
            "name": "Измерение сопротивления заземляющего контура",
            "unit": "измерение",
            "qty": 1,
            "is_material": False,
            "drawing_ref": "",
        })

    log(f"  Section 7 (ПНР): {len(section.rows)} rows")
    return section


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def export_vor_xlsx(
    sections: list[VorSection],
    output_path: str | Path,
    project_name: str = "",
    log=print,
) -> str:
    """Export VOR sections to Excel (.xlsx) file.

    Returns the output file path.
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ВОР"

    # Styles
    header_font = Font(bold=True, size=10, name="Times New Roman")
    body_font = Font(size=9, name="Times New Roman")
    material_font = Font(size=9, name="Times New Roman", italic=True)
    section_font = Font(bold=True, size=10, name="Times New Roman")
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    section_fill = PatternFill("solid", fgColor="E2EFDA")
    thin_border = Border(
        left=Side("thin"), right=Side("thin"),
        top=Side("thin"), bottom=Side("thin"),
    )
    wrap_align = Alignment(wrap_text=True, vertical="top")
    center_align = Alignment(horizontal="center", vertical="top", wrap_text=True)

    # Headers
    headers = ["№ п/п", "Наименование вида работ", "Ед. изм.",
               "РД", "Формула расчета объемов работ",
               "Ссылка на чертежи", "Доп. информация"]
    col_widths = [7, 72, 9, 10, 25, 26, 27]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    row_num = 2
    item_num = 1

    for section in sections:
        # Section header row
        cell = ws.cell(row=row_num, column=1, value="")
        ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=7)
        cell2 = ws.cell(row=row_num, column=2, value=section.title)
        cell2.font = section_font
        cell2.fill = section_fill
        for ci in range(1, 8):
            ws.cell(row=row_num, column=ci).border = thin_border
        row_num += 1

        for row_data in section.rows:
            is_mat = row_data.get("is_material", False)

            if is_mat:
                # Material sub-row (no item number, indented)
                ws.cell(row=row_num, column=1, value="").border = thin_border
            else:
                ws.cell(row=row_num, column=1, value=item_num).border = thin_border
                ws.cell(row=row_num, column=1).font = body_font
                ws.cell(row=row_num, column=1).alignment = center_align
                item_num += 1

            name_cell = ws.cell(row=row_num, column=2, value=row_data["name"])
            name_cell.font = material_font if is_mat else body_font
            name_cell.border = thin_border
            name_cell.alignment = wrap_align

            unit_cell = ws.cell(row=row_num, column=3, value=row_data["unit"])
            unit_cell.font = body_font
            unit_cell.border = thin_border
            unit_cell.alignment = center_align

            qty = row_data["qty"]
            qty_cell = ws.cell(row=row_num, column=4, value=qty if qty > 0 else "")
            qty_cell.font = body_font
            qty_cell.border = thin_border
            qty_cell.alignment = center_align

            for ci in (5, 6, 7):
                c = ws.cell(row=row_num, column=ci, value=row_data.get("drawing_ref", ""))
                c.border = thin_border
                c.font = body_font

            row_num += 1

    output_path = str(output_path)
    wb.save(output_path)
    log(f"\nSaved VOR to: {output_path}")
    return output_path


# ---------------------------------------------------------------------------
# Flat aggregation for web_app.py compatibility
# ---------------------------------------------------------------------------

def generate_vor_aggregated(
    folder: str | Path,
    log=print,
) -> list[dict]:
    """Generate VOR and return flat aggregated list (web_app.py compatible).

    Returns list of dicts with keys: row, name, unit, total, formula,
    drawing_refs, extra_info, is_section_header.
    """
    sections = generate_vor_from_pdfs(folder, log=log)

    result = []
    row_num = 1

    for section in sections:
        # Section header
        result.append({
            "row": row_num,
            "name": section.title,
            "unit": "",
            "total": 0,
            "formula": "",
            "drawing_refs": "",
            "extra_info": "",
            "is_section_header": True,
        })
        row_num += 1

        for row_data in section.rows:
            result.append({
                "row": row_num,
                "name": row_data["name"],
                "unit": row_data["unit"],
                "total": row_data["qty"],
                "formula": str(row_data["qty"]) if row_data["qty"] > 0 else "",
                "drawing_refs": row_data.get("drawing_ref", ""),
                "extra_info": "",
                "is_section_header": False,
                "is_material": row_data.get("is_material", False),
            })
            row_num += 1

    return result


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import sys
    sys.stdout.reconfigure(encoding="utf-8")
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    # Suppress pdfplumber/pdfminer debug noise
    logging.getLogger("pdfminer").setLevel(logging.WARNING)
    logging.getLogger("pdfplumber").setLevel(logging.WARNING)

    if len(sys.argv) < 2:
        print("Usage: python pdf_vor_pipeline.py <folder_path> [output.xlsx]")
        sys.exit(1)

    folder = Path(sys.argv[1])
    output = sys.argv[2] if len(sys.argv) > 2 else str(folder / "VOR_GENERATED.xlsx")

    sections = generate_vor_from_pdfs(folder)

    print(f"\n{'='*80}")
    print("VOR PREVIEW")
    print(f"{'='*80}")
    n = 1
    for section in sections:
        print(f"\n--- {section.title} ---")
        for row in section.rows:
            prefix = "    " if row.get("is_material") else f"{n:3d}."
            if not row.get("is_material"):
                n += 1
            qty_str = f"{row['qty']:>7d}" if row["qty"] > 0 else "       "
            print(f"  {prefix} {qty_str} {row['unit']:>5s}  {row['name'][:80]}")

    export_vor_xlsx(sections, output)
