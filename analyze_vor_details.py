"""
Detailed analysis of VOR Excel files to understand the structure and content.
"""

import sys
import os

# Fix Windows console encoding for Cyrillic characters
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

sys.path.insert(0, '.')

import openpyxl
from pathlib import Path

# VOR Excel files for analysis
vor_files = [
    r"C:\Cursor\TayfaProject\EquipmentCounter\Data\ДБТ разделы для ИИ\01_АБК\Обновленные файлы\ВОР_\ВОР АБК ЭО.xlsx",
    r"C:\Cursor\TayfaProject\EquipmentCounter\Data\ДБТ разделы для ИИ\30. КПП\ВОР поз.30.xlsx",
]


def analyze_vor_detailed(excel_path: str):
    """Read and analyze VOR Excel file in detail."""
    print(f"\n{'='*100}")
    print(f"VOR FILE: {Path(excel_path).name}")
    print(f"{'='*100}")

    if not Path(excel_path).exists():
        print(f"  ❌ FILE NOT FOUND")
        return

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)

        # Find the main sheet
        sheet = None
        for ws in wb.worksheets:
            if "ВОР" in ws.title or "ЭО" in ws.title or ws == wb.worksheets[0]:
                sheet = ws
                break

        if sheet is None:
            print(f"  ⚠️  No suitable sheet found")
            return

        print(f"  Sheet: {sheet.title}")
        print(f"  Total rows: {sheet.max_row}, Total columns: {sheet.max_column}")
        print()

        # Read header row
        header_row_idx = 1
        headers = [cell.value for cell in sheet[header_row_idx]]
        print(f"  COLUMN HEADERS:")
        for i, h in enumerate(headers[:10], 1):
            if h:
                print(f"    Col {i}: {repr(h)}")
        print()

        # Analyze data rows - looking for patterns related to lighting fixtures
        print(f"  SAMPLE DATA ROWS (showing first 30 relevant rows):")
        print(f"  {'='*100}")

        data_rows_shown = 0
        for row_idx in range(2, min(200, sheet.max_row + 1)):
            row_values = [cell.value for cell in sheet[row_idx]]

            # Skip empty rows
            if not any(v for v in row_values if v and str(v).strip()):
                continue

            # Get key columns
            num = row_values[0] if len(row_values) > 0 else None
            name = row_values[1] if len(row_values) > 1 else None
            unit = row_values[2] if len(row_values) > 2 else None
            qty = row_values[3] if len(row_values) > 3 else None
            formula = row_values[4] if len(row_values) > 4 else None if len(headers) > 4 else None
            ref = row_values[5] if len(row_values) > 5 else None if len(headers) > 5 else None
            info = row_values[6] if len(row_values) > 6 else None if len(headers) > 6 else None

            # Format output
            num_str = f"{num:>4}" if num is not None else "    "
            name_str = f"{str(name)[:80]:80}" if name else " " * 80
            unit_str = f"{str(unit)[:6]:6}" if unit else "      "
            qty_str = f"{str(qty):>6}" if qty is not None else "      "

            print(f"  Row {row_idx:4d} | {num_str} | {name_str} | {unit_str} | {qty_str}")

            # Show additional info if available
            if info and str(info).strip():
                print(f"          INFO: {info}")
            if formula and str(formula).strip():
                print(f"          FORMULA: {formula}")
            if ref and str(ref).strip():
                print(f"          REF: {ref}")

            data_rows_shown += 1
            if data_rows_shown >= 30:
                break

        print(f"  {'='*100}")
        print()

        # Look for rows related to светильник (fixtures)
        print(f"  SVETILNIK (FIXTURE) ENTRIES (first 20):")
        print(f"  {'='*100}")

        fixture_count = 0
        for row_idx in range(2, min(500, sheet.max_row + 1)):
            row_values = [cell.value for cell in sheet[row_idx]]

            # Get name column
            name = row_values[1] if len(row_values) > 1 else None

            if name and "светильник" in str(name).lower():
                num = row_values[0] if len(row_values) > 0 else None
                unit = row_values[2] if len(row_values) > 2 else None
                qty = row_values[3] if len(row_values) > 3 else None
                info = row_values[6] if len(row_values) > 6 else None if len(headers) > 6 else None

                print(f"  Row {row_idx:4d} | {num} | {name}")
                print(f"            Unit: {unit}, Qty: {qty}")
                if info and str(info).strip():
                    print(f"            Info: {info}")
                print()

                fixture_count += 1
                if fixture_count >= 20:
                    break

        print(f"  Total fixture entries found (in first 500 rows): {fixture_count}")

    except Exception as e:
        print(f"  ❌ Error reading Excel: {e}")
        import traceback
        traceback.print_exc()


def main():
    print("\n" + "="*100)
    print("DETAILED VOR EXCEL FILE ANALYSIS")
    print("="*100)

    for vor_path in vor_files:
        try:
            analyze_vor_detailed(vor_path)
        except Exception as e:
            print(f"\n{'='*100}")
            print(f"VOR FILE: {vor_path}")
            print(f"{'='*100}")
            print(f"  ❌ Error: {e}")
            import traceback
            traceback.print_exc()


if __name__ == "__main__":
    main()
