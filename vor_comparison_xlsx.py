#!/usr/bin/env python3
"""vor_comparison_xlsx.py -- Generate 3-sheet Excel: Our VOR / Reference VOR / Comparison.

Usage:
    python vor_comparison_xlsx.py "Data/ДБТ разделы для ИИ/03_ГПК_/3-я захватка/02_PDF"
"""

from __future__ import annotations

import logging
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FONT = Font(bold=True, size=10, name="Times New Roman")
_BODY_FONT = Font(size=9, name="Times New Roman")
_MAT_FONT = Font(size=9, name="Times New Roman", italic=True)
_SECTION_FONT = Font(bold=True, size=10, name="Times New Roman")
_HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
_SECTION_FILL = PatternFill("solid", fgColor="E2EFDA")
_GOOD_FILL = PatternFill("solid", fgColor="C6EFCE")    # green
_CLOSE_FILL = PatternFill("solid", fgColor="FFEB9C")    # yellow
_BAD_FILL = PatternFill("solid", fgColor="FFC7CE")      # red
_WRAP = Alignment(wrap_text=True, vertical="top")
_CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)


def _style_header(ws, row, ncols, fill=None):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _HEADER_FONT
        cell.fill = fill or _HEADER_FILL
        cell.border = _BORDER
        cell.alignment = _CENTER


def _write_row(ws, row_num, values, fonts=None, fills=None, aligns=None):
    for ci, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=ci, value=val)
        cell.border = _BORDER
        cell.font = (fonts[ci - 1] if fonts else _BODY_FONT)
        if fills and fills[ci - 1]:
            cell.fill = fills[ci - 1]
        cell.alignment = (aligns[ci - 1] if aligns else _WRAP)


# ---------------------------------------------------------------------------
# Parse reference VOR from PDF
# ---------------------------------------------------------------------------

def _parse_reference_vor(ref_path: str) -> list[dict]:
    """Parse the reference VOR (PDF or XLSX) into flat rows.

    Returns list of dicts: {num, name, unit, qty, is_section, is_material, drawing_ref}
    """
    if ref_path.lower().endswith(".xlsx"):
        return _parse_reference_vor_xlsx(ref_path)
    return _parse_reference_vor_pdf(ref_path)


def _parse_reference_vor_xlsx(xlsx_path: str) -> list[dict]:
    """Parse reference VOR from XLSX file.

    Tries sheet "ЭО" first, then first sheet.
    Expects columns: № п/п | Наименование | Ед.изм | Кол-во | ... | Ссылка
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    # Pick best sheet
    if "ЭО" in wb.sheetnames:
        ws = wb["ЭО"]
    else:
        ws = wb[wb.sheetnames[0]]

    rows: list[dict] = []

    # Detect header row and column layout
    name_col = 1  # B by default
    unit_col = 2  # C
    qty_col = 3   # D
    drawing_col = 5  # F
    header_found = False

    for ri, row_tuple in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
        cells = [str(v).strip() if v is not None else "" for v in row_tuple]

        # Detect header row by looking for "Наименование" keyword
        if not header_found:
            for ci, c in enumerate(cells):
                if "наименование" in c.lower():
                    name_col = ci
                    # Unit is typically next column
                    unit_col = ci + 1
                    qty_col = ci + 2
                    drawing_col = ci + 4 if ci + 4 < len(cells) else len(cells) - 1
                    header_found = True
                    break
            if header_found or ri <= 3:
                continue

        num = cells[0] if len(cells) > 0 else ""
        name = cells[name_col][:120] if len(cells) > name_col else ""
        unit = cells[unit_col] if len(cells) > unit_col else ""
        qty_raw = cells[qty_col] if len(cells) > qty_col else ""
        drawing = cells[drawing_col][:80] if len(cells) > drawing_col else ""

        if not name:
            continue

        # Parse qty
        qty = 0
        qty_str = qty_raw
        if "/" in qty_raw:
            qty_str = qty_raw
            try:
                qty = int(qty_raw.split("/")[0].strip())
            except ValueError:
                qty = 0
        else:
            cleaned = re.sub(r"\s+", "", qty_raw)
            try:
                qty = int(float(cleaned)) if cleaned else 0
            except ValueError:
                qty = 0

        # Detect section headers (no number, no unit, no qty)
        is_section = False
        if not num and not unit and not qty_raw:
            nl = name.lower()
            section_kws = ("щитовое", "светотехническое", "светильник",
                           "электроустановочных", "кабельная продукция",
                           "кабельных лотков", "пвх издели",
                           "пусконаладочные", "заземлен", "молниезащит",
                           "монтаж систем", "прокладка силовых кабелей",
                           "кабель ппг")
            for kw in section_kws:
                if kw in nl:
                    is_section = True
                    break

        is_material = (num == "" and not is_section and name != "")
        # Work items without row numbers: if name starts with a work verb,
        # treat as work (not material). Common in grounding/lightning sections.
        if is_material and name:
            _work_verbs = ("установка", "монтаж", "прокладка", "забивка",
                           "окраска", "засыпка", "разработка", "проверка",
                           "измерение", "определение")
            if name.lower().startswith(_work_verbs):
                is_material = False

        rows.append({
            "num": num,
            "name": name,
            "unit": unit,
            "qty": qty,
            "qty_str": qty_str,
            "is_section": is_section,
            "is_material": is_material,
            "drawing_ref": drawing,
        })

    return rows


def _parse_reference_vor_pdf(pdf_path: str) -> list[dict]:
    """Parse the reference VOR PDF into flat rows.

    Returns list of dicts: {num, name, unit, qty, is_section, is_material, drawing_ref}
    """
    import pdfplumber

    rows = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                for trow in table:
                    if not trow:
                        continue
                    cells = [(c or "").replace("\n", " ").strip() for c in trow]

                    # Skip header rows
                    if cells[0] in ("№ п/п", "1") and cells[1] in ("Наименование вида работ", "2"):
                        continue

                    num = cells[0].strip() if len(cells) > 0 else ""
                    name = cells[1].strip()[:120] if len(cells) > 1 else ""
                    unit = cells[2].strip() if len(cells) > 2 else ""
                    qty_raw = cells[3].strip() if len(cells) > 3 else ""
                    drawing = cells[5].strip()[:80] if len(cells) > 5 else ""

                    if not name:
                        # Page-break artifact: a numbered work row whose name
                        # was on the previous page.  The row starts a new page
                        # with num + empty name + unit + qty.
                        # Example: ['8', '', 'шт', '206', '', '', '']
                        # This is "Монтаж ... на высоте от 20 до 35 метров"
                        # whose text was cut at the page boundary.
                        if num and qty_raw and unit:
                            # Reconstruct the work row.  The previous page's
                            # last WRK row had the title text (e.g., "от 13
                            # до 20 метров") but this row is the NEXT work
                            # item.  We can infer the name by looking at the
                            # previous work row's pattern and incrementing
                            # the height category.
                            prev_wrk = None
                            for prev in reversed(rows):
                                if not prev.get("is_material") and not prev.get("is_section"):
                                    prev_wrk = prev
                                    break

                            inferred_name = ""
                            if prev_wrk:
                                # Try to infer next height category
                                pn = prev_wrk["name"]
                                for cur_h, next_h in [
                                    ("до 5 м", "от 5 до 13 метров"),
                                    ("от 5 до 13", "от 13 до 20 метров"),
                                    ("от 13 до 20", "от 20 до 35 метров"),
                                ]:
                                    if cur_h in pn.lower():
                                        # Replace height in name
                                        base = re.sub(
                                            r"(до 5 м\w*|от 5 до 13 м\w*|от 13 до 20 м\w*|от 20 до 35 м\w*)",
                                            next_h,
                                            pn,
                                            flags=re.IGNORECASE,
                                        )
                                        inferred_name = base.rstrip(":").strip() + ":"
                                        break

                            if inferred_name:
                                cleaned = re.sub(r"\s+", "", qty_raw)
                                try:
                                    qty_val = int(cleaned) if cleaned else 0
                                except ValueError:
                                    qty_val = 0
                                rows.append({
                                    "num": num,
                                    "name": inferred_name[:120],
                                    "unit": unit,
                                    "qty": qty_val,
                                    "qty_str": qty_raw,
                                    "is_section": False,
                                    "is_material": False,
                                    "drawing_ref": "",
                                })
                            continue

                        # Check if qty is in a weird column (page break artifact)
                        if qty_raw and not unit:
                            # Append qty to previous row if exists
                            if rows and rows[-1].get("_partial"):
                                rows[-1]["qty_str"] = qty_raw
                                rows[-1]["_partial"] = False
                            continue
                        continue

                    # Parse qty
                    qty = 0
                    qty_str = qty_raw
                    if "/" in qty_raw:
                        qty_str = qty_raw  # Keep as-is for "35/1" format
                        try:
                            qty = int(qty_raw.split("/")[0].strip())
                        except ValueError:
                            qty = 0
                    else:
                        cleaned = re.sub(r"\s+", "", qty_raw)
                        try:
                            qty = int(cleaned) if cleaned else 0
                        except ValueError:
                            qty = 0

                    # Detect section headers
                    is_section = False
                    if not num and not unit and not qty_raw:
                        # Check known section keywords
                        nl = name.lower()
                        for kw in ("щитовое", "светотехническое", "электроустановочных",
                                   "кабельная продукция", "кабельных лотков",
                                   "пвх издели", "пусконаладочные",
                                   "заземлен", "молниезащит", "монтаж систем",
                                   "светильник", "прокладка силовых кабелей",
                                   "кабель ппг"):
                            if kw in nl:
                                is_section = True
                                break

                    is_material = (num == "" and not is_section and name != "")
                    # Work items without row numbers: if name starts with
                    # a work verb, treat as work (not material).
                    if is_material and name:
                        _work_verbs = ("установка", "монтаж", "прокладка",
                                       "забивка", "окраска", "засыпка",
                                       "разработка", "проверка",
                                       "измерение", "определение")
                        if name.lower().startswith(_work_verbs):
                            is_material = False

                    rows.append({
                        "num": num,
                        "name": name,
                        "unit": unit,
                        "qty": qty,
                        "qty_str": qty_str,
                        "is_section": is_section,
                        "is_material": is_material,
                        "drawing_ref": drawing,
                    })

    return rows


# ---------------------------------------------------------------------------
# Generate our VOR
# ---------------------------------------------------------------------------

def _generate_our_vor(folder: str) -> list[dict]:
    """Generate our VOR and return flat rows."""
    from pdf_vor_pipeline import generate_vor_from_pdfs

    sections = generate_vor_from_pdfs(folder)

    rows = []
    item_num = 1

    for section in sections:
        rows.append({
            "num": "",
            "name": section.title,
            "unit": "",
            "qty": 0,
            "qty_str": "",
            "is_section": True,
            "is_material": False,
            "drawing_ref": "",
        })

        for row in section.rows:
            is_mat = row.get("is_material", False)
            # Use qty_str if present (e.g. "35/1" format for breakers)
            qty_str = row.get("qty_str", str(row["qty"]) if row["qty"] > 0 else "")
            rows.append({
                "num": "" if is_mat else str(item_num),
                "name": row["name"][:120],
                "unit": row["unit"],
                "qty": row["qty"],
                "qty_str": qty_str,
                "is_section": False,
                "is_material": is_mat,
                "drawing_ref": row.get("drawing_ref", ""),
            })
            if not is_mat:
                item_num += 1

    return rows


# ---------------------------------------------------------------------------
# Write Sheet 1: Our VOR
# ---------------------------------------------------------------------------

def _write_vor_sheet(ws, title: str, rows: list[dict]):
    """Write a VOR table to a worksheet."""
    ws.title = title

    headers = ["№", "Наименование вида работ", "Ед.изм.", "Кол-во"]
    col_widths = [6, 80, 8, 12]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _BORDER
        cell.alignment = _CENTER
        ws.column_dimensions[get_column_letter(ci)].width = w

    rn = 2
    for row in rows:
        if row["is_section"]:
            ws.cell(row=rn, column=1, value="").border = _BORDER
            ws.merge_cells(start_row=rn, start_column=2, end_row=rn, end_column=4)
            cell = ws.cell(row=rn, column=2, value=row["name"])
            cell.font = _SECTION_FONT
            cell.fill = _SECTION_FILL
            cell.border = _BORDER
            for c in range(1, 5):
                ws.cell(row=rn, column=c).border = _BORDER
            rn += 1
            continue

        num_val = row["num"]
        name_val = row["name"]
        unit_val = row["unit"]
        qty_val = row.get("qty_str", str(row["qty"]) if row["qty"] else "")

        ws.cell(row=rn, column=1, value=num_val).border = _BORDER
        ws.cell(row=rn, column=1).font = _BODY_FONT
        ws.cell(row=rn, column=1).alignment = _CENTER

        name_cell = ws.cell(row=rn, column=2, value=name_val)
        name_cell.font = _MAT_FONT if row["is_material"] else _BODY_FONT
        name_cell.border = _BORDER
        name_cell.alignment = _WRAP

        ws.cell(row=rn, column=3, value=unit_val).border = _BORDER
        ws.cell(row=rn, column=3).font = _BODY_FONT
        ws.cell(row=rn, column=3).alignment = _CENTER

        qty_cell = ws.cell(row=rn, column=4, value=qty_val)
        qty_cell.border = _BORDER
        qty_cell.font = _BODY_FONT
        qty_cell.alignment = _CENTER

        rn += 1


# ---------------------------------------------------------------------------
# Build comparison data
# ---------------------------------------------------------------------------

def _normalize_for_compare(name: str) -> str:
    """Normalize a work item name for fuzzy matching."""
    txt = name.lower().strip()
    # Remove punctuation (commas, semicolons, parens, brackets, quotes)
    txt = re.sub(r"[,;()\[\]\"']", " ", txt)
    # Remove dashes that are surrounded by spaces (orphaned after paren removal)
    # e.g. "ппгнг- а -hf" → "ппгнг а hf"  (but keep "l-1500", "sds-max")
    txt = re.sub(r"\s+-\s+", " ", txt)
    txt = re.sub(r"\s+-(?=\S)", " ", txt)  # " -hf" → " hf"
    txt = re.sub(r"(?<=\S)-\s+", " ", txt)  # "ппгнг- " → "ппгнг "
    txt = re.sub(r"\s+", " ", txt)
    # Remove trailing colons, periods
    txt = txt.rstrip(":.")
    # Normalize number-unit combos: "22вт" → "22 вт", "2500лм" → "2500 лм"
    txt = re.sub(r"(\d+)(вт|лм|мм|м\b|шт|кв)", r"\1 \2", txt)
    # Truncate to meaningful part
    return txt[:120]


def _extract_height_context(name: str) -> str:
    """Extract height category from a work item name for context matching."""
    nl = name.lower()
    for hc in ("до 5 м", "от 5 до 13", "от 13 до 20", "от 20 до 35"):
        if hc in nl:
            return hc
    return ""


def _work_type(name: str) -> str:
    """Classify work item into a coarse type for matching sanity check."""
    nl = name.lower()
    if "анкерн" in nl:
        return "luminaire_anchor"
    if "настенн" in nl and "светильник" in nl:
        return "luminaire_wall"
    if ("на шпильк" in nl or "к перекрыт" in nl) and "светильник" in nl:
        return "luminaire_ceiling"
    if "светильник" in nl or "cd led" in nl:
        return "luminaire"
    if "указател" in nl and "монтаж" in nl:
        return "indicator_wrk"
    if "указател" in nl or "пиктограмм" in nl or "наклейк" in nl:
        return "indicator"
    # Junction boxes (must be before cable check — "с кабельными вводами")
    if "коробк" in nl or "ltjb" in nl:
        return "junction"
    if "кабел" in nl and "прокладк" in nl and "лотк" in nl:
        return "cable_tray"
    if "кабел" in nl and "прокладк" in nl and "гофр" in nl:
        return "cable_conduit"
    if "кабел" in nl and "прокладк" in nl:
        return "cable_work"
    if "кабел" in nl:
        return "cable_mat"
    if "лоток" in nl and "металлическ" in nl:
        return "tray_work"
    if "лоток" in nl or "лотк" in nl:
        return "tray"
    if "гофр" in nl and "монтаж" in nl:
        return "pvc_work"
    if "гофр" in nl or "пвх" in nl or "труб" in nl:
        return "pvc"
    if "выключател" in nl or "пост управлен" in nl:
        return "switch"
    if "щит" in nl or "цсао" in nl:
        return "panel"
    if "автоматическ" in nl or "проверка" in nl:
        return "pnr"
    if "измерен" in nl or "лаборатори" in nl or "целостност" in nl:
        return "pnr"
    return "other"


# Compatible type groups for matching
_COMPATIBLE_GROUPS = [
    {"luminaire", "luminaire_ceiling", "luminaire_wall", "luminaire_anchor"},
    {"cable_work", "cable_tray", "cable_conduit", "cable_mat"},
    {"indicator", "indicator_wrk"},
    {"tray", "tray_work"},
    {"pvc", "pvc_work"},
]


def _types_compatible(rt: str, ot: str) -> bool:
    """Check if two work types are compatible for matching."""
    if rt == ot:
        return True
    if rt == "other" or ot == "other":
        return True
    for group in _COMPATIBLE_GROUPS:
        if rt in group and ot in group:
            return True
    return False


def _mount_type_match(ref_name: str, our_name: str) -> bool:
    """Check if mount types match (шпильки vs анкер vs настенный)."""
    rt = _work_type(ref_name)
    ot = _work_type(our_name)
    # Only enforce mount type for luminaire sub-types
    lumi_subtypes = {"luminaire_ceiling", "luminaire_wall", "luminaire_anchor"}
    if rt in lumi_subtypes and ot in lumi_subtypes:
        return rt == ot
    return True


# Special name aliases for items that have very different names in spec vs reference
_NAME_ALIASES = {
    "пиктограмма": "наклейк",
    "наклейк": "пиктограмма",
    "пэу": "наклейк",
    "коробки взрывозащищенной соединительной": "коробка клеммная взрывозащищенная",
    "коробка клеммная взрывозащищенная": "коробки взрывозащищенной соединительной",
    "коробки распред": "коробка ответвит",
    "коробка ответвит": "коробки распред",
    "выключателя 1-клавишного в сборе": "выключатель 1-клавишный",
    "выключатель 1-клавишный": "выключателя 1-клавишного",
}

# Keyword-based matching: if both names contain ALL of these keywords → match
_KEYWORD_MATCH_SETS = [
    {"рамк", "atlasdesign", "1"},  # Рамка 1-постовая ATLASDESIGN
    {"рамк", "atlasdesign", "3"},  # Рамка 3-постовая ATLASDESIGN
    {"антикорроз", "лент"},        # Антикоррозийная / антикоррозионная лента
    {"соединител", "45"},          # Соединитель 45х45 мм универсальный
    {"зажим", "крепежн"},          # Зажим крепежный 45х45
    {"заземля", "скоб"},           # Заземляющая скоба на ленте
    {"проводник", "круглы", "8"},     # Проводник круглый диаметром 8 мм
    {"держател", "круглы", "плоск"},  # Держатель для круглых и плоских проводников
    {"плоск", "проводник", "40"},     # Плоский проводник 40x4 мм по периметру
]


def _names_match(ref_name: str, our_name: str) -> float:
    """Score how well two item names match (0..1).

    Returns 1.0 for exact match, 0.9 for strong substring match, etc.
    Refuses to match items of incompatible work types (cable vs luminaire).
    """
    rn = _normalize_for_compare(ref_name)
    on = _normalize_for_compare(our_name)

    # Sanity check: refuse to match incompatible work types
    rt = _work_type(ref_name)
    ot = _work_type(our_name)
    if not _types_compatible(rt, ot):
        return 0.0

    # Refuse to match different luminaire mount types
    if not _mount_type_match(ref_name, our_name):
        return 0.0

    # Refuse to match different conductor types (круглый vs плоский)
    # But allow items that reference both types (e.g., "держатель для круглых и плоских")
    rl = ref_name.lower()
    ol = our_name.lower()
    r_has_round = "круглы" in rl or "круглого" in rl
    r_has_flat = "плоск" in rl
    o_has_round = "круглы" in ol or "круглого" in ol
    o_has_flat = "плоск" in ol
    # Block only when one name is purely round and other purely flat
    if r_has_round and not r_has_flat and o_has_flat and not o_has_round:
        return 0.0
    if r_has_flat and not r_has_round and o_has_round and not o_has_flat:
        return 0.0

    # Refuse to match different luminaire models
    # (prevents CD LED from stealing INSEL row, SLICK from matching ARCTIC, etc.)
    _LUMI_MODELS = ["insel", "slick", "arctic", "cd led", "mercury", "atom", "mars", "luna"]
    r_lumi = [m for m in _LUMI_MODELS if m in rl]
    o_lumi = [m for m in _LUMI_MODELS if m in ol]
    if r_lumi and o_lumi and r_lumi != o_lumi:
        return 0.0

    # Refuse to match different cable brands (ВБШвнг vs ППГнг vs ВВГнг)
    _CABLE_BRANDS = ["вбшвнг", "ппгнг", "ввгнг"]
    r_brands = [b for b in _CABLE_BRANDS if b in rl]
    o_brands = [b for b in _CABLE_BRANDS if b in ol]
    if r_brands and o_brands and r_brands != o_brands:
        return 0.0

    # Refuse to match different cable fire-resistance suffixes within same brand
    # e.g., ВБШвнг-LS ≠ ВБШвнг-FRLS, ППГнг-HF ≠ ППГнг-FRHF
    _FR_SUFFIXES = ["-frhf", "-frls", "-hf", "-ls"]
    r_suf = [s for s in _FR_SUFFIXES if s in rl]
    o_suf = [s for s in _FR_SUFFIXES if s in ol]
    if r_suf and o_suf and r_suf != o_suf:
        return 0.0

    if rn == on:
        return 1.0

    # Substring match (first 40 chars)
    if len(rn) > 10 and len(on) > 10:
        short = rn[:40]
        if short in on or on[:40] in rn:
            # Penalty if one has "Ex" suffix and the other doesn't —
            # distinguishes SLICK30 Ex from SLICK30 non-Ex.
            rn_has_ex = bool(re.search(r'\bex\b', rn, re.IGNORECASE))
            on_has_ex = bool(re.search(r'\bex\b', on, re.IGNORECASE))
            if rn_has_ex != on_has_ex:
                return 0.6  # Reduced score for Ex mismatch
            return 0.9

    # Check name aliases (e.g., "Пиктограмма" ↔ "Наклейка", коробки, выключатели)
    for alias_key, alias_val in _NAME_ALIASES.items():
        if alias_key in rn and alias_val in on:
            return 0.85
    # Also check reversed
    for alias_key, alias_val in _NAME_ALIASES.items():
        if alias_key in on and alias_val in rn:
            return 0.85

    # Check keyword match sets (both names must contain ALL keywords from a set)
    for kw_set in _KEYWORD_MATCH_SETS:
        if all(any(kw in w for w in rn.split()) or kw in rn for kw in kw_set):
            if all(any(kw in w for w in on.split()) or kw in on for kw in kw_set):
                return 0.85

    # Token overlap match
    rw = set(rn.split())
    ow = set(on.split())
    if rw and ow:
        common = len(rw & ow)
        total = max(len(rw), len(ow))
        if common >= 3 and common / total > 0.4:
            return 0.5 + 0.4 * (common / total)

        # Relaxed matching for grounding/lightning items: both names share
        # a work-action prefix (забивка, установка, монтаж, прокладка, окраска)
        # + at least 3 common tokens + ratio > 0.25 (lower threshold because
        # reference names have long catalog number suffixes).
        _ground_lightning_verbs = {
            "забивка", "установка", "монтаж", "прокладка", "окраска",
        }
        shared_verbs = _ground_lightning_verbs & rw & ow
        if shared_verbs and common >= 3 and common / total > 0.25:
            return 0.5 + 0.3 * (common / total)

        # Relaxed matching: same item type keyword + 2 common tokens
        # Handles short spec descriptions vs long reference names
        _item_keywords = {
            "стержней", "стержня", "наконечника", "головки", "соединителя",
            "держателя", "ленты", "спреем", "проводника", "обоймы",
            "держатель", "клик", "токоотвода",
        }
        shared_keywords = _item_keywords & rw & ow
        if shared_keywords and common >= 2 and common / total > 0.2:
            return 0.5 + 0.3 * (common / total)

        # Relaxed matching for materials with key parameters (Вт + Лм)
        if common >= 2:
            # Check if key electrical parameters match (wattage + lumens)
            has_watt = any(w.endswith("вт") or w == "вт" for w in rw & ow)
            has_lumen = any("лм" in w for w in rw & ow)
            has_type = "светильник" in (rw & ow) or "светодиодный" in (rw & ow)
            if has_type and (has_watt or has_lumen) and common >= 2:
                return 0.6

    return 0.0


def _build_comparison(our_rows: list[dict], ref_rows: list[dict]) -> list[dict]:
    """Build comparison rows aligning our VOR with reference.

    Uses positional (sequential) matching that respects height context,
    so that the same material name appearing under different height
    categories is matched to the correct counterpart.

    Returns list of dicts with keys:
      section, name, unit, ref_qty, our_qty, diff, pct_diff, status
    """
    comparison = []

    # ── Build indexed list of our rows with context ──
    # Each entry: (index, row, section_title, height_context)
    our_indexed: list[tuple[int, dict, str, str]] = []
    cur_sec = ""
    for i, r in enumerate(our_rows):
        if r["is_section"]:
            cur_sec = r["name"]
            continue
        hctx = _extract_height_context(r["name"])
        our_indexed.append((i, r, cur_sec, hctx))

    # Track which of our items (by index) were matched
    matched_our_indices: set[int] = set()

    # ── Walk through reference rows ──
    current_section = ""
    ref_height_context = ""  # Track height from WRK rows for their MAT children

    for r in ref_rows:
        if r["is_section"]:
            current_section = r["name"]
            ref_height_context = ""
            comparison.append({
                "section": current_section,
                "name": current_section,
                "unit": "",
                "ref_qty": "",
                "our_qty": "",
                "diff": "",
                "pct_diff": "",
                "status": "section",
            })
            continue

        ref_name = r["name"]
        ref_qty = r["qty"]
        ref_qty_str = r.get("qty_str", str(ref_qty) if ref_qty else "")
        ref_unit = r["unit"]
        is_mat = r["is_material"]

        # Update height context from work rows
        if not is_mat:
            h = _extract_height_context(ref_name)
            if h:
                ref_height_context = h

        # ── Find best matching our row ──
        best_idx = -1
        best_score = 0.0
        best_row = None

        for oi, orow, osec, ohctx in our_indexed:
            if oi in matched_our_indices:
                continue
            # Must match work/material type
            if orow.get("is_material", False) != is_mat:
                continue

            score = _names_match(ref_name, orow["name"])
            if score < 0.5:
                continue

            # Exact name match: strong bonus to prevent height-context
            # mismatch from stealing a perfect match (e.g., INSEL row)
            if score >= 1.0:
                score += 0.5

            # Bonus for matching height context
            if ref_height_context and ohctx:
                if ref_height_context == ohctx:
                    score += 0.3  # Same height = strong bonus
                elif ref_height_context in ohctx or ohctx in ref_height_context:
                    score += 0.2
            elif ref_height_context and not ohctx:
                # For materials, check parent WRK row's height
                # Look back from oi in our_rows to find the nearest WRK row
                parent_height = ""
                for back_i in range(oi - 1, -1, -1):
                    pr = our_rows[back_i]
                    if pr.get("is_section"):
                        break
                    if not pr.get("is_material", False):
                        parent_height = _extract_height_context(pr["name"])
                        break
                if parent_height:
                    if ref_height_context == parent_height:
                        score += 0.3
                    elif ref_height_context in parent_height or parent_height in ref_height_context:
                        score += 0.2

            if score > best_score:
                best_score = score
                best_idx = oi
                best_row = orow

        # Register match
        match = best_row
        if match is not None:
            matched_our_indices.add(best_idx)

        our_qty = 0
        our_qty_str = ""
        if match:
            our_qty = match["qty"]
            our_qty_str = match.get("qty_str", str(our_qty) if our_qty else "")

        # Calculate diff
        diff = ""
        pct_diff = ""
        status = "missing"

        if match is not None:
            if ref_qty == 0 and our_qty == 0:
                status = "exact"
                diff = 0
                pct_diff = "0%"
            elif ref_qty > 0:
                d = our_qty - ref_qty
                diff = d
                pct = abs(d) / ref_qty * 100
                pct_diff = f"{pct:.1f}%"
                if d == 0:
                    status = "exact"
                elif pct <= 10:
                    status = "close"
                elif pct <= 25:
                    status = "moderate"
                else:
                    status = "off"
            else:
                diff = our_qty
                pct_diff = ""
                status = "exact" if our_qty == 0 else "off"
        else:
            diff = -ref_qty if ref_qty else ""
            status = "missing"

        comparison.append({
            "section": current_section,
            "name": ref_name,
            "unit": ref_unit,
            "ref_qty": ref_qty_str if ref_qty_str else "",
            "our_qty": our_qty_str if our_qty_str else "",
            "diff": diff,
            "pct_diff": pct_diff,
            "status": status,
            "is_material": is_mat,
        })

    # ── Add our items that weren't in reference (extras) ──
    for oi, orow, osec, ohctx in our_indexed:
        if oi in matched_our_indices:
            continue
        if orow.get("is_material", False):
            continue
        if orow["qty"] > 0:
            comparison.append({
                "section": "Доп. позиции (нет в эталоне)",
                "name": orow["name"],
                "unit": orow["unit"],
                "ref_qty": "",
                "our_qty": str(orow["qty"]),
                "diff": f"+{orow['qty']}",
                "pct_diff": "NEW",
                "status": "extra",
                "is_material": False,
            })

    return comparison


# ---------------------------------------------------------------------------
# Write Sheet 3: Comparison
# ---------------------------------------------------------------------------

def _write_comparison_sheet(ws, comparison: list[dict]):
    ws.title = "Сравнение"

    headers = ["Секция", "Наименование", "Ед.изм.",
               "Эталон", "Наш", "Разница", "Отклонение", "Статус"]
    col_widths = [25, 75, 8, 12, 12, 12, 12, 12]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _BORDER
        cell.alignment = _CENTER
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Status labels
    status_labels = {
        "exact": "✓ Точно",
        "close": "~ Близко",
        "moderate": "≈ Умеренно",
        "off": "✗ Расхождение",
        "missing": "— Отсутствует",
        "extra": "+ Лишнее",
        "section": "",
    }
    status_fills = {
        "exact": _GOOD_FILL,
        "close": _CLOSE_FILL,
        "moderate": _CLOSE_FILL,
        "off": _BAD_FILL,
        "missing": _BAD_FILL,
        "extra": PatternFill("solid", fgColor="BDD7EE"),
        "section": _SECTION_FILL,
    }

    rn = 2
    prev_section = ""

    # Summary counters
    counts = {"exact": 0, "close": 0, "moderate": 0, "off": 0, "missing": 0, "extra": 0}

    for row in comparison:
        status = row["status"]

        if status == "section":
            # Section header row
            ws.merge_cells(start_row=rn, start_column=1, end_row=rn, end_column=8)
            cell = ws.cell(row=rn, column=1, value=row["name"])
            cell.font = _SECTION_FONT
            cell.fill = _SECTION_FILL
            cell.border = _BORDER
            for c in range(1, 9):
                ws.cell(row=rn, column=c).border = _BORDER
            rn += 1
            continue

        counts[status] = counts.get(status, 0) + 1
        fill = status_fills.get(status)
        is_mat = row.get("is_material", False)
        font = _MAT_FONT if is_mat else _BODY_FONT

        values = [
            "",  # section (shown via headers)
            row["name"],
            row["unit"],
            row["ref_qty"],
            row["our_qty"],
            row["diff"] if row["diff"] != "" else "",
            row["pct_diff"],
            status_labels.get(status, status),
        ]

        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=rn, column=ci, value=val)
            cell.border = _BORDER
            cell.font = font
            cell.alignment = _CENTER if ci >= 3 else _WRAP
            # Color the status column and diff columns
            if ci == 8 and fill:
                cell.fill = fill
            if ci == 6 and isinstance(row["diff"], (int, float)):
                if row["diff"] > 0:
                    cell.font = Font(size=9, name="Times New Roman", color="FF0000")
                elif row["diff"] < 0:
                    cell.font = Font(size=9, name="Times New Roman", color="0000FF")

        rn += 1

    # Summary row
    rn += 1
    ws.merge_cells(start_row=rn, start_column=1, end_row=rn, end_column=2)
    cell = ws.cell(row=rn, column=1, value="ИТОГО")
    cell.font = _HEADER_FONT
    cell.fill = _HEADER_FILL
    cell.border = _BORDER

    rn += 1
    summary_items = [
        ("✓ Точное совпадение", counts.get("exact", 0), _GOOD_FILL),
        ("~ Близко (≤10%)", counts.get("close", 0), _CLOSE_FILL),
        ("≈ Умеренно (≤25%)", counts.get("moderate", 0), _CLOSE_FILL),
        ("✗ Расхождение (>25%)", counts.get("off", 0), _BAD_FILL),
        ("— Отсутствует у нас", counts.get("missing", 0), _BAD_FILL),
        ("+ Лишнее (нет в эталоне)", counts.get("extra", 0),
         PatternFill("solid", fgColor="BDD7EE")),
    ]

    total_items = sum(counts.values())
    total_good = counts.get("exact", 0) + counts.get("close", 0)

    for label, count, fill in summary_items:
        ws.cell(row=rn, column=1, value="").border = _BORDER
        cell = ws.cell(row=rn, column=2, value=label)
        cell.font = _BODY_FONT
        cell.border = _BORDER
        cell.alignment = _WRAP

        cell = ws.cell(row=rn, column=3, value=count)
        cell.font = _HEADER_FONT
        cell.border = _BORDER
        cell.alignment = _CENTER
        cell.fill = fill

        pct = count / total_items * 100 if total_items > 0 else 0
        cell = ws.cell(row=rn, column=4, value=f"{pct:.1f}%")
        cell.font = _BODY_FONT
        cell.border = _BORDER
        cell.alignment = _CENTER

        for c in range(5, 9):
            ws.cell(row=rn, column=c).border = _BORDER

        rn += 1

    # Total row
    rn += 1
    ws.cell(row=rn, column=2, value="ВСЕГО позиций").font = _HEADER_FONT
    ws.cell(row=rn, column=3, value=total_items).font = _HEADER_FONT
    ws.cell(row=rn, column=2).border = _BORDER
    ws.cell(row=rn, column=3).border = _BORDER

    rn += 1
    acc = total_good / total_items * 100 if total_items > 0 else 0
    ws.cell(row=rn, column=2, value="Точность (точно + близко)").font = _HEADER_FONT
    cell = ws.cell(row=rn, column=3, value=f"{acc:.1f}%")
    cell.font = Font(bold=True, size=12, name="Times New Roman",
                     color="006100" if acc > 60 else "9C0006")
    ws.cell(row=rn, column=2).border = _BORDER
    ws.cell(row=rn, column=3).border = _BORDER


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def generate_comparison_xlsx(folder: str, output: str = None):
    """Generate 3-sheet comparison Excel."""
    folder_path = Path(folder)
    if output is None:
        output = str(folder_path / "VOR_COMPARISON.xlsx")

    print("=" * 80)
    print("Генерация сравнительного Excel: Наш ВОР / Эталон / Сравнение")
    print("=" * 80)

    # 1. Find reference VOR (PDF or XLSX) in folder or parent
    ref_file = None
    search_dirs = [folder_path, folder_path.parent]
    for search_dir in search_dirs:
        for ext in ("*.pdf", "*.xlsx"):
            for f in search_dir.glob(ext):
                fl = f.name.lower()
                if "вор" in fl and "generated" not in fl and "comparison" not in fl:
                    ref_file = str(f)
                    break
            if ref_file:
                break
        if ref_file:
            break
    if not ref_file:
        print("ERROR: Эталонный ВОР (PDF/XLSX) не найден в папке или родительской папке!")
        return

    print(f"\nЭталонный ВОР: {Path(ref_file).name}")

    # 2. Parse reference VOR
    print("Парсинг эталонного ВОР...")
    ref_rows = _parse_reference_vor(ref_file)
    print(f"  Извлечено {len(ref_rows)} строк из эталона")

    # 3. Generate our VOR
    print("\nГенерация нашего ВОР...")
    our_rows = _generate_our_vor(str(folder_path))
    print(f"  Сгенерировано {len(our_rows)} строк")

    # 4. Build comparison
    print("\nПостроение сравнения...")
    comparison = _build_comparison(our_rows, ref_rows)
    print(f"  {len(comparison)} строк сравнения")

    # 5. Create Excel workbook
    print("\nСоздание Excel...")
    wb = openpyxl.Workbook()

    # Sheet 1: Our VOR
    ws1 = wb.active
    _write_vor_sheet(ws1, "Наш ВОР", our_rows)

    # Sheet 2: Reference VOR
    ws2 = wb.create_sheet()
    _write_vor_sheet(ws2, "Эталон", ref_rows)

    # Sheet 3: Comparison
    ws3 = wb.create_sheet()
    _write_comparison_sheet(ws3, comparison)

    # Save
    wb.save(output)
    print(f"\nСохранено: {output}")

    # Print summary
    stats = {}
    for row in comparison:
        s = row["status"]
        if s != "section":
            stats[s] = stats.get(s, 0) + 1

    total = sum(stats.values())
    exact = stats.get("exact", 0)
    close = stats.get("close", 0)
    moderate = stats.get("moderate", 0)

    print(f"\n{'='*60}")
    print(f"ИТОГО: {total} позиций")
    print(f"  ✓ Точно:        {exact:>3d} ({exact/total*100:.1f}%)")
    print(f"  ~ Близко (≤10%): {close:>3d} ({close/total*100:.1f}%)")
    print(f"  ≈ Умеренно:     {moderate:>3d} ({moderate/total*100:.1f}%)")
    print(f"  ✗ Расхождение:  {stats.get('off', 0):>3d} ({stats.get('off', 0)/total*100:.1f}%)")
    print(f"  — Отсутствует:  {stats.get('missing', 0):>3d} ({stats.get('missing', 0)/total*100:.1f}%)")
    print(f"  + Лишнее:       {stats.get('extra', 0):>3d} ({stats.get('extra', 0)/total*100:.1f}%)")
    print(f"  Точность (✓+~): {(exact+close)/total*100:.1f}%")


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    logging.getLogger("pdfminer").setLevel(logging.WARNING)
    logging.getLogger("pdfplumber").setLevel(logging.WARNING)

    if len(sys.argv) < 2:
        print("Usage: python vor_comparison_xlsx.py <folder_path> [output.xlsx]")
        sys.exit(1)

    folder = sys.argv[1]
    output = sys.argv[2] if len(sys.argv) > 2 else None
    generate_comparison_xlsx(folder, output)
