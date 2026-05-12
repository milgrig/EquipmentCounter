#!/usr/bin/env python3
"""test_generate_vor_gpk3.py — Generate VOR from PDFs in GPK 3rd capture folder.

Uses the existing pipeline from web_app.py:
  1. _count_equipment_in_pdf: legend extraction + counting
  2. _aggregate_equipment: aggregate counts from multiple PDFs
  3. vor_elevation_grouper: regroup by elevation
  4. vor_height_injector: add vertical cable rows
  5. vor_docx_renderer: render to DOCX (эталонный формат ДБТ)
  6. Export to XLSX using openpyxl

Output files:
  - VOR_TEST_3захватка.docx
  - VOR_TEST_3захватка.xlsx
"""

from __future__ import annotations

import io
import logging
import sys
from pathlib import Path

# Project root
PROJECT_ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(PROJECT_ROOT))

# Import pipeline components from web_app.py
from web_app import _count_equipment_in_pdf, _aggregate_equipment

# Import VOR processing modules
from vor_docx_renderer import render_vor_docx
from vor_elevation_grouper import regroup_aggregate_by_elevation
from vor_height_injector import inject_vertical_cable_rows

# Import openpyxl for XLSX generation
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger("test_generate_vor_gpk3")


# ============================================================================
# Configuration
# ============================================================================

PDF_DIR = Path(r"C:\Cursor\TayfaProject\EquipmentCounter\Data\ДБТ разделы для ИИ\03_ГПК_\3-я захватка\02_PDF")
OUTPUT_DIR = PDF_DIR  # Save outputs in the same folder
OUTPUT_BASE_NAME = "VOR_TEST_3захватка"


# ============================================================================
# Main pipeline
# ============================================================================

def generate_vor_from_pdfs(pdf_dir: Path, output_dir: Path, base_name: str):
    """Generate VOR files (DOCX and XLSX) from PDFs in a folder."""
    log.info("="*80)
    log.info("Starting VOR generation for: %s", pdf_dir)
    log.info("="*80)

    # Step 1: Find all PDF files
    pdf_files = sorted(pdf_dir.glob("*.pdf"))
    if not pdf_files:
        log.error("No PDF files found in: %s", pdf_dir)
        return

    log.info("Found %d PDF files", len(pdf_files))
    for pdf_file in pdf_files:
        log.info("  - %s", pdf_file.name)

    # Step 2: Count equipment in each PDF
    log.info("\n" + "="*80)
    log.info("Step 1: Counting equipment in each PDF...")
    log.info("="*80)
    all_results: dict[str, list[dict]] = {}

    for pdf_path in pdf_files:
        log.info("\nProcessing: %s", pdf_path.name)
        try:
            file_result = _count_equipment_in_pdf(str(pdf_path))
            all_results[pdf_path.name] = file_result
            log.info("  Found %d items", len(file_result))
        except Exception as exc:
            log.error("  Failed: %s", exc, exc_info=True)
            all_results[pdf_path.name] = []

    # Step 3: Aggregate equipment from all PDFs
    log.info("\n" + "="*80)
    log.info("Step 2: Aggregating equipment across all PDFs...")
    log.info("="*80)
    aggregated = _aggregate_equipment(all_results)
    log.info("Aggregated to %d unique items", len(aggregated))

    # Step 4: Regroup by elevation
    log.info("\n" + "="*80)
    log.info("Step 3: Regrouping by elevation...")
    log.info("="*80)
    aggregated = regroup_aggregate_by_elevation(aggregated, labeled=True)
    log.info("Regrouped items: %d", len(aggregated))

    # Step 5: Inject vertical cable rows
    log.info("\n" + "="*80)
    log.info("Step 4: Injecting vertical cable rows...")
    log.info("="*80)
    pdf_paths_str = [str(p) for p in pdf_files]
    aggregated = inject_vertical_cable_rows(aggregated, pdf_paths_str, mode="split")
    log.info("After injection: %d items", len(aggregated))

    # Step 6: Generate DOCX output
    log.info("\n" + "="*80)
    log.info("Step 5: Generating DOCX output...")
    log.info("="*80)
    docx_path = output_dir / f"{base_name}.docx"

    try:
        # Get relative folder name for the document
        rel_folder = pdf_dir.parent.name + " / " + pdf_dir.name

        docx_bytes = render_vor_docx(
            aggregated,
            rel_folder=rel_folder,
            project_name="«Комплекс по глубокой переработке зерна для производства аминокислот, расположенный по адресу: Ростовская область, Батайск»",
            object_name="ГПК (Главный производственный корпус), 3-я захватка",
            section_basis="Рабочие чертежи раздела ЭО, 3-я захватка ГПК",
        )

        with open(docx_path, "wb") as f:
            f.write(docx_bytes)

        log.info("DOCX saved: %s", docx_path)
        log.info("File size: %.1f KB", len(docx_bytes) / 1024)
    except Exception as exc:
        log.error("Failed to generate DOCX: %s", exc, exc_info=True)

    # Step 7: Generate XLSX output
    log.info("\n" + "="*80)
    log.info("Step 6: Generating XLSX output...")
    log.info("="*80)
    xlsx_path = output_dir / f"{base_name}.xlsx"

    try:
        _write_vor_xlsx(aggregated, xlsx_path)
        log.info("XLSX saved: %s", xlsx_path)
    except Exception as exc:
        log.error("Failed to generate XLSX: %s", exc, exc_info=True)

    # Summary
    log.info("\n" + "="*80)
    log.info("VOR generation complete!")
    log.info("="*80)
    log.info("Output files:")
    if docx_path.exists():
        log.info("  DOCX: %s (%.1f KB)", docx_path, docx_path.stat().st_size / 1024)
    if xlsx_path.exists():
        log.info("  XLSX: %s (%.1f KB)", xlsx_path, xlsx_path.stat().st_size / 1024)


def _write_vor_xlsx(aggregated: list[dict], output_path: Path):
    """Write aggregated VOR data to XLSX file.

    Format matches web_app.py api_folder_export_xlsx endpoint.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ВОР"

    # Styling
    header_font = Font(bold=True, size=10)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    thin_border = Border(
        left=Side("thin"),
        right=Side("thin"),
        top=Side("thin"),
        bottom=Side("thin")
    )

    # Column headers and widths
    headers = [
        "№ п/п",
        "Наименование вида работ",
        "Ед. изм.",
        "Объем работ",
        "Формула расчета",
        "Ссылка на чертежи",
        "Доп. информация"
    ]
    col_widths = [7, 72, 9, 12, 20, 26, 27]

    # Write header row
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    # Write data rows
    for idx, row_data in enumerate(aggregated, start=1):
        ri = idx + 1  # +1 because header is row 1
        vals = [
            row_data.get("row", idx),  # Use row field if exists, else use index
            row_data["name"],
            row_data["unit"],
            row_data["total"],
            row_data["formula"],
            row_data["drawing_refs"],
            row_data.get("extra_info", "")
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = thin_border
            # Center-align numeric columns
            if ci in (1, 3, 4):
                cell.alignment = Alignment(horizontal="center")

    # Save workbook
    wb.save(output_path)


# ============================================================================
# Entry point
# ============================================================================

def main():
    """Main entry point."""
    if not PDF_DIR.exists():
        log.error("PDF directory not found: %s", PDF_DIR)
        return 1

    if not OUTPUT_DIR.exists():
        log.error("Output directory not found: %s", OUTPUT_DIR)
        return 1

    generate_vor_from_pdfs(PDF_DIR, OUTPUT_DIR, OUTPUT_BASE_NAME)
    return 0


if __name__ == "__main__":
    sys.exit(main())
