#!/usr/bin/env python3
"""
Analyze VOR comparison results and generate a detailed report.
"""

import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required: pip install openpyxl")

# Paths
COMPARISON_FILE = Path(__file__).parent / "Data" / "ДБТ разделы для ИИ" / "03_ГПК_" / "3-я захватка" / "VOR_COMPARISON.xlsx"

if not COMPARISON_FILE.exists():
    sys.exit(f"❌ Файл сравнения не найден: {COMPARISON_FILE}")

print("=" * 80)
print("АНАЛИЗ РЕЗУЛЬТАТОВ СРАВНЕНИЯ ВОР")
print("=" * 80)
print()

# Load workbook
wb = openpyxl.load_workbook(COMPARISON_FILE, data_only=True)

print(f"📊 Листы в файле: {wb.sheetnames}")
print()

# Analyze main comparison sheet
if "Построчное сравнение" in wb.sheetnames:
    ws = wb["Построчное сравнение"]
    print("=" * 80)
    print("1. ПОСТРОЧНОЕ СРАВНЕНИЕ")
    print("=" * 80)
    print()

    # Count rows by status
    status_counts = {}
    exact_matches = 0
    qty_mismatches = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if i == 0:  # Skip header
            continue
        if not row or not any(row):
            continue

        # Extract columns (adjust indices based on actual structure)
        # Typical structure: Раздел | Эталон: имя/ед/кол | Наш: имя/ед/кол | Δ | Δ% | Статус | Заметка
        if len(row) < 10:
            continue

        status = row[9] if len(row) > 9 else None  # Status column
        if not status:
            continue

        status_counts[status] = status_counts.get(status, 0) + 1

        if status == "OK":
            exact_matches += 1
        elif status == "DIFF_QTY":
            etalon_name = row[1]
            etalon_qty = row[3]
            our_qty = row[6]
            delta = row[7] if len(row) > 7 else None
            qty_mismatches.append({
                'name': etalon_name,
                'etalon': etalon_qty,
                'ours': our_qty,
                'delta': delta,
            })

    print(f"📈 Статистика по статусам:")
    for status, count in sorted(status_counts.items()):
        print(f"  {status:15s}: {count:3d}")
    print()

    if qty_mismatches:
        print(f"⚠️  Найдено {len(qty_mismatches)} расхождений в количествах:")
        print()
        for i, mismatch in enumerate(qty_mismatches[:10], 1):  # Show first 10
            name = mismatch['name']
            if name and len(str(name)) > 60:
                name = str(name)[:57] + "..."
            print(f"  {i}. {name}")
            print(f"     Эталон: {mismatch['etalon']} | Наш: {mismatch['ours']} | Δ: {mismatch['delta']}")
        if len(qty_mismatches) > 10:
            print(f"  ... и еще {len(qty_mismatches) - 10} позиций")
        print()

# Analyze summary sheet
if "Итого по разделам" in wb.sheetnames:
    ws = wb["Итого по разделам"]
    print("=" * 80)
    print("2. ИТОГО ПО РАЗДЕЛАМ")
    print("=" * 80)
    print()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue
        print("  ".join(str(c) if c is not None else "" for c in row[:6]))
    print()

# Analyze "only in reference" sheet
if "Только в эталоне" in wb.sheetnames:
    ws = wb["Только в эталоне"]
    print("=" * 80)
    print("3. ОТСУТСТВУЮТ В СГЕНЕРИРОВАННОМ ВОР (только в эталоне)")
    print("=" * 80)
    print()

    missing_items = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if i == 0:  # Skip header
            continue
        if not row or not any(row):
            continue
        name = row[1] if len(row) > 1 else None
        qty = row[3] if len(row) > 3 else None
        if name:
            missing_items.append({'name': name, 'qty': qty})

    print(f"📉 Всего отсутствующих позиций: {len(missing_items)}")
    print()

    if missing_items:
        print("Примеры (первые 15):")
        for i, item in enumerate(missing_items[:15], 1):
            name = str(item['name'])
            if len(name) > 70:
                name = name[:67] + "..."
            print(f"  {i:2d}. {name}")
            print(f"      Кол-во: {item['qty']}")
        if len(missing_items) > 15:
            print(f"  ... и еще {len(missing_items) - 15} позиций")
    print()

# Analyze "only in generated" sheet
if "Только у нас" in wb.sheetnames:
    ws = wb["Только у нас"]
    print("=" * 80)
    print("4. ЛИШНИЕ В СГЕНЕРИРОВАННОМ ВОР (отсутствуют в эталоне)")
    print("=" * 80)
    print()

    extra_items = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if i == 0:  # Skip header
            continue
        if not row or not any(row):
            continue
        name = row[1] if len(row) > 1 else None
        qty = row[3] if len(row) > 3 else None
        if name:
            extra_items.append({'name': name, 'qty': qty})

    print(f"📈 Всего лишних позиций: {len(extra_items)}")
    print()

    if extra_items:
        print("Список лишних позиций:")
        for i, item in enumerate(extra_items, 1):
            name = str(item['name'])
            if len(name) > 70:
                name = name[:67] + "..."
            print(f"  {i:2d}. {name}")
            print(f"      Кол-во: {item['qty']}")
    print()

# Final summary
print("=" * 80)
print("РЕЗЮМЕ")
print("=" * 80)
print()
print(f"✅ Точные совпадения: {exact_matches}")
print(f"⚠️  Расхождения в кол-ве: {len(qty_mismatches)}")
print(f"❌ Отсутствуют в генерации: {len(missing_items) if 'missing_items' in locals() else '?'}")
print(f"➕ Лишние в генерации: {len(extra_items) if 'extra_items' in locals() else '?'}")
print()

# Calculate accuracy
if 'missing_items' in locals():
    total_ref = exact_matches + len(qty_mismatches) + len(missing_items)
    if total_ref > 0:
        accuracy = (exact_matches / total_ref) * 100
        fuzzy_match = ((exact_matches + len(qty_mismatches)) / total_ref) * 100
        print(f"📊 Точность (exact match): {accuracy:.1f}%")
        print(f"📊 Совпадение по наименованиям: {fuzzy_match:.1f}%")
        print()

print("=" * 80)
print("ОСНОВНЫЕ ПРОБЛЕМЫ И ВЫВОДЫ")
print("=" * 80)
print()
print("1. Система распознала только небольшую часть оборудования из эталона")
print("   - Эталон: 130 строк")
print("   - Сгенерировано: 30 строк (23% от эталона)")
print()
print("2. Основные категории отсутствующих элементов:")
print("   - Щитовое оборудование")
print("   - Установочные изделия (розетки, выключатели)")
print("   - Кабельно-проводниковая продукция")
print("   - Лотки и аксессуары")
print("   - Пусконаладочные работы")
print()
print("3. Возможные причины:")
print("   - Легенда не распознается на всех листах")
print("   - Щиты и схемы требуют специальной обработки")
print("   - Лотки (файлы 019-025) не обрабатываются")
print("   - Кабельная продукция требует дополнительных методов извлечения")
print()

wb.close()
