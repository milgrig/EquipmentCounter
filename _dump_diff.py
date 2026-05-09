import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import load_workbook

wb = load_workbook('Data/ДБТ разделы для ИИ/03_ГПК_/3-я захватка/vor_comparison_GPK3_new.xlsx')
ws = wb['Построчное сравнение']

print('=== DIFF_QTY cases ===')
prev_section = None
for row in ws.iter_rows(min_row=2, values_only=True):
    num, section, et_name, et_unit, et_qty, ou_name, ou_unit, ou_qty, d_abs, d_pct, score, status, note = row
    if status != 'DIFF_QTY':
        continue
    if section != prev_section:
        print('')
        print('--- ' + str(section) + ' ---')
        prev_section = section
    et_n = (et_name or '')[:55]
    ou_n = (ou_name or '')[:55]
    print(f'  ET:{et_qty!s:>7} {et_unit or "":>3} | OUR:{ou_qty!s:>7} {ou_unit or "":>3} | {d_pct!s:>8} | {et_n}')
    if et_n != ou_n:
        print(f'                                                   --> {ou_n}')

print('')
print('=== ONLY_ETALON (пропущено у нас) ===')
prev_section = None
for row in ws.iter_rows(min_row=2, values_only=True):
    num, section, et_name, et_unit, et_qty, ou_name, ou_unit, ou_qty, d_abs, d_pct, score, status, note = row
    if status != 'ONLY_ETALON':
        continue
    if section != prev_section:
        print('')
        print('--- ' + str(section) + ' ---')
        prev_section = section
    print(f'  {et_qty!s:>6} {et_unit or "":>3} | {(et_name or "")[:80]}')

print('')
print('=== ONLY_OURS (лишние у нас) по разделам ===')
from collections import Counter
c = Counter()
for row in ws.iter_rows(min_row=2, values_only=True):
    num, section, et_name, et_unit, et_qty, ou_name, ou_unit, ou_qty, d_abs, d_pct, score, status, note = row
    if status != 'ONLY_OURS':
        continue
    c[section] += 1
for sec, cnt in c.most_common():
    print(f'  {cnt:>3}  {sec}')
