#!/usr/bin/env python3
"""test_vor_cross_format.py -- Cross-format DXF+PDF comparison against reference VOR.

Runs BOTH pipelines (DXF and PDF) on the same test cases and compares each
against the reference VOR Excel.  Produces three comparison axes:

  1. DXF vs Reference VOR  → DXF accuracy (baseline, should be high ~80%+)
  2. PDF vs Reference VOR  → PDF accuracy (what we optimise)
  3. PDF vs DXF            → delta between formats (what PDF loses)

Per-element detail: for every VOR reference item we show whether DXF found it,
whether PDF found it, and whether counts matched.

Usage:
    python test_vor_cross_format.py
    python test_vor_cross_format.py --only abk_eo,pos_28
    python test_vor_cross_format.py --verbose
"""

from __future__ import annotations

import argparse
import difflib
import json
import logging
import re
import sys
import time
import traceback
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

# ── Project root ──────────────────────────────────────────────────────
PROJECT_ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(PROJECT_ROOT))

DATA_DIR = PROJECT_ROOT / "Data"
DBT_DIR = DATA_DIR / "ДБТ разделы для ИИ"

# ── Imports from project ─────────────────────────────────────────────
from pdf_legend_parser import parse_legend  # noqa: E402
from pdf_count_text import count_symbols  # noqa: E402
from pdf_count_cables import extract_cables  # noqa: E402
from pdf_count_visual import match_symbols, detect_pictograms  # noqa: E402
from vor_work_mapping import map_items as vor_map_items  # noqa: E402
from equipment_counter import process_dxf  # noqa: E402

try:
    import openpyxl  # noqa: E402
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

log = logging.getLogger("test_vor_cross_format")


# =====================================================================
#  Test cases — VOR + PDF + DXF directories
# =====================================================================

_ABK = DBT_DIR / "01_АБК" / "Обновленные файлы"

TEST_CASES: list[dict] = [
    {
        "name": "abk_eo",
        "ref_vor": str(_ABK / "ВОР_" / "ВОР АБК ЭО.xlsx"),
        "pdf_dir": str(_ABK / "PDF_" / "ЭО"),
        "dxf_dir": str(_ABK / "DWG" / "ЭО" / "_converted_dxf"),
    },
    {
        "name": "abk_em",
        "ref_vor": str(_ABK / "ВОР_" / "ВОР АБК ЭМ.xlsx"),
        "pdf_dir": str(_ABK / "PDF_" / "ЭМ"),
        "dxf_dir": str(_ABK / "DWG" / "ЭМ" / "_converted_dxf"),
    },
    {
        "name": "abk_eg",
        "ref_vor": str(_ABK / "ВОР_" / "ВОР АБК ЭГ.xlsx"),
        "pdf_dir": str(_ABK / "PDF_" / "ЭГ"),
        "dxf_dir": str(_ABK / "DWG" / "ЭГ" / "_converted_dxf"),
    },
    {
        "name": "pos_12_3",
        "ref_vor": str(DBT_DIR / "12.3 Насосная станция поверхностных стоков"
                       / "ВОР поз. 12.3.xlsx"),
        "pdf_dir": str(DBT_DIR / "12.3 Насосная станция поверхностных стоков"
                       / "02_PDF"),
        "dxf_dir": str(DBT_DIR / "12.3 Насосная станция поверхностных стоков"
                       / "01_DWG" / "_converted_dxf"),
    },
    {
        "name": "pos_16_2",
        "ref_vor": str(DBT_DIR / "16.2_ЖД КПП" / "Обновленные файлы 1"
                       / "ВОР поз. 16.2.xlsx"),
        "pdf_dir": str(DBT_DIR / "16.2_ЖД КПП" / "Обновленные файлы 1"
                       / "02_PDF"),
        "dxf_dir": "",  # no DXF available
    },
    {
        "name": "pos_27",
        "ref_vor": str(DBT_DIR / "27_Склад вспомогательных материалов "
                       "с участком погрузки крытых вагонов"
                       / "Обновленные файлы" / "ВОР поз.27.xlsx"),
        "pdf_dir": str(DBT_DIR / "27_Склад вспомогательных материалов "
                       "с участком погрузки крытых вагонов"
                       / "Обновленные файлы" / "03_PDF"),
        "dxf_dir": str(DBT_DIR / "27_Склад вспомогательных материалов "
                       "с участком погрузки крытых вагонов"
                       / "Обновленные файлы" / "02_DWG" / "_converted_dxf"),
    },
    {
        "name": "pos_28",
        "ref_vor": str(DBT_DIR / "28. Автовесы" / "ВОР поз.28.xlsx"),
        "pdf_dir": str(DBT_DIR / "28. Автовесы" / "02_PDF"),
        "dxf_dir": str(DBT_DIR / "28. Автовесы" / "01_DWG" / "_converted_dxf"),
    },
    {
        "name": "pos_30",
        "ref_vor": str(DBT_DIR / "30. КПП" / "ВОР поз.30.xlsx"),
        "pdf_dir": str(DBT_DIR / "30. КПП" / "03_PDF"),
        "dxf_dir": str(DBT_DIR / "30. КПП" / "02_DWG" / "_converted_dxf"),
    },
    {
        "name": "pos_8_2_em",
        "ref_vor": str(DBT_DIR / "8.2 Участок хранения сульфата аммония"
                       / "ЭМ" / "ВОР ЭМ поз. 8.2.xlsx"),
        "pdf_dir": str(DBT_DIR / "8.2 Участок хранения сульфата аммония"
                       / "ЭМ" / "02_PDF"),
        "dxf_dir": str(DBT_DIR / "8.2 Участок хранения сульфата аммония"
                       / "ЭМ" / "Обновленные файлы" / "01_DWG" / "_converted_dxf"),
    },
    {
        "name": "pos_8_2_eo",
        "ref_vor": str(DBT_DIR / "8.2 Участок хранения сульфата аммония"
                       / "ЭО" / "ВОР поз. 8.2.xlsx"),
        "pdf_dir": str(DBT_DIR / "8.2 Участок хранения сульфата аммония"
                       / "ЭО" / "02_PDF"),
        "dxf_dir": str(DBT_DIR / "8.2 Участок хранения сульфата аммония"
                       / "ЭО" / "01_DWG" / "_converted_dxf"),
    },
]


# =====================================================================
#  Data structures
# =====================================================================

@dataclass
class RefItem:
    """One row from the reference VOR Excel."""
    row_num: str
    name: str
    unit: str
    qty: float
    is_section: bool = False

    @property
    def key(self) -> str:
        return _normalize(self.name)


@dataclass
class GenItem:
    """Aggregated item from any pipeline (PDF or DXF)."""
    name: str
    total: float
    files: list[str] = field(default_factory=list)

    @property
    def key(self) -> str:
        return _normalize(self.name)


@dataclass
class ElementComparison:
    """Per-element comparison: one VOR ref item vs DXF and PDF results."""
    ref_name: str
    ref_qty: float
    # DXF results
    dxf_found: bool
    dxf_name: str = ""
    dxf_qty: float = 0
    dxf_match_ratio: float = 0.0
    dxf_qty_match: bool = False
    # PDF results
    pdf_found: bool = False
    pdf_name: str = ""
    pdf_qty: float = 0
    pdf_match_ratio: float = 0.0
    pdf_qty_match: bool = False


@dataclass
class CaseResult:
    name: str
    status: str  # "ok", "error", "skip"
    error_msg: str = ""
    has_dxf: bool = False
    ref_item_count: int = 0
    dxf_item_count: int = 0
    pdf_item_count: int = 0
    pdf_count: int = 0
    dxf_file_count: int = 0
    elapsed_sec: float = 0.0
    # DXF vs VOR
    dxf_matched: int = 0
    dxf_count_mismatch: int = 0
    dxf_missing: int = 0
    # PDF vs VOR
    pdf_matched: int = 0
    pdf_count_mismatch: int = 0
    pdf_missing: int = 0
    # Per-element details
    elements: list[ElementComparison] = field(default_factory=list)
    # Items found by DXF but not PDF (improvement targets)
    dxf_only: list[dict] = field(default_factory=list)
    # Extra items not in reference
    pdf_extra: list[dict] = field(default_factory=list)
    dxf_extra: list[dict] = field(default_factory=list)

    @property
    def dxf_name_match_pct(self) -> float:
        total = self.dxf_matched + self.dxf_count_mismatch + self.dxf_missing
        return (self.dxf_matched + self.dxf_count_mismatch) / total * 100 if total else 0

    @property
    def dxf_exact_pct(self) -> float:
        total = self.dxf_matched + self.dxf_count_mismatch + self.dxf_missing
        return self.dxf_matched / total * 100 if total else 0

    @property
    def pdf_name_match_pct(self) -> float:
        total = self.pdf_matched + self.pdf_count_mismatch + self.pdf_missing
        return (self.pdf_matched + self.pdf_count_mismatch) / total * 100 if total else 0

    @property
    def pdf_exact_pct(self) -> float:
        total = self.pdf_matched + self.pdf_count_mismatch + self.pdf_missing
        return self.pdf_matched / total * 100 if total else 0


# =====================================================================
#  Normalization (reused from test_vor_accuracy.py)
# =====================================================================

_AUTO_DETECT_RE = re.compile(
    r"^\[Auto-detected\s|^\[Обозначение\s",
    re.IGNORECASE,
)


def _is_auto_detected(name: str) -> bool:
    return bool(_AUTO_DETECT_RE.match(name))


def _normalize(text: str) -> str:
    """Normalize a description for fuzzy comparison."""
    s = text.strip().lower()
    s = re.sub(
        r"^(монтаж|установка|прокладка|забивка|крепление|окраска|подключение)\s+",
        "", s,
    )
    s = re.sub(r"светильник\w*", "светильник", s)
    s = re.sub(r"розетк\w*", "розетка", s)
    s = re.sub(r"выключател\w*", "выключатель", s)
    s = re.sub(r"коробк\w*", "коробка", s)
    s = re.sub(r"датчик\w*", "датчик", s)
    s = re.sub(r"лот[оа]к\w*", "лоток", s)
    s = re.sub(r"указател\w*", "указатель", s)
    s = re.sub(r"щит\w*", "щит", s)
    s = re.sub(r"труб\w*", "труба", s)
    s = re.sub(r"автомат\w*", "автомат", s)
    s = re.sub(r"кабел\w*", "кабель", s)
    s = re.sub(r"провод\w*(?!\S)", "провод", s)
    s = re.sub(r"светодиодн\w*\s+светильник", "светильник", s)
    s = re.sub(r"светильник\s+светодиодн\w*", "светильник", s)
    s = re.sub(r"\s+в\s+подвесн\w*\s+потолк\w*", "", s)
    s = re.sub(r"\s+(?:на\s+)?(?:высоте?\s+)?(?:до|от)\s+\d+\s*(?:до\s+\d+\s*)?(?:метр\w*|м)\b", "", s)
    s = re.sub(r"\s+настенн\w*", "", s)
    s = re.sub(r"\s+потолочн\w*", "", s)
    s = re.sub(r"\s+(?:подвесн\w*|анкерн\w*|на\s+шпильк\w*)", "", s)
    s = re.sub(r"\s+в\s+сборе", "", s)
    s = re.sub(r"щит\s+(?:распределительн\w*|осветительн\w*|силов\w*)\s*", "щит ", s)
    s = re.sub(r"лоток\s+кабельн\w*\s*", "лоток ", s)
    s = re.sub(r"\s*аварийн\w*\s*", " ", s)
    s = re.sub(r",?\s*\d{5,}[\d\s,]*$", "", s)
    s = re.sub(
        r",?\s*(?:DKS|Ostec|EKF|IEK|ABB|Systeme\s+Electric|СЗИМ)\b.*$",
        "", s, flags=re.IGNORECASE,
    )
    s = re.sub(r"\b[A-Z]{2,}\d[\w-]{3,}\b", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\((?:до|от|суммарн|однополюс|двухполюс|масс\w*)[^)]*\)", "", s)
    s = re.sub(r"\(прокладка\)", "", s)
    s = re.sub(
        r",?\s*(?:цвет\s+\S+|белый|серый|чёрный|черный)\s*,?",
        "", s, flags=re.IGNORECASE,
    )
    s = re.sub(r"\d{3,4}\s*[kкК]\s*", "", s)
    s = re.sub(r"\d+\s*[вwВW]т\w*", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\d+\s*[лl]м\w*", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\bсечением?\b\s*", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    s = s.rstrip(",.:; ")
    return s


# =====================================================================
#  Read reference VOR Excel
# =====================================================================

def read_reference_vor(xlsx_path: str) -> list[RefItem]:
    """Read reference VOR Excel and return list of RefItem."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    ws = None
    for sname in wb.sheetnames:
        if "заказчик" in sname.lower():
            ws = wb[sname]
            break
    if ws is None:
        for sname in wb.sheetnames:
            candidate = wb[sname]
            if candidate.max_row and candidate.max_row > 10:
                ws = candidate
                break
    if ws is None:
        ws = wb.active

    items: list[RefItem] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if not row or len(row) < 4:
            continue
        col_a, col_b, col_c, col_d = row[0], row[1], row[2], row[3]
        name = str(col_b or "").strip()
        if not name:
            continue
        if col_d is None and col_a is None:
            items.append(RefItem(row_num="", name=name, unit="", qty=0, is_section=True))
            continue
        try:
            qty = float(col_d) if col_d is not None else 0
        except (ValueError, TypeError):
            m = re.search(r"[\d]+(?:[.,]\d+)?", str(col_d or ""))
            qty = float(m.group().replace(",", ".")) if m else 0
        if qty <= 0:
            continue
        unit = str(col_c or "").strip()
        row_num = str(col_a or "").strip()
        items.append(RefItem(row_num=row_num, name=name, unit=unit, qty=qty))

    wb.close()
    return items


# =====================================================================
#  PDF Pipeline (same as test_vor_accuracy.py)
# =====================================================================

def _count_equipment_in_pdf(pdf_path: str) -> list[dict]:
    """Run legend extraction + counting on a single PDF."""
    legend_result = parse_legend(pdf_path)
    if not legend_result.items:
        return []

    items: list[dict] = []

    text_counts: dict[str, int] = {}
    try:
        text_result = count_symbols(pdf_path, legend_result)
        text_counts = text_result.counts
    except Exception as exc:
        log.warning("Text counting failed for %s: %s", pdf_path, exc)

    visual_counts: dict[int, int] = {}
    symbols_needing_visual = []
    for idx, item in enumerate(legend_result.items):
        sym = item.symbol or ""
        if sym and text_counts.get(sym, 0) == 0:
            symbols_needing_visual.append(idx)
        elif not sym:
            symbols_needing_visual.append(idx)

    if symbols_needing_visual:
        try:
            vis_result = match_symbols(pdf_path, legend_result)
            visual_counts = vis_result.counts
        except Exception as exc:
            log.warning("Visual counting failed for %s: %s", pdf_path, exc)

    for idx, item in enumerate(legend_result.items):
        sym = item.symbol or ""
        name = item.description or ""
        if not name:
            continue
        count = 0
        if sym and text_counts.get(sym, 0) > 0:
            count = text_counts[sym]
        elif idx in visual_counts and visual_counts[idx] > 0:
            count = visual_counts[idx]
        if count <= 0:
            continue
        items.append({
            "symbol": sym, "name": name,
            "count": count, "count_ae": 0, "total": count,
        })

    # Pictogram detection (T149)
    try:
        picto_result = detect_pictograms(pdf_path, legend_result)
        for pname, pcount in picto_result.counts.items():
            if pcount > 0:
                items.append({
                    "symbol": "", "name": pname,
                    "count": pcount, "count_ae": 0, "total": pcount,
                })
    except Exception as exc:
        log.warning("Pictogram detection failed for %s: %s", pdf_path, exc)

    try:
        cable_result = extract_cables(pdf_path, legend_result)
        for entry in cable_result.cable_schedule:
            group = entry.get("group", "")
            panel = entry.get("panel", "")
            cable_types = entry.get("cable_types", [])
            cross_sections = entry.get("cross_sections", [])
            run_count = entry.get("run_count", 0) or 0
            total_length_m = entry.get("total_length_m", 0) or 0
            type_label = (cable_types[0] if cable_types
                          else cross_sections[0] if cross_sections else "")
            if not type_label and not group:
                continue
            if run_count > 0:
                cable_name = f"Кабель {type_label}" if type_label else "Кабель"
                if group:
                    cable_name += f" ({panel}-{group})" if panel else f" ({group})"
                items.append({
                    "symbol": "", "name": cable_name,
                    "count": run_count, "count_ae": 0,
                    "total": run_count, "unit": "шт",
                })
            if total_length_m > 0:
                cable_name_m = (f"Кабель {type_label} (прокладка)"
                                if type_label else "Кабель (прокладка)")
                if group:
                    cable_name_m += f" ({panel}-{group})" if panel else f" ({group})"
                items.append({
                    "symbol": "", "name": cable_name_m,
                    "count": 0, "count_ae": 0,
                    "total": round(total_length_m, 1), "unit": "м",
                })
    except Exception as exc:
        log.warning("Cable extraction failed for %s: %s", pdf_path, exc)

    items = vor_map_items(items)
    return items


def process_pdf_folder(folder_path: str) -> tuple[dict[str, GenItem], int, int]:
    """Process all PDFs in folder via PDF pipeline. Returns (items_by_key, count, errors)."""
    pdf_dir = Path(folder_path)
    pdf_files = sorted(pdf_dir.glob("*.pdf"))
    all_items: dict[str, GenItem] = {}
    processed = 0
    errors = 0

    for pdf_path in pdf_files:
        try:
            items = _count_equipment_in_pdf(str(pdf_path))
            for item in items:
                work_name = item.get("work_name", "").strip()
                raw_name = item.get("name", "").strip()
                name = work_name or raw_name
                if not name or _is_auto_detected(name):
                    continue
                total = item.get("total", item.get("count", 0) + item.get("count_ae", 0))
                if total <= 0:
                    continue
                key = _normalize(name)
                if key in all_items:
                    all_items[key].total += total
                    all_items[key].files.append(pdf_path.name)
                else:
                    all_items[key] = GenItem(name=name, total=total, files=[pdf_path.name])
            processed += 1
        except Exception as e:
            errors += 1
            log.warning("PDF error %s: %s", pdf_path.name, e)

    return all_items, processed, errors


# =====================================================================
#  DXF Pipeline
# =====================================================================

def process_dxf_folder(folder_path: str) -> tuple[dict[str, GenItem], int, int]:
    """Process all DXF files in folder via process_dxf(). Returns (items_by_key, count, errors)."""
    dxf_dir = Path(folder_path)
    if not dxf_dir.exists():
        return {}, 0, 0

    dxf_files = sorted(dxf_dir.glob("*.dxf"))
    all_items: dict[str, GenItem] = {}
    processed = 0
    errors = 0

    for dxf_path in dxf_files:
        try:
            equipment = process_dxf(str(dxf_path))
            # Convert to dicts and apply VOR mapping (same as PDF pipeline)
            raw_items = []
            for eq in equipment:
                name = eq.name or ""
                if not name or name.startswith("[?"):
                    continue
                total = eq.count + eq.count_ae
                if total <= 0:
                    continue
                raw_items.append({
                    "symbol": eq.symbol or "", "name": name,
                    "count": eq.count, "count_ae": eq.count_ae,
                    "total": total,
                })
            # Apply VOR work-name mapping for fair comparison
            mapped_items = vor_map_items(raw_items)
            for item in mapped_items:
                work_name = item.get("work_name", "").strip()
                raw_name = item.get("name", "").strip()
                name = work_name or raw_name
                if not name:
                    continue
                total = item.get("total", item.get("count", 0) + item.get("count_ae", 0))
                if total <= 0:
                    continue
                key = _normalize(name)
                if key in all_items:
                    all_items[key].total += total
                    all_items[key].files.append(dxf_path.name)
                else:
                    all_items[key] = GenItem(name=name, total=total, files=[dxf_path.name])
            processed += 1
        except Exception as e:
            errors += 1
            log.warning("DXF error %s: %s", dxf_path.name, e)

    return all_items, processed, errors


# =====================================================================
#  Fuzzy comparison: generated items vs reference
# =====================================================================

def _fuzzy_compare(
    generated: dict[str, GenItem],
    reference: list[RefItem],
    threshold: float = 0.45,
) -> tuple[list[dict], list[RefItem], list[GenItem]]:
    """Compare generated items against reference using fuzzy matching.

    Returns:
        (matches, missing_refs, extra_gen)
        Each match is a dict with ref_name, ref_qty, gen_name, gen_qty,
        match_ratio, qty_match, delta.
    """
    ref_items = [r for r in reference if not r.is_section]
    # Aggregate reference by normalized key
    ref_by_key: dict[str, RefItem] = {}
    for item in ref_items:
        key = item.key
        if key in ref_by_key:
            ref_by_key[key] = RefItem(
                row_num=ref_by_key[key].row_num, name=ref_by_key[key].name,
                unit=ref_by_key[key].unit, qty=ref_by_key[key].qty + item.qty,
            )
        else:
            ref_by_key[key] = RefItem(
                row_num=item.row_num, name=item.name,
                unit=item.unit, qty=item.qty,
            )

    matches: list[dict] = []
    missing: list[RefItem] = []
    used_gen: set[str] = set()

    # Pass 1: exact key match
    matched_ref: set[str] = set()
    for ref_key, ref_item in ref_by_key.items():
        if ref_key in generated:
            gen = generated[ref_key]
            used_gen.add(ref_key)
            matched_ref.add(ref_key)
            delta = gen.total - ref_item.qty
            matches.append({
                "ref_name": ref_item.name, "ref_qty": ref_item.qty,
                "gen_name": gen.name, "gen_qty": gen.total,
                "match_ratio": 1.0, "qty_match": abs(delta) < 0.01,
                "delta": delta,
            })

    # Pass 2: fuzzy match remaining
    remaining_ref = {k: v for k, v in ref_by_key.items() if k not in matched_ref}
    remaining_gen = {k: v for k, v in generated.items() if k not in used_gen}

    for ref_key, ref_item in remaining_ref.items():
        best_ratio = 0.0
        best_gen_key = None
        for gen_key in remaining_gen:
            if gen_key in used_gen:
                continue
            ratio = difflib.SequenceMatcher(None, ref_key, gen_key).ratio()
            if ratio > best_ratio:
                best_ratio = ratio
                best_gen_key = gen_key

        if best_ratio >= threshold and best_gen_key is not None:
            gen = remaining_gen[best_gen_key]
            used_gen.add(best_gen_key)
            delta = gen.total - ref_item.qty
            matches.append({
                "ref_name": ref_item.name, "ref_qty": ref_item.qty,
                "gen_name": gen.name, "gen_qty": gen.total,
                "match_ratio": best_ratio, "qty_match": abs(delta) < 0.01,
                "delta": delta,
            })
        else:
            missing.append(ref_item)

    extra = [v for k, v in generated.items() if k not in used_gen]
    return matches, missing, extra


# =====================================================================
#  Per-element cross comparison
# =====================================================================

def _build_element_comparison(
    reference: list[RefItem],
    dxf_matches: list[dict],
    dxf_missing: list[RefItem],
    pdf_matches: list[dict],
    pdf_missing: list[RefItem],
) -> tuple[list[ElementComparison], list[dict]]:
    """Build per-element comparison showing DXF and PDF status for each ref item.

    Returns:
        (elements, dxf_only_items)
        dxf_only_items: items found by DXF but not PDF (improvement targets)
    """
    ref_items = [r for r in reference if not r.is_section]
    # Aggregate
    ref_by_key: dict[str, RefItem] = {}
    for item in ref_items:
        key = item.key
        if key in ref_by_key:
            ref_by_key[key] = RefItem(
                row_num=ref_by_key[key].row_num, name=ref_by_key[key].name,
                unit=ref_by_key[key].unit, qty=ref_by_key[key].qty + item.qty,
            )
        else:
            ref_by_key[key] = RefItem(
                row_num=item.row_num, name=item.name,
                unit=item.unit, qty=item.qty,
            )

    # Index DXF matches by normalized ref_name
    dxf_by_ref: dict[str, dict] = {}
    for m in dxf_matches:
        dxf_by_ref[_normalize(m["ref_name"])] = m

    # Index PDF matches by normalized ref_name
    pdf_by_ref: dict[str, dict] = {}
    for m in pdf_matches:
        pdf_by_ref[_normalize(m["ref_name"])] = m

    elements: list[ElementComparison] = []
    dxf_only: list[dict] = []

    for ref_key, ref_item in ref_by_key.items():
        dxf_m = dxf_by_ref.get(ref_key)
        pdf_m = pdf_by_ref.get(ref_key)

        elem = ElementComparison(
            ref_name=ref_item.name,
            ref_qty=ref_item.qty,
            dxf_found=dxf_m is not None,
            dxf_name=dxf_m["gen_name"] if dxf_m else "",
            dxf_qty=dxf_m["gen_qty"] if dxf_m else 0,
            dxf_match_ratio=dxf_m["match_ratio"] if dxf_m else 0,
            dxf_qty_match=dxf_m["qty_match"] if dxf_m else False,
            pdf_found=pdf_m is not None,
            pdf_name=pdf_m["gen_name"] if pdf_m else "",
            pdf_qty=pdf_m["gen_qty"] if pdf_m else 0,
            pdf_match_ratio=pdf_m["match_ratio"] if pdf_m else 0,
            pdf_qty_match=pdf_m["qty_match"] if pdf_m else False,
        )
        elements.append(elem)

        # Track items found by DXF but not by PDF
        if dxf_m and not pdf_m:
            dxf_only.append({
                "ref_name": ref_item.name,
                "ref_qty": ref_item.qty,
                "dxf_name": dxf_m["gen_name"],
                "dxf_qty": dxf_m["gen_qty"],
                "dxf_ratio": dxf_m["match_ratio"],
            })

    return elements, dxf_only


# =====================================================================
#  Run one test case
# =====================================================================

def run_case(case: dict, threshold: float = 0.45) -> CaseResult:
    """Run a single cross-format test case."""
    name = case["name"]
    ref_path = case["ref_vor"]
    pdf_dir = case["pdf_dir"]
    dxf_dir = case.get("dxf_dir", "")

    if not Path(ref_path).exists():
        return CaseResult(name=name, status="skip",
                          error_msg=f"VOR not found: {Path(ref_path).name}")
    if not Path(pdf_dir).exists():
        return CaseResult(name=name, status="skip",
                          error_msg=f"PDF folder not found")

    has_dxf = bool(dxf_dir) and Path(dxf_dir).exists()
    pdf_count = len(list(Path(pdf_dir).glob("*.pdf")))
    dxf_file_count = len(list(Path(dxf_dir).glob("*.dxf"))) if has_dxf else 0

    t0 = time.time()
    try:
        # 1. Read reference VOR
        reference = read_reference_vor(ref_path)
        ref_data = [r for r in reference if not r.is_section]
        if not ref_data:
            return CaseResult(name=name, status="error",
                              error_msg="No valid data in reference VOR",
                              elapsed_sec=time.time() - t0)

        # 2. Run PDF pipeline
        pdf_items, pdf_proc, pdf_errs = process_pdf_folder(pdf_dir)

        # 3. Run DXF pipeline (if available)
        dxf_items: dict[str, GenItem] = {}
        if has_dxf:
            dxf_items, dxf_proc, dxf_errs = process_dxf_folder(dxf_dir)

        # 4. Compare DXF vs VOR
        dxf_matches, dxf_missing_refs, dxf_extra = [], [], []
        if has_dxf:
            dxf_matches, dxf_missing_refs, dxf_extra = _fuzzy_compare(
                dxf_items, reference, threshold)

        # 5. Compare PDF vs VOR
        pdf_matches, pdf_missing_refs, pdf_extra = _fuzzy_compare(
            pdf_items, reference, threshold)

        # 6. Build per-element comparison
        elements, dxf_only = [], []
        if has_dxf:
            elements, dxf_only = _build_element_comparison(
                reference, dxf_matches, dxf_missing_refs,
                pdf_matches, pdf_missing_refs)

        # Count DXF results
        dxf_exact = sum(1 for m in dxf_matches if m["qty_match"])
        dxf_mismatch = len(dxf_matches) - dxf_exact

        # Count PDF results
        pdf_exact = sum(1 for m in pdf_matches if m["qty_match"])
        pdf_mismatch = len(pdf_matches) - pdf_exact

        result = CaseResult(
            name=name, status="ok", has_dxf=has_dxf,
            ref_item_count=len(ref_data),
            dxf_item_count=len(dxf_items),
            pdf_item_count=len(pdf_items),
            pdf_count=pdf_count,
            dxf_file_count=dxf_file_count,
            elapsed_sec=time.time() - t0,
            # DXF vs VOR
            dxf_matched=dxf_exact,
            dxf_count_mismatch=dxf_mismatch,
            dxf_missing=len(dxf_missing_refs),
            # PDF vs VOR
            pdf_matched=pdf_exact,
            pdf_count_mismatch=pdf_mismatch,
            pdf_missing=len(pdf_missing_refs),
            # Per-element
            elements=elements,
            dxf_only=dxf_only,
            pdf_extra=[{"name": g.name, "qty": g.total} for g in pdf_extra[:20]],
            dxf_extra=[{"name": g.name, "qty": g.total} for g in dxf_extra[:20]],
        )
        return result

    except Exception as e:
        return CaseResult(
            name=name, status="error",
            error_msg=f"{e}\n{traceback.format_exc()[-400:]}",
            elapsed_sec=time.time() - t0,
        )


# =====================================================================
#  Main
# =====================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Cross-format DXF+PDF comparison against reference VOR",
    )
    parser.add_argument("--only", default="",
                        help="Comma-separated case names (default: all)")
    parser.add_argument("--verbose", "-v", action="store_true",
                        help="Print per-element details")
    parser.add_argument("--threshold", type=float, default=0.45,
                        help="Fuzzy match threshold (default: 0.45)")
    args = parser.parse_args()

    cases = TEST_CASES
    if args.only:
        names = {n.strip() for n in args.only.split(",")}
        cases = [c for c in cases if c["name"] in names]
        if not cases:
            print(f"No matching cases for: {args.only}")
            print(f"Available: {', '.join(c['name'] for c in TEST_CASES)}")
            sys.exit(1)

    print("=" * 80)
    print("  CROSS-FORMAT TEST: DXF + PDF vs Reference VOR")
    print("=" * 80)
    print(f"  Cases: {len(cases)}  |  Threshold: {args.threshold}")
    print()

    # ── Aggregate counters ──
    agg = {
        "total_ref": 0,
        "dxf_matched": 0, "dxf_mismatch": 0, "dxf_missing": 0,
        "pdf_matched": 0, "pdf_mismatch": 0, "pdf_missing": 0,
        "dxf_only_count": 0,
    }
    results: list[CaseResult] = []

    for idx, case in enumerate(cases, 1):
        print(f"\n{'_' * 70}")
        print(f"[{idx}/{len(cases)}] {case['name']}")

        res = run_case(case, threshold=args.threshold)
        results.append(res)

        if res.status != "ok":
            print(f"  {res.status.upper()}: {res.error_msg}")
            continue

        dxf_tag = f"DXF({res.dxf_file_count})" if res.has_dxf else "no DXF"
        print(f"  Ref: {res.ref_item_count} items  |  "
              f"PDF({res.pdf_count}): {res.pdf_item_count} items  |  "
              f"{dxf_tag}: {res.dxf_item_count} items")

        # DXF vs VOR
        if res.has_dxf:
            print(f"\n  DXF vs VOR:  exact={res.dxf_matched}  "
                  f"mismatch={res.dxf_count_mismatch}  "
                  f"missing={res.dxf_missing}  "
                  f"name_match={res.dxf_name_match_pct:.1f}%  "
                  f"exact={res.dxf_exact_pct:.1f}%")

        # PDF vs VOR
        print(f"  PDF vs VOR:  exact={res.pdf_matched}  "
              f"mismatch={res.pdf_count_mismatch}  "
              f"missing={res.pdf_missing}  "
              f"name_match={res.pdf_name_match_pct:.1f}%  "
              f"exact={res.pdf_exact_pct:.1f}%")

        # DXF-only items (found by DXF but not PDF) — improvement targets
        if res.dxf_only:
            print(f"\n  ⚡ DXF found but PDF missed ({len(res.dxf_only)} items):")
            for item in res.dxf_only[:5]:
                print(f"    - {item['ref_name'][:55]:55s}  "
                      f"ref={item['ref_qty']:.0f}  dxf={item['dxf_qty']:.0f}")

        # Verbose: per-element table
        if args.verbose and res.elements:
            print(f"\n  Per-element comparison:")
            print(f"  {'Ref item':<45s} {'RefQ':>5} "
                  f"{'DXF?':>5} {'DXFQ':>5} "
                  f"{'PDF?':>5} {'PDFQ':>5}")
            print(f"  {'-'*45} {'-'*5} {'-'*5} {'-'*5} {'-'*5} {'-'*5}")
            for elem in res.elements:
                dxf_mark = "Y" if elem.dxf_found else "-"
                pdf_mark = "Y" if elem.pdf_found else "-"
                dxf_q = f"{elem.dxf_qty:.0f}" if elem.dxf_found else "-"
                pdf_q = f"{elem.pdf_qty:.0f}" if elem.pdf_found else "-"
                print(f"  {elem.ref_name[:45]:<45s} {elem.ref_qty:>5.0f} "
                      f"{dxf_mark:>5} {dxf_q:>5} "
                      f"{pdf_mark:>5} {pdf_q:>5}")

        # Accumulate
        agg["total_ref"] += res.ref_item_count
        agg["dxf_matched"] += res.dxf_matched
        agg["dxf_mismatch"] += res.dxf_count_mismatch
        agg["dxf_missing"] += res.dxf_missing
        agg["pdf_matched"] += res.pdf_matched
        agg["pdf_mismatch"] += res.pdf_count_mismatch
        agg["pdf_missing"] += res.pdf_missing
        agg["dxf_only_count"] += len(res.dxf_only)

    # ── Overall summary ──
    print(f"\n{'=' * 80}")
    print("  OVERALL CROSS-FORMAT SUMMARY")
    print(f"{'=' * 80}")

    ok_count = sum(1 for r in results if r.status == "ok")
    total_ref = agg["total_ref"]

    # DXF accuracy (baseline)
    # Only count ref items from cases that have DXF
    dxf_ref = sum(r.ref_item_count for r in results if r.status == "ok" and r.has_dxf)
    dxf_nm = agg["dxf_matched"] + agg["dxf_mismatch"]
    dxf_nm_pct = dxf_nm / dxf_ref * 100 if dxf_ref else 0
    dxf_ex_pct = agg["dxf_matched"] / dxf_ref * 100 if dxf_ref else 0

    # PDF accuracy
    pdf_nm = agg["pdf_matched"] + agg["pdf_mismatch"]
    pdf_nm_pct = pdf_nm / total_ref * 100 if total_ref else 0
    pdf_ex_pct = agg["pdf_matched"] / total_ref * 100 if total_ref else 0

    print(f"\n  {'Metric':<35s} {'DXF':>10s} {'PDF':>10s} {'Delta':>10s}")
    print(f"  {'-'*35} {'-'*10} {'-'*10} {'-'*10}")
    print(f"  {'Ref items (total)':<35s} {dxf_ref:>10d} {total_ref:>10d}")
    print(f"  {'Matched (name+qty)':<35s} {agg['dxf_matched']:>10d} {agg['pdf_matched']:>10d} "
          f"{agg['pdf_matched'] - agg['dxf_matched']:>+10d}")
    print(f"  {'Count mismatch':<35s} {agg['dxf_mismatch']:>10d} {agg['pdf_mismatch']:>10d} "
          f"{agg['pdf_mismatch'] - agg['dxf_mismatch']:>+10d}")
    print(f"  {'Missing':<35s} {agg['dxf_missing']:>10d} {agg['pdf_missing']:>10d} "
          f"{agg['pdf_missing'] - agg['dxf_missing']:>+10d}")
    print(f"  {'Name match rate %':<35s} {dxf_nm_pct:>9.1f}% {pdf_nm_pct:>9.1f}% "
          f"{pdf_nm_pct - dxf_nm_pct:>+9.1f}%")
    print(f"  {'Exact accuracy %':<35s} {dxf_ex_pct:>9.1f}% {pdf_ex_pct:>9.1f}% "
          f"{pdf_ex_pct - dxf_ex_pct:>+9.1f}%")
    print(f"\n  DXF found, PDF missed:            {agg['dxf_only_count']:>10d}")

    # Per-case summary table
    print(f"\n  {'Case':<14s} {'Ref':>4} "
          f"{'DXF%':>6} {'DXFex':>6} "
          f"{'PDF%':>6} {'PDFex':>6} "
          f"{'DOnly':>5} {'Time':>5}")
    print(f"  {'-'*14} {'-'*4} "
          f"{'-'*6} {'-'*6} {'-'*6} {'-'*6} {'-'*5} {'-'*5}")
    for res in results:
        if res.status != "ok":
            print(f"  {res.name:<14s} {'':>4} {'':>6} {'':>6} "
                  f"{'':>6} {'':>6} {'':>5} {res.status:>5}")
            continue
        dxf_nm_str = f"{res.dxf_name_match_pct:.1f}" if res.has_dxf else "-"
        dxf_ex_str = f"{res.dxf_exact_pct:.1f}" if res.has_dxf else "-"
        print(f"  {res.name:<14s} {res.ref_item_count:>4d} "
              f"{dxf_nm_str:>5s}% {dxf_ex_str:>5s}% "
              f"{res.pdf_name_match_pct:>5.1f}% {res.pdf_exact_pct:>5.1f}% "
              f"{len(res.dxf_only):>5d} {res.elapsed_sec:>4.0f}s")

    # ── Top DXF-only items (PDF improvement targets) ──
    all_dxf_only: list[tuple[str, dict]] = []
    for res in results:
        if res.status == "ok":
            for item in res.dxf_only:
                all_dxf_only.append((res.name, item))
    all_dxf_only.sort(key=lambda x: x[1]["ref_qty"], reverse=True)

    if all_dxf_only:
        print(f"\n  TOP PDF IMPROVEMENT TARGETS (DXF found, PDF missed):")
        for case_name, item in all_dxf_only[:15]:
            print(f"    [{case_name:12s}] {item['ref_name'][:50]:50s}  "
                  f"ref={item['ref_qty']:>4.0f}  dxf={item['dxf_qty']:>4.0f}")

    # ── Save JSON report ──
    report = {
        "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
        "threshold": args.threshold,
        "overall": {
            "cases_ok": ok_count,
            "total_ref_items": total_ref,
            "dxf": {
                "ref_items": dxf_ref,
                "matched": agg["dxf_matched"],
                "count_mismatch": agg["dxf_mismatch"],
                "missing": agg["dxf_missing"],
                "name_match_rate_pct": round(dxf_nm_pct, 1),
                "exact_accuracy_pct": round(dxf_ex_pct, 1),
            },
            "pdf": {
                "ref_items": total_ref,
                "matched": agg["pdf_matched"],
                "count_mismatch": agg["pdf_mismatch"],
                "missing": agg["pdf_missing"],
                "name_match_rate_pct": round(pdf_nm_pct, 1),
                "exact_accuracy_pct": round(pdf_ex_pct, 1),
            },
            "dxf_only_count": agg["dxf_only_count"],
        },
        "improvement_targets": [
            {
                "case": cn,
                "ref_name": it["ref_name"],
                "ref_qty": it["ref_qty"],
                "dxf_name": it["dxf_name"],
                "dxf_qty": it["dxf_qty"],
            }
            for cn, it in all_dxf_only[:30]
        ],
        "cases": [],
    }

    for res in results:
        entry: dict = {"name": res.name, "status": res.status}
        if res.status == "ok":
            entry.update({
                "has_dxf": res.has_dxf,
                "ref_item_count": res.ref_item_count,
                "dxf_item_count": res.dxf_item_count,
                "pdf_item_count": res.pdf_item_count,
                "pdf_count": res.pdf_count,
                "dxf_file_count": res.dxf_file_count,
                "elapsed_sec": round(res.elapsed_sec, 1),
                "dxf_vs_vor": {
                    "matched": res.dxf_matched,
                    "count_mismatch": res.dxf_count_mismatch,
                    "missing": res.dxf_missing,
                    "name_match_pct": round(res.dxf_name_match_pct, 1),
                    "exact_pct": round(res.dxf_exact_pct, 1),
                },
                "pdf_vs_vor": {
                    "matched": res.pdf_matched,
                    "count_mismatch": res.pdf_count_mismatch,
                    "missing": res.pdf_missing,
                    "name_match_pct": round(res.pdf_name_match_pct, 1),
                    "exact_pct": round(res.pdf_exact_pct, 1),
                },
                "dxf_only": res.dxf_only[:15],
                "elements": [
                    {
                        "ref_name": e.ref_name,
                        "ref_qty": e.ref_qty,
                        "dxf_found": e.dxf_found,
                        "dxf_name": e.dxf_name,
                        "dxf_qty": e.dxf_qty,
                        "dxf_qty_match": e.dxf_qty_match,
                        "pdf_found": e.pdf_found,
                        "pdf_name": e.pdf_name,
                        "pdf_qty": e.pdf_qty,
                        "pdf_qty_match": e.pdf_qty_match,
                    }
                    for e in res.elements
                ],
            })
        else:
            entry["error"] = res.error_msg
        report["cases"].append(entry)

    report_path = PROJECT_ROOT / "test_vor_cross_format_report.json"
    with open(report_path, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2, default=str)
    print(f"\n  Report saved: {report_path}")
    print("=" * 80)


if __name__ == "__main__":
    main()
