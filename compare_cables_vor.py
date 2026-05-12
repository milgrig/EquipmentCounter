"""
compare_cables_vor.py — Compare PDF-derived cable lengths against the
                       VOR (Ведомость объёмов работ) Excel.

Reads:
  * a PDF lighting plan (e.g. КПП-007), processed by
    ``pdf_cables_by_method.measure_cables`` to obtain length-by-method.
  * the matching VOR .xlsx file: extracts every "Прокладка кабеля …"
    and "Прокладка провода …" row with its quantity in metres.

Outputs a side-by-side table:

   method                    PDF, м    VOR (РД), м    ratio
   po_konstrukciyam            160.9          696.0     0.23
   gofra/truba                  17.2          641.0     0.03
   lotok                         6.5            0.0      —

Usage:
    python compare_cables_vor.py <pdf> <vor.xlsx> [--scale-mm-per-pt 35.29]
"""

from __future__ import annotations

import io
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from typing import Optional

import openpyxl

from pdf_cables_by_method import measure_cables, MM_PER_PT


# ---------------------------------------------------------------------------
# VOR parsing
# ---------------------------------------------------------------------------

@dataclass
class VorRow:
    sheet: str
    row_index: int
    name: str
    unit: str
    qty_p: float = 0.0   # column "Объем работ П"
    qty_rd: float = 0.0  # column "РД"


# Map free-text names from the VOR work-name column to a coarse method
# label compatible with pdf_cables_by_method.
_METHOD_KEYWORDS = [
    # (pattern, method tag)
    (re.compile(r"в\s+гофре", re.IGNORECASE),       "gofra/truba"),
    (re.compile(r"в\s+трубе",  re.IGNORECASE),       "gofra/truba"),
    (re.compile(r"в\s+лотке",  re.IGNORECASE),       "lotok"),
    (re.compile(r"в\s+земле",  re.IGNORECASE),       "v_zemle"),
    (re.compile(r"по\s+конструкци", re.IGNORECASE),  "po_konstrukciyam"),
    (re.compile(r"по\s+стен",       re.IGNORECASE),  "po_konstrukciyam"),
    (re.compile(r"по\s+потолк",     re.IGNORECASE),  "po_konstrukciyam"),
    (re.compile(r"по\s+кровле",     re.IGNORECASE),  "po_konstrukciyam"),
    (re.compile(r"молниепри",       re.IGNORECASE),  "molniezashita"),
    (re.compile(r"заземлител",      re.IGNORECASE),  "zazemlenie"),
]

_LAYING_HINT_RE = re.compile(r"прокладк", re.IGNORECASE)


def _parse_method_from_name(name: str) -> str:
    """Map a VOR work-name to a coarse laying-method tag."""
    if not name:
        return ""
    for pat, tag in _METHOD_KEYWORDS:
        if pat.search(name):
            return tag
    if _LAYING_HINT_RE.search(name) and (
        "кабел" in name.lower() or "провод" in name.lower()
    ):
        # Fallback: a "Прокладка кабеля/провода" line without explicit
        # method keyword — treat as "по конструкциям" by default.
        return "po_konstrukciyam"
    return ""


def _looks_like_cable_row(name: str) -> bool:
    if not name:
        return False
    low = name.lower()
    return ("кабел" in low or "провод" in low) and "прокладк" in low


def parse_vor(xlsx_path: str) -> list[VorRow]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    out: list[VorRow] = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        # Header row: find the column indices for "Наименование", "Ед. изм",
        # "Объем работ П", "РД". We search the first 5 rows for headers.
        col_name = col_unit = col_p = col_rd = None
        for r in range(1, 6):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if not isinstance(v, str):
                    continue
                low = v.lower().replace("\n", " ")
                if "наименование" in low and col_name is None:
                    col_name = c
                if "ед" in low and "изм" in low and col_unit is None:
                    col_unit = c
                if "объем" in low and "п" == low.strip()[-1:] and col_p is None:
                    col_p = c
                elif "объем" in low and col_p is None:
                    col_p = c
                if low.strip() == "рд" and col_rd is None:
                    col_rd = c
        if not (col_name and col_unit):
            continue

        for r in range(1, ws.max_row + 1):
            name = ws.cell(row=r, column=col_name).value
            if not isinstance(name, str):
                continue
            unit = ws.cell(row=r, column=col_unit).value
            if isinstance(unit, str):
                unit = unit.strip()
            qty_p = ws.cell(row=r, column=col_p).value if col_p else 0
            qty_rd = ws.cell(row=r, column=col_rd).value if col_rd else 0
            try:
                qty_p_f = float(qty_p) if qty_p is not None else 0.0
            except (TypeError, ValueError):
                qty_p_f = 0.0
            try:
                qty_rd_f = float(qty_rd) if qty_rd is not None else 0.0
            except (TypeError, ValueError):
                qty_rd_f = 0.0
            out.append(VorRow(
                sheet=sheet, row_index=r, name=name.strip(),
                unit=str(unit) if unit else "",
                qty_p=qty_p_f, qty_rd=qty_rd_f,
            ))
    return out


def aggregate_vor_cables(rows: list[VorRow]) -> dict[str, dict]:
    """Sum cable lengths from VOR rows, grouped by laying method.

    Returns:
        {method: {"qty_p": float, "qty_rd": float, "rows": [VorRow, ...]}}
    """
    agg: dict[str, dict] = defaultdict(
        lambda: {"qty_p": 0.0, "qty_rd": 0.0, "rows": []}
    )
    # Use only the first sheet ('Для сметного отдела') to avoid double-count
    seen_sheets = set()
    primary_sheet = None
    for r in rows:
        if primary_sheet is None:
            primary_sheet = r.sheet
        if r.sheet != primary_sheet:
            continue
        if r.unit and "м" not in r.unit.lower():
            continue
        # The "Прокладка кабеля в гофре …" header rows usually carry the
        # cumulative qty, but the cable-spec sub-rows that follow repeat
        # the same value broken down by section. We take the *header* row
        # (the one that explicitly mentions "Прокладка") and skip the
        # sub-rows that are bare "Кабель …" without "Прокладка".
        method = _parse_method_from_name(r.name)
        if not method:
            continue
        if not _looks_like_cable_row(r.name):
            continue
        slot = agg[method]
        slot["qty_p"] += r.qty_p
        slot["qty_rd"] += r.qty_rd
        slot["rows"].append(r)
    return dict(agg)


# ---------------------------------------------------------------------------
# Comparison
# ---------------------------------------------------------------------------

@dataclass
class ComparisonRow:
    method: str
    pdf_m: float
    vor_p_m: float
    vor_rd_m: float


def compare(pdf_path: str, vor_path: str,
            scale_mm_per_pt: Optional[float] = None,
            page_index: int = 0,
            ) -> tuple[list[ComparisonRow], dict]:
    rep = measure_cables(pdf_path, page_index=page_index,
                         scale_mm_per_pt=scale_mm_per_pt)
    pdf_by_method: dict[str, float] = defaultdict(float)
    for (_color, method), L in rep.totals_m.items():
        pdf_by_method[method] += L

    vor_rows = parse_vor(vor_path)
    vor_by_method = aggregate_vor_cables(vor_rows)

    methods = sorted(set(pdf_by_method.keys()) | set(vor_by_method.keys()))
    out = []
    for m in methods:
        out.append(ComparisonRow(
            method=m,
            pdf_m=round(pdf_by_method.get(m, 0.0), 1),
            vor_p_m=round(vor_by_method.get(m, {}).get("qty_p", 0.0), 1),
            vor_rd_m=round(vor_by_method.get(m, {}).get("qty_rd", 0.0), 1),
        ))

    meta = dict(
        pdf_path=pdf_path,
        vor_path=vor_path,
        scale_mm_per_pt=rep.scale_mm_per_pt,
        scale_source=rep.scale_source,
        drawing_scale=f"1:{rep.scale_mm_per_pt / MM_PER_PT:.0f}",
        polylines=len(rep.polylines),
        markers=rep.circle_markers,
        tray_rects=rep.tray_rects,
        vor_rows_total=len(vor_rows),
        vor_methods=list(vor_by_method.keys()),
    )
    return out, meta


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    if sys.platform == "win32":
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer,
                                      encoding="utf-8", errors="replace")

    args = sys.argv[1:]
    if len(args) < 2:
        print("Usage: python compare_cables_vor.py <pdf> <vor.xlsx> "
              "[--scale-mm-per-pt 35.29] [--page N]")
        sys.exit(1)
    pdf_path, vor_path = args[0], args[1]
    scale = None
    page = 0
    i = 2
    while i < len(args):
        if args[i] == "--scale-mm-per-pt" and i + 1 < len(args):
            scale = float(args[i + 1]); i += 2
        elif args[i] == "--page" and i + 1 < len(args):
            page = int(args[i + 1]); i += 2
        else:
            i += 1

    rows, meta = compare(pdf_path, vor_path,
                         scale_mm_per_pt=scale, page_index=page)

    print(f"PDF: {meta['pdf_path']}")
    print(f"VOR: {meta['vor_path']}")
    print(f"Scale: {meta['scale_mm_per_pt']} mm/pt   "
          f"({meta['scale_source']}, {meta['drawing_scale']})")
    print(f"PDF polylines={meta['polylines']}, "
          f"markers={meta['markers']}, trays={meta['tray_rects']}")
    print(f"VOR rows total={meta['vor_rows_total']}; "
          f"methods seen: {meta['vor_methods']}")
    print()
    print(f"{'method':<22} {'PDF, м':>10} {'VOR-П, м':>10} "
          f"{'VOR-РД, м':>10} {'PDF/РД':>8}")
    print("-" * 64)
    sum_pdf = sum_p = sum_rd = 0.0
    for r in rows:
        ratio = (r.pdf_m / r.vor_rd_m) if r.vor_rd_m > 0 else 0.0
        ratio_s = f"{ratio:.2f}" if r.vor_rd_m > 0 else "—"
        print(f"{r.method:<22} {r.pdf_m:>10.1f} {r.vor_p_m:>10.1f} "
              f"{r.vor_rd_m:>10.1f} {ratio_s:>8}")
        sum_pdf += r.pdf_m
        sum_p += r.vor_p_m
        sum_rd += r.vor_rd_m
    print("-" * 64)
    final_ratio = (sum_pdf / sum_rd) if sum_rd > 0 else 0
    final_s = f"{final_ratio:.2f}" if sum_rd > 0 else "—"
    print(f"{'TOTAL':<22} {sum_pdf:>10.1f} {sum_p:>10.1f} "
          f"{sum_rd:>10.1f} {final_s:>8}")


if __name__ == "__main__":
    main()
