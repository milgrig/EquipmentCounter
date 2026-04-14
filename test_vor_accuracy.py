#!/usr/bin/env python3
"""test_vor_accuracy.py -- Compare generated VOR (from PDFs) vs reference Excel VOR files.

Finds all reference VOR Excel files in Data/, processes the corresponding PDF
folders through the S019 pipeline (parse_legend + count_symbols + match_symbols
+ extract_cables + vor_map_items), aggregates results, and compares with the
reference VOR.  Reports accuracy: matched items, missing items, extra items,
count mismatches.

Usage:
    python test_vor_accuracy.py
    python test_vor_accuracy.py --only abk_eo,pos_28
    python test_vor_accuracy.py --verbose
"""

from __future__ import annotations

import argparse
import difflib
import json
import logging
import re
import sys
import time
import traceback
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

# ── Project root ──────────────────────────────────────────────────────
PROJECT_ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(PROJECT_ROOT))

DATA_DIR = PROJECT_ROOT / "Data"
DBT_DIR = DATA_DIR / "ДБТ разделы для ИИ"

# ── Imports from project (S017 pipeline) ─────────────────────────────
from pdf_legend_parser import parse_legend, LegendResult  # noqa: E402
from pdf_count_text import count_symbols  # noqa: E402
from pdf_count_cables import extract_cables  # noqa: E402
from pdf_count_visual import match_symbols  # noqa: E402
from vor_work_mapping import map_equipment_to_work, map_items as vor_map_items  # noqa: E402

try:
    import openpyxl  # noqa: E402
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

# ── Baseline metrics (S017 full-run results, from test_vor_report_s017.json) ──
BASELINE = {
    "exact_accuracy_pct": 0.3,
    "name_match_rate_pct": 9.7,
    "total_ref_items": 965,
    "matched": 3,
    "count_mismatch": 91,
    "missing": 690,
    "extra": 31,
}

log = logging.getLogger("test_vor_accuracy")


# =====================================================================
#  Test cases — manually mapped VOR Excel -> PDF folder
# =====================================================================
#
# Each entry maps a reference VOR Excel to the PDF folder that contains
# the corresponding engineering drawings.
#
# Mapping rules (discovered by inspecting the folder tree):
#   01_ABK:  VOR_ subfolder has 3 VOR files (ЭО, ЭМ, ЭГ),
#            PDFs are in PDF_/ЭО, PDF_/ЭМ, PDF_/ЭГ
#   Others:  VOR xlsx sits next to or near 02_PDF / 03_PDF folder

_ABK = DBT_DIR / "01_АБК" / "Обновленные файлы"

TEST_CASES: list[dict[str, str]] = [
    {
        "name": "abk_eo",
        "ref_vor": str(_ABK / "ВОР_" / "ВОР АБК ЭО.xlsx"),
        "pdf_dir": str(_ABK / "PDF_" / "ЭО"),
    },
    {
        "name": "abk_em",
        "ref_vor": str(_ABK / "ВОР_" / "ВОР АБК ЭМ.xlsx"),
        "pdf_dir": str(_ABK / "PDF_" / "ЭМ"),
    },
    {
        "name": "abk_eg",
        "ref_vor": str(_ABK / "ВОР_" / "ВОР АБК ЭГ.xlsx"),
        "pdf_dir": str(_ABK / "PDF_" / "ЭГ"),
    },
    {
        "name": "pos_12_3",
        "ref_vor": str(DBT_DIR / "12.3 Насосная станция поверхностных стоков"
                       / "ВОР поз. 12.3.xlsx"),
        "pdf_dir": str(DBT_DIR / "12.3 Насосная станция поверхностных стоков"
                       / "02_PDF"),
    },
    {
        "name": "pos_16_2",
        "ref_vor": str(DBT_DIR / "16.2_ЖД КПП" / "Обновленные файлы 1"
                       / "ВОР поз. 16.2.xlsx"),
        "pdf_dir": str(DBT_DIR / "16.2_ЖД КПП" / "Обновленные файлы 1"
                       / "02_PDF"),
    },
    {
        "name": "pos_27",
        "ref_vor": str(DBT_DIR / "27_Склад вспомогательных материалов "
                       "с участком погрузки крытых вагонов"
                       / "Обновленные файлы" / "ВОР поз.27.xlsx"),
        "pdf_dir": str(DBT_DIR / "27_Склад вспомогательных материалов "
                       "с участком погрузки крытых вагонов"
                       / "Обновленные файлы" / "03_PDF"),
    },
    {
        "name": "pos_28",
        "ref_vor": str(DBT_DIR / "28. Автовесы" / "ВОР поз.28.xlsx"),
        "pdf_dir": str(DBT_DIR / "28. Автовесы" / "02_PDF"),
    },
    {
        "name": "pos_30",
        "ref_vor": str(DBT_DIR / "30. КПП" / "ВОР поз.30.xlsx"),
        "pdf_dir": str(DBT_DIR / "30. КПП" / "03_PDF"),
    },
    {
        "name": "pos_8_2_em",
        "ref_vor": str(DBT_DIR / "8.2 Участок хранения сульфата аммония"
                       / "ЭМ" / "ВОР ЭМ поз. 8.2.xlsx"),
        "pdf_dir": str(DBT_DIR / "8.2 Участок хранения сульфата аммония"
                       / "ЭМ" / "02_PDF"),
    },
    {
        "name": "pos_8_2_eo",
        "ref_vor": str(DBT_DIR / "8.2 Участок хранения сульфата аммония"
                       / "ЭО" / "ВОР поз. 8.2.xlsx"),
        "pdf_dir": str(DBT_DIR / "8.2 Участок хранения сульфата аммония"
                       / "ЭО" / "02_PDF"),
    },
]


# =====================================================================
#  Data structures
# =====================================================================

@dataclass
class RefItem:
    """One row from the reference VOR Excel."""
    row_num: str      # column A (may be empty for sub-items/materials)
    name: str         # column B
    unit: str         # column C
    qty: float        # column D
    is_section: bool = False  # section header (no qty, no row num)

    @property
    def key(self) -> str:
        return _normalize(self.name)


@dataclass
class GenItem:
    """Aggregated item from PDF processing."""
    name: str
    total: int
    files: list[str] = field(default_factory=list)

    @property
    def key(self) -> str:
        return _normalize(self.name)


@dataclass
class MatchResult:
    ref_name: str
    ref_qty: float
    gen_name: str
    gen_qty: float
    match_ratio: float
    qty_match: bool       # quantities match exactly
    delta: float          # gen_qty - ref_qty


@dataclass
class CaseResult:
    name: str
    status: str           # "ok", "error", "skip"
    error_msg: str = ""
    ref_vor_path: str = ""
    pdf_dir_path: str = ""
    ref_item_count: int = 0
    gen_item_count: int = 0
    pdf_count: int = 0
    pdf_errors: int = 0
    elapsed_sec: float = 0.0
    matched: list[MatchResult] = field(default_factory=list)
    count_mismatch: list[MatchResult] = field(default_factory=list)
    missing: list[RefItem] = field(default_factory=list)
    extra: list[GenItem] = field(default_factory=list)

    @property
    def accuracy_pct(self) -> float:
        total = len(self.matched) + len(self.count_mismatch) + len(self.missing)
        if total == 0:
            return 0.0
        return len(self.matched) / total * 100

    @property
    def fuzzy_match_pct(self) -> float:
        """Percentage of ref items that were matched (regardless of qty)."""
        total = len(self.matched) + len(self.count_mismatch) + len(self.missing)
        if total == 0:
            return 0.0
        return (len(self.matched) + len(self.count_mismatch)) / total * 100


# =====================================================================
#  Normalization for fuzzy matching
# =====================================================================

# Patterns for items auto-detected by process_pdf when no legend is found.
# These have no real equipment names and should be excluded from comparison.
_AUTO_DETECT_RE = re.compile(
    r"^\[Auto-detected\s|^\[Обозначение\s",
    re.IGNORECASE,
)


def _is_auto_detected(name: str) -> bool:
    """Check if an item name is auto-generated (no real legend found)."""
    return bool(_AUTO_DETECT_RE.match(name))


def _normalize(text: str) -> str:
    """Normalize a description for comparison.

    Handles the domain gap between:
      - Generated work items: "Монтаж светильников светодиодный UNI/R"
      - VOR work items:       "Монтаж светильника в подвесных потолках..."
      - VOR material rows:    "Светодиодный светильник 4000К 40Вт UNI/R..."
    """
    s = text.strip().lower()
    # Remove leading action verbs (VOR prepends these)
    s = re.sub(
        r"^(монтаж|установка|прокладка|забивка|крепление|окраска|подключение)\s+",
        "", s,
    )
    # Normalize noun forms: singular/plural/genitive for equipment
    # светильника/светильников/светильник -> светильник
    s = re.sub(r"светильник\w*", "светильник", s)
    # розетки/розеток/розетка -> розетка
    s = re.sub(r"розетк\w*", "розетка", s)
    # выключатели/выключателей/выключатель -> выключатель
    s = re.sub(r"выключател\w*", "выключатель", s)
    # коробки/коробок/коробка -> коробка
    s = re.sub(r"коробк\w*", "коробка", s)
    # датчиков/датчики/датчик -> датчик
    s = re.sub(r"датчик\w*", "датчик", s)
    # лотков/лотка/лоток -> лоток
    s = re.sub(r"лот[оа]к\w*", "лоток", s)
    # указателей/указатель -> указатель
    s = re.sub(r"указател\w*", "указатель", s)
    # щита/щитов -> щит
    s = re.sub(r"щит\w*", "щит", s)
    # труб/трубы -> труба
    s = re.sub(r"труб\w*", "труба", s)
    # автоматов/автомат -> автомат
    s = re.sub(r"автомат\w*", "автомат", s)
    # кабеля/кабелей/кабель -> кабель
    s = re.sub(r"кабел\w*", "кабель", s)
    # провода/проводов/провод -> провод
    s = re.sub(r"провод\w*(?!\S)", "провод", s)  # don't mangle "проводник"

    # Normalize "светодиодный светильник" <-> "светильник светодиодный"
    s = re.sub(r"светодиодн\w*\s+светильник", "светильник", s)
    s = re.sub(r"светильник\s+светодиодн\w*", "светильник", s)

    # Remove height/mounting context: "в подвесных потолках", "на высоте до 5 м"
    s = re.sub(r"\s+в\s+подвесн\w*\s+потолк\w*", "", s)
    s = re.sub(r"\s+(?:на\s+)?(?:высоте?\s+)?(?:до|от)\s+\d+\s*(?:до\s+\d+\s*)?(?:метр\w*|м)\b", "", s)
    s = re.sub(r"\s+настенн\w*", "", s)
    s = re.sub(r"\s+потолочн\w*", "", s)
    # Remove "подвесной/анкерный/на шпильках" mounting qualifiers
    s = re.sub(r"\s+(?:подвесн\w*|анкерн\w*|на\s+шпильк\w*)", "", s)
    # Remove "в сборе"
    s = re.sub(r"\s+в\s+сборе", "", s)
    # Remove "распределительн/осветительн/силов" after щит
    s = re.sub(r"щит\s+(?:распределительн\w*|осветительн\w*|силов\w*)\s*", "щит ", s)
    # Remove "кабельного" after лоток
    s = re.sub(r"лоток\s+кабельн\w*\s*", "лоток ", s)
    # Remove "аварийн" (already categorized)
    s = re.sub(r"\s*аварийн\w*\s*", " ", s)

    # Remove catalog/article numbers at the end (5+ digit codes)
    s = re.sub(r",?\s*\d{5,}[\d\s,]*$", "", s)
    # Remove brand/model codes at end
    s = re.sub(
        r",?\s*(?:DKS|Ostec|EKF|IEK|ABB|Systeme\s+Electric|СЗИМ)\b.*$",
        "", s, flags=re.IGNORECASE,
    )
    # Remove model codes like PA16-044B, ATN440101
    s = re.sub(r"\b[A-Z]{2,}\d[\w-]{3,}\b", "", s, flags=re.IGNORECASE)
    # Remove specs in parentheses that don't help matching
    # Keep (595x595), IP55, but remove (до 5 м), (суммарное сечение...)
    s = re.sub(r"\((?:до|от|суммарн|однополюс|двухполюс|масс\w*)[^)]*\)", "", s)
    # Remove "(прокладка)" marker
    s = re.sub(r"\(прокладка\)", "", s)
    # Remove color/packaging info
    s = re.sub(
        r",?\s*(?:цвет\s+\S+|белый|серый|чёрный|черный)\s*,?",
        "", s, flags=re.IGNORECASE,
    )
    # Remove power/flux specs like "4000К 40Вт 5200Лм"
    s = re.sub(r"\d{3,4}\s*[kкК]\s*", "", s)  # color temp
    s = re.sub(r"\d+\s*[вwВW]т\w*", "", s, flags=re.IGNORECASE)  # watts
    s = re.sub(r"\d+\s*[лl]м\w*", "", s, flags=re.IGNORECASE)  # lumens
    # Remove "сечением"
    s = re.sub(r"\bсечением?\b\s*", "", s)
    # Normalize whitespace
    s = re.sub(r"\s+", " ", s).strip()
    # Remove trailing punctuation
    s = s.rstrip(",.:; ")
    return s


# =====================================================================
#  Read reference VOR Excel
# =====================================================================

def read_reference_vor(xlsx_path: str) -> list[RefItem]:
    """Read reference VOR Excel and return list of RefItem.

    Handles multiple sheets -- uses the first sheet with actual data,
    preferring sheets named like 'Для заказчика' or the active sheet.
    Structure:
      Row 1: headers
      Row 2: column numbers (skip)
      Row 3+: data (section headers, work items, material sub-items)
    Column A (0): row number (int or str)
    Column B (1): description
    Column C (2): unit
    Column D (3): quantity
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    # Pick the best sheet
    ws = None
    for sname in wb.sheetnames:
        if "заказчик" in sname.lower():
            ws = wb[sname]
            break
    if ws is None:
        # Use first sheet with data rows > 10
        for sname in wb.sheetnames:
            candidate = wb[sname]
            if candidate.max_row and candidate.max_row > 10:
                ws = candidate
                break
    if ws is None:
        ws = wb.active

    items: list[RefItem] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if not row or len(row) < 4:
            continue

        col_a = row[0]  # row number
        col_b = row[1]  # description
        col_c = row[2]  # unit
        col_d = row[3]  # quantity

        name = str(col_b or "").strip()
        if not name:
            continue

        # Section header: has a name but no quantity and no row number
        if col_d is None and col_a is None:
            items.append(RefItem(
                row_num="", name=name, unit="", qty=0, is_section=True,
            ))
            continue

        # Parse quantity
        try:
            qty = float(col_d) if col_d is not None else 0
        except (ValueError, TypeError):
            # Try extracting number from string
            m = re.search(r"[\d]+(?:[.,]\d+)?", str(col_d or ""))
            qty = float(m.group().replace(",", ".")) if m else 0

        if qty <= 0:
            continue

        unit = str(col_c or "").strip()
        row_num = str(col_a or "").strip()

        items.append(RefItem(
            row_num=row_num, name=name, unit=unit, qty=qty,
        ))

    wb.close()
    return items


# =====================================================================
#  S017 Pipeline: _count_equipment_in_pdf (replicated from web_app.py)
# =====================================================================

def _count_equipment_in_pdf(pdf_path: str) -> list[dict]:
    """Run legend extraction + counting methods on a single PDF.

    Replicates the pipeline from web_app._count_equipment_in_pdf():
    1. Parse legend via pdf_legend_parser (gives LegendResult with items)
    2. Run count_symbols (text method) to get actual counts per symbol
    3. For symbols with 0 text counts, fall back to match_symbols (visual)
    4. Run extract_cables to find cable items with quantities
    5. Apply VOR work-name mapping
    6. Return enriched item dicts with actual counts
    """
    # Step 1: parse legend
    legend_result = parse_legend(pdf_path)
    if not legend_result.items:
        return []

    items: list[dict] = []

    # Step 2: run text counting
    text_counts: dict[str, int] = {}
    try:
        text_result = count_symbols(pdf_path, legend_result)
        text_counts = text_result.counts  # symbol -> count
    except Exception as exc:
        log.warning("Text counting failed for %s: %s", pdf_path, exc)

    # Step 3: run visual counting as fallback
    visual_counts: dict[int, int] = {}  # symbol_index -> count
    symbols_needing_visual = []
    for idx, item in enumerate(legend_result.items):
        sym = item.symbol or ""
        if sym and text_counts.get(sym, 0) == 0:
            symbols_needing_visual.append(idx)
        elif not sym:
            # Graphical-only items always need visual
            symbols_needing_visual.append(idx)

    if symbols_needing_visual:
        try:
            vis_result = match_symbols(pdf_path, legend_result)
            visual_counts = vis_result.counts  # symbol_index -> count
        except Exception as exc:
            log.warning("Visual counting failed for %s: %s", pdf_path, exc)

    # Step 4: build enriched items from legend + counts
    for idx, item in enumerate(legend_result.items):
        sym = item.symbol or ""
        name = item.description or ""
        if not name:
            continue

        # Determine count: prefer text count, fall back to visual
        count = 0
        if sym and text_counts.get(sym, 0) > 0:
            count = text_counts[sym]
        elif idx in visual_counts and visual_counts[idx] > 0:
            count = visual_counts[idx]

        if count <= 0:
            continue

        items.append({
            "symbol": sym,
            "name": name,
            "count": count,
            "count_ae": 0,
            "total": count,
        })

    # Step 5: extract cables and add to items
    try:
        cable_result = extract_cables(pdf_path, legend_result)
        for entry in cable_result.cable_schedule:
            group = entry.get("group", "")
            panel = entry.get("panel", "")
            cable_types = entry.get("cable_types", [])
            cross_sections = entry.get("cross_sections", [])
            run_count = entry.get("run_count", 0) or 0
            total_length_m = entry.get("total_length_m", 0) or 0

            # Use cable_type if available, otherwise cross_section
            type_label = (cable_types[0] if cable_types
                          else cross_sections[0] if cross_sections
                          else "")
            if not type_label and not group:
                continue

            if run_count > 0:
                cable_name = f"Кабель {type_label}" if type_label else "Кабель"
                if group:
                    cable_name += f" ({panel}-{group})" if panel else f" ({group})"
                items.append({
                    "symbol": "",
                    "name": cable_name,
                    "count": run_count,
                    "count_ae": 0,
                    "total": run_count,
                    "unit": "шт",
                })
            if total_length_m > 0:
                cable_name_m = (f"Кабель {type_label} (прокладка)"
                                if type_label else "Кабель (прокладка)")
                if group:
                    cable_name_m += f" ({panel}-{group})" if panel else f" ({group})"
                items.append({
                    "symbol": "",
                    "name": cable_name_m,
                    "count": 0,
                    "count_ae": 0,
                    "total": round(total_length_m, 1),
                    "unit": "м",
                })
    except Exception as exc:
        log.warning("Cable extraction failed for %s: %s", pdf_path, exc)

    # Step 6: apply VOR work-name mapping
    items = vor_map_items(items)

    return items


# =====================================================================
#  Process PDF folder (S017 pipeline)
# =====================================================================

def process_pdf_folder(folder_path: str, verbose: bool = False) -> tuple[
    dict[str, GenItem], int, int
]:
    """Process all PDFs in folder via S017 pipeline (_count_equipment_in_pdf).

    Returns:
        (items_by_key, processed_count, error_count)
    """
    pdf_dir = Path(folder_path)
    pdf_files = sorted(pdf_dir.glob("*.pdf"))

    all_items: dict[str, GenItem] = {}
    processed = 0
    errors = 0

    for pdf_path in pdf_files:
        try:
            items = _count_equipment_in_pdf(str(pdf_path))
            real_count = 0
            for item in items:
                # Use work_name (VOR description) if available, fall back to name
                work_name = item.get("work_name", "").strip()
                raw_name = item.get("name", "").strip()
                name = work_name or raw_name
                if not name:
                    continue
                # Skip auto-detected items (no real legend found)
                if _is_auto_detected(name):
                    continue
                total = item.get("total", item.get("count", 0) + item.get("count_ae", 0))
                if total <= 0:
                    continue
                key = _normalize(name)
                if key in all_items:
                    all_items[key].total += total
                    all_items[key].files.append(pdf_path.name)
                else:
                    all_items[key] = GenItem(
                        name=name, total=total,
                        files=[pdf_path.name],
                    )
                real_count += 1
            processed += 1
            if verbose:
                print(f"    OK  {pdf_path.name}: {len(items)} items"
                      f" ({real_count} real)")
        except Exception as e:
            errors += 1
            print(f"    ERR {pdf_path.name}: {e}")

    return all_items, processed, errors


# =====================================================================
#  Fuzzy comparison
# =====================================================================

def compare(
    generated: dict[str, GenItem],
    reference: list[RefItem],
    threshold: float = 0.45,
) -> tuple[
    list[MatchResult],       # exact qty matches
    list[MatchResult],       # count mismatches (name matched, qty differs)
    list[RefItem],           # missing from generated
    list[GenItem],           # extra in generated
]:
    """Compare generated items against reference items using fuzzy matching.

    Two-pass approach:
      Pass 1: Exact normalized key match
      Pass 2: Fuzzy SequenceMatcher on normalized keys
    """
    # Filter out section headers from reference
    ref_items = [r for r in reference if not r.is_section]

    # Aggregate reference items by normalized key
    ref_by_key: dict[str, RefItem] = {}
    for item in ref_items:
        key = item.key
        if key in ref_by_key:
            ref_by_key[key] = RefItem(
                row_num=ref_by_key[key].row_num,
                name=ref_by_key[key].name,
                unit=ref_by_key[key].unit,
                qty=ref_by_key[key].qty + item.qty,
            )
        else:
            ref_by_key[key] = RefItem(
                row_num=item.row_num,
                name=item.name,
                unit=item.unit,
                qty=item.qty,
            )

    matched: list[MatchResult] = []
    count_mismatch: list[MatchResult] = []
    missing: list[RefItem] = []
    used_gen_keys: set[str] = set()

    gen_keys = list(generated.keys())

    # ── Pass 1: Exact key match ──
    for ref_key, ref_item in ref_by_key.items():
        if ref_key in generated:
            gen_item = generated[ref_key]
            used_gen_keys.add(ref_key)
            delta = gen_item.total - ref_item.qty
            qty_ok = abs(delta) < 0.01
            mr = MatchResult(
                ref_name=ref_item.name, ref_qty=ref_item.qty,
                gen_name=gen_item.name, gen_qty=gen_item.total,
                match_ratio=1.0, qty_match=qty_ok, delta=delta,
            )
            if qty_ok:
                matched.append(mr)
            else:
                count_mismatch.append(mr)

    # ── Pass 2: Fuzzy match remaining ──
    unmatched_ref_keys = [k for k in ref_by_key if k not in used_gen_keys
                          and k not in {r.key for r in
                                        [ref_by_key[k2] for k2 in ref_by_key
                                         if k2 in used_gen_keys]}]
    # Actually let's redo this properly
    matched_ref_keys: set[str] = set()
    for mr in matched:
        matched_ref_keys.add(_normalize(mr.ref_name))
    for mr in count_mismatch:
        matched_ref_keys.add(_normalize(mr.ref_name))

    remaining_ref = {k: v for k, v in ref_by_key.items() if k not in matched_ref_keys}
    remaining_gen = {k: v for k, v in generated.items() if k not in used_gen_keys}

    for ref_key, ref_item in remaining_ref.items():
        best_ratio = 0.0
        best_gen_key = None

        for gen_key in remaining_gen:
            if gen_key in used_gen_keys:
                continue
            ratio = difflib.SequenceMatcher(None, ref_key, gen_key).ratio()
            if ratio > best_ratio:
                best_ratio = ratio
                best_gen_key = gen_key

        if best_ratio >= threshold and best_gen_key is not None:
            gen_item = remaining_gen[best_gen_key]
            used_gen_keys.add(best_gen_key)
            delta = gen_item.total - ref_item.qty
            qty_ok = abs(delta) < 0.01
            mr = MatchResult(
                ref_name=ref_item.name, ref_qty=ref_item.qty,
                gen_name=gen_item.name, gen_qty=gen_item.total,
                match_ratio=best_ratio, qty_match=qty_ok, delta=delta,
            )
            if qty_ok:
                matched.append(mr)
            else:
                count_mismatch.append(mr)
        else:
            missing.append(ref_item)

    # Extra items: in generated but not matched to any reference
    extra = [v for k, v in generated.items() if k not in used_gen_keys]

    return matched, count_mismatch, missing, extra


# =====================================================================
#  Run one test case
# =====================================================================

def run_case(case: dict[str, str], verbose: bool = False,
             threshold: float = 0.45) -> CaseResult:
    """Run a single test case: read ref VOR, process PDFs, compare."""
    name = case["name"]
    ref_path = case["ref_vor"]
    pdf_dir = case["pdf_dir"]

    # Validate paths
    if not Path(ref_path).exists():
        return CaseResult(
            name=name, status="skip",
            error_msg=f"Reference VOR not found: {Path(ref_path).name}",
            ref_vor_path=ref_path, pdf_dir_path=pdf_dir,
        )
    if not Path(pdf_dir).exists():
        return CaseResult(
            name=name, status="skip",
            error_msg=f"PDF folder not found: {pdf_dir}",
            ref_vor_path=ref_path, pdf_dir_path=pdf_dir,
        )

    pdf_count = len(list(Path(pdf_dir).glob("*.pdf")))
    if pdf_count == 0:
        return CaseResult(
            name=name, status="skip",
            error_msg="No PDF files in folder",
            ref_vor_path=ref_path, pdf_dir_path=pdf_dir,
        )

    t0 = time.time()

    try:
        # 1. Read reference VOR
        reference = read_reference_vor(ref_path)
        ref_data_items = [r for r in reference if not r.is_section]

        if not ref_data_items:
            return CaseResult(
                name=name, status="error",
                error_msg="No valid data items in reference VOR",
                ref_vor_path=ref_path, pdf_dir_path=pdf_dir,
                elapsed_sec=time.time() - t0,
            )

        # 2. Process PDFs
        generated, processed, errors = process_pdf_folder(pdf_dir, verbose)

        # 3. Compare
        m, mm, mi, ex = compare(generated, reference, threshold=threshold)

        return CaseResult(
            name=name, status="ok",
            ref_vor_path=ref_path, pdf_dir_path=pdf_dir,
            ref_item_count=len(ref_data_items),
            gen_item_count=len(generated),
            pdf_count=pdf_count, pdf_errors=errors,
            elapsed_sec=time.time() - t0,
            matched=m, count_mismatch=mm, missing=mi, extra=ex,
        )

    except Exception as e:
        return CaseResult(
            name=name, status="error",
            error_msg=f"{e}\n{traceback.format_exc()[-400:]}",
            ref_vor_path=ref_path, pdf_dir_path=pdf_dir,
            elapsed_sec=time.time() - t0,
        )


# =====================================================================
#  Main
# =====================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Compare generated VOR (from PDFs) vs reference Excel VOR",
    )
    parser.add_argument(
        "--only", default="",
        help="Comma-separated case names to run (default: all)",
    )
    parser.add_argument(
        "--verbose", "-v", action="store_true",
        help="Print per-PDF processing details",
    )
    parser.add_argument(
        "--threshold", type=float, default=0.45,
        help="Fuzzy match threshold (default: 0.45)",
    )
    args = parser.parse_args()

    cases = TEST_CASES
    if args.only:
        names = {n.strip() for n in args.only.split(",")}
        cases = [c for c in cases if c["name"] in names]
        if not cases:
            print(f"No matching cases for: {args.only}")
            print(f"Available: {', '.join(c['name'] for c in TEST_CASES)}")
            sys.exit(1)

    print("=" * 80)
    print("  MASS PDF TESTING: Generated VOR vs Reference Excel VOR")
    print("=" * 80)
    print(f"  Cases: {len(cases)}")
    print(f"  Fuzzy threshold: {args.threshold}")
    print()

    overall = {
        "total_ref": 0, "matched": 0, "mismatch": 0,
        "missing": 0, "extra": 0,
    }
    results: list[CaseResult] = []

    for idx, case in enumerate(cases, 1):
        print(f"\n{'_' * 70}")
        print(f"[{idx}/{len(cases)}] {case['name']}")
        print(f"  VOR: {Path(case['ref_vor']).name}")
        print(f"  PDF: {Path(case['pdf_dir']).name if Path(case['pdf_dir']).exists() else '(not found)'}")

        res = run_case(case, verbose=args.verbose, threshold=args.threshold)
        results.append(res)

        if res.status == "skip":
            print(f"  SKIPPED: {res.error_msg}")
            continue
        elif res.status == "error":
            print(f"  ERROR: {res.error_msg[:200]}")
            continue

        # Print results
        m = len(res.matched)
        mm = len(res.count_mismatch)
        mi = len(res.missing)
        ex = len(res.extra)

        print(f"  PDFs processed: {res.pdf_count} ({res.pdf_errors} errors)")
        print(f"  Reference items: {res.ref_item_count}")
        print(f"  Generated items: {res.gen_item_count}")
        print()
        print(f"  RESULTS:")
        print(f"    Matched (name+qty):   {m:4d}")
        print(f"    Count mismatch:       {mm:4d}")
        print(f"    Missing (not found):  {mi:4d}")
        print(f"    Extra (not in ref):   {ex:4d}")
        print(f"    Name match rate:      {res.fuzzy_match_pct:5.1f}%")
        print(f"    Exact accuracy:       {res.accuracy_pct:5.1f}%")
        print(f"    Time: {res.elapsed_sec:.1f}s")

        if res.count_mismatch:
            print(f"\n  COUNT MISMATCHES (top 10):")
            for item in sorted(res.count_mismatch,
                                key=lambda x: abs(x.delta), reverse=True)[:10]:
                rname = item.ref_name[:45]
                gname = item.gen_name[:45]
                print(f"    ref: {rname}")
                print(f"    gen: {gname}")
                print(f"         ref={item.ref_qty:6.0f}  gen={item.gen_qty:6.0f}"
                      f"  delta={item.delta:+.0f}  ratio={item.match_ratio:.0%}")
                print()

        if res.missing:
            print(f"\n  MISSING from generated (top 10):")
            for item in res.missing[:10]:
                print(f"    {item.name[:65]:65s}  qty={item.qty:.0f}")

        if args.verbose and res.extra:
            print(f"\n  EXTRA in generated (top 10):")
            for item in res.extra[:10]:
                print(f"    {item.name[:65]:65s}  qty={item.total}")

        overall["total_ref"] += res.ref_item_count
        overall["matched"] += m
        overall["mismatch"] += mm
        overall["missing"] += mi
        overall["extra"] += ex

    # ── Overall summary ──
    print(f"\n{'=' * 80}")
    print("  OVERALL SUMMARY (S019 Pipeline)")
    print(f"{'=' * 80}")

    ok_count = sum(1 for r in results if r.status == "ok")
    skip_count = sum(1 for r in results if r.status == "skip")
    err_count = sum(1 for r in results if r.status == "error")
    total_ref = overall["total_ref"]
    total_acc = (overall["matched"] / total_ref * 100) if total_ref > 0 else 0
    total_fuzzy = (
        (overall["matched"] + overall["mismatch"]) / total_ref * 100
        if total_ref > 0 else 0
    )
    total_elapsed = sum(r.elapsed_sec for r in results if r.status == "ok")

    print(f"  Cases OK / Skip / Error:  {ok_count} / {skip_count} / {err_count}")
    print(f"  Total ref items:          {total_ref}")
    print(f"  Matched (name+qty):       {overall['matched']}")
    print(f"  Count mismatches:         {overall['mismatch']}")
    print(f"  Missing:                  {overall['missing']}")
    print(f"  Extra:                    {overall['extra']}")
    print(f"  Name match rate:          {total_fuzzy:.1f}%")
    print(f"  Exact accuracy:           {total_acc:.1f}%")
    print(f"  Total time:               {total_elapsed:.0f}s")

    # ── Baseline comparison ──
    print(f"\n  {'':─<70}")
    print(f"  BASELINE COMPARISON (S017 vs S019)")
    print(f"  {'-' * 70}")
    b = BASELINE
    delta_acc = total_acc - b["exact_accuracy_pct"]
    delta_name = total_fuzzy - b["name_match_rate_pct"]
    delta_matched = overall["matched"] - b["matched"]
    delta_mismatch = overall["mismatch"] - b["count_mismatch"]
    delta_missing = overall["missing"] - b["missing"]

    print(f"  {'Metric':<28s} {'Before':>10s} {'After':>10s} {'Delta':>10s}")
    print(f"  {'-'*28} {'-'*10} {'-'*10} {'-'*10}")
    print(f"  {'Exact accuracy %':<28s} {b['exact_accuracy_pct']:>9.1f}% {total_acc:>9.1f}% {delta_acc:>+9.1f}%")
    print(f"  {'Name match rate %':<28s} {b['name_match_rate_pct']:>9.1f}% {total_fuzzy:>9.1f}% {delta_name:>+9.1f}%")
    print(f"  {'Matched (name+qty)':<28s} {b['matched']:>10d} {overall['matched']:>10d} {delta_matched:>+10d}")
    print(f"  {'Count mismatches':<28s} {b['count_mismatch']:>10d} {overall['mismatch']:>10d} {delta_mismatch:>+10d}")
    print(f"  {'Missing':<28s} {b['missing']:>10d} {overall['missing']:>10d} {delta_missing:>+10d}")

    # Per-case summary table
    print(f"\n  {'Case':<16s} {'Ref':>5s} {'Gen':>5s} {'Match':>5s}"
          f" {'MisM':>5s} {'Miss':>5s} {'Xtra':>5s}"
          f" {'NmRt%':>6s} {'Acc%':>6s}")
    print(f"  {'-'*16} {'-'*5} {'-'*5} {'-'*5}"
          f" {'-'*5} {'-'*5} {'-'*5} {'-'*6} {'-'*6}")
    for res in results:
        if res.status != "ok":
            print(f"  {res.name:<16s} {'':>5s} {'':>5s} {'':>5s}"
                  f" {'':>5s} {'':>5s} {'':>5s}"
                  f" {'':>6s} {res.status:>6s}")
            continue
        print(f"  {res.name:<16s} {res.ref_item_count:>5d}"
              f" {res.gen_item_count:>5d}"
              f" {len(res.matched):>5d}"
              f" {len(res.count_mismatch):>5d}"
              f" {len(res.missing):>5d}"
              f" {len(res.extra):>5d}"
              f" {res.fuzzy_match_pct:>5.1f}%"
              f" {res.accuracy_pct:>5.1f}%")

    # ── Top 10 best matches (highest similarity) ──
    all_matches: list[tuple[str, MatchResult]] = []
    for res in results:
        if res.status == "ok":
            for m in res.matched:
                all_matches.append((res.name, m))
            for m in res.count_mismatch:
                all_matches.append((res.name, m))
    all_matches.sort(key=lambda x: x[1].match_ratio, reverse=True)

    print(f"\n  TOP 10 BEST MATCHES:")
    for case_name, m in all_matches[:10]:
        qty_mark = "=" if m.qty_match else "~"
        print(f"    [{case_name}] ratio={m.match_ratio:.0%} {qty_mark}")
        print(f"      ref: {m.ref_name[:70]}")
        print(f"      gen: {m.gen_name[:70]}")
        print(f"      qty: ref={m.ref_qty:.0f}  gen={m.gen_qty:.0f}  delta={m.delta:+.0f}")

    # ── Top 10 worst misses (reference items not found at all) ──
    all_missing: list[tuple[str, RefItem]] = []
    for res in results:
        if res.status == "ok":
            for m in res.missing:
                all_missing.append((res.name, m))
    # Sort by quantity (biggest misses first)
    all_missing.sort(key=lambda x: x[1].qty, reverse=True)

    print(f"\n  TOP 10 WORST MISSES (ref items not found):")
    for case_name, m in all_missing[:10]:
        print(f"    [{case_name}] {m.name[:65]:65s}  qty={m.qty:.0f}")

    # ── Save JSON report ──
    report_data = {
        "sprint": "S019",
        "pipeline": "parse_legend + count_symbols + match_symbols + extract_cables + vor_map_items",
        "baseline": BASELINE,
        "overall": {
            "cases_ok": ok_count,
            "cases_skip": skip_count,
            "cases_error": err_count,
            "total_ref_items": total_ref,
            "matched": overall["matched"],
            "count_mismatch": overall["mismatch"],
            "missing": overall["missing"],
            "extra": overall["extra"],
            "name_match_rate_pct": round(total_fuzzy, 1),
            "exact_accuracy_pct": round(total_acc, 1),
            "total_elapsed_sec": round(total_elapsed, 1),
        },
        "delta_vs_baseline": {
            "exact_accuracy_pct": round(delta_acc, 1),
            "name_match_rate_pct": round(delta_name, 1),
            "matched": delta_matched,
            "count_mismatch": delta_mismatch,
            "missing": delta_missing,
        },
        "top_10_best_matches": [
            {
                "case": case_name,
                "ref_name": m.ref_name,
                "gen_name": m.gen_name,
                "ref_qty": m.ref_qty,
                "gen_qty": m.gen_qty,
                "match_ratio": round(m.match_ratio, 2),
                "qty_match": m.qty_match,
            }
            for case_name, m in all_matches[:10]
        ],
        "top_10_worst_misses": [
            {
                "case": case_name,
                "name": m.name,
                "qty": m.qty,
            }
            for case_name, m in all_missing[:10]
        ],
        "results": [],
    }

    for res in results:
        entry: dict = {
            "name": res.name,
            "status": res.status,
            "ref_vor": res.ref_vor_path,
            "pdf_dir": res.pdf_dir_path,
        }
        if res.status == "ok":
            entry.update({
                "ref_item_count": res.ref_item_count,
                "gen_item_count": res.gen_item_count,
                "pdf_count": res.pdf_count,
                "pdf_errors": res.pdf_errors,
                "matched": len(res.matched),
                "count_mismatch": len(res.count_mismatch),
                "missing": len(res.missing),
                "extra": len(res.extra),
                "name_match_rate_pct": round(res.fuzzy_match_pct, 1),
                "exact_accuracy_pct": round(res.accuracy_pct, 1),
                "elapsed_sec": round(res.elapsed_sec, 1),
                "count_mismatch_details": [
                    {
                        "ref_name": m.ref_name,
                        "ref_qty": m.ref_qty,
                        "gen_name": m.gen_name,
                        "gen_qty": m.gen_qty,
                        "delta": m.delta,
                        "match_ratio": round(m.match_ratio, 2),
                    }
                    for m in res.count_mismatch
                ],
                "missing_details": [
                    {"name": m.name, "qty": m.qty}
                    for m in res.missing[:50]
                ],
                "extra_details": [
                    {"name": g.name, "qty": g.total, "files": g.files[:5]}
                    for g in res.extra[:50]
                ],
                "matched_details": [
                    {
                        "ref_name": m.ref_name,
                        "ref_qty": m.ref_qty,
                        "gen_name": m.gen_name,
                        "gen_qty": m.gen_qty,
                        "match_ratio": round(m.match_ratio, 2),
                    }
                    for m in res.matched
                ],
            })
        elif res.status != "ok":
            entry["error"] = res.error_msg

        report_data["results"].append(entry)

    report_path = PROJECT_ROOT / "test_vor_report_s019.json"
    with open(report_path, "w", encoding="utf-8") as f:
        json.dump(report_data, f, ensure_ascii=False, indent=2)
    print(f"\n  Report saved: {report_path}")
    print("=" * 80)


if __name__ == "__main__":
    main()
