"""
web_app.py — FastAPI web application for PDF legend analysis.

Provides endpoints for browsing PDFs, rendering pages, parsing legends,
and debugging word extraction.

Usage:
    uvicorn web_app:app --host 0.0.0.0 --port 8050 --reload
    # or
    python web_app.py
"""

from __future__ import annotations

import asyncio
import hashlib
import io
import json as json_mod
import math
import os
import time
from datetime import datetime
from pathlib import Path
from typing import Optional

import fitz  # PyMuPDF
import pdfplumber
from fastapi import FastAPI, HTTPException, Query, UploadFile, File
from fastapi.responses import HTMLResponse, JSONResponse, Response, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.requests import Request

import re as re_mod

from equipment_counter import process_pdf as ec_process_pdf
from pdf_legend_parser import parse_legend, LegendResult
from pdf_count_text import count_symbols
from pdf_count_cables import extract_cables
from pdf_count_geometry import measure_cables
from cable_length import measure_cable_lengths_raster
import height_bucketer
import route_classifier
import thickness_extractor
from pdf_count_visual import match_symbols, detect_pictograms, _extract_symbol_images, build_equipment_cluster_bboxes
from vor_work_mapping import map_items as vor_map_items
from legend_validator import validate_legend_symbols

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "Data"
WEB_DIR = BASE_DIR / "web"
VOR_DIR = BASE_DIR / "vor_output"   # saved VOR xlsx files
VOR_DIR.mkdir(exist_ok=True)
UPLOADS_DIR = BASE_DIR / "uploads"  # uploaded projects via web interface
UPLOADS_DIR.mkdir(exist_ok=True)

# ---------------------------------------------------------------------------
# FastAPI app
# ---------------------------------------------------------------------------

app = FastAPI(
    title="PDF Legend Viewer",
    description="Web viewer for PDF legend analysis in electrical drawings",
    version="1.0.0",
)

# Mount static files and templates
app.mount("/static", StaticFiles(directory=str(WEB_DIR / "static")), name="static")
templates = Jinja2Templates(directory=str(WEB_DIR / "templates"))


# ---------------------------------------------------------------------------
# Cache: store VOR results after SSE processing so Excel export is instant
# ---------------------------------------------------------------------------
_vor_results_cache: dict[str, dict[str, list[dict]]] = {}  # folder_id → {filename: items}
_vor_xlsx_cache: dict[str, bytes] = {}  # folder_id → ready xlsx bytes


def _vor_folder_for(rel_folder: str) -> Path:
    """Return (and create) the vor_output sub-folder for a given project folder."""
    safe = rel_folder.replace("/", "_").replace("\\", "_")
    safe = "".join(c if c.isalnum() or c in "._- " else "_" for c in safe) or "root"
    p = VOR_DIR / safe
    p.mkdir(parents=True, exist_ok=True)
    return p


def _save_vor_xlsx(xlsx_bytes: bytes, rel_folder: str) -> Path:
    """Save VOR Excel to disk with timestamp. Returns the saved file path."""
    dest_dir = _vor_folder_for(rel_folder)
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    fname = f"VOR_{ts}.xlsx"
    path = dest_dir / fname
    path.write_bytes(xlsx_bytes)
    return path


def _list_vor_files(rel_folder: str) -> list[dict]:
    """List all saved VOR xlsx files for a folder, newest first."""
    dest_dir = _vor_folder_for(rel_folder)
    files = sorted(dest_dir.glob("VOR_*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    result = []
    for f in files:
        stat = f.stat()
        result.append({
            "filename": f.name,
            "size_kb": round(stat.st_size / 1024, 1),
            "created": datetime.fromtimestamp(stat.st_mtime).strftime("%d.%m.%Y %H:%M"),
        })
    return result

# ---------------------------------------------------------------------------
# Helpers: folder operations
# ---------------------------------------------------------------------------

def _folder_id(rel_folder: str) -> str:
    """Generate stable URL-safe folder ID from relative folder path."""
    return hashlib.sha256(("FOLDER:" + rel_folder).encode("utf-8")).hexdigest()[:16]


def _id_to_folder(folder_id: str) -> Optional[str]:
    """Resolve folder ID back to relative folder path.

    Scans both DATA_DIR and UPLOADS_DIR (with '_uploads/' prefix).
    """
    seen: set[str] = set()
    # Scan DATA_DIR
    for pdf_path in DATA_DIR.rglob("*.pdf"):
        rel = str(pdf_path.relative_to(DATA_DIR)).replace("\\", "/")
        parts = rel.rsplit("/", 1)
        folder = parts[0] if len(parts) > 1 else ""
        if folder not in seen:
            seen.add(folder)
            if _folder_id(folder) == folder_id:
                return folder
    # Scan UPLOADS_DIR
    if UPLOADS_DIR.exists():
        for pdf_path in UPLOADS_DIR.rglob("*.pdf"):
            rel = str(pdf_path.relative_to(UPLOADS_DIR)).replace("\\", "/")
            parts = rel.rsplit("/", 1)
            sub = parts[0] if len(parts) > 1 else ""
            folder = "_uploads/" + sub if sub else "_uploads"
            if folder not in seen:
                seen.add(folder)
                if _folder_id(folder) == folder_id:
                    return folder
    return None


def _folder_files(rel_folder: str) -> list[Path]:
    """Return all PDF files in a specific folder under DATA_DIR or UPLOADS_DIR."""
    if rel_folder.startswith("_uploads/"):
        folder_path = UPLOADS_DIR / rel_folder[len("_uploads/"):]
    elif rel_folder == "_uploads":
        folder_path = UPLOADS_DIR
    else:
        folder_path = DATA_DIR / rel_folder
    if not folder_path.is_dir():
        return []
    return sorted(folder_path.glob("*.pdf"))


templates.env.globals["folder_id"] = _folder_id


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _file_id(rel_path: str) -> str:
    """Generate a stable, URL-safe file ID from a relative path."""
    return hashlib.sha256(rel_path.encode("utf-8")).hexdigest()[:16]


def _id_to_path(file_id: str) -> Optional[Path]:
    """Resolve file ID back to an absolute path by scanning all PDFs.

    Searches both DATA_DIR and UPLOADS_DIR.
    """
    for pdf_path in DATA_DIR.rglob("*.pdf"):
        rel = str(pdf_path.relative_to(DATA_DIR))
        if _file_id(rel) == file_id:
            return pdf_path
    # Search uploads
    if UPLOADS_DIR.exists():
        for pdf_path in UPLOADS_DIR.rglob("*.pdf"):
            rel = "_uploads/" + str(pdf_path.relative_to(UPLOADS_DIR)).replace("\\", "/")
            if _file_id(rel) == file_id:
                return pdf_path
    return None


def _guess_type(filename: str) -> str:
    """Guess drawing type from filename/path segments."""
    upper = filename.upper()
    # Check path segments and filename
    if "/ЭО/" in upper or "\\ЭО\\" in upper or "ЭО" in upper:
        return "ЭО"
    if "/ЭМ/" in upper or "\\ЭМ\\" in upper or "ЭМ" in upper:
        return "ЭМ"
    if "/ЭГ/" in upper or "\\ЭГ\\" in upper or "ЭГ" in upper:
        return "ЭГ"
    if "/ЭС/" in upper or "\\ЭС\\" in upper or "ЭС" in upper:
        return "ЭС"
    return ""


def _detect_section_type(folder_path: str) -> str:
    """Detect electrical section type from folder path.

    Returns "ЭО", "ЭМ", or "ЭГ" based on folder name segments.
    Default: "ЭО" (most common — electrical lighting).
    """
    # Normalize separators
    normalized = folder_path.replace("\\", "/")
    # Check for section markers in path segments
    # Use segment boundaries to avoid false matches (e.g. "ЭОС" should not match "ЭО")
    for segment in normalized.split("/"):
        seg = segment.strip()
        if seg == "ЭО" or seg.startswith("ЭО "):
            return "ЭО"
        if seg == "ЭМ" or seg.startswith("ЭМ "):
            return "ЭМ"
        if seg == "ЭГ" or seg.startswith("ЭГ "):
            return "ЭГ"
    # Fallback: check if any segment contains ЭМ or ЭГ as a substring
    upper = normalized.upper()
    if "/ЭМ/" in upper or "/ЭМ" == upper[-3:]:
        return "ЭМ"
    if "/ЭГ/" in upper or "/ЭГ" == upper[-3:]:
        return "ЭГ"
    return "ЭО"


def _scan_pdfs() -> list[dict]:
    """Recursively scan Data/ and uploads/ for PDF files and return metadata."""
    results = []

    # Scan DATA_DIR
    if DATA_DIR.exists():
        for pdf_path in sorted(DATA_DIR.rglob("*.pdf")):
            try:
                stat = pdf_path.stat()
            except OSError:
                continue

            rel = str(pdf_path.relative_to(DATA_DIR))
            fid = _file_id(rel)

            # Compute folder group
            rel_posix = rel.replace("\\", "/")
            parts = rel_posix.rsplit("/", 1)
            folder = parts[0] if len(parts) > 1 else ""

            results.append({
                "id": fid,
                "filename": pdf_path.name,
                "path": rel,
                "folder": folder,
                "size_kb": round(stat.st_size / 1024, 1),
                "modified": time.strftime(
                    "%Y-%m-%d %H:%M", time.localtime(stat.st_mtime)
                ),
                "type_guess": _guess_type(rel),
            })

    # Scan UPLOADS_DIR
    if UPLOADS_DIR.exists():
        for pdf_path in sorted(UPLOADS_DIR.rglob("*.pdf")):
            try:
                stat = pdf_path.stat()
            except OSError:
                continue

            rel_upload = str(pdf_path.relative_to(UPLOADS_DIR)).replace("\\", "/")
            rel = "_uploads/" + rel_upload
            fid = _file_id(rel)

            # Compute folder group (with _uploads/ prefix)
            parts = rel.rsplit("/", 1)
            folder = parts[0] if len(parts) > 1 else "_uploads"

            results.append({
                "id": fid,
                "filename": pdf_path.name,
                "path": rel,
                "folder": folder,
                "size_kb": round(stat.st_size / 1024, 1),
                "modified": time.strftime(
                    "%Y-%m-%d %H:%M", time.localtime(stat.st_mtime)
                ),
                "type_guess": _guess_type(rel),
                "is_upload": True,
            })

    return results


# ---------------------------------------------------------------------------
# 1. GET / — main page
# ---------------------------------------------------------------------------

@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    """Render the main page with file list."""
    files = _scan_pdfs()
    return templates.TemplateResponse(
        request=request,
        name="index.html",
        context={"request": request, "files": files, "total": len(files)},
    )


# ---------------------------------------------------------------------------
# GET /viewer/{file_id} — viewer page for a specific file
# ---------------------------------------------------------------------------

@app.get("/viewer/{file_id}", response_class=HTMLResponse)
async def viewer(request: Request, file_id: str):
    """Render the PDF viewer page for a specific file."""
    pdf_path = _id_to_path(file_id)
    if pdf_path is None:
        raise HTTPException(status_code=404, detail="File not found")

    # Get page count
    try:
        doc = fitz.open(str(pdf_path))
        page_count = len(doc)
        doc.close()
    except Exception:
        page_count = 1

    try:
        rel = str(pdf_path.relative_to(DATA_DIR))
    except ValueError:
        # File is in UPLOADS_DIR
        rel = "_uploads/" + str(pdf_path.relative_to(UPLOADS_DIR)).replace("\\", "/")
    return templates.TemplateResponse(
        request=request,
        name="viewer.html",
        context={
            "request": request,
            "file_id": file_id,
            "filename": pdf_path.name,
            "filepath": rel,
            "page_count": page_count,
        },
    )


# ---------------------------------------------------------------------------
# 2. GET /api/files — JSON file list
# ---------------------------------------------------------------------------

@app.get("/api/files")
async def api_files():
    """Return JSON list of all PDFs in Data/ directory."""
    files = _scan_pdfs()
    return JSONResponse(content=files)


# ---------------------------------------------------------------------------
# 3. GET /api/file/{id}/render — render PDF page as PNG
# ---------------------------------------------------------------------------

@app.get("/api/file/{file_id}/render")
async def api_render(
    file_id: str,
    page: int = Query(0, ge=0, description="Page index (0-based)"),
    dpi: int = Query(150, ge=72, le=600, description="Render DPI"),
):
    """Render a PDF page as a PNG image using PyMuPDF."""
    pdf_path = _id_to_path(file_id)
    if pdf_path is None:
        raise HTTPException(status_code=404, detail="File not found")

    try:
        doc = fitz.open(str(pdf_path))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Cannot open PDF: {e}")

    if page >= len(doc):
        doc.close()
        raise HTTPException(
            status_code=400,
            detail=f"Page {page} out of range (0-{len(doc) - 1})",
        )

    zoom = dpi / 72.0
    mat = fitz.Matrix(zoom, zoom)
    pix = doc[page].get_pixmap(matrix=mat, alpha=False)

    img_bytes = pix.tobytes("png")
    doc.close()

    return Response(content=img_bytes, media_type="image/png")


# ---------------------------------------------------------------------------
# 4. GET /api/file/{id}/legend — parse legend
# ---------------------------------------------------------------------------

@app.get("/api/file/{file_id}/legend")
async def api_legend(file_id: str):
    """Parse legend from PDF and return structured JSON result."""
    pdf_path = _id_to_path(file_id)
    if pdf_path is None:
        raise HTTPException(status_code=404, detail="File not found")

    try:
        result = parse_legend(str(pdf_path))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Legend parse error: {e}")

    # Determine legend_type
    has_numbered = any(item.symbol and item.symbol[0].isdigit() for item in result.items)
    has_graphical = any(not item.symbol for item in result.items)
    if has_numbered and has_graphical:
        legend_type = "mixed"
    elif has_numbered:
        legend_type = "numbered"
    elif has_graphical:
        legend_type = "graphical"
    else:
        legend_type = "numbered"

    # Count raw words for debug info
    raw_words_count = 0
    try:
        with pdfplumber.open(str(pdf_path)) as pdf:
            for page in pdf.pages:
                words = page.extract_words(x_tolerance=3, y_tolerance=3) or []
                raw_words_count += len(words)
    except Exception:
        pass

    return JSONResponse(content={
        "legend_found": len(result.items) > 0,
        "legend_bbox": {
            "x0": round(result.legend_bbox[0], 1),
            "y0": round(result.legend_bbox[1], 1),
            "x1": round(result.legend_bbox[2], 1),
            "y1": round(result.legend_bbox[3], 1),
        },
        "page": result.page_index,
        "items": [
            {
                "symbol": item.symbol,
                "description": item.description,
                "category": item.category,
                "color": item.color,
                "bbox": {
                    "x0": round(item.bbox[0], 1),
                    "y0": round(item.bbox[1], 1),
                    "x1": round(item.bbox[2], 1),
                    "y1": round(item.bbox[3], 1),
                },
                "image_url": f"/api/file/{file_id}/symbol_image/{i}",
            }
            for i, item in enumerate(result.items)
        ],
        "legend_type": legend_type,
        "raw_words_count": raw_words_count,
        "columns_detected": result.columns_detected,
    })


# ---------------------------------------------------------------------------
# 4b. GET /api/file/{id}/validate_legend — validate legend symbol uniqueness
# ---------------------------------------------------------------------------

@app.get("/api/file/{file_id}/validate_legend")
async def api_validate_legend(file_id: str):
    """Validate that each legend item can be uniquely identified.

    Checks all pairs of legend items for potential conflicts:
    - Same text marker -> ERROR
    - Similar visual templates with no distinguishing features -> PROBLEM
    - Similar visuals but different colors -> OK (color distinguishes)
    - Different text markers -> OK

    Returns per-item status and conflict details.
    """
    pdf_path = _id_to_path(file_id)
    if pdf_path is None:
        raise HTTPException(status_code=404, detail="File not found")

    try:
        import time as _time
        t0 = _time.time()
        result = await asyncio.to_thread(
            validate_legend_symbols, str(pdf_path)
        )
        elapsed = round(_time.time() - t0, 2)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Validation error: {e}")

    # Serialize items
    items_json = []
    for sv in result.items:
        items_json.append({
            "index": sv.index,
            "symbol": sv.symbol,
            "description": sv.description,
            "color": sv.color,
            "has_text_marker": sv.has_text_marker,
            "has_visual_template": sv.has_visual_template,
            "conflicts": sv.conflicts,
            "status": sv.status,
            "notes": sv.notes,
        })

    # Serialize conflicts
    conflicts_json = []
    for cp in result.conflicts:
        conflicts_json.append({
            "index_a": cp.index_a,
            "index_b": cp.index_b,
            "symbol_a": cp.symbol_a,
            "symbol_b": cp.symbol_b,
            "description_a": cp.description_a,
            "description_b": cp.description_b,
            "conflict_type": cp.conflict_type,
            "visual_similarity": cp.visual_similarity,
            "color_a": cp.color_a,
            "color_b": cp.color_b,
            "distinguishable": cp.distinguishable,
            "resolution": cp.resolution,
        })

    return JSONResponse(content={
        "total": result.total,
        "ok_count": result.ok_count,
        "conflict_count": result.conflict_count,
        "unresolvable_count": result.unresolvable_count,
        "items": items_json,
        "conflicts": conflicts_json,
        "elapsed_s": elapsed,
    })


# ---------------------------------------------------------------------------
# 5. GET /api/file/{id}/render_with_overlay — render with legend highlight
# ---------------------------------------------------------------------------

@app.get("/api/file/{file_id}/render_with_overlay")
async def api_render_with_overlay(
    file_id: str,
    page: int = Query(0, ge=0, description="Page index (0-based)"),
    dpi: int = Query(150, ge=72, le=600, description="Render DPI"),
):
    """Render a PDF page with the legend bbox highlighted as a semi-transparent overlay."""
    pdf_path = _id_to_path(file_id)
    if pdf_path is None:
        raise HTTPException(status_code=404, detail="File not found")

    # First, parse legend to get bbox and page
    try:
        legend_result = parse_legend(str(pdf_path))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Legend parse error: {e}")

    # Open PDF with PyMuPDF
    try:
        doc = fitz.open(str(pdf_path))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Cannot open PDF: {e}")

    # Use the legend page if found, otherwise use requested page
    target_page = legend_result.page_index if legend_result.items else page
    if target_page >= len(doc):
        doc.close()
        raise HTTPException(
            status_code=400,
            detail=f"Page {target_page} out of range (0-{len(doc) - 1})",
        )

    zoom = dpi / 72.0

    # Get the page and add legend highlight annotation
    fitz_page = doc[target_page]

    if legend_result.items:
        bbox = legend_result.legend_bbox
        # pdfplumber coordinates: origin at top-left, Y increases downward
        # fitz coordinates: same convention for page rendering
        rect = fitz.Rect(bbox[0], bbox[1], bbox[2], bbox[3])

        # Draw a semi-transparent rectangle
        shape = fitz_page.new_shape()
        shape.draw_rect(rect)
        shape.finish(
            color=(1, 0, 0),        # red border
            fill=(1, 0.9, 0.8),     # light orange fill
            fill_opacity=0.25,
            width=3,
        )
        shape.commit()

        # Also highlight individual item bboxes
        for item in legend_result.items:
            ib = item.bbox
            item_rect = fitz.Rect(ib[0], ib[1], ib[2], ib[3])
            item_shape = fitz_page.new_shape()
            item_shape.draw_rect(item_rect)
            item_shape.finish(
                color=(0, 0.4, 0.8),     # blue border
                fill=(0.8, 0.9, 1.0),    # light blue fill
                fill_opacity=0.15,
                width=1,
            )
            item_shape.commit()

    mat = fitz.Matrix(zoom, zoom)
    pix = fitz_page.get_pixmap(matrix=mat, alpha=False)
    img_bytes = pix.tobytes("png")
    doc.close()

    return Response(content=img_bytes, media_type="image/png")


# ---------------------------------------------------------------------------
# 6. GET /api/file/{id}/debug_words — debug word extraction
# ---------------------------------------------------------------------------

@app.get("/api/file/{file_id}/debug_words")
async def api_debug_words(
    file_id: str,
    page: int = Query(0, ge=0, description="Page index (0-based)"),
):
    """Return all pdfplumber words with coordinates for debugging."""
    pdf_path = _id_to_path(file_id)
    if pdf_path is None:
        raise HTTPException(status_code=404, detail="File not found")

    try:
        with pdfplumber.open(str(pdf_path)) as pdf:
            if page >= len(pdf.pages):
                raise HTTPException(
                    status_code=400,
                    detail=f"Page {page} out of range (0-{len(pdf.pages) - 1})",
                )

            p = pdf.pages[page]
            words = p.extract_words(x_tolerance=3, y_tolerance=3) or []

            return JSONResponse(content={
                "page": page,
                "page_width": round(p.width, 1),
                "page_height": round(p.height, 1),
                "words_count": len(words),
                "words": [
                    {
                        "text": w["text"],
                        "x0": round(w["x0"], 1),
                        "top": round(w["top"], 1),
                        "x1": round(w["x1"], 1),
                        "bottom": round(w["bottom"], 1),
                    }
                    for w in words
                ],
            })
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Word extraction error: {e}")


# ---------------------------------------------------------------------------
# 6b. GET /api/folders — JSON list of folders
# ---------------------------------------------------------------------------

@app.get("/api/folders")
async def api_folders():
    """Return JSON list of folders containing PDFs with IDs and counts."""
    files = _scan_pdfs()
    folder_map: dict[str, dict] = {}
    for f in files:
        folder = f["folder"]
        if folder not in folder_map:
            folder_map[folder] = {"id": _folder_id(folder), "path": folder,
                                   "name": folder.rsplit("/", 1)[-1] if "/" in folder else folder,
                                   "count": 0, "types": set()}
        folder_map[folder]["count"] += 1
        if f["type_guess"]:
            folder_map[folder]["types"].add(f["type_guess"])
    result = []
    for info in sorted(folder_map.values(), key=lambda x: x["path"]):
        info["types"] = sorted(info["types"])
        result.append(info)
    return JSONResponse(content=result)


# ---------------------------------------------------------------------------
# 6c. POST /api/upload — upload PDF files to create a new project
# ---------------------------------------------------------------------------

@app.post("/api/upload")
async def api_upload(
    request: Request,
    files: list[UploadFile] = File(...),
):
    """Upload PDFs and create a new project folder in uploads/.

    Accepts multipart form with 'files' (PDF blobs) and optional 'paths'
    (relative paths preserving folder structure, e.g. "MyProject/02_PDF/plan.pdf").
    """
    # Parse optional 'paths' list from the same multipart form
    form = await request.form()
    raw_paths = form.getlist("paths")

    ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    dest = UPLOADS_DIR / ts
    dest.mkdir(parents=True, exist_ok=True)

    saved = 0
    for i, f in enumerate(files):
        if not f.filename or not f.filename.lower().endswith(".pdf"):
            continue

        # Use relative path from frontend if available, else just filename
        rel_path = raw_paths[i] if i < len(raw_paths) and raw_paths[i] else f.filename
        # Sanitize: resolve to pure posix, strip leading slashes / ".."
        rel_path = rel_path.replace("\\", "/")
        parts = [p for p in rel_path.split("/") if p and p != ".."]
        if not parts:
            continue
        safe_path = Path(*parts)

        target = dest / safe_path
        target.parent.mkdir(parents=True, exist_ok=True)
        content = await f.read()
        target.write_bytes(content)
        saved += 1

    if saved == 0:
        import shutil
        shutil.rmtree(dest, ignore_errors=True)
        raise HTTPException(400, "Нет PDF файлов")

    rel = "_uploads/" + ts
    return JSONResponse({
        "folder_id": _folder_id(rel),
        "folder_name": ts,
        "files_count": saved,
    })


# ---------------------------------------------------------------------------
# Helpers: equipment aggregation
# ---------------------------------------------------------------------------

def _aggregate_equipment(results: dict[str, list[dict]]) -> list[dict]:
    """Aggregate equipment items from multiple files into VOR table.

    Uses 'work_name' (VOR work description) for aggregation if available,
    falling back to 'name'. Preserves 'equipment_name' (original equipment
    name from the PDF legend) for the 'Доп. информация' column.

    T081 (S019-wire-attrs): preserves six attribute fields populated by
    Steps 8-10 (height_bucketer / route_classifier / thickness_extractor)
    so the downstream vor_compose.compose_vor_table can split rows by
    installation context.  KB-008: agg key is extended from name only to
    (name, height_bucket, route, mount) so that the same legend name
    appearing on different floor elevations or routes is NOT silently
    merged into one row.  Scalar fields (cross_section / section_mm2 /
    diameter_mm) are projected as the first non-empty value seen because
    they describe the cable kind, not the installation context.
    """
    import re
    agg: dict[tuple, dict] = {}  # (key, bucket, route, mount) -> {...}

    # T088: strip [UNMATCHED-LEGEND] prefix BEFORE aggregation so that
    # matched and unmatched variants of the same luminaire model merge
    # into one row instead of producing two split rows per model.
    # Idempotent + case-insensitive; mirrors the strip already applied at
    # render time (T078) but applied earlier so the aggregation key
    # collapses correctly.
    _UNMATCHED_LEGEND_RE = re.compile(r"\[UNMATCHED-LEGEND\]\s*", re.IGNORECASE)
    # T088: also collapse "Монтаж светильника Светильник X"
    # → "Монтаж светильника X".  This pattern arises because matched
    # items have work_name="Монтаж светильника <X>" while [UNMATCHED]
    # items fall back to raw_name="Светильник <X>"; after stripping the
    # [UNMATCHED-LEGEND] tag the work-name prefix gets re-prepended
    # downstream and produces the doubled "Монтаж светильника Светильник".
    # Collapsing here lets the matched and unmatched paths produce the
    # same aggregation key.
    _DOUBLED_LUM_RE = re.compile(
        r"^(Монтаж\s+светильника)\s+Светильник\s+",
        re.IGNORECASE,
    )

    def _strip_unmatched(s: str) -> str:
        if not s:
            return s
        out = _UNMATCHED_LEGEND_RE.sub("", s).strip()
        out = _DOUBLED_LUM_RE.sub(r"\1 ", out).strip()
        return out

    for filename, items in results.items():
        drawing_ref = filename.replace(".pdf", "")
        for item in items:
            # Prefer work_name for VOR display; fall back to name
            work_name = _strip_unmatched(item.get("work_name", "").strip())
            raw_name = _strip_unmatched(item.get("name", "").strip())
            display_name = work_name or raw_name
            if not display_name:
                continue
            key_name = re.sub(r"\s+", " ", display_name).strip().lower()
            total = item.get("total", item.get("count", 0) + item.get("count_ae", 0))
            unit = item.get("unit", "шт")
            if total <= 0:
                continue
            # T081: extract 6 attribute fields populated by Steps 8-10.
            height_bucket = item.get("height_bucket") or None
            route = item.get("route") or None
            mount = item.get("mount") or None
            cross_section = item.get("cross_section") or None
            section_mm2 = item.get("section_mm2")
            diameter_mm = item.get("diameter_mm")
            # KB-008: agg key includes attribute context so same-name-
            # different-bucket/route/mount rows are not merged.
            key = (key_name, height_bucket, route, mount)
            if key not in agg:
                # equipment_name is the original name from PDF legend
                equip_name = _strip_unmatched(
                    item.get("equipment_name", raw_name)
                )
                agg[key] = {
                    "name": display_name, "unit": unit, "total": 0,
                    "per_file": {}, "files": [],
                    "equipment_names": set(),
                    # T081: preserved attribute context
                    "height_bucket": height_bucket,
                    "route": route,
                    "mount": mount,
                    "cross_section": cross_section,
                    "section_mm2": section_mm2,
                    "diameter_mm": diameter_mm,
                }
                if equip_name:
                    agg[key]["equipment_names"].add(equip_name)
            else:
                equip_name = _strip_unmatched(
                    item.get("equipment_name", raw_name)
                )
                if equip_name:
                    agg[key]["equipment_names"].add(equip_name)
                # Scalar projection: take first non-empty value seen.
                if not agg[key].get("cross_section") and cross_section:
                    agg[key]["cross_section"] = cross_section
                if agg[key].get("section_mm2") in (None, 0) and section_mm2:
                    agg[key]["section_mm2"] = section_mm2
                if agg[key].get("diameter_mm") in (None, 0) and diameter_mm:
                    agg[key]["diameter_mm"] = diameter_mm
            agg[key]["total"] += total
            agg[key]["per_file"][drawing_ref] = agg[key]["per_file"].get(drawing_ref, 0) + total
            if drawing_ref not in agg[key]["files"]:
                agg[key]["files"].append(drawing_ref)

    result = []
    for i, (key, info) in enumerate(sorted(agg.items(), key=lambda x: x[1]["name"]), 1):
        formula_parts = [str(v) for v in info["per_file"].values()]
        formula = "+".join(formula_parts) if len(formula_parts) > 1 else (formula_parts[0] if formula_parts else "")
        # Build extra info from original equipment names
        equip_names = sorted(info.get("equipment_names", set()))
        extra_info = "; ".join(equip_names) if equip_names else ""
        result.append({
            "row": i, "name": info["name"], "unit": info["unit"],
            "total": info["total"], "formula": formula,
            "drawing_refs": ", ".join(info["files"]),
            "extra_info": extra_info,
            # T081: project preserved attributes to output rows so
            # vor_compose.compose_vor_table and downstream renderers
            # can group/split by installation context.
            "height_bucket": info.get("height_bucket"),
            "route": info.get("route"),
            "mount": info.get("mount"),
            "cross_section": info.get("cross_section"),
            "section_mm2": info.get("section_mm2"),
            "diameter_mm": info.get("diameter_mm"),
        })
    return result


def _detect_section_type(pdf_path: str) -> str:
    """Detect the drawing section type from file/folder name.

    Returns one of: "ЭО", "ЭМ", "ЭГ", "ЭОМ", or "unknown".

    Detection order:
      1. Folder name containing section code (e.g. .../ЭМ/..., .../ЭГ/...)
      2. Filename containing section code (e.g. 1Д-24-1-ЭМ.pdf)
    """
    import re
    path_str = str(pdf_path).replace("\\", "/")
    # Check from most specific to least; ЭОМ before ЭО to avoid false match
    for code in ("ЭОМ", "ЭО", "ЭМ", "ЭГ"):
        # Folder match: /ЭМ/ or /ЭМ\  or path ends with /ЭМ
        if re.search(rf"[/\\]{code}(?:[/\\]|$)", path_str):
            return code
        # Filename match: -ЭМ.pdf, -ЭМ_, _ЭМ.pdf, _ЭМ_, " ЭМ "
        basename = path_str.rsplit("/", 1)[-1]
        if re.search(rf"[-_\s]{code}(?:[-_.\s]|$)", basename, re.IGNORECASE):
            return code
        # Also match if basename STARTS with the section code
        if basename.upper().startswith(code):
            return code
    return "unknown"


_LUMINAIRE_NAME_RE = re_mod.compile(r"светильник", re_mod.IGNORECASE)
_LEVEL_MARK_RE = re_mod.compile(r"^[+\-]?\d{1,2}[.,]\d{3}$")
_LEVEL_FROM_NAME_RE = re_mod.compile(r"отм\.\s*([+\-]?\d{1,2}[.,]\d{3})", re_mod.IGNORECASE)
_HEIGHT_INLINE_RE = re_mod.compile(
    r"^[HНhн]\s*[:=]?\s*(\d{1,2}(?:[.,]\d{1,3})?)$"
)
_HEIGHT_VALUE_RE = re_mod.compile(r"^\d{1,2}(?:[.,]\d{1,3})?$")
_HEIGHT_VALUE_WITH_UNIT_RE = re_mod.compile(
    r"^(\d{1,2}(?:[.,]\d{1,3})?)\s*[mм]$"
)
_HEIGHT_PREFIX_TOKENS = {"H", "Н", "h", "н", "H=", "Н=", "h=", "н="}


def _to_float_maybe(raw: str) -> Optional[float]:
    try:
        return float(raw.replace(",", "."))
    except Exception:
        return None


def _detect_sheet_level_elevation(pdf_path: Path, page_index: int) -> tuple[Optional[float], str]:
    """Detect sheet elevation level (e.g. +4.200) from filename/page."""
    m = _LEVEL_FROM_NAME_RE.search(pdf_path.name)
    if m:
        level = _to_float_maybe(m.group(1))
        if level is not None:
            return level, "filename"

    try:
        with pdfplumber.open(str(pdf_path)) as pdf:
            if page_index >= len(pdf.pages):
                return None, ""
            words = pdf.pages[page_index].extract_words() or []
    except Exception:
        return None, ""

    candidates: list[float] = []
    for w in words:
        t = (w.get("text") or "").strip()
        if not _LEVEL_MARK_RE.match(t):
            continue
        val = _to_float_maybe(t)
        if val is None:
            continue
        if -10.0 <= val <= 120.0:
            candidates.append(val)

    if not candidates:
        return None, ""

    # Pick the median-like representative to avoid outlier labels.
    candidates.sort()
    return candidates[len(candidates) // 2], "page"


def _extract_mount_height_near_anchor(
    words: list[dict],
    anchor_x: float,
    anchor_y: float,
    search_radius_pt: float = 90.0,
) -> tuple[Optional[float], str, float, str]:
    """Extract mount height text nearest to an anchor point."""
    nearby: list[tuple[int, dict, float]] = []
    for i, w in enumerate(words):
        x0 = float(w.get("x0", 0))
        y0 = float(w.get("top", 0))
        x1 = float(w.get("x1", x0))
        y1 = float(w.get("bottom", y0))
        cx = (x0 + x1) / 2.0
        cy = (y0 + y1) / 2.0
        dist = math.sqrt((cx - anchor_x) ** 2 + (cy - anchor_y) ** 2)
        if dist <= search_radius_pt:
            nearby.append((i, w, dist))

    if not nearby:
        return None, "", 0.0, ""

    nearby.sort(key=lambda t: t[2])

    for i, w, _dist in nearby:
        t = (w.get("text") or "").strip()
        m = _HEIGHT_INLINE_RE.match(t)
        if m:
            val = _to_float_maybe(m.group(1))
            if val is not None and 1.0 <= val <= 20.0:
                return val, "nearby_text:inline", 0.95, t

        m = _HEIGHT_VALUE_WITH_UNIT_RE.match(t)
        if m:
            val = _to_float_maybe(m.group(1))
            if val is not None and 1.0 <= val <= 20.0:
                return val, "nearby_text:value+unit", 0.85, t

        if t in _HEIGHT_PREFIX_TOKENS and i + 1 < len(words):
            t2 = (words[i + 1].get("text") or "").strip()
            if _HEIGHT_VALUE_RE.match(t2):
                v2 = _to_float_maybe(t2)
                if v2 is not None and 1.0 <= v2 <= 20.0:
                    return v2, "nearby_text:prefix+next", 0.75, f"{t} {t2}"

        # Common compact token like "Н2"
        if len(t) >= 2 and t[0] in ("Н", "н", "H", "h") and t[1:].isdigit():
            val = _to_float_maybe(t[1:])
            if val is not None and 1.0 <= val <= 20.0:
                return val, "nearby_text:compact", 0.65, t

    return None, "", 0.0, ""


def _detect_default_mount_height(words: list[dict]) -> tuple[Optional[float], str]:
    """Detect page-level default mount height token (e.g. repeated 'Н2')."""
    counts: dict[float, int] = {}
    for w in words:
        t = (w.get("text") or "").strip()
        val: Optional[float] = None
        if len(t) >= 2 and t[0] in ("Н", "н", "H", "h") and t[1:].isdigit():
            val = _to_float_maybe(t[1:])
        else:
            m = _HEIGHT_INLINE_RE.match(t)
            if m:
                val = _to_float_maybe(m.group(1))
        if val is None or not (1.0 <= val <= 20.0):
            continue
        key = round(val, 3)
        counts[key] = counts.get(key, 0) + 1

    if not counts:
        return None, ""

    best_val, best_cnt = max(counts.items(), key=lambda kv: kv[1])
    if best_cnt < 1:
        return None, ""
    return float(best_val), "page_default"


def extract_luminaire_heights(pdf_path: str) -> dict:
    """Extract per-luminaire mount height candidates from one PDF page."""
    path = Path(pdf_path)
    legend = parse_legend(pdf_path)
    if not legend.items:
        return {
            "page_index": 0,
            "level_elevation": None,
            "level_source": "",
            "total_anchors": 0,
            "with_mount_height": 0,
            "rows": [],
        }

    page_index = legend.page_index
    level_elevation, level_source = _detect_sheet_level_elevation(path, page_index)

    # Detect luminaires in legend.
    lum_idx_to_item: dict[int, object] = {}
    lum_sym_to_name: dict[str, str] = {}
    for idx, item in enumerate(legend.items):
        name = (item.description or "").strip()
        sym = (item.symbol or "").strip()
        if not name or not _LUMINAIRE_NAME_RE.search(name):
            continue
        lum_idx_to_item[idx] = item
        if sym:
            lum_sym_to_name[sym] = name

    # Build anchors from text positions + visual matches.
    anchors: list[dict] = []
    try:
        text_result = count_symbols(pdf_path, legend)
        for p in text_result.positions:
            if p.symbol not in lum_sym_to_name:
                continue
            anchors.append({
                "symbol": p.symbol,
                "name": lum_sym_to_name[p.symbol],
                "x": float(p.x),
                "y": float(p.y),
                "method": "text",
            })
    except Exception:
        pass

    try:
        vis_result = match_symbols(pdf_path, legend)
        for m in vis_result.matches:
            if m.symbol_index not in lum_idx_to_item:
                continue
            item = lum_idx_to_item[m.symbol_index]
            anchors.append({
                "symbol": (item.symbol or "").strip(),
                "name": item.description or "",
                "x": float(m.x),
                "y": float(m.y),
                "method": "visual",
            })
    except Exception:
        pass

    # Deduplicate anchors by coarse coordinate key.
    deduped: list[dict] = []
    seen: set[tuple[str, int, int]] = set()
    for a in anchors:
        key = (a["symbol"], int(round(a["x"] / 3.0)), int(round(a["y"] / 3.0)))
        if key in seen:
            continue
        seen.add(key)
        deduped.append(a)
    anchors = deduped

    # Read page words once for nearby-height lookup.
    words: list[dict] = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            if page_index < len(pdf.pages):
                words = pdf.pages[page_index].extract_words() or []
    except Exception:
        words = []
    default_mount_height, default_mount_source = _detect_default_mount_height(words)

    rows: list[dict] = []
    with_mount_height = 0
    for a in anchors:
        mount_h, mount_src, conf, raw = _extract_mount_height_near_anchor(
            words, a["x"], a["y"],
        )
        if mount_h is None and default_mount_height is not None:
            mount_h = default_mount_height
            mount_src = default_mount_source
            conf = 0.35
            raw = ""
        if mount_h is not None:
            with_mount_height += 1
        rows.append({
            "symbol": a["symbol"],
            "name": a["name"],
            "x": round(a["x"], 1),
            "y": round(a["y"], 1),
            "method": a["method"],
            "page_index": page_index,
            "level_elevation": level_elevation,
            "level_source": level_source,
            "mount_height": mount_h,
            "mount_height_source": mount_src,
            "confidence": round(conf, 2),
            "raw_text": raw,
        })

    return {
        "page_index": page_index,
        "level_elevation": level_elevation,
        "level_source": level_source,
        "total_anchors": len(rows),
        "with_mount_height": with_mount_height,
        "rows": rows,
    }


# ---------------------------------------------------------------------------
# T058 / B051: Reverse label -> legend channel
# ---------------------------------------------------------------------------
# T057 recon established that on-drawing engineer labels (e.g. PU6, PU7
# for "Post upravleniya"; SHO3-Gr.15 for working-lighting groups) are
# stored as pdfplumber chars with explicit non-default non_stroking_color
# (blue (0,0,1) for working circuits, red (1,0,0) for emergency).  These
# labels never reach the equipment count because:
#   - Phase 2 (count_symbols) only looks for the legend symbol token.
#   - Phase 3 (match_symbols) needs a distinct glyph in the legend.
#   - Phase 5 (extract_cables) only catches cable schedule entries.
# This helper extracts those colored labels, groups char-level data into
# words, classifies each word into a recognised label family (PU\d+,
# VKL, POST, etc.), then fuzzy-matches the label text against the
# descriptions of legend rows that NO producer stage has covered yet.
# When a match scores above threshold, all labels in that group are
# attributed to the matched legend index.

_REVERSE_BLUE_RGB = (0.0, 0.0, 1.0)
_REVERSE_RED_RGB = (1.0, 0.0, 0.0)
_REVERSE_COLOR_TOL = 0.05  # tuples are exact in this corpus, give tiny slack
# Tight regex for short labels that map back to legend equipment.
# - PU\d+  : control post (Cyrillic Pe-U + digit)
# - VKL\d* : switch ("VKL" = vyklyuchatel in Cyrillic)
# - POST   : standalone POST keyword
# - PULT\w*: pult control
# - DV\d+  : motion sensor / DataVid family
_REVERSE_EQUIP_LABEL_RE = re_mod.compile(
    r"^("
    r"\u041f\u0423\d+"       # PU<n>
    r"|\u0412\u041a\u041b\d*"  # VKL[<n>]
    r"|\u041f\u041e\u0421\u0422"  # POST
    r"|\u041f\u0423\u041b\u042c\w*"  # PULT...
    r"|\u0414\u0412\d+"      # DV<n>
    r")$"
)
# Threshold for fuzzy match between label/keyword and legend description.
# Tuned from T057 recon: "PU" vs "Post upravleniya rabochim osveshcheniem"
# token_set_ratio is ~38; we use a keyword-tag lookup table instead so the
# threshold here is the secondary fallback for plain-text labels.
_REVERSE_FUZZY_THRESHOLD = 65

# Label -> legend-description keyword hint table.  Each label family is
# tied to a set of Cyrillic keywords that MUST appear in the legend
# description for the match to be considered.  This prevents PU<n>
# matching any legend that happens to fuzzy-score high.
_REVERSE_LABEL_KEYWORDS: dict[str, tuple[str, ...]] = {
    "PU":   ("\u043f\u043e\u0441\u0442", "\u0443\u043f\u0440\u0430\u0432\u043b"),   # post + upravlen
    "VKL":  ("\u0432\u044b\u043a\u043b\u044e\u0447\u0430\u0442\u0435\u043b",),       # vyklyuchatel
    "POST": ("\u043f\u043e\u0441\u0442",),                                            # post
    "PULT": ("\u043f\u0443\u043b\u044c\u0442",),                                      # pult
    "DV":   ("\u0434\u0430\u0442\u0447\u0438\u043a",),                                # datchik (sensor)
}


def _reverse_label_family(text: str) -> str | None:
    """Map a label string like 'PU7' to its family key ('PU', 'VKL', ...)."""
    if not text:
        return None
    # Strip trailing digits to get family stem
    stem = re_mod.sub(r"\d+$", "", text)
    stem_upper = stem.upper()
    # Cyrillic -> Latin equivalent mapping for stems we care about
    cyr_to_lat = {
        "\u041f\u0423": "PU",
        "\u0412\u041a\u041b": "VKL",
        "\u041f\u041e\u0421\u0422": "POST",
        "\u041f\u0423\u041b\u042c": "PULT",
        "\u0414\u0412": "DV",
    }
    return cyr_to_lat.get(stem_upper) or (stem_upper if stem_upper in _REVERSE_LABEL_KEYWORDS else None)


def _reverse_extract_colored_words(pdf_path: str, page_index: int) -> list[dict]:
    """Read pdfplumber chars on the given page and return word-level groups
    that have a non-default non_stroking_color (blue or red).

    Output dict shape: {text, x0, y0, x1, y1, color, family}.
    """
    out: list[dict] = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            if page_index >= len(pdf.pages):
                return out
            page = pdf.pages[page_index]
            chars = page.chars or []
    except Exception:
        return out

    def _is_target_color(c) -> bool:
        col = c.get("non_stroking_color")
        if col is None:
            return False
        # pdfplumber returns either a tuple of length 1 (gray), 3 (RGB)
        # or 4 (CMYK).  We only care about RGB blue/red.
        if not isinstance(col, (tuple, list)) or len(col) != 3:
            return False
        cb = tuple(round(float(v), 4) for v in col)
        for target in (_REVERSE_BLUE_RGB, _REVERSE_RED_RGB):
            if all(abs(cb[i] - target[i]) <= _REVERSE_COLOR_TOL for i in range(3)):
                return True
        return False

    colored = [c for c in chars if _is_target_color(c)]
    if not colored:
        return out

    # Group chars to words: y-snap line clustering then x-gap word splitting
    from collections import defaultdict
    y_snap = 1.5
    gap_factor = 0.55
    lines: dict[float, list[dict]] = defaultdict(list)
    for c in colored:
        ykey = round(float(c.get("y0", 0)) / y_snap) * y_snap
        lines[ykey].append(c)

    for _, line_chars in lines.items():
        line_chars.sort(key=lambda x: float(x.get("x0", 0)))
        cur: list[dict] = []

        def _flush(buf: list[dict]):
            if not buf:
                return
            text = "".join(str(cc.get("text", "")) for cc in buf)
            bx0 = min(float(cc.get("x0", 0)) for cc in buf)
            bx1 = max(float(cc.get("x1", 0)) for cc in buf)
            by0 = min(float(cc.get("y0", 0)) for cc in buf)
            by1 = max(float(cc.get("y1", 0)) for cc in buf)
            col = buf[0].get("non_stroking_color")
            out.append({
                "text": text,
                "x0": bx0, "y0": by0, "x1": bx1, "y1": by1,
                "color": tuple(round(float(v), 4) for v in col) if col else None,
                "family": _reverse_label_family(text),
            })

        for c in line_chars:
            if cur:
                last_x1 = float(cur[-1].get("x1", 0))
                size = float(c.get("size", 6) or 6)
                gap_thresh = size * gap_factor
                if (float(c.get("x0", 0)) - last_x1) > gap_thresh:
                    _flush(cur)
                    cur = []
            cur.append(c)
        _flush(cur)

    return out


def _reverse_match_labels_to_legend(
    words: list[dict],
    legend_items,
    covered_idx: set[int],
) -> dict[int, list[dict]]:
    """For each colored word that classifies as an equipment label,
    find the best uncovered legend index whose description contains the
    label-family's keyword set.  Returns idx -> list of label dicts.
    """
    result: dict[int, list[dict]] = {}
    if not words or not legend_items:
        return result

    # Try rapidfuzz, fall back to difflib if not installed
    try:
        from rapidfuzz import fuzz as _fuzz
        def _score(a: str, b: str) -> float:
            return _fuzz.token_set_ratio(a, b)
    except Exception:
        import difflib
        def _score(a: str, b: str) -> float:
            return difflib.SequenceMatcher(None, a, b).ratio() * 100.0

    for word in words:
        text = (word.get("text") or "").strip()
        family = word.get("family")
        if not text or not family:
            continue
        # Tight regex guard: only match exact equip-label forms,
        # excluding e.g. legend footnote 'PUE p.1.1.29' substrings.
        if not _REVERSE_EQUIP_LABEL_RE.match(text):
            continue
        keywords = _REVERSE_LABEL_KEYWORDS.get(family, ())
        if not keywords:
            continue

        # Primary gate: ALL keywords for this family must appear in the
        # legend description (Cyrillic substring check).  When that holds
        # we accept the smallest matching idx; the fuzzy score below is
        # used only to disambiguate when MULTIPLE uncovered legend rows
        # satisfy the keyword test (e.g. several variants of "switch"
        # rows for a VKL\d+ label).
        candidates: list[tuple[int, float]] = []
        for idx, item in enumerate(legend_items):
            if idx in covered_idx:
                continue
            desc = (item.description or "").strip()
            if not desc:
                continue
            desc_lower = desc.lower()
            if not all(kw in desc_lower for kw in keywords):
                continue
            # Secondary fuzzy score on the legend description against
            # itself + keyword block (deterministic tie-breaker).
            kw_blob = " ".join(keywords)
            score = _score(kw_blob, desc_lower)
            candidates.append((idx, score))

        if candidates:
            # Pick the highest-scoring candidate; ties broken by lowest idx.
            candidates.sort(key=lambda t: (-t[1], t[0]))
            best_idx, best_score = candidates[0]
            result.setdefault(best_idx, []).append({
                "text": text,
                "x0": word.get("x0"),
                "y0": word.get("y0"),
                "x1": word.get("x1"),
                "y1": word.get("y1"),
                "family": family,
                "score": round(best_score, 1),
            })

    return result


def _count_equipment_in_pdf(pdf_path: str) -> list[dict]:
    """Run legend extraction + counting methods on a single PDF.

    Supports ЭО, ЭМ, and ЭГ section types:
      - ЭО: full legend + text/visual counting + cables
      - ЭМ: legend (panels/equipment) + cables (heavy cable schedule)
      - ЭГ: cables (grounding conductors) + geometric measurement
      - unknown: same as ЭО (generic)

    Steps:
      1. Detect section type from filename/folder
      2. Parse legend via pdf_legend_parser
      3. Run count_symbols (text) + match_symbols (visual fallback)
      4. Build items from legend + counts
      5. Extract cables (all section types)
      6. For ЭГ: also run measure_cables for geometric lengths
      7. Apply VOR work-name mapping
    """
    import logging
    log = logging.getLogger("web_app._count_equipment_in_pdf")

    section = _detect_section_type(pdf_path)
    log.info("Processing %s (section=%s)", pdf_path, section)

    # Step 1: parse legend
    legend_result = parse_legend(pdf_path)
    has_legend = bool(legend_result.items)

    items: list[dict] = []
    # T054 / KB-015: track which legend indices reached output so we can
    # emit an [UNMATCHED-LEGEND] warning row for any item that no
    # producer stage covered.  Without this audit, symbol-less legend
    # rows (switches, posts, cable trasses) are silently dropped — the
    # exact failure mode reported as B048 on 007-Plans osvescheniya PDF.
    covered_legend_idx: set[int] = set()

    # Step 2-4: legend-based equipment counting (skip if no legend found)
    if has_legend:
        # Step 2: run VISUAL counting first (primary method)
        visual_counts: dict[int, int] = {}  # symbol_index -> count
        disambiguated_indices: set[int] = set()
        try:
            vis_result = match_symbols(pdf_path, legend_result)
            visual_counts = vis_result.counts  # symbol_index -> count
            disambiguated_indices = vis_result.disambiguated_indices
        except Exception as exc:
            log.warning("Visual counting failed for %s: %s", pdf_path, exc)

        # Step 2b: build equipment cluster zones for text filtering
        equip_zones: dict[str, list] | None = None
        try:
            equip_zones = build_equipment_cluster_bboxes(
                pdf_path, legend_result.page_index
            )
        except Exception as exc:
            log.warning("Equipment zone detection failed: %s", exc)

        # Step 3: run text counting as secondary/fallback
        # Pass equipment zones so standalone digit markers are only
        # accepted near coloured equipment clusters.
        text_counts: dict[str, int] = {}
        try:
            text_result = count_symbols(
                pdf_path, legend_result, equipment_zones=equip_zones,
            )
            text_counts = text_result.counts  # symbol -> count
        except Exception as exc:
            log.warning("Text counting failed for %s: %s", pdf_path, exc)

        # Step 4: build enriched items from legend + counts
        for idx, item in enumerate(legend_result.items):
            sym = item.symbol or ""
            name = item.description or ""
            if not name:
                continue

            # Determine count: smart priority between visual and text.
            #
            # Compound markers (containing Cyrillic, e.g. '7АЭ', '8АЭ'):
            # prefer visual count when available so pictogram/template
            # detections participate in final totals; fall back to text if
            # visual has no hits.
            #
            # For standalone digit symbols: visual is preferred, but if
            # suspiciously high vs text (>3×), fall back to text when
            # txt >= 5 (T147 guard against unreliable small text counts).
            #
            # When txt > vis and both > 0: visual may be undercounting
            # due to similar-template cross-symbol NMS suppression;
            # take max(vis, txt) as a better estimate.  (T148)
            vis_count = visual_counts.get(idx, 0)
            txt_count = text_counts.get(sym, 0) if sym else 0
            is_compound = bool(sym and re_mod.search(r'[А-Яа-яЁё]', sym))
            is_simple_compound = bool(sym and re_mod.fullmatch(r'\d{1,2}[А-Яа-яЁё]', sym))
            is_disambiguated = idx in disambiguated_indices
            count = 0
            if is_compound and vis_count > 0:
                # For simple one-letter variants (e.g. 1А), visual matching can
                # flood with false positives on dense sheets. If visual is much
                # higher than text, prefer text as a safer estimate.
                if is_simple_compound and txt_count > 0 and vis_count > txt_count * 3:
                    count = txt_count
                else:
                    # For more specific compound markers (e.g. 7АЭ), keep visual
                    # as primary when it exists.
                    count = vis_count
            elif is_compound and txt_count > 0:
                count = txt_count
            elif is_disambiguated and vis_count > 0:
                # Symbol was resolved by text-aided disambiguation (S021).
                # Each visual match has already been verified against nearby
                # PDF text — trust visual count directly.
                count = vis_count
            elif (vis_count > 0 and txt_count >= 3
                    and vis_count > txt_count * 3):
                count = txt_count  # visual likely has false positives
            elif vis_count > 0 and txt_count > vis_count:
                # Text found more than visual — visual may be under-
                # counting due to cross-symbol NMS suppression between
                # similar templates.  Trust text as the better estimate.
                count = txt_count
            elif vis_count > 0:
                # Guard: if visual is very high but text found nothing,
                # visual likely has massive false positives from a generic
                # template matching background noise (S021).
                if txt_count == 0 and vis_count > 20 and not is_compound:
                    log.debug(
                        "skip sym=%s vis=%d txt=0 — likely FP flood",
                        sym, vis_count,
                    )
                    continue
                count = vis_count
            elif txt_count > 0:
                count = txt_count

            if count <= 0:
                continue

            items.append({
                "symbol": sym,
                "name": name,
                "count": count,
                "count_ae": 0,
                "total": count,
            })
            covered_legend_idx.add(idx)

    # Step 4b: detect pictograms — text labels like "ВЫХОД" on the drawing
    # that have no legend entry and no visual template (T149).
    try:
        picto_result = detect_pictograms(pdf_path, legend_result)
        for name, count in picto_result.counts.items():
            if count > 0:
                # Use "ВЫХОД" symbol to match DXF ground truth naming.
                items.append({
                    "symbol": "ВЫХОД",
                    "name": name,
                    "count": count,
                    "count_ae": 0,
                    "total": count,
                })
    except Exception as exc:
        log.warning("Pictogram detection failed for %s: %s", pdf_path, exc)

    # Step 5: extract cables and add to items (ALL section types)
    try:
        cable_result = extract_cables(pdf_path, legend_result)
        for entry in cable_result.cable_schedule:
            group = entry.get("group", "")
            panel = entry.get("panel", "")
            cable_types = entry.get("cable_types", [])
            cross_sections = entry.get("cross_sections", [])
            run_count = entry.get("run_count", 0) or 0
            total_length_m = entry.get("total_length_m", 0) or 0

            # Use cable_type if available, otherwise cross_section
            type_label = (cable_types[0] if cable_types
                          else cross_sections[0] if cross_sections
                          else "")
            if not type_label and not group:
                continue

            if run_count > 0:
                cable_name = f"Кабель {type_label}" if type_label else "Кабель"
                if group:
                    cable_name += f" ({panel}-{group})" if panel else f" ({group})"
                items.append({
                    "symbol": "",
                    "name": cable_name,
                    "count": run_count,
                    "count_ae": 0,
                    "total": run_count,
                    "unit": "шт",
                })
            if total_length_m > 0:
                cable_name_m = (f"Кабель {type_label} (прокладка)"
                                if type_label else "Кабель (прокладка)")
                if group:
                    cable_name_m += f" ({panel}-{group})" if panel else f" ({group})"
                items.append({
                    "symbol": "",
                    "name": cable_name_m,
                    "count": 0,
                    "count_ae": 0,
                    "total": round(total_length_m, 1),
                    "unit": "м",
                })
    except Exception as exc:
        log.warning("Cable extraction failed for %s: %s", pdf_path, exc)

    # Step 6: for ЭГ section — use geometric measurement for cable lengths
    # ЭГ drawings often have grounding conductors measured by line geometry
    # rather than cable schedule tables
    if section == "ЭГ":
        try:
            geo_result = measure_cables(pdf_path, legend_result)
            # Add red-line measurements (typically grounding conductor runs)
            if geo_result.total_red_length_m > 0:
                items.append({
                    "symbol": "",
                    "name": "Проводник заземления (горизонтальный)",
                    "count": 0,
                    "count_ae": 0,
                    "total": round(geo_result.total_red_length_m, 1),
                    "unit": "м",
                })
            # Add blue-line measurements (typically equipotential bonding)
            if geo_result.total_blue_length_m > 0:
                items.append({
                    "symbol": "",
                    "name": "Проводник уравнивания потенциалов",
                    "count": 0,
                    "count_ae": 0,
                    "total": round(geo_result.total_blue_length_m, 1),
                    "unit": "м",
                })
            # Add individual route details by linewidth if available
            for lw_key, lw_info in geo_result.by_linewidth.items():
                length_m = lw_info.get("length_m", 0)
                segments = lw_info.get("segments", 0)
                if length_m > 0 and segments > 0:
                    log.info("ЭГ geometry: linewidth=%s  length=%.1f m  segments=%d",
                             lw_key, length_m, segments)
        except Exception as exc:
            log.warning("Geometric measurement failed for %s: %s", pdf_path, exc)

    # Step 6a (T058 / B051): reverse label -> legend channel.
    # Extract colored on-drawing labels (blue PU<n>, VKL, POST etc.) via
    # pdfplumber chars and match each label family to the descriptions of
    # legend rows that are still UNCOVERED after the earlier stages.
    # Marks the matched legend idx as covered so the [UNMATCHED-LEGEND]
    # audit below does NOT also flag them.  See T057 recon for the
    # color/encoding evidence on 007-Plans osvescheniya PDF.
    if has_legend:
        try:
            colored_words = _reverse_extract_colored_words(
                pdf_path, legend_result.page_index,
            )
            label_groups = _reverse_match_labels_to_legend(
                colored_words, legend_result.items, covered_legend_idx,
            )
            for idx, labels in label_groups.items():
                item = legend_result.items[idx]
                qty = len(labels)
                items.append({
                    "symbol": (item.symbol or ""),
                    "name": item.description or "",
                    "count": qty,
                    "count_ae": 0,
                    "total": qty,
                    "unit": "шт",
                    "category": "reverse_label_match",
                    "source": "reverse_label_match",
                })
                covered_legend_idx.add(idx)
                log.info(
                    "reverse_label_match idx=%d desc=%r qty=%d family=%s",
                    idx, (item.description or "")[:40], qty,
                    labels[0].get("family", ""),
                )
        except Exception as exc:
            log.warning(
                "Reverse label channel failed for %s: %s", pdf_path, exc,
            )

    # Step 6b (T068 / S016-cable-length): raster polyline detection and
    # length engine.  The vector pipeline (Step 6 measure_cables, only
    # ran on section 'ЭГ' above) misses cable trasses rendered as
    # embedded raster fills, which is the case on most GPK3 lighting
    # sheets and was the single biggest accuracy gap (519 m generated
    # vs ~14 000 m reference on 3-zahvatka).
    #
    # The raster engine renders each plan page at 300 DPI, builds an
    # HSV mask per cable-trace legend category (idx 10..14 on 007-style
    # legends: emergency / working cable trasses + 3 provodka classes),
    # skeletonises and sums pixel runs, then divides by the per-page
    # px-per-metre scale derived from titleblock axis pairs.  See
    # cable_length.measure_cable_lengths_raster.
    #
    # Items emitted carry source='cable_length_raster' so downstream
    # VOR mapping can attribute them to the correct height bucket via
    # T065/T060 rules; we deliberately do NOT mark covered_legend_idx
    # for these because the UNMATCHED audit below should still flag
    # the symbol-less legend rows that lost their visual template match
    # under KB-015.
    if has_legend:
        try:
            cl_pages = [legend_result.page_index] if legend_result is not None else None
            cl_rep = measure_cable_lengths_raster(
                pdf_path, pages=cl_pages, legend_result=legend_result,
            )
            for cl_item in cl_rep.items:
                items.append({
                    "symbol": cl_item.get("symbol", ""),
                    "name": cl_item.get("name", ""),
                    "count": 0,
                    "count_ae": 0,
                    "total": cl_item.get("total", 0.0),
                    "unit": cl_item.get("unit", "\u043c"),
                    "category": "cable_length_raster",
                    "source": "cable_length_raster",
                })
            if cl_rep.total_length_m > 0:
                log.info(
                    "cable_length_raster: %.1f m across %d entries on %s",
                    cl_rep.total_length_m, len(cl_rep.entries), pdf_path,
                )
            # When the raster engine matched a legend idx for a category
            # that produced a non-trivial length (>=0.5 m), mark that
            # idx as covered so the UNMATCHED audit does not double-emit
            # the same row as a [UNMATCHED-LEGEND] warning.
            for e in cl_rep.entries:
                if e.legend_idx >= 0 and e.length_m >= 0.5:
                    covered_legend_idx.add(e.legend_idx)
        except Exception as exc:
            log.warning(
                "Raster cable-length engine failed for %s: %s",
                pdf_path, exc,
            )

    # Step 6c (T054 / KB-015): emit warning rows for legend items that
    # no producer stage covered.  The [UNMATCHED-LEGEND] name prefix
    # surfaces the gap to the user so symbol-less switches, posts, and
    # cable trasses (B048 reproducer on 007-Plans osvescheniya) appear
    # in the VOR even when count_text + match_visual found nothing.
    #
    # Quantity policy:
    #   * line patterns (cable trasse / provodka) — qty=0 because count
    #     is meaningless for these (length is the real metric, deferred
    #     to follow-up B049 cable polyline detection).
    #   * everything else (switches, posts, control devices) — qty=1
    #     so the VOR row at least registers that ONE legend mention
    #     existed; the user can refine count manually.
    #
    # Conservative guard: skip the audit on legends where ZERO producer
    # stages matched anything, because such "legends" are usually
    # title-block or notes-table false positives and would otherwise
    # flood the output with spurious rows.
    if has_legend and covered_legend_idx:
        _line_categories = {
            "кабельная трасса", "проводка", "линия связи",
            "кабельная", "трасса", "wire", "cable", "trasse",
        }
        for idx, item in enumerate(legend_result.items):
            if idx in covered_legend_idx:
                continue
            desc = (item.description or "").strip()
            if not desc:
                continue
            cat = (item.category or "").strip().lower()
            desc_lower = desc.lower()
            is_line = (
                cat in _line_categories
                or "трасс" in desc_lower
                or "прокладыв" in desc_lower
                or "провод" in desc_lower
                or "кабельн" in desc_lower
            )
            warn_count = 0 if is_line else 1
            items.append({
                "symbol": (item.symbol or ""),
                "name": f"[UNMATCHED-LEGEND] {desc}",
                "count": warn_count,
                "count_ae": 0,
                "total": warn_count,
                "unit": "шт",
                "category": "legend_unmatched",
                "source": "legend_coverage_audit",
            })

    # Step 7: apply VOR work-name mapping
    items = vor_map_items(items)

    # Step 8 (T069 / S016-height-bucket): tag every item with its
    # height_bucket key derived from the PDF filename's "\u043e\u0442\u043c."
    # otmetka.  Per T065 recon Q1, the reference VOR groups every
    # installation row into one of 4 buckets ("\u0434\u043e 5 \u043c.",
    # "\u043e\u0442 5 \u0434\u043e 13 \u043c.", "\u043e\u0442 13 \u0434\u043e 20 \u043c.",
    # "\u043e\u0442 20 \u0434\u043e 35 \u043c.") by the floor elevation
    # encoded in the PDF title block.  When the filename carries no
    # otmetka marker (e.g. 001 general-data sheets, 003/004 panel
    # schematics) the bucket falls back to "unknown" so downstream
    # vor_composer can still group those rows separately.
    try:
        height_bucketer.attribute_items(items, pdf_path)
    except Exception as exc:
        log.warning("height-bucket attribution failed for %s: %s", pdf_path, exc)

    # Step 9 (T070 / S016-route-classify): tag every cable-trace item
    # with route in {tray, pipe_hidden, pipe_open, unknown} and every
    # luminaire item with mount in {wall, shpilka, anker, unknown}.
    # Decision rules live in route_classifier; the call is idempotent
    # so pre-tagged items (from a future per-symbol pipeline) are
    # preserved.
    try:
        route_classifier.attribute_items(items)
    except Exception as exc:
        log.warning("route-classify attribution failed for %s: %s", pdf_path, exc)

    # Step 10 (T072 / S016-thickness): extract dimensional metadata --
    # cable cross_section (e.g. "3\u04451,5"), gofra diameter_mm, lotok
    # width_mm x height_mm.  Reuses the cross-section regex hardened
    # against KB-007 (Cyrillic \u0445 / Latin x / Unicode MULT \u00d7).
    # Idempotent over reruns; pre-set fields are preserved.
    try:
        thickness_extractor.attribute_items(items)
    except Exception as exc:
        log.warning("thickness attribution failed for %s: %s", pdf_path, exc)

    return items


# ---------------------------------------------------------------------------
# 6c. Convert pipeline VorSection objects → flat aggregated list for UI
# ---------------------------------------------------------------------------

def _sections_to_aggregated(sections) -> list[dict]:
    """Convert list of VorSection into flat list of dicts for the UI table."""
    result = []
    row_num = 1
    for section in sections:
        # Section header row
        result.append({
            "row": row_num,
            "name": section.title,
            "unit": "",
            "total": "",
            "formula": "",
            "drawing_refs": "",
            "extra_info": "",
            "is_section_header": True,
        })
        row_num += 1
        for row_data in section.rows:
            result.append({
                "row": row_num,
                "name": row_data["name"],
                "unit": row_data["unit"],
                "total": row_data["qty"],
                "formula": str(row_data["qty"]) if row_data.get("qty", 0) > 0 else "",
                "drawing_refs": row_data.get("drawing_ref", ""),
                "extra_info": "",
                "is_section_header": False,
                "is_material": row_data.get("is_material", False),
            })
            row_num += 1
    return result


# ---------------------------------------------------------------------------
# 6c. GET /api/folder/{folder_id}/process — SSE batch processing
# ---------------------------------------------------------------------------

@app.get("/api/folder/{folder_id}/process")
async def api_folder_process(folder_id: str):
    """Process all PDFs in a folder via SSE stream with progress.

    Uses the full VOR pipeline (pdf_vor_pipeline.py) which:
    - Parses specs (СО), plans, schemas, binding plans
    - Builds height distribution ratios
    - Generates proper VOR sections with materials
    """
    from pdf_vor_pipeline import generate_vor_from_pdfs, export_vor_xlsx, generate_vor_aggregated

    rel_folder = _id_to_folder(folder_id)
    if rel_folder is None:
        raise HTTPException(status_code=404, detail="Folder not found")
    pdf_files = _folder_files(rel_folder)
    if not pdf_files:
        raise HTTPException(status_code=404, detail="No PDFs in folder")

    if rel_folder.startswith("_uploads/"):
        folder_path = UPLOADS_DIR / rel_folder[len("_uploads/"):]
    elif rel_folder == "_uploads":
        folder_path = UPLOADS_DIR
    else:
        folder_path = DATA_DIR / rel_folder

    async def event_stream():
        total = len(pdf_files)
        yield f"event: start\ndata: {json_mod.dumps({'total': total, 'folder': rel_folder}, ensure_ascii=False)}\n\n"

        # Collect log messages from pipeline to stream as SSE events
        log_messages: list[str] = []
        step_count = 0

        def pipeline_log(msg: str):
            nonlocal step_count
            log_messages.append(msg)
            step_count += 1

        # Run full VOR pipeline in a thread (CPU-bound) — single pass
        try:
            sections = await asyncio.to_thread(
                generate_vor_from_pdfs, str(folder_path), pipeline_log
            )
        except Exception as e:
            yield f"event: file_done\ndata: {json_mod.dumps({'filename': 'pipeline', 'items': 0, 'status': 'error', 'error': str(e), 'current': 1, 'total': 1}, ensure_ascii=False)}\n\n"
            yield f"event: done\ndata: {json_mod.dumps({'aggregated': [], 'files_processed': 0, 'errors': [{'filename': 'pipeline', 'error': str(e)}], 'total_files': total}, ensure_ascii=False)}\n\n"
            return

        # Send log messages as progress events
        for i, msg in enumerate(log_messages):
            if msg.strip():
                yield f"event: file_done\ndata: {json_mod.dumps({'filename': msg.strip()[:120], 'items': 0, 'status': 'ok', 'current': i + 1, 'total': len(log_messages)}, ensure_ascii=False)}\n\n"

        # Build aggregated list from sections (for UI table)
        aggregated = _sections_to_aggregated(sections)

        # Save Excel to disk using pipeline's own export (proper formatting)
        saved_filename = ""
        try:
            dest_dir = _vor_folder_for(rel_folder)
            ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            xlsx_path = dest_dir / f"VOR_{ts}.xlsx"
            export_vor_xlsx(sections, str(xlsx_path))
            saved_filename = xlsx_path.name
        except Exception:
            # Fallback: build simple xlsx
            try:
                xlsx_bytes = _build_vor_xlsx(aggregated, rel_folder)
                saved_path = _save_vor_xlsx(xlsx_bytes, rel_folder)
                saved_filename = saved_path.name
            except Exception:
                pass

        yield f"event: done\ndata: {json_mod.dumps({'aggregated': aggregated, 'files_processed': total, 'errors': [], 'total_files': total, 'xlsx_file': saved_filename}, ensure_ascii=False)}\n\n"

    return StreamingResponse(event_stream(), media_type="text/event-stream",
                            headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


# ---------------------------------------------------------------------------
# 6d. Helper: build VOR xlsx bytes from aggregated data
# ---------------------------------------------------------------------------

def _build_vor_xlsx(aggregated: list[dict], rel_folder: str) -> bytes:
    """Build VOR Excel workbook and return raw bytes."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ВОР"

    header_font = Font(bold=True, size=10)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    thin_border = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

    headers = ["№ п/п", "Наименование вида работ", "Ед. изм.", "Объем работ",
               "Формула расчета", "Ссылка на чертежи", "Доп. информация"]
    col_widths = [7, 72, 9, 12, 20, 26, 27]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    for row_data in aggregated:
        ri = row_data["row"] + 1
        vals = [row_data["row"], row_data["name"], row_data["unit"],
                row_data["total"], row_data["formula"], row_data["drawing_refs"],
                row_data.get("extra_info", "")]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = thin_border
            if ci in (1, 3, 4):
                cell.alignment = Alignment(horizontal="center")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# 6d. VOR file management: list saved files + download
# ---------------------------------------------------------------------------

@app.get("/api/folder/{folder_id}/vor_files")
async def api_folder_vor_files(folder_id: str):
    """List all saved VOR xlsx files for a folder."""
    rel_folder = _id_to_folder(folder_id)
    if rel_folder is None:
        raise HTTPException(status_code=404, detail="Folder not found")
    return JSONResponse({"files": _list_vor_files(rel_folder)})


@app.get("/api/folder/{folder_id}/vor_download/{filename}")
async def api_folder_vor_download(folder_id: str, filename: str):
    """Download a specific saved VOR xlsx file from disk."""
    from urllib.parse import quote

    rel_folder = _id_to_folder(folder_id)
    if rel_folder is None:
        raise HTTPException(status_code=404, detail="Folder not found")

    # Sanitize filename — only allow VOR_*.xlsx pattern
    if not filename.startswith("VOR_") or not filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Invalid filename")

    file_path = _vor_folder_for(rel_folder) / filename
    if not file_path.is_file():
        raise HTTPException(status_code=404, detail="Файл не найден")

    xlsx_bytes = file_path.read_bytes()
    encoded_name = quote(filename)

    return Response(content=xlsx_bytes,
                   media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                   headers={"Content-Disposition": f"attachment; filename=\"{filename}\"; filename*=UTF-8''{encoded_name}"})


# ---------------------------------------------------------------------------
# Helpers: color analysis
# ---------------------------------------------------------------------------

# Known color labels for electrical drawings
_KNOWN_COLORS: dict[tuple, str] = {
    (0, 0, 0): "Чёрный / Рамка",
    (1, 0, 0): "Красный / Аварийное",
    (0, 0, 1): "Синий / Рабочее",
    (0, 1, 0): "Зелёный",
    (1, 1, 0): "Жёлтый",
    (1, 0, 1): "Пурпурный",
    (0, 1, 1): "Голубой",
}


def _normalize_color(c) -> tuple | None:
    """Normalize a pdfplumber color value to a tuple of floats, or None."""
    if c is None:
        return None
    if isinstance(c, (int, float)):
        # Grayscale
        v = float(c)
        return (v, v, v)
    if isinstance(c, (tuple, list)):
        if len(c) == 3:
            return tuple(round(float(x), 4) for x in c)
        if len(c) == 4:
            # CMYK → RGB approximation
            cc, m, y, k = [float(x) for x in c]
            r = (1 - cc) * (1 - k)
            g = (1 - m) * (1 - k)
            b = (1 - y) * (1 - k)
            return (round(r, 4), round(g, 4), round(b, 4))
        if len(c) == 1:
            v = float(c[0])
            return (v, v, v)
    return None


def _color_to_hex(rgb: tuple) -> str:
    """Convert (r,g,b) floats 0-1 to hex string like 'FF0000'."""
    r = max(0, min(255, int(round(rgb[0] * 255))))
    g = max(0, min(255, int(round(rgb[1] * 255))))
    b = max(0, min(255, int(round(rgb[2] * 255))))
    return f"{r:02X}{g:02X}{b:02X}"


def _label_color(rgb: tuple) -> str:
    """Auto-label a color if it matches known colors (with tolerance)."""
    # Exact match first
    rounded = tuple(round(x, 2) for x in rgb)
    for known, label in _KNOWN_COLORS.items():
        if all(abs(a - b) < 0.05 for a, b in zip(rounded, known)):
            return label
    # Gray detection
    if abs(rgb[0] - rgb[1]) < 0.05 and abs(rgb[1] - rgb[2]) < 0.05:
        level = rgb[0]
        if level < 0.15:
            return "Чёрный / Рамка"
        if level > 0.85:
            return "Белый / Фон"
        return f"Серый ({int(level * 100)}%)"
    return ""


# ---------------------------------------------------------------------------
# 7. GET /api/file/{id}/colors — color palette analysis
# ---------------------------------------------------------------------------

@app.get("/api/file/{file_id}/colors")
async def api_colors(
    file_id: str,
    page: int = Query(0, ge=0, description="Page index (0-based)"),
):
    """Analyze PDF page color palette using pdfplumber."""
    pdf_path = _id_to_path(file_id)
    if pdf_path is None:
        raise HTTPException(status_code=404, detail="File not found")

    try:
        with pdfplumber.open(str(pdf_path)) as pdf:
            if page >= len(pdf.pages):
                raise HTTPException(
                    status_code=400,
                    detail=f"Page {page} out of range (0-{len(pdf.pages) - 1})",
                )

            p = pdf.pages[page]

            # Collect color stats: rgb_tuple -> {lines, rects, chars}
            color_stats: dict[tuple, dict] = {}

            def _inc(rgb, kind):
                if rgb is None:
                    return
                if rgb not in color_stats:
                    color_stats[rgb] = {"lines": 0, "rects": 0, "chars": 0}
                color_stats[rgb][kind] += 1

            # Lines
            for line in (p.lines or []):
                rgb = _normalize_color(line.get("stroking_color"))
                _inc(rgb, "lines")

            # Rects
            for rect in (p.rects or []):
                rgb_s = _normalize_color(rect.get("stroking_color"))
                rgb_f = _normalize_color(rect.get("non_stroking_color"))
                _inc(rgb_s, "rects")
                if rgb_f and rgb_f != rgb_s:
                    _inc(rgb_f, "rects")

            # Chars
            for ch in (p.chars or []):
                rgb = _normalize_color(ch.get("non_stroking_color"))
                _inc(rgb, "chars")

            # Build response
            colors = []
            for rgb, stats in sorted(
                color_stats.items(),
                key=lambda kv: kv[1]["lines"] + kv[1]["rects"] + kv[1]["chars"],
                reverse=True,
            ):
                total = stats["lines"] + stats["rects"] + stats["chars"]
                hex_val = _color_to_hex(rgb)
                label = _label_color(rgb)
                colors.append({
                    "rgb": list(rgb),
                    "hex": hex_val,
                    "label": label,
                    "lines": stats["lines"],
                    "rects": stats["rects"],
                    "chars": stats["chars"],
                    "total": total,
                })

            return JSONResponse(content={
                "page": page,
                "colors": colors,
            })
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Color analysis error: {e}")


# ---------------------------------------------------------------------------
# 8. GET /api/file/{id}/render_filtered — render with color filter
# ---------------------------------------------------------------------------

@app.get("/api/file/{file_id}/render_filtered")
async def api_render_filtered(
    file_id: str,
    page: int = Query(0, ge=0, description="Page index (0-based)"),
    dpi: int = Query(150, ge=72, le=600, description="Render DPI"),
    show: str = Query("", description="Comma-separated hex colors to show (e.g. FF0000,0000FF)"),
    hide: str = Query("", description="Comma-separated hex colors to hide"),
):
    """Render PDF page showing only elements matching specified colors.

    Uses PyMuPDF to render the full page, then composites with pdfplumber
    color data to show/hide elements by color.
    """
    pdf_path = _id_to_path(file_id)
    if pdf_path is None:
        raise HTTPException(status_code=404, detail="File not found")

    # Parse color filters
    show_set: set[str] = set()
    hide_set: set[str] = set()
    if show.strip():
        show_set = {c.strip().upper() for c in show.split(",") if c.strip()}
    if hide.strip():
        hide_set = {c.strip().upper() for c in hide.split(",") if c.strip()}

    if not show_set and not hide_set:
        # No filter — just render normally
        return await api_render(file_id, page, dpi)

    try:
        # Step 1: Get element data from pdfplumber
        with pdfplumber.open(str(pdf_path)) as pdf:
            if page >= len(pdf.pages):
                raise HTTPException(
                    status_code=400,
                    detail=f"Page {page} out of range (0-{len(pdf.pages) - 1})",
                )
            p = pdf.pages[page]
            pg_w, pg_h = float(p.width), float(p.height)

            # Collect elements grouped by visibility
            visible_lines = []
            visible_rects = []
            visible_chars = []

            def _is_visible(rgb_tuple) -> bool:
                if rgb_tuple is None:
                    return not bool(show_set)
                hex_val = _color_to_hex(rgb_tuple)
                if show_set:
                    return hex_val in show_set
                if hide_set:
                    return hex_val not in hide_set
                return True

            for line in (p.lines or []):
                rgb = _normalize_color(line.get("stroking_color"))
                if _is_visible(rgb):
                    visible_lines.append(line)

            for rect in (p.rects or []):
                rgb_s = _normalize_color(rect.get("stroking_color"))
                rgb_f = _normalize_color(rect.get("non_stroking_color"))
                if _is_visible(rgb_s) or _is_visible(rgb_f):
                    visible_rects.append(rect)

            for ch in (p.chars or []):
                rgb = _normalize_color(ch.get("non_stroking_color"))
                if _is_visible(rgb):
                    visible_chars.append(ch)

        # Step 2: Render using PyMuPDF shapes on a blank page
        doc = fitz.open(str(pdf_path))
        if page >= len(doc):
            doc.close()
            raise HTTPException(status_code=400, detail=f"Page out of range")

        src_page = doc[page]
        # Create a new blank document with same page size
        new_doc = fitz.open()
        new_page = new_doc.new_page(width=pg_w, height=pg_h)

        # Draw visible lines
        if visible_lines:
            shape = new_page.new_shape()
            for line in visible_lines:
                x0, y0 = float(line["x0"]), float(line["top"])
                x1, y1 = float(line["x1"]), float(line["bottom"])
                lw = float(line.get("linewidth", 1) or 1)
                rgb = _normalize_color(line.get("stroking_color"))
                color = rgb if rgb else (0, 0, 0)
                shape.draw_line(fitz.Point(x0, y0), fitz.Point(x1, y1))
                shape.finish(color=color, width=max(0.3, lw))
            shape.commit()

        # Draw visible rects
        if visible_rects:
            shape = new_page.new_shape()
            for rect in visible_rects:
                x0, y0 = float(rect["x0"]), float(rect["top"])
                x1, y1 = float(rect["x1"]), float(rect["bottom"])
                lw = float(rect.get("linewidth", 0.5) or 0.5)
                rgb_s = _normalize_color(rect.get("stroking_color"))
                rgb_f = _normalize_color(rect.get("non_stroking_color"))
                s_color = rgb_s if rgb_s else (0, 0, 0)
                f_color = rgb_f if rgb_f else None
                shape.draw_rect(fitz.Rect(x0, y0, x1, y1))
                shape.finish(
                    color=s_color,
                    fill=f_color,
                    width=max(0.1, lw),
                )
            shape.commit()

        # Draw visible chars (group by position for efficiency)
        for ch in visible_chars:
            rgb = _normalize_color(ch.get("non_stroking_color"))
            color = rgb if rgb else (0, 0, 0)
            x0 = float(ch["x0"])
            y0 = float(ch["top"])
            y1 = float(ch["bottom"])
            font_size = y1 - y0
            text = ch.get("text", "")
            if not text or not text.strip():
                continue
            try:
                new_page.insert_text(
                    fitz.Point(x0, y1 - font_size * 0.15),
                    text,
                    fontsize=max(1, font_size * 0.85),
                    color=color,
                )
            except Exception:
                pass

        zoom = dpi / 72.0
        mat = fitz.Matrix(zoom, zoom)
        pix = new_page.get_pixmap(matrix=mat, alpha=False)
        img_bytes = pix.tobytes("png")

        new_doc.close()
        doc.close()

        return Response(content=img_bytes, media_type="image/png")

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Filtered render error: {e}")


# ---------------------------------------------------------------------------
# 9. Find equipment positions (for interactive highlight)
# ---------------------------------------------------------------------------

@app.get("/api/file/{file_id}/find/{row_index}")
async def api_find_positions(file_id: str, row_index: int):
    """Find all instances of a legend item on the drawing.

    Returns positions from Method A (text markers) and optionally
    Method D (visual template matching) if available.
    """
    pdf_path = _id_to_path(file_id)
    if pdf_path is None:
        raise HTTPException(status_code=404, detail="File not found")

    try:
        legend = await asyncio.to_thread(_get_legend, pdf_path, file_id)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Legend parse error: {e}")

    if row_index < 0 or row_index >= len(legend.items):
        raise HTTPException(
            status_code=400,
            detail=f"Row index {row_index} out of range (0-{len(legend.items) - 1})",
        )

    item = legend.items[row_index]
    symbol = item.symbol or ""
    description = item.description or ""

    positions: list[dict] = []
    methods_used: list[str] = []
    excluded_zones: list[dict] = []

    # Method D: visual template matching — PRIMARY
    try:
        t0 = time.time()
        vis_result = await asyncio.to_thread(
            match_symbols, str(pdf_path), legend
        )
        elapsed_vis = round(time.time() - t0, 2)
        methods_used.append("visual")

        for m in vis_result.matches:
            if m.symbol_index == row_index:
                positions.append({
                    "x": round(m.x, 1),
                    "y": round(m.y, 1),
                    "width": 20,
                    "height": 20,
                    "confidence": round(m.confidence, 3),
                    "method": "visual",
                })
    except Exception:
        pass

    # Method A: text markers — SECONDARY (add non-duplicate positions)
    if symbol:
        try:
            # Build equipment zones for spatial filtering
            _eq_zones = None
            try:
                _eq_zones = await asyncio.to_thread(
                    build_equipment_cluster_bboxes,
                    str(pdf_path), legend.page_index,
                )
            except Exception:
                pass

            t0 = time.time()
            text_result = await asyncio.to_thread(
                count_symbols, str(pdf_path), legend, _eq_zones,
            )
            elapsed_text = round(time.time() - t0, 2)
            methods_used.append("text")

            for p in text_result.positions:
                if p.symbol == symbol:
                    # Skip if already found by visual (within 15pt)
                    duplicate = any(
                        abs(existing["x"] - p.x) < 15 and abs(existing["y"] - p.y) < 15
                        for existing in positions
                    )
                    if not duplicate:
                        positions.append({
                            "x": round(p.x, 1),
                            "y": round(p.y, 1),
                            "width": 12,
                            "height": 12,
                            "confidence": 1.0,
                            "method": "text",
                        })

            # Build exclusion zones
            excluded_zones = [
                {
                    "x0": round(z[1][0], 1), "y0": round(z[1][1], 1),
                    "x1": round(z[1][2], 1), "y1": round(z[1][3], 1),
                    "reason": z[0],
                }
                for z in text_result.exclusion_zones
            ]
        except Exception:
            pass

    needs_visual = len(positions) == 0

    return JSONResponse(content={
        "symbol": symbol,
        "description": description,
        "category": item.category or "",
        "row_index": row_index,
        "positions": positions,
        "count": len(positions),
        "excluded_zones": excluded_zones if symbol else [],
        "methods_used": methods_used,
        "needs_visual": needs_visual,
    })


# ---------------------------------------------------------------------------
# 9b. Cable highlight analysis (T107)
# ---------------------------------------------------------------------------

def _extract_cable_segments(pdf_path: str, legend_result, pages):
    """Extract cable line segments grouped by color/linewidth with connected routes.

    Returns a dict with segments, routes, annotations, and scale info.
    """
    from pdf_count_geometry import (
        measure_cables as geo_measure,
        _classify_line_color, _segment_length,
        _build_routes, _build_exclusion_zones, _detect_scale,
        MIN_SEGMENT_LENGTH_PT, ENDPOINT_TOLERANCE,
    )
    from pdf_count_cables import extract_cables as cable_extract

    legend_bbox = None
    legend_page = -1
    if legend_result and legend_result.items:
        legend_bbox = legend_result.legend_bbox
        legend_page = legend_result.page_index

    segments_out = []  # all cable line segments for overlay
    routes_out = []    # connected route polylines
    annotations_out = []  # cable run annotations (text-based)
    scale_info = None

    with pdfplumber.open(pdf_path) as pdf:
        scan_pages = pages if pages is not None else list(range(len(pdf.pages)))

        for page_idx in scan_pages:
            if page_idx >= len(pdf.pages):
                continue
            page = pdf.pages[page_idx]
            pdf_lines = page.lines or []
            words = page.extract_words(x_tolerance=3, y_tolerance=3) or []

            if not pdf_lines:
                continue

            # Detect scale
            page_scale = _detect_scale(page, words, pdf_lines)
            if scale_info is None or (
                scale_info.get("confidence") != "high"
                and page_scale.confidence == "high"
            ):
                scale_info = {
                    "mm_per_pt": round(page_scale.mm_per_pt, 4),
                    "source": page_scale.source,
                    "confidence": page_scale.confidence,
                }

            # Build exclusion zones
            lb = legend_bbox if page_idx == legend_page else None
            zones = _build_exclusion_zones(page, pdf_lines, lb)

            # Classify colored lines
            red_segs = []
            blue_segs = []

            for ln in pdf_lines:
                # Check exclusion
                mx = (ln["x0"] + ln["x1"]) / 2
                my = (ln["top"] + ln["bottom"]) / 2
                excluded = False
                for _, zb in zones:
                    if zb[0] <= mx <= zb[2] and zb[1] <= my <= zb[3]:
                        excluded = True
                        break
                if excluded:
                    continue

                seg_len = _segment_length(ln)
                if seg_len < MIN_SEGMENT_LENGTH_PT:
                    continue

                color = _classify_line_color(ln.get("stroking_color"))
                if color == "other":
                    continue

                lw = round(ln.get("linewidth", 0), 3)
                seg = {
                    "x0": round(ln["x0"], 1), "y0": round(ln["top"], 1),
                    "x1": round(ln["x1"], 1), "y1": round(ln["bottom"], 1),
                    "color": color, "lw": lw, "page": page_idx,
                }
                segments_out.append(seg)

                if color == "red":
                    red_segs.append(ln)
                else:
                    blue_segs.append(ln)

            # Build connected routes
            for color, segs in [("red", red_segs), ("blue", blue_segs)]:
                if not segs:
                    continue
                route_groups = _build_routes(segs, ENDPOINT_TOLERANCE)
                mm_per_pt = scale_info["mm_per_pt"] if scale_info else 35.0

                for ri, route_segs in enumerate(route_groups):
                    total_pt = sum(_segment_length(s) for s in route_segs)
                    total_m = total_pt * mm_per_pt / 1000.0

                    # Compute bounding box of route
                    all_x = []
                    all_y = []
                    for s in route_segs:
                        all_x.extend([s["x0"], s["x1"]])
                        all_y.extend([s["top"], s["bottom"]])

                    routes_out.append({
                        "id": len(routes_out),
                        "color": color,
                        "segment_count": len(route_segs),
                        "length_pt": round(total_pt, 1),
                        "length_m": round(total_m, 1),
                        "bbox": {
                            "x0": round(min(all_x), 1),
                            "y0": round(min(all_y), 1),
                            "x1": round(max(all_x), 1),
                            "y1": round(max(all_y), 1),
                        },
                        "page": page_idx,
                    })

    # Get cable annotations (text-based)
    try:
        cable_result = cable_extract(pdf_path, legend_result, pages)
        for r in cable_result.runs:
            annotations_out.append({
                "panel": r.panel, "group": r.group,
                "group_full": r.group_full,
                "cross_section": r.cross_section,
                "length_m": r.length_m,
                "cable_type": r.cable_type,
                "x": r.position[0], "y": r.position[1],
                "color": r.color,
                "page": r.page_index,
            })
    except Exception:
        pass

    return {
        "segments": segments_out,
        "routes": routes_out,
        "annotations": annotations_out,
        "scale": scale_info,
        "total_segments": len(segments_out),
        "total_routes": len(routes_out),
        "red_segments": sum(1 for s in segments_out if s["color"] == "red"),
        "blue_segments": sum(1 for s in segments_out if s["color"] == "blue"),
    }


@app.get("/api/file/{file_id}/cables")
async def api_cables(
    file_id: str,
    all_pages: bool = Query(False, description="Scan all pages"),
):
    """Cable highlight data: line segments, connected routes, annotations.

    Three filter modes supported by the frontend:
      1. By type: red=emergency, blue=working
      2. By group: ЩО1-Гр.7
      3. By route: connected polyline ID
    """
    pdf_path = _id_to_path(file_id)
    if pdf_path is None:
        raise HTTPException(status_code=404, detail="File not found")

    try:
        t0 = time.time()
        legend = await asyncio.to_thread(_get_legend, pdf_path, file_id)
        pages = None if all_pages else (
            [legend.page_index] if legend.items else None
        )
        data = await asyncio.to_thread(
            _extract_cable_segments, str(pdf_path), legend, pages
        )
        elapsed = round(time.time() - t0, 2)
        data["elapsed_s"] = elapsed
        return JSONResponse(content=data)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Cable analysis error: {e}")


# ---------------------------------------------------------------------------
# 10. Counting API endpoints
# ---------------------------------------------------------------------------

# Cache for legend results (keyed by file_id)
_legend_cache: dict[str, LegendResult] = {}


def _get_legend(pdf_path: Path, file_id: str) -> LegendResult:
    """Get cached or freshly parsed legend result."""
    if file_id not in _legend_cache:
        _legend_cache[file_id] = parse_legend(str(pdf_path))
    return _legend_cache[file_id]


# Cache for symbol images extracted from legend (keyed by file_id)
import numpy as np
import cv2

_symbol_image_cache: dict[str, list] = {}


def _get_symbol_images(pdf_path: Path, file_id: str, legend_result: LegendResult) -> list:
    """Get cached or freshly extracted symbol images from legend."""
    if file_id not in _symbol_image_cache:
        _symbol_image_cache[file_id] = _extract_symbol_images(
            str(pdf_path), legend_result
        )
    return _symbol_image_cache[file_id]


def _make_symbol_png_transparent(img: np.ndarray) -> bytes:
    """Convert BGR symbol image to PNG with transparent background."""
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    # White/near-white pixels become transparent
    _, mask = cv2.threshold(gray, 235, 255, cv2.THRESH_BINARY)
    # Convert BGR to BGRA (add alpha channel)
    bgra = cv2.cvtColor(img, cv2.COLOR_BGR2BGRA)
    # Set alpha: white pixels -> 0 (transparent), rest -> 255 (opaque)
    bgra[:, :, 3] = 255 - mask
    _, png_buf = cv2.imencode(".png", bgra)
    return png_buf.tobytes()


@app.get("/api/file/{file_id}/symbol_image/{row_index}")
async def api_symbol_image(file_id: str, row_index: int):
    """Render a single legend symbol cell as PNG with transparent background."""
    pdf_path = _id_to_path(file_id)
    if pdf_path is None:
        raise HTTPException(status_code=404, detail="File not found")

    legend = await asyncio.to_thread(_get_legend, pdf_path, file_id)
    images = await asyncio.to_thread(_get_symbol_images, pdf_path, file_id, legend)

    # Find image for the requested row index
    for idx, _item, img in images:
        if idx == row_index and img is not None:
            png_bytes = _make_symbol_png_transparent(img)
            return Response(
                content=png_bytes,
                media_type="image/png",
                headers={"Cache-Control": "max-age=3600"},
            )

    # No image for this index — return 204
    return Response(status_code=204)


@app.get("/api/file/{file_id}/count/text")
async def api_count_text(file_id: str):
    """Run Method A: count text markers on drawing."""
    pdf_path = _id_to_path(file_id)
    if pdf_path is None:
        raise HTTPException(status_code=404, detail="File not found")

    try:
        t0 = time.time()
        legend = await asyncio.to_thread(_get_legend, pdf_path, file_id)
        result = await asyncio.to_thread(count_symbols, str(pdf_path), legend)
        elapsed = round(time.time() - t0, 2)

        positions_by_sym: dict[str, list[dict]] = {}
        for p in result.positions:
            if p.symbol not in positions_by_sym:
                positions_by_sym[p.symbol] = []
            positions_by_sym[p.symbol].append({
                "x": round(p.x, 1), "y": round(p.y, 1),
                "merged": p.merged,
            })

        return JSONResponse(content={
            "method": "text",
            "page_index": result.page_index,
            "counts": result.counts,
            "positions": positions_by_sym,
            "total_found": sum(result.counts.values()),
            "symbols_searched": result.symbols_searched,
            "elapsed_s": elapsed,
        })
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Text count error: {e}")


@app.get("/api/file/{file_id}/count/cables")
async def api_count_cables(
    file_id: str,
    all_pages: bool = Query(False, description="Scan all pages"),
):
    """Run Method B: extract cable annotations."""
    pdf_path = _id_to_path(file_id)
    if pdf_path is None:
        raise HTTPException(status_code=404, detail="File not found")

    try:
        t0 = time.time()
        legend = await asyncio.to_thread(_get_legend, pdf_path, file_id)
        pages = None if all_pages else (
            [legend.page_index] if legend.items else None
        )
        result = await asyncio.to_thread(
            extract_cables, str(pdf_path), legend, pages
        )
        elapsed = round(time.time() - t0, 2)

        runs_json = []
        for r in result.runs:
            runs_json.append({
                "panel": r.panel, "group": r.group,
                "group_full": r.group_full,
                "cross_section": r.cross_section,
                "length_m": r.length_m, "cable_type": r.cable_type,
                "position": {"x": r.position[0], "y": r.position[1]},
                "color": r.color, "page_index": r.page_index,
                "is_reversed": r.is_reversed,
            })

        schedule_json = []
        for entry in result.cable_schedule:
            schedule_json.append({
                "group": entry["group"],
                "panel": entry["panel"],
                "cross_sections": entry["cross_sections"],
                "cable_types": entry["cable_types"],
                "run_count": entry["run_count"],
                "total_length_m": entry["total_length_m"],
                "colors": entry.get("colors", []),
            })

        return JSONResponse(content={
            "method": "cables",
            "total_runs": result.total_runs,
            "runs": runs_json,
            "panels": {k: len(v) for k, v in result.panels.items()},
            "cable_schedule": schedule_json,
            "pages_scanned": result.pages_scanned,
            "elapsed_s": elapsed,
        })
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Cable extraction error: {e}")


@app.get("/api/file/{file_id}/count/geometry")
async def api_count_geometry(
    file_id: str,
    all_pages: bool = Query(False, description="Scan all pages"),
):
    """Run Method C: measure cable routes by geometry."""
    pdf_path = _id_to_path(file_id)
    if pdf_path is None:
        raise HTTPException(status_code=404, detail="File not found")

    try:
        t0 = time.time()
        legend = await asyncio.to_thread(_get_legend, pdf_path, file_id)
        pages = None if all_pages else (
            [legend.page_index] if legend.items else None
        )
        result = await asyncio.to_thread(
            measure_cables, str(pdf_path), legend, pages
        )
        elapsed = round(time.time() - t0, 2)

        routes_json = []
        for r in result.routes:
            routes_json.append({
                "color": r.color, "linewidth": r.linewidth,
                "total_length_pt": round(r.total_length_pt, 1),
                "total_length_m": round(r.total_length_m, 1),
                "segment_count": r.segment_count,
                "route_count": r.route_count,
                "page_index": r.page_index,
            })

        scale_info = None
        if result.scale:
            scale_info = {
                "mm_per_pt": round(result.scale.mm_per_pt, 4),
                "source": result.scale.source,
                "confidence": result.scale.confidence,
            }

        return JSONResponse(content={
            "method": "geometry",
            "routes": routes_json,
            "scale": scale_info,
            "pages_scanned": result.pages_scanned,
            "elapsed_s": elapsed,
        })
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Geometry measurement error: {e}")


@app.get("/api/file/{file_id}/count/visual")
async def api_count_visual(
    file_id: str,
    page: Optional[int] = Query(None, description="Page to scan (0-based)"),
    threshold: float = Query(0.75, ge=0.3, le=1.0),
):
    """Run Method D: visual symbol template matching."""
    pdf_path = _id_to_path(file_id)
    if pdf_path is None:
        raise HTTPException(status_code=404, detail="File not found")

    try:
        t0 = time.time()
        legend = await asyncio.to_thread(_get_legend, pdf_path, file_id)
        result = await asyncio.to_thread(
            match_symbols, str(pdf_path), legend, page, threshold
        )
        elapsed = round(time.time() - t0, 2)

        matches_json = []
        for m in result.matches:
            matches_json.append({
                "symbol_index": m.symbol_index,
                "description": m.description,
                "x": m.x, "y": m.y,
                "confidence": m.confidence,
                "scale": m.scale, "rotation": m.rotation,
                "color": m.color, "page_index": m.page_index,
            })

        return JSONResponse(content={
            "method": "visual",
            "counts": result.counts,
            "descriptions": result.descriptions,
            "matches": matches_json,
            "symbols_extracted": result.symbols_extracted,
            "page_index": result.page_index,
            "threshold": result.threshold,
            "elapsed_s": elapsed,
        })
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Visual matching error: {e}")


@app.get("/api/file/{file_id}/luminaire_heights")
async def api_luminaire_heights(file_id: str):
    """Extract mount-height hints near detected luminaires."""
    pdf_path = _id_to_path(file_id)
    if pdf_path is None:
        raise HTTPException(status_code=404, detail="File not found")

    try:
        t0 = time.time()
        result = await asyncio.to_thread(extract_luminaire_heights, str(pdf_path))
        result["elapsed_s"] = round(time.time() - t0, 2)
        return JSONResponse(content=result)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Luminaire height extraction error: {e}")


@app.get("/api/file/{file_id}/count/all")
async def api_count_all(file_id: str):
    """Run ALL counting methods and return combined results."""
    pdf_path = _id_to_path(file_id)
    if pdf_path is None:
        raise HTTPException(status_code=404, detail="File not found")

    t0 = time.time()
    results = {}
    errors = {}

    # Parse legend once
    try:
        legend = await asyncio.to_thread(_get_legend, pdf_path, file_id)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Legend parse error: {e}")

    legend_page = legend.page_index if legend.items else 0

    # Method A: Text markers
    try:
        t1 = time.time()
        text_result = await asyncio.to_thread(
            count_symbols, str(pdf_path), legend
        )
        positions_by_sym: dict[str, list[dict]] = {}
        for p in text_result.positions:
            if p.symbol not in positions_by_sym:
                positions_by_sym[p.symbol] = []
            positions_by_sym[p.symbol].append({
                "x": round(p.x, 1), "y": round(p.y, 1),
            })
        results["text"] = {
            "counts": text_result.counts,
            "positions": positions_by_sym,
            "total": sum(text_result.counts.values()),
            "elapsed_s": round(time.time() - t1, 2),
        }
    except Exception as e:
        errors["text"] = str(e)

    # Method B: Cable annotations
    try:
        t1 = time.time()
        cable_result = await asyncio.to_thread(
            extract_cables, str(pdf_path), legend, [legend_page]
        )
        cable_runs = []
        for r in cable_result.runs:
            cable_runs.append({
                "panel": r.panel, "group_full": r.group_full,
                "cross_section": r.cross_section,
                "length_m": r.length_m, "cable_type": r.cable_type,
                "color": r.color,
                "position": {"x": r.position[0], "y": r.position[1]},
            })
        results["cables"] = {
            "total_runs": cable_result.total_runs,
            "runs": cable_runs,
            "panels": {k: len(v) for k, v in cable_result.panels.items()},
            "schedule": cable_result.cable_schedule,
            "elapsed_s": round(time.time() - t1, 2),
        }
    except Exception as e:
        errors["cables"] = str(e)

    # Method C: Geometry
    try:
        t1 = time.time()
        geo_result = await asyncio.to_thread(
            measure_cables, str(pdf_path), legend, [legend_page]
        )
        routes_json = []
        for r in geo_result.routes:
            routes_json.append({
                "color": r.color,
                "total_length_m": round(r.total_length_m, 1),
                "segment_count": r.segment_count,
            })
        scale_info = None
        if geo_result.scale:
            scale_info = {
                "mm_per_pt": round(geo_result.scale.mm_per_pt, 4),
                "source": geo_result.scale.source,
            }
        results["geometry"] = {
            "routes": routes_json,
            "scale": scale_info,
            "elapsed_s": round(time.time() - t1, 2),
        }
    except Exception as e:
        errors["geometry"] = str(e)

    # Method D: Visual matching
    try:
        t1 = time.time()
        vis_result = await asyncio.to_thread(
            match_symbols, str(pdf_path), legend
        )
        vis_matches = []
        for m in vis_result.matches:
            vis_matches.append({
                "symbol_index": m.symbol_index,
                "description": m.description,
                "x": m.x, "y": m.y,
                "confidence": m.confidence, "color": m.color,
            })
        results["visual"] = {
            "counts": vis_result.counts,
            "descriptions": vis_result.descriptions,
            "matches": vis_matches,
            "symbols_extracted": vis_result.symbols_extracted,
            "elapsed_s": round(time.time() - t1, 2),
        }
    except Exception as e:
        errors["visual"] = str(e)

    # Method E: Reverse label-to-legend channel (T060/B051)
    # Per-file path mirror of the channel added by T058 in
    # _count_equipment_in_pdf.  Picks up on-drawing colored engineer labels
    # (e.g. blue PU6/PU7 for "Post upravleniya") that none of the
    # text/visual stages covered, and attributes them to uncovered legend
    # rows.  Without this, UI "Zapusk" button never surfaces symbol-less
    # legend items even though _count_equipment_in_pdf does.
    try:
        t1 = time.time()
        # Build covered_legend_idx from visual counts only.  This is
        # intentionally a SIMPLER set than what _count_equipment_in_pdf
        # builds at L1225-1300, because per-file results carry visual
        # counts by symbol_index but not the visual/text reconciliation
        # logic.  Reverse channel only fires for legend rows nothing else
        # covered, so over-conservative covered_idx (only visual hits)
        # produces at most one extra reverse_label row per family which
        # is acceptable.
        covered_legend_idx_pf: set[int] = set()
        vis_counts_obj = results.get("visual", {}).get("counts") or {}
        if isinstance(vis_counts_obj, dict):
            for k, v in vis_counts_obj.items():
                try:
                    if int(v) > 0:
                        covered_legend_idx_pf.add(int(k))
                except (TypeError, ValueError):
                    continue
        # Run extraction + match
        colored_words = await asyncio.to_thread(
            _reverse_extract_colored_words, str(pdf_path), legend_page,
        )
        label_groups = await asyncio.to_thread(
            _reverse_match_labels_to_legend,
            colored_words, legend.items, covered_legend_idx_pf,
        )
        reverse_items = []
        for idx, labels in label_groups.items():
            try:
                desc = legend.items[idx].description or ""
            except (IndexError, AttributeError):
                desc = f"legend[{idx}]"
            reverse_items.append({
                "legend_index": idx,
                "name": desc,
                "count": len(labels),
                "source": "reverse_label_match",
                "labels": labels,
            })
        results["reverse"] = {
            "items": reverse_items,
            "blue_words_total": len(colored_words),
            "elapsed_s": round(time.time() - t1, 2),
        }
    except Exception as e:
        errors["reverse"] = str(e)

    total_elapsed = round(time.time() - t0, 2)

    return JSONResponse(content={
        "results": results,
        "errors": errors,
        "legend_page": legend_page,
        "legend_items": len(legend.items),
        "elapsed_s": total_elapsed,
    })


# ---------------------------------------------------------------------------
# 9b. GET /api/file/{file_id}/count/stream — SSE step-by-step counting
# ---------------------------------------------------------------------------

def _sse(event: str, data: dict) -> str:
    """Format a single SSE message."""
    return f"event: {event}\ndata: {json_mod.dumps(data, ensure_ascii=False)}\n\n"


@app.get("/api/file/{file_id}/count/stream")
async def api_count_stream(file_id: str):
    """SSE stream: step-by-step equipment counting with real-time progress."""
    pdf_path = _id_to_path(file_id)
    if pdf_path is None:
        raise HTTPException(status_code=404, detail="File not found")

    async def event_stream():
        t0 = time.time()
        results: dict = {}
        errors: dict = {}
        step_counter = 0  # running step number

        def next_step():
            nonlocal step_counter
            step_counter += 1
            return step_counter

        # ── Step 1: Parse legend ──────────────────────────────
        s = next_step()
        yield _sse("progress", {"step": s, "label": "Парсинг легенды", "status": "running"})
        try:
            legend = await asyncio.to_thread(_get_legend, pdf_path, file_id)
        except Exception as e:
            yield _sse("error", {"step": s, "label": "Парсинг легенды", "error": str(e)})
            yield _sse("done", {"total_elapsed_s": round(time.time() - t0, 2), "ok": False})
            return
        legend_page = legend.page_index if legend.items else 0
        n_items = len(legend.items)

        yield _sse("start", {
            "legend_items": n_items,
        })
        yield _sse("step_done", {
            "step": s,
            "label": "Парсинг легенды",
            "count": n_items,
            "elapsed_s": round(time.time() - t0, 2),
        })

        # ── Visual matching per symbol ────────────────────────
        # Use progress_callback + asyncio.Queue to stream per-symbol results
        queue: asyncio.Queue = asyncio.Queue()
        loop = asyncio.get_event_loop()

        def on_symbol_progress(idx, item, count):
            """Called from worker thread for each symbol."""
            asyncio.run_coroutine_threadsafe(
                queue.put(("symbol_done", idx, item.symbol or "?",
                           (item.description or "")[:50], count)),
                loop,
            )

        vis_error = None

        async def run_visual():
            nonlocal vis_error
            try:
                result = await asyncio.to_thread(
                    match_symbols, str(pdf_path), legend,
                    progress_callback=on_symbol_progress,
                )
                await queue.put(("visual_done", result))
            except Exception as e:
                vis_error = str(e)
                await queue.put(("visual_done", None))

        # Emit "running" for first symbol
        if n_items > 0:
            s = next_step()
            yield _sse("progress", {
                "step": s, "label": "Подготовка визуального поиска...",
                "status": "running",
            })

        vis_task = asyncio.create_task(run_visual())

        while not vis_task.done() or not queue.empty():
            try:
                msg = await asyncio.wait_for(queue.get(), timeout=0.3)
            except asyncio.TimeoutError:
                continue

            if msg[0] == "symbol_done":
                _, idx, sym, desc, count = msg
                label = f"Поиск: {sym} — {desc}" if sym != "?" else f"Поиск: {desc}"
                yield _sse("step_done", {
                    "step": s,
                    "type": "symbol",
                    "label": label,
                    "symbol": sym, "description": desc,
                    "symbol_index": idx, "count": count,
                })
                # Advance step for next symbol
                s = next_step()
                yield _sse("progress", {
                    "step": s, "label": "Визуальный поиск...",
                    "status": "running",
                })

            elif msg[0] == "visual_done":
                vis_result = msg[1]
                break

        # Wait for task to fully complete
        await vis_task

        # Process visual results
        if vis_result is not None:
            vis_matches = []
            for m in vis_result.matches:
                vis_matches.append({
                    "symbol_index": m.symbol_index,
                    "description": m.description,
                    "x": m.x, "y": m.y,
                    "confidence": m.confidence, "color": m.color,
                })
            results["visual"] = {
                "counts": vis_result.counts,
                "descriptions": vis_result.descriptions,
                "matches": vis_matches,
                "symbols_extracted": vis_result.symbols_extracted,
            }
        elif vis_error:
            errors["visual"] = vis_error

        # ── Text markers ──────────────────────────────────────
        s = next_step()
        yield _sse("progress", {
            "step": s, "label": "Текстовый поиск маркеров", "status": "running",
        })
        try:
            t1 = time.time()
            text_result = await asyncio.to_thread(
                count_symbols, str(pdf_path), legend,
            )
            positions_by_sym: dict[str, list[dict]] = {}
            for p in text_result.positions:
                if p.symbol not in positions_by_sym:
                    positions_by_sym[p.symbol] = []
                positions_by_sym[p.symbol].append({
                    "x": round(p.x, 1), "y": round(p.y, 1),
                })
            results["text"] = {
                "counts": text_result.counts,
                "positions": positions_by_sym,
                "total": sum(text_result.counts.values()),
                "elapsed_s": round(time.time() - t1, 2),
            }
            yield _sse("step_done", {
                "step": s,
                "type": "text", "label": "Текстовый поиск маркеров",
                "count": sum(text_result.counts.values()),
                "elapsed_s": round(time.time() - t1, 2),
            })
        except Exception as e:
            errors["text"] = str(e)
            yield _sse("step_done", {
                "step": s,
                "type": "text", "label": "Текстовый поиск маркеров",
                "error": str(e),
            })

        # ── Cables ────────────────────────────────────────────
        s = next_step()
        yield _sse("progress", {
            "step": s, "label": "Анализ кабелей", "status": "running",
        })
        try:
            t1 = time.time()
            cable_result = await asyncio.to_thread(
                extract_cables, str(pdf_path), legend, [legend_page],
            )
            cable_runs = []
            for r in cable_result.runs:
                cable_runs.append({
                    "panel": r.panel, "group_full": r.group_full,
                    "cross_section": r.cross_section,
                    "length_m": r.length_m, "cable_type": r.cable_type,
                    "color": r.color,
                    "position": {"x": r.position[0], "y": r.position[1]},
                })
            results["cables"] = {
                "total_runs": cable_result.total_runs,
                "runs": cable_runs,
                "panels": {k: len(v) for k, v in cable_result.panels.items()},
                "schedule": cable_result.cable_schedule,
                "elapsed_s": round(time.time() - t1, 2),
            }
            yield _sse("step_done", {
                "step": s,
                "type": "cables", "label": "Анализ кабелей",
                "count": cable_result.total_runs,
                "elapsed_s": round(time.time() - t1, 2),
            })
        except Exception as e:
            errors["cables"] = str(e)
            yield _sse("step_done", {
                "step": s,
                "type": "cables", "label": "Анализ кабелей",
                "error": str(e),
            })

        # ── Geometry ──────────────────────────────────────────
        s = next_step()
        yield _sse("progress", {
            "step": s, "label": "Измерение маршрутов", "status": "running",
        })
        try:
            t1 = time.time()
            geo_result = await asyncio.to_thread(
                measure_cables, str(pdf_path), legend, [legend_page],
            )
            routes_json = []
            for r in geo_result.routes:
                routes_json.append({
                    "color": r.color,
                    "total_length_m": round(r.total_length_m, 1),
                    "segment_count": r.segment_count,
                })
            scale_info = None
            if geo_result.scale:
                scale_info = {
                    "mm_per_pt": round(geo_result.scale.mm_per_pt, 4),
                    "source": geo_result.scale.source,
                }
            results["geometry"] = {
                "routes": routes_json,
                "scale": scale_info,
                "elapsed_s": round(time.time() - t1, 2),
            }
            yield _sse("step_done", {
                "step": s,
                "type": "geometry", "label": "Измерение маршрутов",
                "elapsed_s": round(time.time() - t1, 2),
            })
        except Exception as e:
            errors["geometry"] = str(e)
            yield _sse("step_done", {
                "step": s,
                "type": "geometry", "label": "Измерение маршрутов",
                "error": str(e),
            })

        # ── Done ─────────────────────────────────────────────
        total_elapsed = round(time.time() - t0, 2)
        yield _sse("done", {
            "total_elapsed_s": total_elapsed,
            "total_steps": step_counter,
            "ok": True,
            "results": results,
            "errors": errors,
            "legend_page": legend_page,
            "legend_items": n_items,
        })

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("web_app:app", host="0.0.0.0", port=8051, reload=False)
